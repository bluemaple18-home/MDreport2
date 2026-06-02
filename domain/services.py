from __future__ import annotations

import csv
import json
import os
import hashlib
import re
import calendar
import base64
import uuid
from datetime import date, datetime, timedelta
from dataclasses import replace
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from infra.dsp_api import DspApiClient, normalize_dsp_report_rows, resolve_dsp_api_settings
from infra.dsp_rules import classify_dsp_row
from infra.sqlite.repository import SQLiteRepository
from infra.ssp_api import (
    SSP_MONTHLY_COUNTRY_DIMENSIONS,
    SSP_MONTHLY_ZONE_SIZE_DIMENSIONS,
    SspApiClient,
    normalize_ssp_ad_group_report_rows,
    normalize_ssp_monthly_country_rows,
    normalize_ssp_monthly_zone_campaign_size_rows,
    normalize_ssp_report_rows,
    resolve_ssp_api_settings,
)

DSP_TEMPLATE_SHEET_NAMES = [
    "2025年_MF_合作績效統計總表",
    "2025_外部+行政_合作績效統計總表 ",
    "mF投資量_總表",
    "各經銷商明細",
    "北流進單追蹤",
]

MONTH_AMOUNT_COL_START = 5  # E
MONTH_COUNT = 12
DETAIL_YEAR_ROWS = (5, 24, 44, 63, 82)
DETAIL_INPUT_ROWS = (
    7, 8, 9, 10, 11, 12, 13,
    26, 27, 28, 29, 30, 31, 32,
    46, 47, 48, 49, 50, 51, 52,
    65, 66, 67, 68, 69, 70, 71,
    84, 85, 86, 87, 88, 89, 90,
)
DATE_PREFIX_RE = re.compile(r"^(\d{4})[-/](\d{1,2})")
TEMPLATE_YEAR_PREFIX_RE = re.compile(r"^(\d{4})")
TEMPLATE_MMDD_RANGE_RE = re.compile(r"_(\d{4})-(\d{4})$")
WORKBOOK_ILLEGAL_CHAR_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")

TAB4_MONTH_LABELS = [f"{idx + 1}月" for idx in range(MONTH_COUNT)]
MONTHLY_P4_MANUAL_INPUTS = [
    {"key": "external_io_momo", "label": "IO momo"},
    {"key": "external_io_live", "label": "IO 直播"},
    {"key": "hb_revenue", "label": "串接收入 (HB)"},
    {"key": "external_beiliu_io", "label": "外部經銷商 北流委刊IO"},
    {"key": "remaining_traffic_revenue", "label": "剩餘流量變現(無成本)"},
    {"key": "data_monetization_adjustment", "label": "數據變現補值"},
]
MONTHLY_P4_TEST_TEMPLATE_KINDS = {
    "base": "基礎模板",
    "check": "檢核模板",
}
MONTHLY_P4_SNAPSHOT_ROW_SPECS = [
    ("product_total", ["產品處廣告總營收", "廣告總營收"]),
    ("mf_marketing", ["內經銷商營銷處", "營銷處"]),
    ("mf_strategy", ["內經銷商策略部", "策略部"]),
    ("external_total", ["外經銷商自操", "外經銷商自操io", "外經銷商"]),
    ("hb_revenue", ["串接收入hb", "串接收入"]),
    ("external_beiliu_io", ["北流委刊io", "北流"]),
    ("data_fee", ["數據費", "數據變現"]),
    ("remaining_traffic_revenue", ["剩餘流量變現"]),
    ("mf_total", ["mltiforce總目標", "mtliforce總目標", "multiforce總目標"]),
]
MONTHLY_P4_FIXED_ROW_MAP = {
    4: ("product_total", "target"),
    5: ("product_total", "actual"),
    6: ("product_total", "rate"),
    11: ("mf_marketing", "target"),
    12: ("mf_marketing", "actual"),
    13: ("mf_marketing", "rate"),
    14: ("mf_strategy", "target"),
    15: ("mf_strategy", "actual"),
    16: ("mf_strategy", "rate"),
    17: ("external_total", "target"),
    18: ("external_total", "actual"),
    19: ("external_total", "rate"),
    20: ("hb_revenue", "target"),
    21: ("hb_revenue", "actual"),
    22: ("hb_revenue", "rate"),
    23: ("external_beiliu_io", "target"),
    24: ("external_beiliu_io", "actual"),
    25: ("external_beiliu_io", "rate"),
    26: ("mf_total", "target"),
    27: ("mf_total", "actual"),
    28: ("mf_total", "rate"),
    30: ("data_fee", "target"),
    31: ("data_fee", "actual"),
    32: ("data_fee", "rate"),
    33: ("remaining_traffic_revenue", "target"),
    34: ("remaining_traffic_revenue", "actual"),
    35: ("remaining_traffic_revenue", "rate"),
    37: ("other_total", "target"),
    38: ("other_total", "actual"),
    39: ("other_total", "rate"),
}
MONTHLY_P4_MONTH_LABELS = {
    "jan": "2026-01", "january": "2026-01", "1月": "2026-01", "01月": "2026-01",
    "feb": "2026-02", "february": "2026-02", "2月": "2026-02", "02月": "2026-02",
    "mar": "2026-03", "march": "2026-03", "3月": "2026-03", "03月": "2026-03",
    "apr": "2026-04", "april": "2026-04", "4月": "2026-04", "04月": "2026-04",
    "may": "2026-05", "5月": "2026-05", "05月": "2026-05",
    "jun": "2026-06", "june": "2026-06", "6月": "2026-06", "06月": "2026-06",
    "jul": "2026-07", "july": "2026-07", "7月": "2026-07", "07月": "2026-07",
    "aug": "2026-08", "august": "2026-08", "8月": "2026-08", "08月": "2026-08",
    "sep": "2026-09", "sept": "2026-09", "september": "2026-09", "9月": "2026-09", "09月": "2026-09",
    "oct": "2026-10", "october": "2026-10", "10月": "2026-10",
    "nov": "2026-11", "november": "2026-11", "11月": "2026-11",
    "dec": "2026-12", "december": "2026-12", "12月": "2026-12",
    "total": "total", "ttl": "total",
}
MONTHLY_DSP_ARCHIVE_ORDER_MARKER_PREFIX = "MONTHLY_ARCHIVE_DSP"
SSP_AD_GROUP_CATALOG = [
    {"id": 330, "name": "文中創意 高價版位", "format": "文中創意", "tier": "高"},
    {"id": 329, "name": "文中創意 中價版位", "format": "文中創意", "tier": "中"},
    {"id": 328, "name": "文中創意 低價版位", "format": "文中創意", "tier": "低"},
    {"id": 327, "name": "純蓋板 高價版位", "format": "純蓋板", "tier": "高"},
    {"id": 326, "name": "純蓋板 中價版位", "format": "純蓋板", "tier": "中"},
    {"id": 325, "name": "純蓋板 低價版位", "format": "純蓋板", "tier": "低"},
    {"id": 324, "name": "置底展開 高價版位", "format": "置底展開", "tier": "高"},
    {"id": 323, "name": "置底展開 中價版位", "format": "置底展開", "tier": "中"},
    {"id": 322, "name": "置底展開 低價版位", "format": "置底展開", "tier": "低"},
    {"id": 321, "name": "置底非展開 高價版位", "format": "置底非展開", "tier": "高"},
    {"id": 320, "name": "置底非展開 中價版位", "format": "置底非展開", "tier": "中"},
    {"id": 319, "name": "置底非展開 低價版位", "format": "置底非展開", "tier": "低"},
    {"id": 318, "name": "展示型 高價版位", "format": "展示型", "tier": "高"},
    {"id": 317, "name": "展示型 中價版位", "format": "展示型", "tier": "中"},
    {"id": 316, "name": "展示型 低價版位", "format": "展示型", "tier": "低"},
    {"id": 335, "name": "知名媒體 高價版位 BN", "format": "知名媒體 BN", "tier": "高"},
    {"id": 334, "name": "知名媒體 中價版位 BN", "format": "知名媒體 BN", "tier": "中"},
    {"id": 333, "name": "知名媒體 低價版位 BN", "format": "知名媒體 BN", "tier": "低"},
]
SSP_AD_GROUP_CATALOG_BY_ID = {int(item["id"]): item for item in SSP_AD_GROUP_CATALOG}
SSP_AD_GROUP_METRICS = ["request", "impress", "click", "ctr", "ecpm", "ecpc", "advertiser_mu"]
MONTHLY_PRESENTATION_VIDEO_FORMATS = {"影音摩天", "preroll"}
MONTHLY_CREATIVE_TRAFFIC_EXCLUDED_FORMATS = {"一般廣告", "影音摩天", "preroll", "DOOH北流"}
MONTHLY_CREATIVE_TRAFFIC_EXCLUDED_TOKENS = (
    "橫幅",
    "banner",
    "影音廣告",
    "原生廣告",
    "vast",
    "pre-roll",
    "pre roll",
    "preroll",
    "影音摩天",
)
MONTHLY_CREATIVE_TRAFFIC_INCLUDED_SIZE_IDS = {
    "36",
    "55",
    "56",
    "59",
    "67",
    "73",
    "77",
    "81",
    "86",
    "91",
    "93",
    "94",
    "95",
    "96",
    "101",
    "102",
    "103",
    "104",
    "106",
    "109",
    "116",
    "120",
    "121",
    "146",
    "154",
    "155",
    "160",
    "168",
    "184",
    "201",
    "203",
    "210",
    "213",
    "227",
    "228",
    "237",
}
TAB4_DETAIL_SECTION_SPECS = [
    {
        "id": "marketing",
        "year_row": 5,
        "total_row": 6,
        "total_label_a": "營銷處 DSP投資額 總計",
        "total_label_d": "",
        "detail_label_a": "營銷事業處\n分項績效",
        "detail_labels": [
            {"b": "三螢", "c": "一般廣告", "d": ""},
            {"b": "三螢", "c": "創意", "d": "蓋板/置底(展開&不展)/文中"},
            {"b": "三螢", "c": "影音", "d": "影音摩天(outstream)"},
            {"b": "三螢", "c": "影音", "d": "preroll (instream)"},
            {"b": "DOOH外部", "c": "影音", "d": "前線媒體/presco"},
            {"b": "DOOH北流", "c": "影音", "d": "北流"},
            {"b": "CTV", "c": "影音", "d": ""},
        ],
    },
    {
        "id": "strategy",
        "year_row": 24,
        "total_row": 25,
        "total_label_a": "策略部 DSP投資額 總計",
        "total_label_d": "",
        "detail_label_a": "策略部\n分項績效",
        "detail_labels": [
            {"b": "三螢", "c": "一般廣告", "d": ""},
            {"b": "三螢", "c": "創意", "d": "蓋板/置底(展開&不展)/文中"},
            {"b": "三螢", "c": "影音", "d": "影音摩天(outstream)"},
            {"b": "三螢", "c": "影音", "d": "preroll (instream)"},
            {"b": "DOOH外部", "c": "影音", "d": "前線媒體/presco"},
            {"b": "DOOH北流", "c": "影音", "d": "北流"},
            {"b": "CTV", "c": "影音", "d": ""},
        ],
    },
    {
        "id": "external_self",
        "year_row": 44,
        "total_row": 45,
        "total_label_a": "外部經銷(自操) DSP投資額 總計",
        "total_label_d": "玩藝/春樹/ADGeek等系統自操",
        "detail_label_a": "外部經銷(自操)\n分項績效",
        "detail_labels": [
            {"b": "三螢", "c": "一般廣告", "d": ""},
            {"b": "三螢", "c": "創意", "d": "蓋板/置底(展開&不展)/文中"},
            {"b": "三螢", "c": "影音", "d": "影音摩天(outstream)"},
            {"b": "三螢", "c": "影音", "d": "preroll (instream)"},
            {"b": "DOOH外部", "c": "影音", "d": "前線媒體/presco"},
            {"b": "DOOH北流", "c": "影音", "d": "北流"},
            {"b": "CTV", "c": "影音", "d": ""},
        ],
    },
    {
        "id": "external_io",
        "year_row": 63,
        "total_row": 64,
        "total_label_a": "外部IO委刊 DSP投資額 總計",
        "total_label_d": "MOMO、DOOH委刊",
        "detail_label_a": "外部IO委刊 \n分項績效",
        "detail_labels": [
            {"b": "三螢", "c": "一般廣告", "d": ""},
            {"b": "三螢", "c": "創意", "d": "蓋板/置底(展開&不展)/文中"},
            {"b": "三螢", "c": "影音", "d": "影音摩天(outstream)"},
            {"b": "三螢", "c": "影音", "d": "preroll (instream)"},
            {"b": "DOOH外部", "c": "影音", "d": "前線媒體/presco"},
            {"b": "DOOH北流", "c": "影音", "d": "北流"},
            {"b": "CTV", "c": "影音", "d": ""},
        ],
    },
    {
        "id": "hb_bridge",
        "year_row": 82,
        "total_row": 83,
        "total_label_a": "HB串接 DSP投資額 總計",
        "total_label_d": "Appier/宇匯Bridgewell /Criteo/ RTBhouse /Teads/ucfunnel少許",
        "detail_label_a": "HB 串接\n分項績效",
        "detail_labels": [
            {"b": "三螢", "c": "一般廣告", "d": ""},
            {"b": "三螢", "c": "創意", "d": "蓋板/置底(展開&不展)/文中"},
            {"b": "三螢", "c": "影音", "d": "影音摩天(outstream)"},
            {"b": "三螢", "c": "影音", "d": "preroll (instream)"},
            {"b": "DOOH外部", "c": "影音", "d": "前線媒體/presco"},
            {"b": "DOOH北流", "c": "影音", "d": "北流"},
            {"b": "CTV", "c": "影音", "d": ""},
        ],
    },
]


def _pick_category(row: dict, keys: list[str]) -> str:
    for key in keys:
        text = str(row.get(key) or "").strip()
        if text:
            return text
    return ""


def _is_internal_distributor_level(value: str) -> bool:
    return value in {"內部經銷商", "內經銷商"}


def _is_external_distributor_level(value: str) -> bool:
    return value in {"外部經銷商", "外經銷商"}


def _monthly_report_row_ad_format(row: dict[str, object]) -> str:
    creative_size_id = str(row.get("creative_size_id") or "")
    size_id = _monthly_creative_size_id(creative_size_id)
    if size_id == "231":
        return "DOOH北流"
    if size_id == "179":
        return "DOOH北流"
    if _is_monthly_special_video_ad(creative_size_id, size_id):
        classification = classify_dsp_row(
            {
                "訂單": "",
                "素材": "",
                "廣告形式": creative_size_id,
                "尺寸": creative_size_id,
                "素材樣板": creative_size_id,
                "cpm": row.get("dsp_ecpm"),
            }
        )
        return _monthly_presentation_ad_format(classification["最終廣告形式"])
    ad_format_rule = str(row.get("ad_format_rule") or "")
    ad_format = str(row.get("ad_format") or "").strip()
    trusted_table_hit = ad_format_rule.startswith(("table:material:", "table:size_id:"))
    if not size_id and trusted_table_hit and ad_format and ad_format != "DOOH北流":
        return _monthly_presentation_ad_format(ad_format)
    classification = classify_dsp_row(
        {
            "訂單": "",
            "素材": "",
            "廣告形式": creative_size_id,
            "尺寸": creative_size_id,
            "素材樣板": creative_size_id,
            "cpm": row.get("dsp_ecpm"),
        }
    )
    return _monthly_presentation_ad_format(classification["最終廣告形式"])


def _monthly_presentation_ad_format(value: object) -> str:
    ad_format = str(value or "").strip()
    if ad_format in MONTHLY_PRESENTATION_VIDEO_FORMATS:
        return "影音摩天"
    return ad_format


def _is_monthly_special_video_ad(creative_size_id: object, size_id: str) -> bool:
    if size_id == "176":
        return True
    token = _monthly_creative_size_token(creative_size_id)
    return token in {"影音廣告", "16:9影音廣告", "169影音廣告"}


def _monthly_creative_size_token(value: object) -> str:
    text = str(value or "").strip().lower().replace("\u200b", "")
    return re.sub(r"[\s_（）()/-]+", "", text)


def _monthly_creative_size_id(value: object) -> str:
    match = re.search(r"\((\d+)\)", str(value or ""))
    return match.group(1) if match else ""


def _is_taiwan_country(value: object) -> bool:
    token = str(value or "").strip().lower().replace(" ", "").replace("_", "").replace("-", "")
    return token in {"tw", "twn", "taiwan", "台灣", "臺灣", "taiwan,provinceofchina"}


def _is_monthly_creative_traffic_row(row: dict[str, object]) -> bool:
    creative_size_id = str(row.get("creative_size_id") or "").strip()
    if not creative_size_id:
        return _monthly_report_row_ad_format(row) not in MONTHLY_CREATIVE_TRAFFIC_EXCLUDED_FORMATS
    size_id = _monthly_creative_size_id(creative_size_id)
    if size_id:
        return size_id in MONTHLY_CREATIVE_TRAFFIC_INCLUDED_SIZE_IDS
    token = _monthly_creative_size_token(creative_size_id)
    if not token:
        return False
    if any(_monthly_creative_size_token(excluded) in token for excluded in MONTHLY_CREATIVE_TRAFFIC_EXCLUDED_TOKENS):
        return False
    return _monthly_report_row_ad_format(row) not in MONTHLY_CREATIVE_TRAFFIC_EXCLUDED_FORMATS


def _monthly_request_summary_row(row: dict[str, object]) -> dict[str, object]:
    request = float(row.get("request") or 0.0)
    summary = dict(row)
    summary["source"] = f"{row.get('source') or ''}:pb0_request"
    summary["campaign_id"] = ""
    summary["campaign_name"] = ""
    summary["request"] = request
    summary["request_including_padding"] = request
    summary["request_excluding_padding"] = 0.0
    for key in (
        "impress",
        "active_view",
        "active_view_rate",
        "click",
        "ctr",
        "ecpm",
        "ecpc",
        "invalid_impress",
        "invalid_click",
        "profit",
        "site_mu",
        "advertiser_mu",
        "dsp_ecpm",
        "dsp_ecpc",
    ):
        summary[key] = 0.0
    return summary


def _monthly_delivery_detail_row(row: dict[str, object]) -> dict[str, object]:
    request = float(row.get("request") or 0.0)
    detail = dict(row)
    detail["request"] = 0.0
    detail["request_including_padding"] = 0.0
    detail["request_excluding_padding"] = request
    return detail


def _resolve_year_month(row: dict) -> tuple[int, int] | None:
    raw = str(row.get("日期時間") or "").strip()
    matched = DATE_PREFIX_RE.match(raw)
    if not matched:
        return None
    year = int(matched.group(1))
    month = int(matched.group(2))
    if month < 1 or month > 12:
        return None
    return year, month - 1


def _canonical_day_text(row: dict) -> str:
    raw = str(row.get("日期時間") or "").strip()
    return raw[:10] if len(raw) >= 10 else raw


def _inclusive_day_texts(start_day: str, end_day: str) -> set[str]:
    start = date.fromisoformat(start_day)
    end = date.fromisoformat(end_day)
    if start > end:
        raise ValueError("start_day cannot be after end_day")
    days: set[str] = set()
    current = start
    while current <= end:
        days.add(current.isoformat())
        current += timedelta(days=1)
    return days


def _month_text_from_date_text(value: str) -> str:
    raw = str(value or "").strip().replace("/", "-")
    return raw[:7] if re.match(r"^\d{4}-\d{2}", raw) else ""


def _shift_month(month_text: str, offset: int) -> str:
    year, month = [int(part) for part in month_text.split("-")]
    month_index = year * 12 + (month - 1) + offset
    next_year = month_index // 12
    next_month = month_index % 12 + 1
    return f"{next_year:04d}-{next_month:02d}"


def _year_month_from_month_text(month_text: str) -> tuple[int, int]:
    year, month = [int(part) for part in month_text.split("-")]
    return year, month - 1


def _month_window_from_period(*, week_start: str | None, week_end: str | None) -> tuple[str, list[str]]:
    anchor = _month_text_from_date_text(week_end or "") or _month_text_from_date_text(week_start or "")
    if not anchor:
        today = date.today()
        month_index = today.year * 12 + today.month - 1
        anchor = f"{month_index // 12:04d}-{month_index % 12 + 1:02d}"
    return anchor, [_shift_month(anchor, -2), _shift_month(anchor, -1), anchor]


def _month_year_window(anchor_month: str) -> list[str]:
    year, _month_idx = _year_month_from_month_text(anchor_month)
    return [f"{year:04d}-{month_idx:02d}" for month_idx in range(1, 13)]


def _month_date_range(month_text: str) -> tuple[str, str]:
    year, month = [int(part) for part in month_text.split("-")]
    last_day = calendar.monthrange(year, month)[1]
    return f"{year:04d}-{month:02d}-01", f"{year:04d}-{month:02d}-{last_day:02d}"


def _to_number(value: object) -> float:
    if isinstance(value, (int, float)):
        out = float(value)
        return out if out == out and out not in (float("inf"), float("-inf")) else 0.0
    raw = str(value or "").strip()
    if not raw:
        return 0.0
    negative_by_paren = raw.startswith("(") and raw.endswith(")")
    normalized = raw.replace(",", "").replace("$", "").replace("%", "").replace(" ", "").replace("(", "").replace(")", "")
    try:
        out = float(normalized)
    except Exception:
        return 0.0
    if negative_by_paren:
        return -abs(out)
    return out


def _sanitize_workbook_cell_value(value: object) -> object:
    if isinstance(value, str):
        return WORKBOOK_ILLEGAL_CHAR_RE.sub("", value)
    return value


def _is_formula(value: object) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _same_cell_value(left: object, right: object, *, tol: float = 1e-6) -> bool:
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        lf = float(left)
        rf = float(right)
        scale = max(1.0, abs(lf), abs(rf))
        return abs(lf - rf) <= tol * scale
    return left == right


def _normalize_p4_token(value: object) -> str:
    return re.sub(r"[^0-9a-zA-Z\u4e00-\u9fff]+", "", str(value or "").strip()).lower()


def _coerce_p4_numeric(value: object, *, metric: str) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        text = str(value).strip()
        if not text or text in {"-", "－"}:
            return None
        is_percent = "%" in text
        cleaned = re.sub(r"[^0-9.\-]", "", text)
        if cleaned in {"", "-", "."}:
            return None
        try:
            number = float(cleaned)
        except ValueError:
            return None
        if is_percent:
            return number
    if metric == "rate" and abs(number) <= 10:
        return number * 100
    return number


def _display_p4_value(value: float, *, metric: str) -> float:
    return float(Decimal(str(value + 1e-6)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


class CanonicalService:
    def __init__(
        self,
        repo: SQLiteRepository,
        *,
        feature_flags: dict[str, bool] | None = None,
        monthly_test_repo: SQLiteRepository | None = None,
    ) -> None:
        self.repo = repo
        self.monthly_test_repo = monthly_test_repo or repo
        self._field_contract = repo.field_contract
        self._feature_flags = feature_flags or {}

    def _trace_marker(self, *, workflow: str, run_type: str, run_id: str) -> str:
        if not self._feature_flags.get("enable_trace_markers", False):
            return ""
        return f"{workflow}:{run_type}:{run_id}"

    def _extra_debug_payload(self) -> dict[str, object]:
        if not self._feature_flags.get("enable_test_hooks", False):
            return {}
        return {"test_hooks_enabled": True}

    def _target_map_for_year(self, conn, target_year: int) -> dict[str, dict[int, float]]:
        out: dict[str, dict[int, float]] = {}
        for row in self.repo.read_monthly_p4_targets_in_tx(conn, target_year):
            out.setdefault(str(row["item_key"]), {})[int(row["month_index"])] = float(row["target_value"] or 0.0)
        return out

    def _target_value(self, targets: dict[str, dict[int, float]], item_key: str, month_text: str) -> float:
        month_index = int(month_text.split("-")[1])
        return float(targets.get(item_key, {}).get(month_index, 0.0))

    def _target_value_for_month(
        self,
        targets_by_year: dict[int, dict[str, dict[int, float]]],
        item_key: str,
        month_text: str,
    ) -> float:
        target_year, _month_idx = _year_month_from_month_text(month_text)
        return self._target_value(targets_by_year.get(target_year, {}), item_key, month_text)

    def _monthly_p4_entry_key(self, item_key: str, metric: str, month: str) -> str:
        return f"{item_key}.{metric}.{month}"

    def _monthly_p4_snapshot_from_payloads(self, month_payloads: list[dict[str, object]]) -> dict[str, object]:
        entries: dict[str, dict[str, object]] = {}
        item_keys = [
            "product_total",
            "mf_marketing",
            "mf_strategy",
            "external_total",
            "hb_revenue",
            "external_beiliu_io",
            "data_fee",
            "remaining_traffic_revenue",
            "mf_total",
        ]
        for raw_month in month_payloads:
            month = str(raw_month.get("month") or "")
            if not month:
                continue
            targets = raw_month.get("targets") if isinstance(raw_month.get("targets"), dict) else {}
            actuals = raw_month.get("actuals") if isinstance(raw_month.get("actuals"), dict) else {}
            for item_key in item_keys:
                target = float(targets.get(item_key, 0.0) or 0.0)
                actual = float(actuals.get(item_key, 0.0) or 0.0)
                for metric, value in (
                    ("target", target),
                    ("actual", actual),
                    ("rate", (actual / target * 100.0) if target else 0.0),
                ):
                    entries[self._monthly_p4_entry_key(item_key, metric, month)] = {
                        "itemKey": item_key,
                        "metric": metric,
                        "month": month,
                        "value": _display_p4_value(value, metric=metric),
                        "source": "runtime",
                    }
        months = [str(item.get("month") or "") for item in month_payloads if str(item.get("month") or "")]
        for item_key in item_keys:
            target_total = 0.0
            actual_total = 0.0
            for month in months:
                target_total += float(entries.get(self._monthly_p4_entry_key(item_key, "target", month), {}).get("value", 0.0) or 0.0)
                actual_total += float(entries.get(self._monthly_p4_entry_key(item_key, "actual", month), {}).get("value", 0.0) or 0.0)
            for metric, value in (
                ("target", target_total),
                ("actual", actual_total),
                ("rate", (actual_total / target_total * 100.0) if target_total else 0.0),
            ):
                entries[self._monthly_p4_entry_key(item_key, metric, "total")] = {
                    "itemKey": item_key,
                    "metric": metric,
                    "month": "total",
                    "value": _display_p4_value(value, metric=metric),
                    "source": "runtime",
                }
        return {"entries": entries, "entryCount": len(entries), "source": "runtime"}

    def _apply_monthly_p4_base_snapshot(
        self,
        month_payloads: list[dict[str, object]],
        base_snapshot: dict[str, object] | None,
        *,
        anchor_month: str,
    ) -> list[dict[str, object]]:
        if not base_snapshot:
            return month_payloads
        base_entries = base_snapshot.get("entries") if isinstance(base_snapshot.get("entries"), dict) else {}
        if not base_entries:
            return month_payloads
        out: list[dict[str, object]] = []
        for month_payload in month_payloads:
            month = str(month_payload.get("month") or "")
            if not month or month >= anchor_month:
                out.append(month_payload)
                continue
            next_payload = dict(month_payload)
            targets = dict(next_payload.get("targets") if isinstance(next_payload.get("targets"), dict) else {})
            actuals = dict(next_payload.get("actuals") if isinstance(next_payload.get("actuals"), dict) else {})
            for key, entry in base_entries.items():
                if not isinstance(entry, dict) or entry.get("month") != month:
                    continue
                item_key = str(entry.get("itemKey") or "")
                metric = str(entry.get("metric") or "")
                value = float(entry.get("value") or 0.0)
                if metric == "target":
                    targets[item_key] = value
                elif metric == "actual":
                    actuals[item_key] = value
            mf_keys = ("mf_marketing", "mf_strategy", "external_total", "hb_revenue", "external_beiliu_io")
            other_keys = ("data_fee", "remaining_traffic_revenue")
            targets["mf_total"] = sum(float(targets.get(key, 0.0) or 0.0) for key in mf_keys)
            actuals["mf_total"] = sum(float(actuals.get(key, 0.0) or 0.0) for key in mf_keys)
            targets["other_total"] = sum(float(targets.get(key, 0.0) or 0.0) for key in other_keys)
            actuals["other_total"] = sum(float(actuals.get(key, 0.0) or 0.0) for key in other_keys)
            targets["product_total"] = float(targets["mf_total"]) + float(targets["other_total"])
            actuals["product_total"] = float(actuals["mf_total"]) + float(actuals["other_total"])
            next_payload["targets"] = targets
            next_payload["actuals"] = actuals
            out.append(next_payload)
        return out

    def _parse_monthly_p4_workbook_snapshot(self, file_bytes: bytes, *, filename: str) -> dict[str, object]:
        workbook = load_workbook(BytesIO(file_bytes), read_only=False, data_only=True)
        try:
            target_sheet_name = "績效追蹤 p4 5 (j)"
            best_sheet = workbook[target_sheet_name] if target_sheet_name in workbook.sheetnames else None
            if best_sheet is None:
                return {
                    "source": "excel",
                    "filename": filename,
                    "sheet": "",
                    "entries": {},
                    "entryCount": 0,
                    "warnings": [f"找不到指定頁籤：{target_sheet_name}"],
                }
            best_months: dict[int, str] = {}
            for row in best_sheet.iter_rows():
                for cell in row:
                    token = _normalize_p4_token(cell.value)
                    if token in MONTHLY_P4_MONTH_LABELS:
                        best_months[int(cell.column)] = MONTHLY_P4_MONTH_LABELS[token]
            if not [month for month in best_months.values() if month != "total"]:
                return {
                    "source": "excel",
                    "filename": filename,
                    "sheet": best_sheet.title,
                    "entries": {},
                    "entryCount": 0,
                    "warnings": [f"{target_sheet_name} 找不到月份欄位"],
                }

            entries: dict[str, dict[str, object]] = {}
            for row_number, (item_key, metric) in MONTHLY_P4_FIXED_ROW_MAP.items():
                for column_number, month in best_months.items():
                    cell = best_sheet.cell(row=row_number, column=column_number)
                    number = _coerce_p4_numeric(cell.value, metric=metric)
                    if number is None:
                        continue
                    key = self._monthly_p4_entry_key(item_key, metric, month)
                    entries[key] = {
                        "itemKey": item_key,
                        "metric": metric,
                        "month": month,
                        "value": _display_p4_value(number, metric=metric),
                        "cell": f"{best_sheet.title}!{cell.coordinate}",
                        "source": "excel",
                    }
            return {
                "source": "excel",
                "filename": filename,
                "sheet": best_sheet.title,
                "months": list(dict.fromkeys(best_months.values())),
                "entries": entries,
                "entryCount": len(entries),
                "warnings": [],
            }
        finally:
            workbook.close()

    def _monthly_p4_diff(self, candidate_snapshot: dict[str, object], answer_snapshot: dict[str, object] | None) -> dict[str, object]:
        if not answer_snapshot:
            return {"status": "missing_answer", "diffs": [], "diffCount": 0}
        candidate_entries = candidate_snapshot.get("entries") if isinstance(candidate_snapshot.get("entries"), dict) else {}
        answer_entries = answer_snapshot.get("entries") if isinstance(answer_snapshot.get("entries"), dict) else {}
        diffs: list[dict[str, object]] = []
        for key in sorted(set(candidate_entries.keys()) | set(answer_entries.keys())):
            candidate = candidate_entries.get(key)
            answer = answer_entries.get(key)
            if not isinstance(candidate, dict) and not isinstance(answer, dict):
                continue
            source = answer if isinstance(answer, dict) else candidate
            if not isinstance(source, dict):
                continue
            candidate_value = candidate.get("value") if isinstance(candidate, dict) else None
            answer_value = answer.get("value") if isinstance(answer, dict) else None
            if candidate_value == answer_value:
                continue
            if not isinstance(candidate, dict):
                reason = "missing_in_candidate"
            elif not isinstance(answer, dict):
                reason = "missing_in_check_template"
            else:
                reason = "value_mismatch"
            diffs.append({
                "key": key,
                "reason": reason,
                "itemKey": source.get("itemKey", ""),
                "metric": source.get("metric", ""),
                "month": source.get("month", ""),
                "candidate": candidate_value,
                "answer": answer_value,
                "delta": (float(candidate_value or 0.0) - float(answer_value or 0.0)) if candidate_value is not None and answer_value is not None else None,
                "cell": answer.get("cell", "") if isinstance(answer, dict) else "",
            })
        status = "matched" if not diffs else "mismatch"
        return {"status": status, "diffs": diffs, "diffCount": len(diffs)}

    def _monthly_p4_computed_amounts(self, rows: list[dict], months: list[str]) -> dict[str, dict[str, float]]:
        row_to_key = {
            0: "mf_marketing",
            1: "mf_strategy",
            2: "external_self_operated",
        }
        out: dict[str, dict[str, float]] = {}
        rows_by_month: dict[str, list[dict]] = {month: [] for month in months}
        for row in rows:
            resolved = _resolve_year_month(row)
            if resolved is None:
                continue
            row_month = f"{resolved[0]:04d}-{resolved[1] + 1:02d}"
            if row_month in rows_by_month:
                rows_by_month[row_month].append(row)

        for month_text in months:
            year, month_idx = _year_month_from_month_text(month_text)
            summary, _detail = self.build_dsp_tab4_preview_payload(rows=rows_by_month[month_text], fallback_year=year)
            tab4_rows = summary.get("rows") or []
            for row_idx, key in row_to_key.items():
                monthly_amounts = list((tab4_rows[row_idx] or {}).get("monthlyAmounts") or []) if row_idx < len(tab4_rows) else []
                amount = monthly_amounts[month_idx] if month_idx < len(monthly_amounts) else 0.0
                out.setdefault(month_text, {})[key] = float(amount or 0.0)
        return out

    def _monthly_archive_detail_row_meta(self) -> dict[int, dict[str, str]]:
        def section_block(spec_id: str) -> tuple[str, str, str]:
            if spec_id == "marketing":
                return "內部經銷商", "營銷事業處", "營銷事業處"
            if spec_id == "strategy":
                return "內部經銷商", "策略部", "策略部"
            if spec_id == "external_self":
                return "外部經銷商", "經銷推廣", "外部經銷商"
            if spec_id == "external_io":
                return "外部經銷商", "IO委刊", "IO委刊"
            if spec_id == "hb_bridge":
                return "HB串接", "MD", "HB串接"
            return spec_id, "", spec_id

        row_meta: dict[int, dict[str, str]] = {}
        for spec in TAB4_DETAIL_SECTION_SPECS:
            total_row = int(spec["total_row"])
            block_b, block_c, distributor = section_block(str(spec["id"]))
            for idx, label in enumerate(spec["detail_labels"]):
                row_idx = total_row + 1 + idx
                label_b = str(label.get("b") or "")
                label_c = str(label.get("c") or "")
                label_d = str(label.get("d") or "")
                ad_format = label_d or label_c or label_b
                if label_b.upper() == "CTV":
                    ad_format = "CTV"
                row_meta[row_idx] = {
                    "distributor": distributor,
                    "block_b": block_b,
                    "block_c": block_c,
                    "label_b": label_b,
                    "label_c": label_c,
                    "label_d": label_d,
                    "ad_format": ad_format,
                    "section": str(spec["total_label_a"]),
                }
        return row_meta

    def archive_dsp_month(self, *, month: str, force: bool = False) -> dict:
        month_text = _month_text_from_date_text(f"{month}-01")
        if not month_text:
            raise ValueError("month must be YYYY-MM")
        year, month_idx = _year_month_from_month_text(month_text)
        marker = f"{MONTHLY_DSP_ARCHIVE_ORDER_MARKER_PREFIX}_{month_text}"
        row_meta = self._monthly_archive_detail_row_meta()
        with self.repo.connect() as conn:
            self.repo._ensure_monthly_dsp_archive_table(conn)
            existing = conn.execute(
                "SELECT archive_row_count FROM monthly_dsp_archives WHERE month = ? AND status = 'ok'",
                (month_text,),
            ).fetchone()
            if existing and not force:
                return {
                    "month": month_text,
                    "marker": marker,
                    "status": "skipped",
                    "reason": "already_archived",
                    "archive_row_count": int(existing[0] or 0),
                }

            rows = self.repo.read_canonical_rows_in_tx(conn, "dsp")
            source_rows = [
                row for row in rows
                if _month_text_from_date_text(str(row.get("日期時間") or "")) == month_text
                and str(row.get("訂單") or "") != marker
                and not str(row.get("訂單") or "").startswith(MONTHLY_DSP_ARCHIVE_ORDER_MARKER_PREFIX)
            ]
            _summary_year, detail_amounts = self._build_detail_matrix_values(rows=source_rows, fallback_year=year)
            source_total = sum(_to_number(row.get("執行金額")) for row in source_rows)
            archive_rows: list[dict[str, object]] = []
            for row_idx in DETAIL_INPUT_ROWS:
                amount = float(detail_amounts.get(row_idx, [0.0 for _ in range(MONTH_COUNT)])[month_idx] or 0.0)
                if abs(amount) <= 1e-9:
                    continue
                meta = row_meta[row_idx]
                archive_rows.append(
                    {
                        "日期時間": _month_date_range(month_text)[1],
                        "經銷商": meta["distributor"],
                        "訂單": marker,
                        "素材": f"row{row_idx}_month{month_idx + 1}",
                        "廣告形式": meta["ad_format"],
                        "尺寸": "",
                        "素材樣板": meta["label_d"] or meta["ad_format"],
                        "執行金額": amount,
                        "系統營收": amount,
                        "媒體費用": amount,
                        "原始經銷商": meta["distributor"],
                        "原始廣告形式": meta["ad_format"],
                        "最終經銷商": meta["distributor"],
                        "規則命中_經銷商": "monthly_dsp_archive",
                        "最終來源_經銷商": "monthly_dsp_archive",
                        "分類層級B": meta["block_b"],
                        "分類層級C": meta["block_c"],
                        "分類層級D": "CTV" if meta["label_b"].upper() == "CTV" else meta["label_d"] or meta["label_c"] or meta["label_b"],
                        "最終廣告形式": meta["ad_format"],
                        "規則命中_廣告形式": "monthly_dsp_archive",
                        "最終來源_廣告形式": "monthly_dsp_archive",
                    }
                )

            deleted = conn.execute(
                """
                DELETE FROM canonical_raw
                WHERE workflow = 'dsp'
                  AND substr("日期時間", 1, 7) = ?
                """,
                (month_text,),
            ).rowcount
            max_order = conn.execute(
                "SELECT COALESCE(MAX(row_order), -1) FROM canonical_raw WHERE workflow = 'dsp'"
            ).fetchone()[0]
            now = datetime.now().isoformat(timespec="seconds")
            insert_columns = ["workflow", "row_order", *self.repo.canonical_columns, "updated_at"]
            sql = (
                "INSERT INTO canonical_raw("
                + ", ".join(insert_columns)
                + ") VALUES ("
                + ", ".join("?" for _ in insert_columns)
                + ")"
            )
            for offset, row in enumerate(archive_rows, start=1):
                values = [
                    "dsp",
                    int(max_order) + offset,
                    *[row.get(col, self.repo.field_contract.by_name[col].default) for col in self.repo.canonical_columns],
                    now,
                ]
                conn.execute(sql, values)

            archive_total = sum(float(row.get("執行金額") or 0.0) for row in archive_rows)
            detail = {
                "deleted_rows": int(deleted or 0),
                "retention_policy": "raw data keeps recent two months; closed older months are archived to monthly detail rows",
                "source": "canonical_raw",
            }
            conn.execute(
                """
                INSERT OR REPLACE INTO monthly_dsp_archives(
                  month, workflow, marker, source_row_count, archive_row_count,
                  source_total, archive_total, status, detail_json, archived_at
                ) VALUES (?, 'dsp', ?, ?, ?, ?, ?, 'ok', ?, ?)
                """,
                (
                    month_text,
                    marker,
                    len(source_rows),
                    len(archive_rows),
                    float(source_total),
                    float(archive_total),
                    json.dumps(detail, ensure_ascii=False, sort_keys=True),
                    now,
                ),
            )
            self.repo.append_audit_event(
                conn,
                event_type="monthly_dsp_archive",
                scope=f"dsp:{month_text}",
                status="ok",
                payload={
                    "workflow": "dsp",
                    "month": month_text,
                    "marker": marker,
                    "source_row_count": len(source_rows),
                    "archive_row_count": len(archive_rows),
                    "source_total": float(source_total),
                    "archive_total": float(archive_total),
                },
            )
            conn.commit()

        return {
            "month": month_text,
            "marker": marker,
            "status": "ok",
            "source_row_count": len(source_rows),
            "archive_row_count": len(archive_rows),
            "source_total": float(source_total),
            "archive_total": float(archive_total),
        }

    def build_monthly_p4_snapshot(
        self,
        *,
        week_start: str | None = None,
        week_end: str | None = None,
        manual_source: str = "formal",
        test_id: str = "default",
    ) -> dict:
        anchor_month, months = _month_window_from_period(week_start=week_start, week_end=week_end)
        available_months = _month_year_window(anchor_month)
        payload_months = list(dict.fromkeys([*available_months, *months]))
        target_years = sorted({_year_month_from_month_text(month)[0] for month in payload_months})
        with self.repo.connect() as conn:
            targets_by_year = {target_year: self._target_map_for_year(conn, target_year) for target_year in target_years}
            if manual_source == "test":
                with self.monthly_test_repo.connect() as test_conn:
                    manual_inputs = self.monthly_test_repo.read_monthly_p4_test_inputs_in_tx(
                        test_conn,
                        payload_months,
                        test_id=test_id,
                    )
                    test_templates = self.monthly_test_repo.read_monthly_p4_test_templates_in_tx(test_conn, test_id=test_id)
            else:
                manual_inputs = self.repo.read_monthly_p4_manual_inputs_in_tx(conn, payload_months)
                test_templates = {}
            canonical_rows = self.repo.read_canonical_rows_in_tx(conn, "dsp")

        computed = self._monthly_p4_computed_amounts(canonical_rows, payload_months)

        def manual(month: str, key: str) -> float:
            return float(manual_inputs.get(month, {}).get(key, 0.0))

        month_payloads: list[dict[str, object]] = []
        for month in payload_months:
            month_computed = computed.get(month, {})
            mf_marketing_base = float(month_computed.get("mf_marketing", 0.0))
            mf_strategy_base = float(month_computed.get("mf_strategy", 0.0))
            external_self = float(month_computed.get("external_self_operated", 0.0))
            marketing_io_actual = manual(month, "external_io_momo") + manual(month, "external_io_live")
            external_total_actual = external_self
            data_fee_actual = external_self * 0.05 + manual(month, "data_monetization_adjustment")
            remaining_actual = manual(month, "remaining_traffic_revenue")
            mf_actuals = {
                "mf_marketing": mf_marketing_base + marketing_io_actual,
                "mf_strategy": mf_strategy_base,
                "external_total": external_total_actual,
                "hb_revenue": manual(month, "hb_revenue"),
                "external_beiliu_io": manual(month, "external_beiliu_io"),
            }
            other_actuals = {
                "data_fee": data_fee_actual,
                "remaining_traffic_revenue": remaining_actual,
            }
            mf_target = sum(self._target_value_for_month(targets_by_year, key, month) for key in mf_actuals)
            mf_actual = sum(mf_actuals.values())
            other_target = sum(self._target_value_for_month(targets_by_year, key, month) for key in other_actuals)
            other_actual = sum(other_actuals.values())
            total_target = mf_target + other_target
            total_actual = mf_actual + other_actual
            month_payloads.append(
                {
                    "month": month,
                    "dateRange": _month_date_range(month),
                    "targets": {
                        "mf_marketing": self._target_value_for_month(targets_by_year, "mf_marketing", month),
                        "mf_strategy": self._target_value_for_month(targets_by_year, "mf_strategy", month),
                        "external_total": self._target_value_for_month(targets_by_year, "external_total", month),
                        "hb_revenue": self._target_value_for_month(targets_by_year, "hb_revenue", month),
                        "external_beiliu_io": self._target_value_for_month(targets_by_year, "external_beiliu_io", month),
                        "data_fee": self._target_value_for_month(targets_by_year, "data_fee", month),
                        "remaining_traffic_revenue": self._target_value_for_month(targets_by_year, "remaining_traffic_revenue", month),
                        "mf_total": mf_target,
                        "other_total": other_target,
                        "product_total": total_target,
                    },
                    "computed": {
                        "mf_marketing": mf_marketing_base,
                        "mf_strategy": mf_strategy_base,
                        "external_self_operated": external_self,
                    },
                    "manualInputs": manual_inputs.get(month, {}),
                    "actuals": {
                        **mf_actuals,
                        **other_actuals,
                        "mf_total": mf_actual,
                        "other_total": other_actual,
                        "product_total": total_actual,
                    },
                }
            )

        base_snapshot = None
        if isinstance(test_templates.get("base"), dict):
            raw_base = test_templates["base"].get("snapshot")
            if isinstance(raw_base, dict):
                base_snapshot = raw_base
        if manual_source == "test":
            month_payloads = self._apply_monthly_p4_base_snapshot(
                month_payloads,
                base_snapshot,
                anchor_month=anchor_month,
            )
        candidate_snapshot = self._monthly_p4_snapshot_from_payloads(month_payloads)
        answer_snapshot = None
        if isinstance(test_templates.get("check"), dict):
            raw_answer = test_templates["check"].get("snapshot")
            if isinstance(raw_answer, dict):
                answer_snapshot = raw_answer

        return {
            "anchorMonth": anchor_month,
            "months": months,
            "availableMonths": available_months,
            "manualInputDefinitions": MONTHLY_P4_MANUAL_INPUTS,
            "monthPayloads": month_payloads,
            "candidateSnapshot": candidate_snapshot if manual_source == "test" else {},
            "diff": self._monthly_p4_diff(candidate_snapshot, answer_snapshot) if manual_source == "test" else {},
            "source": "monthly_p4_test_runtime" if manual_source == "test" else "monthly_p4_runtime",
            "testDbPath": str(self.monthly_test_repo.db_path) if manual_source == "test" else "",
            "testTemplates": test_templates,
            "note": "P4(J) 月報表格，手 key 欄位存檔後即時重算。",
        }

    def _monthly_p4_mf_actuals_by_month(self, months: list[str]) -> dict[str, float]:
        month_keys = sorted({str(month or "").strip() for month in months if str(month or "").strip()})
        if not month_keys:
            return {}
        with self.repo.connect() as conn:
            metrics_by_month = self.repo.read_monthly_p4_closed_metrics_in_tx(
                conn,
                month_keys,
                metric_keys=["mf_total_actual"],
            )
        out: dict[str, float] = {}
        for month_text in month_keys:
            metric = metrics_by_month.get(month_text, {}).get("mf_total_actual")
            if not metric:
                continue
            out[month_text] = float(metric.get("value") or 0.0)
        return out

    def _missing_monthly_p4_manual_input_keys(self, manual_inputs: object) -> list[str]:
        values = manual_inputs if isinstance(manual_inputs, dict) else {}
        present = {str(key) for key in values.keys()}
        return [str(item["key"]) for item in MONTHLY_P4_MANUAL_INPUTS if str(item["key"]) not in present]

    def close_monthly_p4_month(
        self,
        *,
        month: str,
        template_version: str,
        rule_version: str,
    ) -> dict[str, object]:
        month_text = str(month or "").strip()
        if not re.match(r"^\d{4}-\d{2}$", month_text):
            raise ValueError("month must be YYYY-MM")
        start_day, end_day = _month_date_range(month_text)
        snapshot = self.build_monthly_p4_snapshot(week_start=start_day, week_end=end_day)
        payload = next(
            (item for item in snapshot.get("monthPayloads", []) if isinstance(item, dict) and item.get("month") == month_text),
            None,
        )
        if not payload:
            raise ValueError(f"找不到 {month_text} 的 P4 月報資料")
        missing_input_keys = self._missing_monthly_p4_manual_input_keys(payload.get("manualInputs"))
        if missing_input_keys:
            raise ValueError(f"monthly P4 manual inputs missing for {month_text}: {', '.join(missing_input_keys)}")
        actuals = payload.get("actuals") if isinstance(payload.get("actuals"), dict) else {}
        targets = payload.get("targets") if isinstance(payload.get("targets"), dict) else {}
        mf_actual = float(actuals.get("mf_total") or 0.0)
        mf_target = float(targets.get("mf_total") or 0.0)
        metrics = {
            "mf_total_actual": {
                "value": mf_actual,
                "source": "monthly_p4_close",
                "sourceFile": "",
                "sourceCell": "",
                "payload": {
                    "month": month_text,
                    "target": mf_target,
                    "template_version": template_version,
                    "rule_version": rule_version,
                },
            }
        }
        with self.repo.connect() as conn:
            written = self.repo.replace_monthly_p4_closed_metrics_in_tx(conn, month_text, metrics)
            self.repo.append_audit_event(
                conn,
                event_type="monthly_p4_close",
                scope=f"monthly:{month_text}",
                status="ok",
                payload={
                    "workflow": "monthly",
                    "month": month_text,
                    "metric_count": written,
                    "mf_total_actual": mf_actual,
                    "template_version": template_version,
                    "rule_version": rule_version,
                    **self._extra_debug_payload(),
                },
            )
        return {
            "status": "ok",
            "month": month_text,
            "metric_count": written,
            "mf_total_actual": round(mf_actual, 2),
            "source": "monthly_p4_close",
        }

    def import_monthly_p4_closed_workbook(
        self,
        *,
        workbook_path: str | Path,
        through_month: str,
        template_version: str,
        rule_version: str,
    ) -> dict[str, object]:
        path = Path(workbook_path)
        if not path.exists():
            raise FileNotFoundError(f"monthly P4 workbook not found: {path}")
        through_text = str(through_month or "").strip()
        if not re.match(r"^\d{4}-\d{2}$", through_text):
            raise ValueError("through_month must be YYYY-MM")
        with path.open("rb") as fh:
            snapshot = self._parse_monthly_p4_workbook_snapshot(fh.read(), filename=path.name)
        entries = snapshot.get("entries") if isinstance(snapshot.get("entries"), dict) else {}
        metrics_by_month: dict[str, dict[str, dict[str, object]]] = {}
        imported: list[dict[str, object]] = []
        for entry in entries.values():
            if not isinstance(entry, dict):
                continue
            if entry.get("itemKey") != "mf_total" or entry.get("metric") != "actual":
                continue
            month_text = str(entry.get("month") or "")
            if not re.match(r"^\d{4}-\d{2}$", month_text) or month_text > through_text:
                continue
            value = float(entry.get("value") or 0.0)
            metric = {
                "value": value,
                "source": "monthly_p4_closed_workbook",
                "sourceFile": path.name,
                "sourceCell": str(entry.get("cell") or ""),
                "payload": {
                    "sheet": str(snapshot.get("sheet") or ""),
                    "template_version": template_version,
                    "rule_version": rule_version,
                },
            }
            metrics_by_month.setdefault(month_text, {})["mf_total_actual"] = metric
            imported.append({"month": month_text, "metricKey": "mf_total_actual", "value": value, "cell": metric["sourceCell"]})
        with self.repo.connect() as conn:
            written = 0
            for month_text, metrics in metrics_by_month.items():
                written += self.repo.replace_monthly_p4_closed_metrics_in_tx(conn, month_text, metrics)
            self.repo.append_audit_event(
                conn,
                event_type="monthly_p4_closed_workbook_import",
                scope=f"monthly:through:{through_text}",
                status="ok",
                payload={
                    "workflow": "monthly",
                    "through_month": through_text,
                    "source_file": path.name,
                    "metric_count": written,
                    "template_version": template_version,
                    "rule_version": rule_version,
                    **self._extra_debug_payload(),
                },
            )
        return {
            "status": "ok",
            "throughMonth": through_text,
            "sourceFile": path.name,
            "sheet": str(snapshot.get("sheet") or ""),
            "metric_count": written,
            "metrics": imported,
            "warnings": snapshot.get("warnings") if isinstance(snapshot.get("warnings"), list) else [],
        }

    def save_monthly_p4_manual_inputs(
        self,
        *,
        month: str,
        inputs: dict[str, object],
        template_version: str,
        rule_version: str,
    ) -> dict:
        if not re.match(r"^\d{4}-\d{2}$", month or ""):
            raise ValueError("month must be YYYY-MM")
        allowed = {str(item["key"]) for item in MONTHLY_P4_MANUAL_INPUTS}
        filtered = {key: value for key, value in inputs.items() if key in allowed}
        with self.repo.connect() as conn:
            written = self.repo.replace_monthly_p4_manual_inputs_in_tx(conn, month, filtered)
            self.repo.append_audit_event(
                conn,
                event_type="monthly_p4_save",
                scope=f"monthly:{month}",
                status="ok",
                payload={
                    "workflow": "monthly",
                    "month": month,
                    "input_count": written,
                    "template_version": template_version,
                    "rule_version": rule_version,
                    **self._extra_debug_payload(),
                },
            )
        return {"status": "ok", "month": month, "input_count": written}

    def save_monthly_p4_test_inputs(
        self,
        *,
        month: str,
        inputs: dict[str, object],
        template_version: str,
        rule_version: str,
        test_id: str = "default",
    ) -> dict:
        if not re.match(r"^\d{4}-\d{2}$", month or ""):
            raise ValueError("month must be YYYY-MM")
        allowed = {str(item["key"]) for item in MONTHLY_P4_MANUAL_INPUTS}
        filtered = {key: value for key, value in inputs.items() if key in allowed}
        with self.monthly_test_repo.connect() as conn:
            written = self.monthly_test_repo.replace_monthly_p4_test_inputs_in_tx(conn, month, filtered, test_id=test_id)
            self.monthly_test_repo.append_audit_event(
                conn,
                event_type="monthly_p4_test_save",
                scope=f"monthly-test:{test_id}:{month}",
                status="ok",
                payload={
                    "workflow": "monthly",
                    "test_id": test_id,
                    "month": month,
                    "input_count": written,
                    "template_version": template_version,
                    "rule_version": rule_version,
                    **self._extra_debug_payload(),
                },
            )
        return {"status": "ok", "month": month, "test_id": test_id, "input_count": written}

    def save_monthly_p4_test_template(
        self,
        *,
        template_kind: str,
        filename: str,
        content_base64: str,
        template_version: str,
        rule_version: str,
        test_id: str = "default",
    ) -> dict:
        kind = str(template_kind or "").strip()
        if kind not in MONTHLY_P4_TEST_TEMPLATE_KINDS:
            raise ValueError("template_kind must be base or check")
        original_filename = Path(str(filename or "").strip()).name
        if not original_filename:
            raise ValueError("filename required")
        suffix = Path(original_filename).suffix.lower()
        if suffix not in {".xlsx", ".xlsm"}:
            raise ValueError("只支援 .xlsx / .xlsm")
        raw_base64 = str(content_base64 or "").strip()
        if "," in raw_base64 and raw_base64.lower().startswith("data:"):
            raw_base64 = raw_base64.split(",", 1)[1]
        file_bytes = base64.b64decode(raw_base64, validate=True)
        if not file_bytes:
            raise ValueError("上傳檔案是空的")
        if len(file_bytes) > 25 * 1024 * 1024:
            raise ValueError("檔案超過 25MB")
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=False)
        try:
            sheet_names = list(workbook.sheetnames)
        finally:
            workbook.close()
        parsed_snapshot = self._parse_monthly_p4_workbook_snapshot(file_bytes, filename=original_filename)

        target_dir = self.monthly_test_repo.db_path.parent / "monthly_p4_test_templates"
        target_dir.mkdir(parents=True, exist_ok=True)
        stored_name = f"{test_id}_{kind}{suffix}"
        stored_path = (target_dir / stored_name).resolve()
        stored_path.write_bytes(file_bytes)
        with self.monthly_test_repo.connect() as conn:
            self.monthly_test_repo.replace_monthly_p4_test_template_in_tx(
                conn,
                test_id=test_id,
                template_kind=kind,
                original_filename=original_filename,
                stored_path=str(stored_path),
                file_size=len(file_bytes),
                sheet_names=sheet_names,
                snapshot=parsed_snapshot,
            )
            self.monthly_test_repo.append_audit_event(
                conn,
                event_type="monthly_p4_test_template_upload",
                scope=f"monthly-test:{test_id}:{kind}",
                status="ok",
                payload={
                    "workflow": "monthly",
                    "test_id": test_id,
                    "template_kind": kind,
                    "filename": original_filename,
                    "file_size": len(file_bytes),
                    "sheet_names": sheet_names,
                    "parsed_entry_count": parsed_snapshot.get("entryCount", 0),
                    "template_version": template_version,
                    "rule_version": rule_version,
                    **self._extra_debug_payload(),
                },
            )
        return {
            "status": "ok",
            "test_id": test_id,
            "template_kind": kind,
            "template_label": MONTHLY_P4_TEST_TEMPLATE_KINDS[kind],
            "filename": original_filename,
            "stored_path": str(stored_path),
            "file_size": len(file_bytes),
            "sheet_names": sheet_names,
            "parsed_entry_count": parsed_snapshot.get("entryCount", 0),
            "parsed_sheet": parsed_snapshot.get("sheet", ""),
        }

    def resolve_ssp_effective_snapshot(self) -> dict[str, object]:
        with self.repo.connect() as conn:
            return self._resolve_ssp_effective_snapshot_in_tx(conn)

    def _resolve_ssp_effective_snapshot_in_tx(self, conn) -> dict[str, object]:
        ssp_rows = self.repo.read_ssp_raw_rows_in_tx(conn)
        if ssp_rows:
            field_names = list(self.repo.workflow_columns("ssp"))
            return {
                "source": "ssp_raw",
                "rows": ssp_rows,
                "columns": ["row_order", *field_names, "updated_at"],
                "field_names": field_names,
                "manual_fields": [],
            }

        field_names = list(self.repo.canonical_columns)
        canonical_rows = self.repo.read_canonical_rows_in_tx(conn, "ssp")
        return {
            "source": "canonical_raw",
            "rows": canonical_rows,
            "columns": ["row_order", *field_names, "updated_at"],
            "field_names": field_names,
            "manual_fields": list(self.repo.modify_allowed_columns),
        }

    def _resolve_export_period(self, *, week_start: str | None, week_end: str | None) -> tuple[str, str]:
        has_start = bool(week_start)
        has_end = bool(week_end)
        if has_start != has_end:
            raise ValueError("week_start and week_end must be provided together")
        if not has_start and not has_end:
            today = date.today()
            this_week_start = today - timedelta(days=today.weekday())
            previous_week_start = this_week_start - timedelta(days=7)
            previous_week_end = this_week_start - timedelta(days=1)
            return previous_week_start.isoformat(), previous_week_end.isoformat()
        assert week_start is not None
        assert week_end is not None
        try:
            week_start_date = date.fromisoformat(week_start)
            week_end_date = date.fromisoformat(week_end)
        except ValueError as exc:
            raise ValueError("week_start and week_end must be YYYY-MM-DD") from exc
        if week_start_date > week_end_date:
            raise ValueError("week_start must be <= week_end")
        return week_start_date.isoformat(), week_end_date.isoformat()

    def _dsp_template_candidate_groups(self) -> list[list[Path]]:
        groups: list[list[Path]] = []
        env_path = os.getenv("MDREP_DSP_TAB4_TEMPLATE_PATH", "").strip()
        if env_path:
            groups.append([Path(env_path).expanduser()])
        if self.repo.project_root is not None:
            groups.append(
                [
                    self.repo.project_root / "templates" / "dsp_tab4_template.xlsx",
                    self.repo.project_root / "templates" / "2026 DSP投資量報表_0101-0503.xlsx",
                ]
            )
        return groups

    def _extract_template_period_window_from_sidecar(self, template_path: Path) -> tuple[date, date] | None:
        sidecar_path = template_path.with_name(f"{template_path.name}.period.json")
        if not sidecar_path.exists():
            return None
        try:
            payload = json.loads(sidecar_path.read_text(encoding="utf-8"))
        except Exception as exc:
            raise ValueError(f"invalid dsp template period sidecar: {sidecar_path}") from exc
        if not isinstance(payload, dict):
            raise ValueError(f"invalid dsp template period sidecar shape: {sidecar_path}")
        raw_start = str(payload.get("week_start") or payload.get("period_start") or "").strip()
        raw_end = str(payload.get("week_end") or payload.get("period_end") or "").strip()
        if not raw_start or not raw_end:
            raise ValueError(f"dsp template period sidecar requires week_start/week_end: {sidecar_path}")
        try:
            start_date = date.fromisoformat(raw_start)
            end_date = date.fromisoformat(raw_end)
        except ValueError as exc:
            raise ValueError(f"dsp template period sidecar must be YYYY-MM-DD: {sidecar_path}") from exc
        if start_date > end_date:
            raise ValueError(f"dsp template period sidecar week_start > week_end: {sidecar_path}")
        return start_date, end_date

    def _extract_template_period_window_from_filename(self, template_path: Path) -> tuple[date, date] | None:
        name = template_path.stem
        year_match = TEMPLATE_YEAR_PREFIX_RE.match(name)
        range_match = TEMPLATE_MMDD_RANGE_RE.search(name)
        if year_match is None or range_match is None:
            return None
        year = int(year_match.group(1))
        start_mmdd = range_match.group(1)
        end_mmdd = range_match.group(2)
        try:
            start_month = int(start_mmdd[:2])
            start_day = int(start_mmdd[2:])
            end_month = int(end_mmdd[:2])
            end_day = int(end_mmdd[2:])
            start_date = date(year, start_month, start_day)
            end_year = year if (end_month, end_day) >= (start_month, start_day) else year + 1
            end_date = date(end_year, end_month, end_day)
        except ValueError as exc:
            raise ValueError(f"invalid dsp template filename period window: {template_path.name}") from exc
        return start_date, end_date

    def _resolve_dsp_template_period_window(self, template_path: Path) -> tuple[date, date] | None:
        sidecar_window = self._extract_template_period_window_from_sidecar(template_path)
        if sidecar_window is not None:
            return sidecar_window
        return self._extract_template_period_window_from_filename(template_path)

    def _pick_dsp_template_from_candidates(
        self,
        *,
        candidates: list[Path],
        week_start_date: date,
        week_end_date: date,
        week_start: str,
        week_end: str,
    ) -> Path | None:
        period_bound_candidates: list[tuple[Path, date, date]] = []
        generic_candidates: list[Path] = []
        for candidate in candidates:
            resolved = candidate.resolve()
            if not (resolved.exists() and resolved.is_file()):
                continue
            try:
                wb = load_workbook(resolved, read_only=True, data_only=True)
                try:
                    if all(sheet_name in wb.sheetnames for sheet_name in DSP_TEMPLATE_SHEET_NAMES):
                        period_window = self._resolve_dsp_template_period_window(resolved)
                        if period_window is None:
                            generic_candidates.append(resolved)
                            continue
                        period_start, period_end = period_window
                        period_bound_candidates.append((resolved, period_start, period_end))
                        if week_start_date >= period_start and week_end_date <= period_end:
                            return resolved
                finally:
                    wb.close()
            except Exception:
                continue
        if period_bound_candidates:
            available_windows = ", ".join(
                f"{path.name}[{start.isoformat()}..{end.isoformat()}]"
                for path, start, end in period_bound_candidates
            )
            raise ValueError(
                "dsp period has no matching base template: "
                f"period={week_start}..{week_end}; available={available_windows}"
            )
        if generic_candidates:
            return generic_candidates[0]
        return None

    def _resolve_dsp_export_template_path(self, *, week_start: str, week_end: str) -> Path:
        week_start_date = date.fromisoformat(week_start)
        week_end_date = date.fromisoformat(week_end)
        candidate_groups = self._dsp_template_candidate_groups()
        for candidates in candidate_groups:
            resolved = self._pick_dsp_template_from_candidates(
                candidates=candidates,
                week_start_date=week_start_date,
                week_end_date=week_end_date,
                week_start=week_start,
                week_end=week_end,
            )
            if resolved is not None:
                return resolved

        flat_candidates = [str(path) for group in candidate_groups for path in group]
        expected = "\n".join(f"- {candidate}" for candidate in flat_candidates)
        raise FileNotFoundError(
            "找不到 DSP Tab4 匯出模板，請提供模板檔。\n"
            "可用方式：\n"
            "1) 設定 MDREP_DSP_TAB4_TEMPLATE_PATH\n"
            "2) 放在 <project_root>/templates/dsp_tab4_template.xlsx\n"
            f"已檢查路徑：\n{expected}"
        )

    def _build_dsp_export_filename(self, week_start: str, week_end: str) -> str:
        start = date.fromisoformat(week_start)
        end = date.fromisoformat(week_end)
        return f"{end.year} DSP投資量報表_{start:%m%d}-{end:%m%d}.xlsx"

    def _resolve_dsp_weekly_baseline_path(self, *, week_start: str) -> Path | None:
        week_start_date = date.fromisoformat(week_start)
        baseline_week_end = (week_start_date - timedelta(days=1)).isoformat()
        seed_root_name = "data_seed_test" if self._feature_flags.get("enable_test_hooks", False) else "data_seed"
        candidates = list(dict.fromkeys([
            self.repo.project_root / seed_root_name / "dsp_weekly_baselines",
            self.repo.project_root / "data_seed" / "dsp_weekly_baselines",
        ]))
        baseline_feature_active = any(
            (root / "manifest.json").exists()
            for root in candidates
        )
        for root in candidates:
            manifest_path = root / "manifest.json"
            if not manifest_path.exists():
                continue
            try:
                manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            except json.JSONDecodeError as exc:
                raise ValueError(f"invalid dsp weekly baseline manifest: {manifest_path}") from exc
            entries = manifest.get("files")
            if entries is None:
                entries = manifest.get("baselines")
            if not isinstance(entries, list):
                raise ValueError(f"invalid dsp weekly baseline manifest entries: {manifest_path}")
            for item in entries:
                if not isinstance(item, dict):
                    continue
                if str(item.get("week_end") or "") != baseline_week_end:
                    continue
                rel_path = str(item.get("path") or item.get("file") or "").strip()
                if not rel_path:
                    raise ValueError(f"dsp weekly baseline manifest item missing path: {manifest_path}")
                resolved = (root / rel_path).resolve()
                try:
                    resolved.relative_to(root.resolve())
                except ValueError as exc:
                    raise PermissionError("dsp weekly baseline path out of baseline root") from exc
                if not resolved.exists():
                    raise FileNotFoundError(f"dsp weekly baseline workbook missing: {resolved}")
                return resolved
        if baseline_feature_active:
            checked = ", ".join(str((root / "manifest.json")) for root in candidates)
            raise FileNotFoundError(
                f"找不到 DSP 週報基底 workbook: baseline_week_end={baseline_week_end}; 已檢查 {checked}"
            )
        return None

    def _filter_rows_by_period(self, rows: list[dict], *, week_start: str, week_end: str) -> list[dict]:
        start = date.fromisoformat(week_start)
        end = date.fromisoformat(week_end)
        out: list[dict] = []
        for row in rows:
            raw_date = str(row.get("日期時間") or row.get("date") or "").strip()[:10]
            try:
                row_date = date.fromisoformat(raw_date)
            except ValueError:
                continue
            if start <= row_date <= end:
                out.append(row)
        return out

    def _filter_rows_through_period_end(self, rows: list[dict], *, week_end: str) -> list[dict]:
        end = date.fromisoformat(week_end)
        start = date(end.year, 1, 1)
        out: list[dict] = []
        for row in rows:
            raw_date = str(row.get("日期時間") or row.get("date") or "").strip()[:10]
            try:
                row_date = date.fromisoformat(raw_date)
            except ValueError:
                continue
            if start <= row_date <= end:
                out.append(row)
        return out

    def _hydrate_dsp_template_workbook(
        self,
        *,
        template_path: Path,
        artifact_path: Path,
        rows: list[dict],
        week_start: str,
        week_end: str,
    ) -> None:
        baseline_path = self._resolve_dsp_weekly_baseline_path(week_start=week_start)
        workbook_source = baseline_path or template_path
        wb = load_workbook(workbook_source)
        try:
            if wb.sheetnames != DSP_TEMPLATE_SHEET_NAMES:
                raise ValueError(
                    "DSP template 工作表結構不符，預期順序: "
                    + ", ".join(repr(name) for name in DSP_TEMPLATE_SHEET_NAMES)
                )

            ws_summary = wb["mF投資量_總表"]
            ws_detail = wb["各經銷商明細"]
            ws_tracking = wb["北流進單追蹤"]

            week_end_date = date.fromisoformat(week_end)
            for row_idx in DETAIL_YEAR_ROWS:
                ws_detail[f"A{row_idx}"] = week_end_date.year
            ws_tracking["A1"] = f"{week_end_date.year}年{week_end_date.month}月份_北流進單狀態"

            if baseline_path is not None:
                period_rows = self._filter_rows_by_period(rows, week_start=week_start, week_end=week_end)
                summary_year, detail_monthly_amounts = self._build_detail_matrix_values(
                    rows=period_rows,
                    fallback_year=week_end_date.year,
                )
                self._add_template_input_cells(
                    ws_summary=ws_summary,
                    ws_detail=ws_detail,
                    year=summary_year,
                    detail_monthly_amounts=detail_monthly_amounts,
                )
            else:
                ytd_rows = self._filter_rows_through_period_end(rows, week_end=week_end)
                summary_year, detail_monthly_amounts = self._build_detail_matrix_values(
                    rows=ytd_rows,
                    fallback_year=week_end_date.year,
                )
                self._write_template_input_cells(
                    ws_summary=ws_summary,
                    ws_detail=ws_detail,
                    year=summary_year,
                    detail_monthly_amounts=detail_monthly_amounts,
                )

            wb.save(artifact_path)
        finally:
            wb.close()
        self._assert_dsp_export_matches_template(
            template_path=workbook_source,
            artifact_path=artifact_path,
        )

    def _dsp_template_mutable_cells(self) -> dict[str, set[str]]:
        mutable: dict[str, set[str]] = {
            "mF投資量_總表": {"A1"},
            "各經銷商明細": {f"A{row_idx}" for row_idx in DETAIL_YEAR_ROWS},
            "北流進單追蹤": {"A1"},
        }
        month_amount_cols = [MONTH_AMOUNT_COL_START + (idx * 2) for idx in range(MONTH_COUNT)]
        detail_cells = mutable["各經銷商明細"]
        for row_idx in DETAIL_INPUT_ROWS:
            for col_idx in month_amount_cols:
                detail_cells.add(f"{get_column_letter(col_idx)}{row_idx}")
        return mutable

    def _assert_dsp_export_matches_template(self, *, template_path: Path, artifact_path: Path) -> None:
        template_wb = load_workbook(template_path, data_only=False)
        export_wb = load_workbook(artifact_path, data_only=False)
        try:
            if template_wb.sheetnames != export_wb.sheetnames:
                raise ValueError("DSP export workbook sheetnames mismatch template")
            mutable_cells_by_sheet = self._dsp_template_mutable_cells()
            for sheet_name in template_wb.sheetnames:
                template_ws = template_wb[sheet_name]
                export_ws = export_wb[sheet_name]
                self._assert_dsp_sheet_layout_matches_template(
                    sheet_name=sheet_name,
                    template_ws=template_ws,
                    export_ws=export_ws,
                )
                self._assert_dsp_sheet_cells_match_template(
                    sheet_name=sheet_name,
                    template_ws=template_ws,
                    export_ws=export_ws,
                    mutable_cells=mutable_cells_by_sheet.get(sheet_name, set()),
                )
        finally:
            export_wb.close()
            template_wb.close()

    def _assert_dsp_sheet_layout_matches_template(self, *, sheet_name: str, template_ws, export_ws) -> None:
        if str(template_ws.sheet_state) != str(export_ws.sheet_state):
            raise ValueError(f"DSP export sheet_state mismatch: {sheet_name}")
        if str(template_ws.freeze_panes or "") != str(export_ws.freeze_panes or ""):
            raise ValueError(f"DSP export freeze_panes mismatch: {sheet_name}")
        template_merged = sorted(str(item) for item in template_ws.merged_cells.ranges)
        export_merged = sorted(str(item) for item in export_ws.merged_cells.ranges)
        if template_merged != export_merged:
            raise ValueError(f"DSP export merged ranges mismatch: {sheet_name}")
        if repr(template_ws.sheet_properties.tabColor) != repr(export_ws.sheet_properties.tabColor):
            raise ValueError(f"DSP export tab color mismatch: {sheet_name}")

        template_hidden_rows = sorted(idx for idx, dim in template_ws.row_dimensions.items() if bool(getattr(dim, "hidden", False)))
        export_hidden_rows = sorted(idx for idx, dim in export_ws.row_dimensions.items() if bool(getattr(dim, "hidden", False)))
        if template_hidden_rows != export_hidden_rows:
            raise ValueError(f"DSP export hidden rows mismatch: {sheet_name}")

        template_hidden_cols = sorted(name for name, dim in template_ws.column_dimensions.items() if bool(getattr(dim, "hidden", False)))
        export_hidden_cols = sorted(name for name, dim in export_ws.column_dimensions.items() if bool(getattr(dim, "hidden", False)))
        if template_hidden_cols != export_hidden_cols:
            raise ValueError(f"DSP export hidden columns mismatch: {sheet_name}")

        template_col_widths = {
            name: float(dim.width)
            for name, dim in template_ws.column_dimensions.items()
            if dim.width is not None
        }
        export_col_widths = {
            name: float(dim.width)
            for name, dim in export_ws.column_dimensions.items()
            if dim.width is not None
        }
        if template_col_widths != export_col_widths:
            raise ValueError(f"DSP export column widths mismatch: {sheet_name}")

    def _assert_dsp_sheet_cells_match_template(
        self,
        *,
        sheet_name: str,
        template_ws,
        export_ws,
        mutable_cells: set[str],
    ) -> None:
        template_cells = getattr(template_ws, "_cells", {})
        export_cells = getattr(export_ws, "_cells", {})
        all_coords = sorted(set(template_cells.keys()) | set(export_cells.keys()))
        for coord in all_coords:
            template_cell = template_ws.cell(row=coord[0], column=coord[1])
            export_cell = export_ws.cell(row=coord[0], column=coord[1])
            coordinate = template_cell.coordinate

            if (
                repr(template_cell.font) != repr(export_cell.font)
                or repr(template_cell.fill) != repr(export_cell.fill)
                or repr(template_cell.border) != repr(export_cell.border)
                or repr(template_cell.alignment) != repr(export_cell.alignment)
                or repr(template_cell.protection) != repr(export_cell.protection)
            ):
                raise ValueError(f"DSP export style mismatch: {sheet_name}!{coordinate}")
            if str(template_cell.number_format or "") != str(export_cell.number_format or ""):
                raise ValueError(f"DSP export number format mismatch: {sheet_name}!{coordinate}")

            template_formula = _is_formula(template_cell.value)
            export_formula = _is_formula(export_cell.value)
            if template_formula != export_formula:
                raise ValueError(f"DSP export formula marker mismatch: {sheet_name}!{coordinate}")
            if template_formula and str(template_cell.value) != str(export_cell.value):
                raise ValueError(f"DSP export formula text mismatch: {sheet_name}!{coordinate}")

            if coordinate in mutable_cells:
                continue
            if not _same_cell_value(template_cell.value, export_cell.value):
                raise ValueError(f"DSP export static cell mismatch: {sheet_name}!{coordinate}")

    def _build_detail_matrix_values(
        self,
        *,
        rows: list[dict],
        fallback_year: int,
    ) -> tuple[int, dict[int, list[float]]]:
        years: list[int] = []
        detail_monthly_amounts: dict[int, list[float]] = {
            row_idx: [0.0 for _ in range(MONTH_COUNT)]
            for row_idx in DETAIL_INPUT_ROWS
        }

        for row in rows:
            resolved = _resolve_year_month(row)
            if resolved is None:
                continue
            year, month_idx = resolved
            years.append(year)
            amount = _to_number(row.get("執行金額"))
            target_row = self._detail_input_row(row)
            detail_monthly_amounts[target_row][month_idx] += amount

        summary_year = max(years) if years else fallback_year
        return summary_year, detail_monthly_amounts

    def _detail_input_row(self, row: dict) -> int:
        block_base = self._detail_block_base_row(row)
        offset = self._detail_metric_offset(row)
        return block_base + offset

    def _detail_block_base_row(self, row: dict) -> int:
        b = _pick_category(row, ["分類層級B", "最終經銷商", "經銷商"])
        c = _pick_category(row, ["分類層級C", "最終廣告形式", "廣告形式"])
        distributor = _pick_category(row, ["最終經銷商", "經銷商", "原始經銷商"])
        haystack = f"{b} {c} {distributor}"

        if _is_internal_distributor_level(b) and c == "策略部":
            return 26
        if _is_external_distributor_level(b) and c == "經銷推廣":
            return 46
        if _is_external_distributor_level(b) and c == "IO委刊":
            return 65
        if b == "HB串接":
            return 84
        if "策略" in haystack:
            return 26
        if "IO委刊" in haystack or "MOMO" in haystack.upper() or "DOOH委刊" in haystack:
            return 65
        if "外部" in haystack or "經銷推廣" in haystack:
            return 46
        if "HB" in haystack.upper() or "串接" in haystack:
            return 84
        return 7

    def _detail_metric_offset(self, row: dict) -> int:
        b = _pick_category(row, ["分類層級B", "最終廣告形式", "廣告形式"])
        c = _pick_category(row, ["分類層級C", "最終廣告形式", "廣告形式"])
        d = _pick_category(row, ["分類層級D", "素材樣板", "素材", "訂單"])
        ad_format = _pick_category(row, ["最終廣告形式", "廣告形式", "素材樣板"])
        order = _pick_category(row, ["訂單", "素材"])
        text = f"{b} {c} {d} {ad_format} {order}".lower()

        if "ctv" in text:
            return 6
        if "北流" in text:
            return 5
        if "dooh外部" in text or "presco" in text or "前線媒體" in text:
            return 4
        if "pre roll" in text or "preroll" in text or "instream" in text:
            return 3
        if "影音摩天" in text or "outstream" in text:
            return 2
        if "創意" in text or "蓋板" in text or "置底" in text or "文中" in text:
            return 1
        return 0

    def _write_template_input_cells(
        self,
        *,
        ws_summary,
        ws_detail,
        year: int,
        detail_monthly_amounts: dict[int, list[float]],
    ) -> None:
        ws_summary["A1"] = year
        month_amount_cols = [MONTH_AMOUNT_COL_START + (idx * 2) for idx in range(MONTH_COUNT)]

        for row_idx in DETAIL_INPUT_ROWS:
            monthly_amounts = detail_monthly_amounts.get(row_idx, [0.0 for _ in range(MONTH_COUNT)])
            for month_idx, col in enumerate(month_amount_cols):
                cell = ws_detail.cell(row=row_idx, column=col)
                if not _is_formula(cell.value):
                    cell.value = monthly_amounts[month_idx]

    def _add_template_input_cells(
        self,
        *,
        ws_summary,
        ws_detail,
        year: int,
        detail_monthly_amounts: dict[int, list[float]],
    ) -> None:
        ws_summary["A1"] = year
        month_amount_cols = [MONTH_AMOUNT_COL_START + (idx * 2) for idx in range(MONTH_COUNT)]

        for row_idx in DETAIL_INPUT_ROWS:
            monthly_amounts = detail_monthly_amounts.get(row_idx, [0.0 for _ in range(MONTH_COUNT)])
            for month_idx, col in enumerate(month_amount_cols):
                delta = float(monthly_amounts[month_idx] or 0.0)
                if delta == 0:
                    continue
                cell = ws_detail.cell(row=row_idx, column=col)
                if _is_formula(cell.value):
                    continue
                cell.value = _to_number(cell.value) + delta

    def _read_template_input_cells(self, *, ws_detail) -> dict[int, list[float]]:
        month_amount_cols = [MONTH_AMOUNT_COL_START + (idx * 2) for idx in range(MONTH_COUNT)]
        out: dict[int, list[float]] = {}
        for row_idx in DETAIL_INPUT_ROWS:
            monthly_amounts: list[float] = []
            for col in month_amount_cols:
                cell = ws_detail.cell(row=row_idx, column=col)
                monthly_amounts.append(0.0 if _is_formula(cell.value) else _to_number(cell.value))
            out[row_idx] = monthly_amounts
        return out

    def build_dsp_tab4_preview_payload(self, *, rows: list[dict], fallback_year: int) -> tuple[dict, dict]:
        preview_year, detail_monthly_amounts = self._build_detail_matrix_values(
            rows=rows,
            fallback_year=fallback_year,
        )
        return self._build_dsp_tab4_preview_payload_from_amounts(
            preview_year=preview_year,
            detail_monthly_amounts=detail_monthly_amounts,
            source="canonical_raw",
        )

    def build_dsp_tab4_preview_payload_for_period(
        self,
        *,
        rows: list[dict],
        week_start: str,
        week_end: str,
        fallback_year: int,
    ) -> tuple[dict, dict]:
        baseline_path = self._resolve_dsp_weekly_baseline_path(week_start=week_start)
        if baseline_path is not None:
            period_rows = self._filter_rows_by_period(rows, week_start=week_start, week_end=week_end)
            preview_year, period_monthly_amounts = self._build_detail_matrix_values(
                rows=period_rows,
                fallback_year=fallback_year,
            )
            wb = load_workbook(baseline_path, data_only=False)
            try:
                baseline_monthly_amounts = self._read_template_input_cells(ws_detail=wb["各經銷商明細"])
            finally:
                wb.close()
            for row_idx in DETAIL_INPUT_ROWS:
                base = baseline_monthly_amounts.get(row_idx, [0.0 for _ in range(MONTH_COUNT)])
                delta = period_monthly_amounts.get(row_idx, [0.0 for _ in range(MONTH_COUNT)])
                baseline_monthly_amounts[row_idx] = [
                    float(base[month_idx] or 0.0) + float(delta[month_idx] or 0.0)
                    for month_idx in range(MONTH_COUNT)
                ]
            detail_monthly_amounts = baseline_monthly_amounts
            source = "weekly_baseline_plus_period_delta"
        else:
            ytd_rows = self._filter_rows_through_period_end(rows, week_end=week_end)
            preview_year, detail_monthly_amounts = self._build_detail_matrix_values(
                rows=ytd_rows,
                fallback_year=fallback_year,
            )
            source = "canonical_raw"
        return self._build_dsp_tab4_preview_payload_from_amounts(
            preview_year=preview_year,
            detail_monthly_amounts=detail_monthly_amounts,
            source=source,
        )

    def _build_dsp_tab4_preview_payload_from_amounts(
        self,
        *,
        preview_year: int,
        detail_monthly_amounts: dict[int, list[float]],
        source: str,
    ) -> tuple[dict, dict]:
        sections: list[dict] = []
        total_row_monthly_amounts: dict[int, list[float]] = {}
        for spec in TAB4_DETAIL_SECTION_SPECS:
            total_row = int(spec["total_row"])
            detail_row_indices = list(range(total_row + 1, total_row + 8))
            total_monthly_amounts = [
                sum(detail_monthly_amounts[row_idx][month_idx] for row_idx in detail_row_indices)
                for month_idx in range(MONTH_COUNT)
            ]
            total_annual_amount = sum(total_monthly_amounts)
            total_row_monthly_amounts[total_row] = total_monthly_amounts
            rows_out: list[dict] = []
            for idx, row_idx in enumerate(detail_row_indices):
                monthly_amounts = detail_monthly_amounts[row_idx]
                annual_amount = sum(monthly_amounts)
                monthly_rates = [
                    (amount / total_monthly_amounts[month_idx]) if total_monthly_amounts[month_idx] > 0 else 0.0
                    for month_idx, amount in enumerate(monthly_amounts)
                ]
                rows_out.append(
                    {
                        "excelRow": row_idx,
                        "labelA": str(spec["detail_label_a"]) if idx == 0 else "",
                        "labelB": str((spec["detail_labels"][idx] or {}).get("b") or ""),
                        "labelC": str((spec["detail_labels"][idx] or {}).get("c") or ""),
                        "labelD": str((spec["detail_labels"][idx] or {}).get("d") or ""),
                        "monthlyAmounts": monthly_amounts,
                        "monthlyRates": monthly_rates,
                        "annualAmount": annual_amount,
                        "annualRate": (annual_amount / total_annual_amount) if total_annual_amount > 0 else 0.0,
                    }
                )

            sections.append(
                {
                    "id": str(spec["id"]),
                    "year": preview_year,
                    "monthLabels": TAB4_MONTH_LABELS,
                    "total": {
                        "excelRow": total_row,
                        "labelA": str(spec["total_label_a"]),
                        "labelB": "",
                        "labelC": "",
                        "labelD": str(spec["total_label_d"]),
                        "monthlyAmounts": total_monthly_amounts,
                        "monthlyRates": [1.0 if value > 0 else 0.0 for value in total_monthly_amounts],
                        "annualAmount": total_annual_amount,
                        "annualRate": 1.0 if total_annual_amount > 0 else 0.0,
                    },
                    "rows": rows_out,
                }
            )

        month_totals = [
            sum(section["total"]["monthlyAmounts"][month_idx] for section in sections)
            for month_idx in range(MONTH_COUNT)
        ]
        annual_total = sum(month_totals)

        def _row_monthly_from_row_index(row_index: int) -> list[float]:
            if row_index in total_row_monthly_amounts:
                return list(total_row_monthly_amounts[row_index])
            return list(detail_monthly_amounts.get(row_index, [0.0 for _ in range(MONTH_COUNT)]))

        summary_row_defs = [
            ("r3", _row_monthly_from_row_index(6), False),
            ("r4", _row_monthly_from_row_index(25), False),
            ("r5", _row_monthly_from_row_index(45), False),
            ("r6", _row_monthly_from_row_index(64), False),
            ("r7", _row_monthly_from_row_index(83), False),
            ("r8", [0.0 for _ in range(MONTH_COUNT)], True),
            ("r9", _row_monthly_from_row_index(7), False),
            ("r10", _row_monthly_from_row_index(8), False),
            ("r11", _row_monthly_from_row_index(9), False),
            ("r12", _row_monthly_from_row_index(10), False),
            ("r13", _row_monthly_from_row_index(11), False),
            ("r14", _row_monthly_from_row_index(12), False),
            ("r15", _row_monthly_from_row_index(13), False),
        ]

        summary_rows: list[dict] = []
        for idx, (_row_id, monthly_amounts, note_only) in enumerate(summary_row_defs):
            annual_amount = sum(monthly_amounts)
            monthly_rates = [None for _ in range(MONTH_COUNT)] if note_only else [
                (amount / month_totals[month_idx]) if month_totals[month_idx] > 0 else 0.0
                for month_idx, amount in enumerate(monthly_amounts)
            ]
            summary_rows.append(
                {
                    "excelRow": idx + 3,
                    "monthlyAmounts": monthly_amounts,
                    "monthlyRates": monthly_rates,
                    "annualAmount": annual_amount,
                    "annualRate": None if note_only else ((annual_amount / annual_total) if annual_total > 0 else 0.0),
                }
            )

        summary_payload = {
            "source": source,
            "year": preview_year,
            "monthTotals": month_totals,
            "monthTotalRates": [1.0 if value > 0 else 0.0 for value in month_totals],
            "annualTotal": annual_total,
            "annualRate": 1.0 if annual_total > 0 else 0.0,
            "rows": summary_rows,
        }
        detail_payload = {
            "source": source,
            "monthLabels": TAB4_MONTH_LABELS,
            "kpiRows": [
                {
                    "excelRow": 2,
                    "label": "全體經銷 總投資量目標 & 達成率 (含北流)",
                    "monthlyAmounts": month_totals,
                    "monthlyRates": [1.0 if value > 0 else 0.0 for value in month_totals],
                    "annualAmount": annual_total,
                    "annualRate": 1.0 if annual_total > 0 else 0.0,
                },
                {
                    "excelRow": 3,
                    "label": "營銷事業處 總投資量目標 & 達成率 (含北流)",
                    "monthlyAmounts": [0.0 for _ in range(MONTH_COUNT)],
                    "monthlyRates": [0.0 for _ in range(MONTH_COUNT)],
                    "annualAmount": 0.0,
                    "annualRate": 0.0,
                },
                {
                    "excelRow": 4,
                    "label": "營銷事業處 北流投資量目標 & 達成率",
                    "monthlyAmounts": [0.0 for _ in range(MONTH_COUNT)],
                    "monthlyRates": [0.0 for _ in range(MONTH_COUNT)],
                    "annualAmount": 0.0,
                    "annualRate": 0.0,
                },
            ],
            "sections": sections,
        }
        return summary_payload, detail_payload

    def save(
        self,
        *,
        workflow: str,
        rows: list[dict],
        template_version: str,
        rule_version: str,
        week_start: str | None = None,
        week_end: str | None = None,
    ) -> dict:
        resolved_week_start: str | None = None
        resolved_week_end: str | None = None
        if workflow == "dsp" and (week_start or week_end):
            resolved_week_start, resolved_week_end = self._resolve_export_period(
                week_start=week_start,
                week_end=week_end,
            )
            self._resolve_dsp_export_template_path(
                week_start=resolved_week_start,
                week_end=resolved_week_end,
            )
        normalized_rows = self._field_contract.validate_and_normalize_save_rows(rows)
        with self.repo.connect() as conn:
            # fail-fast: 先驗證 template/rule binding 合法，再寫 canonical
            self.repo.resolve_trace_binding(conn, workflow, template_version, rule_version)
            written = self.repo.save_canonical_rows(conn, workflow, normalized_rows)
            trace = self.repo.build_trace_meta(conn, workflow, template_version, rule_version)
            run_id = self.repo.insert_run_log(
                conn,
                run_type="save",
                workflow=workflow,
                status="ok",
                trace=trace,
                detail={
                    "row_count": written,
                    "week_start": resolved_week_start or "",
                    "week_end": resolved_week_end or "",
                },
            )
            marker = self._trace_marker(workflow=workflow, run_type="save", run_id=run_id)
            audit_payload = {
                "workflow": workflow,
                "run_id": run_id,
                "template_version": template_version,
                "rule_version": rule_version,
                "canonical_token": trace.canonical_token,
                "row_count": written,
                "week_start": resolved_week_start or "",
                "week_end": resolved_week_end or "",
                **self._extra_debug_payload(),
            }
            if marker:
                audit_payload["trace_marker"] = marker
            self.repo.append_audit_event(
                conn,
                event_type="save",
                scope="service",
                status="ok",
                payload=audit_payload,
            )
        out = {"run_id": run_id, "row_count": written}
        if marker:
            out["trace_marker"] = marker
        if self._feature_flags.get("enable_test_hooks", False):
            out["test_hooks_enabled"] = True
        return out

    def save_ssp_media_slots(
        self,
        *,
        runtime_env: str,
        slots: list[dict],
        template_version: str,
        rule_version: str,
    ) -> dict:
        resolved_runtime_env = str(runtime_env or "").strip() or "prod"
        with self.repo.connect() as conn:
            self.repo.resolve_trace_binding(conn, "ssp", template_version, rule_version)
            written = self.repo.replace_ssp_media_slots_in_tx(conn, resolved_runtime_env, slots)
            trace = self.repo.build_trace_meta(conn, "ssp", template_version, rule_version)
            run_id = self.repo.insert_run_log(
                conn,
                run_type="ssp_media_save",
                workflow="ssp",
                status="ok",
                trace=trace,
                detail={
                    "runtime_env": resolved_runtime_env,
                    "row_count": written,
                },
            )
            marker = self._trace_marker(workflow="ssp", run_type="ssp_media_save", run_id=run_id)
            audit_payload = {
                "workflow": "ssp",
                "run_id": run_id,
                "template_version": template_version,
                "rule_version": rule_version,
                "canonical_token": trace.canonical_token,
                "runtime_env": resolved_runtime_env,
                "row_count": written,
                **self._extra_debug_payload(),
            }
            if marker:
                audit_payload["trace_marker"] = marker
            self.repo.append_audit_event(
                conn,
                event_type="ssp_media_save",
                scope="service",
                status="ok",
                payload=audit_payload,
            )
        out = {
            "status": "ok",
            "runtime_env": resolved_runtime_env,
            "row_count": written,
            "run_id": run_id,
        }
        if marker:
            out["trace_marker"] = marker
        if self._feature_flags.get("enable_test_hooks", False):
            out["test_hooks_enabled"] = True
        return out

    def modify(self, *, workflow: str, updates: list[dict], template_version: str, rule_version: str) -> dict:
        self._field_contract.validate_modify_updates(updates)
        with self.repo.connect() as conn:
            self.repo.resolve_trace_binding(conn, workflow, template_version, rule_version)
            changed = self.repo.apply_modifications(conn, workflow, updates)
            trace = self.repo.build_trace_meta(conn, workflow, template_version, rule_version)
            run_id = self.repo.insert_run_log(
                conn,
                run_type="modify",
                workflow=workflow,
                status="ok",
                trace=trace,
                detail={"changed_count": changed},
            )
            adjustment_count = self.repo.insert_override_adjustments(
                conn,
                workflow=workflow,
                updates=updates,
                template_version=template_version,
                rule_version=rule_version,
                run_id=run_id,
            )
            marker = self._trace_marker(workflow=workflow, run_type="modify", run_id=run_id)
            audit_payload = {
                "workflow": workflow,
                "run_id": run_id,
                "template_version": template_version,
                "rule_version": rule_version,
                "canonical_token": trace.canonical_token,
                "changed_count": changed,
                "adjustment_count": adjustment_count,
                **self._extra_debug_payload(),
            }
            if marker:
                audit_payload["trace_marker"] = marker
            self.repo.append_audit_event(
                conn,
                event_type="modify",
                scope="service",
                status="ok",
                payload=audit_payload,
            )
        out = {"run_id": run_id, "changed_count": changed, "adjustment_count": adjustment_count}
        if marker:
            out["trace_marker"] = marker
        if self._feature_flags.get("enable_test_hooks", False):
            out["test_hooks_enabled"] = True
        return out

    def fetch_ssp_api(
        self,
        *,
        start_day: str,
        end_day: str,
        template_version: str,
        rule_version: str,
        email: str | None = None,
        password: str | None = None,
        scope_check_url: str | None = None,
        api_base_url: str | None = None,
        auth_decrypt_key: str | None = None,
        service_id: int | None = None,
        source_name: str | None = None,
        timeout_seconds: int | None = None,
    ) -> dict:
        settings = resolve_ssp_api_settings(
            email=email,
            password=password,
            scope_check_url=scope_check_url,
            api_base_url=api_base_url,
            auth_decrypt_key=auth_decrypt_key,
            service_id=service_id,
            source_name=source_name,
            timeout_seconds=timeout_seconds,
        )
        bundle = SspApiClient(settings).fetch_report_bundle(start_day=start_day, end_day=end_day)
        rows = normalize_ssp_report_rows(
            [row for row in bundle["rows"] if isinstance(row, dict)],
            source_name=settings.source_name,
        )
        auth = bundle.get("auth") if isinstance(bundle.get("auth"), dict) else {}
        auth_user = auth.get("user") if isinstance(auth.get("user"), dict) else {}
        login = bundle.get("login") if isinstance(bundle.get("login"), dict) else {}
        login_user_id = int(auth_user.get("id") or login.get("id") or 0)
        login_email = str(auth_user.get("email") or login.get("email") or "")

        with self.repo.connect() as conn:
            self.repo.resolve_trace_binding(conn, "ssp", template_version, rule_version)
            changed = self.repo.save_ssp_raw_rows(conn, rows)
            self.repo.save_canonical_rows(conn, "ssp", [])
            trace = self.repo.build_trace_meta(conn, "ssp", template_version, rule_version)
            detail = {
                "start_day": start_day,
                "end_day": end_day,
                "row_count": changed,
                "records_total": int(bundle.get("records_total") or 0),
                "report_id": int(bundle.get("report_id") or 0),
                "report_ids": list(bundle.get("report_ids") or []),
                "daily": list(bundle.get("daily") or []),
                "chunk_mode": str(bundle.get("chunk_mode") or "single"),
                "chunk_days": int(bundle.get("chunk_days") or 1),
                "service_id": int(auth.get("service_id") or 0),
                "source_name": settings.source_name,
                "login_user_id": login_user_id,
                "login_email": login_email,
            }
            run_id = self.repo.insert_run_log(
                conn,
                run_type="fetch_ssp_api",
                workflow="ssp",
                status="ok",
                trace=trace,
                detail=detail,
            )
            self.repo.append_audit_event(
                conn,
                event_type="fetch_ssp_api",
                scope="service",
                status="ok",
                payload={"run_id": run_id, **detail},
            )
            conn.commit()

        return {
            "status": "ok",
            "workflow": "ssp",
            "run_id": run_id,
            "start_day": start_day,
            "end_day": end_day,
            "row_count": changed,
            "records_total": int(bundle.get("records_total") or 0),
            "report_id": int(bundle.get("report_id") or 0),
            "report_ids": list(bundle.get("report_ids") or []),
            "daily": list(bundle.get("daily") or []),
            "chunk_mode": str(bundle.get("chunk_mode") or "single"),
            "chunk_days": int(bundle.get("chunk_days") or 1),
            "service_id": int(auth.get("service_id") or 0),
            "login_user_id": login_user_id,
            "login_email": login_email,
            "source_name": settings.source_name,
            "sum_row": bundle.get("sum_row") if isinstance(bundle.get("sum_row"), dict) else {},
        }

    def fetch_monthly_report_ssp_regular_api(
        self,
        *,
        start_day: str,
        end_day: str,
        pb: int = 1,
        email: str | None = None,
        password: str | None = None,
        scope_check_url: str | None = None,
        api_base_url: str | None = None,
        auth_decrypt_key: str | None = None,
        service_id: int | None = None,
        source_name: str | None = None,
        timeout_seconds: int | None = None,
    ) -> dict:
        settings = resolve_ssp_api_settings(
            email=email,
            password=password,
            scope_check_url=scope_check_url,
            api_base_url=api_base_url,
            auth_decrypt_key=auth_decrypt_key,
            service_id=service_id,
            source_name=source_name,
            timeout_seconds=timeout_seconds,
        )
        client = SspApiClient(settings)
        delivery_bundle = client.fetch_monthly_zone_campaign_size_bundle(
            start_day=start_day,
            end_day=end_day,
            pb=1,
        )
        request_bundle = client.fetch_monthly_zone_campaign_size_bundle(
            start_day=start_day,
            end_day=end_day,
            pb=0,
            dimensions=SSP_MONTHLY_ZONE_SIZE_DIMENSIONS,
        )
        country_bundle = client.fetch_monthly_country_bundle(
            start_day=start_day,
            end_day=end_day,
            pb=0,
        )
        child_country_bundle = client.fetch_monthly_country_bundle(
            start_day=start_day,
            end_day=end_day,
            pb=0,
            zone_group_id=117,
        )
        delivery_rows = normalize_ssp_monthly_zone_campaign_size_rows(
            [row for row in delivery_bundle["rows"] if isinstance(row, dict)],
            source_name=settings.source_name,
        )
        request_rows = normalize_ssp_monthly_zone_campaign_size_rows(
            [row for row in request_bundle["rows"] if isinstance(row, dict)],
            source_name=settings.source_name,
        )
        country_rows = normalize_ssp_monthly_country_rows(
            [row for row in country_bundle["rows"] if isinstance(row, dict)],
            source_name=settings.source_name,
            country_scope="total",
        )
        child_country_rows = normalize_ssp_monthly_country_rows(
            [row for row in child_country_bundle["rows"] if isinstance(row, dict)],
            source_name=settings.source_name,
            country_scope="child",
            zone_group_id=117,
        )
        rows = [
            *[_monthly_request_summary_row(row) for row in request_rows],
            *[_monthly_delivery_detail_row(row) for row in delivery_rows],
        ]
        auth = delivery_bundle.get("auth") if isinstance(delivery_bundle.get("auth"), dict) else {}
        auth_user = auth.get("user") if isinstance(auth.get("user"), dict) else {}
        login = delivery_bundle.get("login") if isinstance(delivery_bundle.get("login"), dict) else {}
        login_user_id = int(auth_user.get("id") or login.get("id") or 0)
        login_email = str(auth_user.get("email") or login.get("email") or "")
        delivery_condition = (
            delivery_bundle.get("report_condition") if isinstance(delivery_bundle.get("report_condition"), dict) else {}
        )
        request_condition = (
            request_bundle.get("report_condition") if isinstance(request_bundle.get("report_condition"), dict) else {}
        )
        country_condition = (
            country_bundle.get("report_condition") if isinstance(country_bundle.get("report_condition"), dict) else {}
        )
        child_country_condition = (
            child_country_bundle.get("report_condition") if isinstance(child_country_bundle.get("report_condition"), dict) else {}
        )
        delivery_result = delivery_bundle.get("report_result") if isinstance(delivery_bundle.get("report_result"), dict) else {}
        request_result = request_bundle.get("report_result") if isinstance(request_bundle.get("report_result"), dict) else {}
        country_result = country_bundle.get("report_result") if isinstance(country_bundle.get("report_result"), dict) else {}
        child_country_result = (
            child_country_bundle.get("report_result") if isinstance(child_country_bundle.get("report_result"), dict) else {}
        )
        sum_row = delivery_bundle.get("sum_row") if isinstance(delivery_bundle.get("sum_row"), dict) else {}
        report_kind = "ssp_regular_monthly_zone_campaign_size"
        run_id = f"monthly-report-{uuid.uuid4().hex}"

        with self.repo.connect_monthly_report() as conn:
            changed = self.repo.save_monthly_report_rows(
                conn,
                run_id=run_id,
                report_kind=report_kind,
                start_day=start_day,
                end_day=end_day,
                report_id=int(delivery_bundle.get("report_id") or 0),
                records_total=int(delivery_bundle.get("records_total") or 0)
                + int(request_bundle.get("records_total") or 0)
                + int(country_bundle.get("records_total") or 0)
                + int(child_country_bundle.get("records_total") or 0),
                source=settings.source_name,
                pb=1,
                request_payload={
                    "delivery_pb1": delivery_condition,
                    "request_pb0": request_condition,
                    "country_pb0": country_condition,
                    "child_country_pb0": child_country_condition,
                },
                response_payload={
                    "delivery_pb1": delivery_result,
                    "request_pb0": request_result,
                    "country_pb0": country_result,
                    "child_country_pb0": child_country_result,
                },
                sum_row=sum_row,
                rows=rows,
            )
            self.repo.save_monthly_country_rows(conn, run_id=run_id, rows=[*country_rows, *child_country_rows])
            conn.commit()

        return {
            "status": "ok",
            "workflow": "monthly",
            "run_id": run_id,
            "report_kind": report_kind,
            "monthly_report_db_path": str(self.repo.monthly_report_db_path),
            "start_day": start_day,
            "end_day": end_day,
            "row_count": changed,
            "records_total": int(delivery_bundle.get("records_total") or 0)
            + int(request_bundle.get("records_total") or 0)
            + int(country_bundle.get("records_total") or 0)
            + int(child_country_bundle.get("records_total") or 0),
            "report_id": int(delivery_bundle.get("report_id") or 0),
            "report_ids": [
                *list(delivery_bundle.get("report_ids") or []),
                *list(request_bundle.get("report_ids") or []),
                *list(country_bundle.get("report_ids") or []),
                *list(child_country_bundle.get("report_ids") or []),
            ],
            "chunk_mode": str(delivery_bundle.get("chunk_mode") or "single"),
            "chunk_days": int(delivery_bundle.get("chunk_days") or 1),
            "pb": 1,
            "request_pb": 0,
            "delivery_row_count": len(delivery_rows),
            "request_row_count": len(request_rows),
            "country_row_count": len(country_rows),
            "child_country_row_count": len(child_country_rows),
            "service_id": int(auth.get("service_id") or 0),
            "login_user_id": login_user_id,
            "login_email": login_email,
            "source_name": settings.source_name,
            "sum_row": sum_row,
        }

    def build_monthly_media_cost_analysis(self, *, month: str) -> dict[str, object]:
        month_text = _month_text_from_date_text(f"{month}-01" if len(str(month)) == 7 else str(month))
        if not month_text:
            raise ValueError("month must be YYYY-MM")
        start_day, end_day = _month_date_range(month_text)
        rows = self.repo.read_monthly_report_rows(month=month_text)
        latest_run = self.repo.read_latest_monthly_report_run(
            report_kind="ssp_regular_monthly_zone_campaign_size",
            month=month_text,
        ) or {}
        media_cost = sum(float(row.get("profit") or 0.0) for row in rows)
        fallback_investment = sum(float(row.get("advertiser_mu") or 0.0) for row in rows)
        p4_mf_actual = float(self._monthly_p4_mf_actuals_by_month([month_text]).get(month_text, 0.0) or 0.0)
        total_investment = p4_mf_actual
        gross_profit = total_investment - media_cost if total_investment > 0 else 0.0
        media_cost_rate = (media_cost / total_investment * 100.0) if total_investment > 0 else 0.0
        snapshot = {
            "chartKey": "media_cost_analysis",
            "month": month_text,
            "startDay": start_day,
            "endDay": end_day,
            "sourceRunId": str(latest_run.get("run_id") or ""),
            "source": str(latest_run.get("source") or ""),
            "rowCount": len(rows),
            "metrics": {
                "mediaCost": round(media_cost, 2),
                "totalInvestment": round(total_investment, 2),
                "p4MfActual": round(p4_mf_actual, 2),
                "fallbackInvestment": round(fallback_investment, 2),
                "grossProfit": round(gross_profit, 2),
                "mediaCostRate": round(media_cost_rate, 4),
            },
        }
        snapshot_id = f"monthly-chart-{uuid.uuid4().hex}"
        self.repo.save_monthly_chart_snapshot(
            snapshot_id=snapshot_id,
            chart_key="media_cost_analysis",
            month=month_text,
            start_day=start_day,
            end_day=end_day,
            source_run_id=str(latest_run.get("run_id") or ""),
            payload=snapshot,
        )
        return {
            "status": "ok",
            "snapshot_id": snapshot_id,
            "monthly_report_db_path": str(self.repo.monthly_report_db_path),
            **snapshot,
        }

    def build_monthly_dimension_summary(self, *, month: str, limit: int = 20) -> dict[str, object]:
        month_text = _month_text_from_date_text(f"{month}-01" if len(str(month)) == 7 else str(month))
        if not month_text:
            raise ValueError("month must be YYYY-MM")
        start_day, end_day = _month_date_range(month_text)
        rows = self.repo.read_monthly_report_rows(month=month_text)
        latest_run = self.repo.read_latest_monthly_report_run(
            report_kind="ssp_regular_monthly_zone_campaign_size",
            month=month_text,
        ) or {}

        def add_metric(target: dict[str, object], row: dict[str, object]) -> None:
            for key in ("request", "impress", "click", "profit", "advertiser_mu"):
                target[key] = float(target.get(key) or 0.0) + float(row.get(key) or 0.0)

        by_zone: dict[str, dict[str, object]] = {}
        by_campaign: dict[str, dict[str, object]] = {}
        by_format: dict[str, dict[str, object]] = {}
        for row in rows:
            zone_key = str(row.get("zone_id") or "")
            zone = by_zone.setdefault(
                zone_key,
                {"zone_id": int(row.get("zone_id") or 0), "zone_name": str(row.get("zone_name") or "")},
            )
            add_metric(zone, row)

            campaign_key = str(row.get("campaign_id") or "")
            if campaign_key:
                campaign = by_campaign.setdefault(
                    campaign_key,
                    {
                        "campaign_id": campaign_key,
                        "campaign_name": str(row.get("campaign_name") or ""),
                    },
                )
                add_metric(campaign, row)

            format_key = _monthly_report_row_ad_format(row) or "未分類"
            ad_format = by_format.setdefault(
                format_key,
                {
                    "ad_format": format_key,
                    "ad_format_rule": str(row.get("ad_format_rule") or ""),
                },
            )
            add_metric(ad_format, row)

        def enrich(items: list[dict[str, object]]) -> list[dict[str, object]]:
            out: list[dict[str, object]] = []
            for item in items:
                impress = float(item.get("impress") or 0.0)
                click = float(item.get("click") or 0.0)
                advertiser_mu = float(item.get("advertiser_mu") or 0.0)
                item["ctr"] = round((click / impress * 100.0) if impress > 0 else 0.0, 4)
                item["dsp_ecpm"] = round((advertiser_mu / impress * 1000.0) if impress > 0 else 0.0, 4)
                item["dsp_ecpc"] = round((advertiser_mu / click) if click > 0 else 0.0, 4)
                out.append(item)
            return out

        top_limit = max(1, int(limit or 20))
        snapshot = {
            "chartKey": "monthly_dimension_summary",
            "month": month_text,
            "startDay": start_day,
            "endDay": end_day,
            "sourceRunId": str(latest_run.get("run_id") or ""),
            "source": str(latest_run.get("source") or ""),
            "rowCount": len(rows),
            "topZones": enrich(sorted(by_zone.values(), key=lambda item: float(item.get("advertiser_mu") or 0.0), reverse=True)[:top_limit]),
            "topCampaigns": enrich(sorted(by_campaign.values(), key=lambda item: float(item.get("advertiser_mu") or 0.0), reverse=True)[:top_limit]),
            "adFormats": enrich(sorted(by_format.values(), key=lambda item: float(item.get("advertiser_mu") or 0.0), reverse=True)),
        }
        snapshot_id = f"monthly-chart-{uuid.uuid4().hex}"
        self.repo.save_monthly_chart_snapshot(
            snapshot_id=snapshot_id,
            chart_key="monthly_dimension_summary",
            month=month_text,
            start_day=start_day,
            end_day=end_day,
            source_run_id=str(latest_run.get("run_id") or ""),
            payload=snapshot,
        )
        return {
            "status": "ok",
            "snapshot_id": snapshot_id,
            "monthly_report_db_path": str(self.repo.monthly_report_db_path),
            **snapshot,
        }

    def import_monthly_zone_group_csv(self, *, csv_path: str | Path, group_id: int, group_name: str) -> dict[str, object]:
        path = Path(csv_path)
        if not path.exists():
            raise FileNotFoundError(f"zone group CSV not found: {path}")

        zone_ids: list[int] = []
        skipped_rows: list[dict[str, object]] = []
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            for row_index, row in enumerate(csv.reader(fh), start=1):
                if not row:
                    continue
                raw = str(row[0] or "").strip()
                if not raw:
                    continue
                try:
                    zone_ids.append(int(raw))
                except ValueError:
                    skipped_rows.append({"row": row_index, "value": raw})

        duplicate_count = len(zone_ids) - len(set(zone_ids))
        result = self.repo.replace_monthly_zone_group(
            group_id=int(group_id),
            group_name=str(group_name or ""),
            zone_ids=zone_ids,
        )
        return {
            "status": "ok",
            "workflow": "monthly",
            "monthly_report_db_path": str(self.repo.monthly_report_db_path),
            "source_csv": str(path),
            "input_row_count": len(zone_ids),
            "duplicate_count": duplicate_count,
            "skipped_rows": skipped_rows[:20],
            "skipped_row_count": len(skipped_rows),
            **result,
        }

    def build_monthly_charts_snapshot(self, *, months: list[str] | None = None, limit: int = 12) -> dict[str, object]:
        with self.repo.connect_monthly_report() as conn:
            self.repo._ensure_monthly_report_tables(conn)
            if months:
                month_keys = sorted({str(month or "").strip() for month in months if str(month or "").strip()})
            else:
                month_keys = [
                    str(row[0] or "")
                    for row in conn.execute(
                        "SELECT DISTINCT month FROM monthly_report_rows WHERE month != '' ORDER BY month ASC"
                    ).fetchall()
                ]
        top_limit = max(1, int(limit or 12))
        child_group = self.repo.read_monthly_zone_group(group_id=117)
        child_zone_ids = set(child_group.get("zone_ids") or set())
        child_group_name = str(child_group.get("group_name") or "子聯播網")
        p4_mf_actuals = self._monthly_p4_mf_actuals_by_month(month_keys)

        def empty_month(month_text: str) -> dict[str, object]:
            start_day, end_day = _month_date_range(month_text)
            days = max(1, (date.fromisoformat(end_day) - date.fromisoformat(start_day)).days + 1)
            return {
                "month": month_text,
                "startDay": start_day,
                "endDay": end_day,
                "days": days,
                "request": 0.0,
                "impress": 0.0,
                "click": 0.0,
                "profit": 0.0,
                "advertiser_mu": 0.0,
                "mediaCostInvestment": 0.0,
            "p4MfActual": 0.0,
            "p4Closed": False,
            "p4InvestmentSource": "unclosed",
                "grossProfit": 0.0,
                "mediaCostRate": 0.0,
                "ctr": 0.0,
                "dspEcpm": 0.0,
                "dspEcpc": 0.0,
                "dailyInvestment": 0.0,
                "dailyRequest": 0.0,
                "dailyImpress": 0.0,
                "dailyClick": 0.0,
            }

        monthly_rows: list[dict[str, object]] = []
        ad_formats_by_month: dict[str, list[dict[str, object]]] = {}
        top_zones_by_month: dict[str, list[dict[str, object]]] = {}
        top_campaigns_by_month: dict[str, list[dict[str, object]]] = {}
        creative_daily_rows: list[dict[str, object]] = []
        creative_traffic_rows: list[dict[str, object]] = []
        network_usage_rows: list[dict[str, object]] = []
        for month_text in month_keys:
            rows = self.repo.read_monthly_report_rows(month=month_text)
            monthly = empty_month(month_text)
            child_network = empty_month(month_text)
            creative_traffic = empty_month(month_text)
            tw_country = empty_month(month_text)
            tw_child_country = empty_month(month_text)
            by_format: dict[str, dict[str, object]] = {}
            by_zone: dict[str, dict[str, object]] = {}
            by_campaign: dict[str, dict[str, object]] = {}

            def add_metric(target: dict[str, object], row: dict[str, object]) -> None:
                for key in ("request", "impress", "click", "profit", "advertiser_mu"):
                    target[key] = float(target.get(key) or 0.0) + float(row.get(key) or 0.0)

            for row in rows:
                zone_id = int(row.get("zone_id") or 0)
                add_metric(monthly, row)
                if zone_id in child_zone_ids:
                    add_metric(child_network, row)
                if _is_monthly_creative_traffic_row(row):
                    add_metric(creative_traffic, row)
                fmt_key = _monthly_report_row_ad_format(row) or "未分類"
                fmt = by_format.setdefault(fmt_key, {"month": month_text, "adFormat": fmt_key})
                add_metric(fmt, row)

                zone_key = str(row.get("zone_id") or "")
                zone = by_zone.setdefault(
                    zone_key,
                    {"month": month_text, "zoneId": int(row.get("zone_id") or 0), "zoneName": str(row.get("zone_name") or "")},
                )
                add_metric(zone, row)

                campaign_key = str(row.get("campaign_id") or "")
                if campaign_key:
                    campaign = by_campaign.setdefault(
                        campaign_key,
                        {"month": month_text, "campaignId": campaign_key, "campaignName": str(row.get("campaign_name") or "")},
                    )
                    add_metric(campaign, row)

            country_rows = self.repo.read_monthly_country_rows(month=month_text)
            for row in country_rows:
                if _is_taiwan_country(row.get("country")):
                    target_country = (
                        tw_child_country
                        if str(row.get("country_scope") or "") == "child" and int(row.get("zone_group_id") or 0) == 117
                        else tw_country
                    )
                    target_country["request"] = float(target_country.get("request") or 0.0) + float(row.get("request") or 0.0)
                    target_country["impress"] = float(target_country.get("impress") or 0.0) + float(row.get("impress") or 0.0)

            p4_mf_actual = float(p4_mf_actuals.get(month_text, 0.0) or 0.0)
            monthly["p4MfActual"] = p4_mf_actual
            monthly["mediaCostInvestment"] = p4_mf_actual
            monthly["p4Closed"] = month_text in p4_mf_actuals
            monthly["p4InvestmentSource"] = "monthly_p4_closed_metrics" if month_text in p4_mf_actuals else "unclosed"

            def enrich(item: dict[str, object], days: int) -> dict[str, object]:
                request = float(item.get("request") or 0.0)
                impress = float(item.get("impress") or 0.0)
                click = float(item.get("click") or 0.0)
                profit = float(item.get("profit") or 0.0)
                advertiser_mu = float(item.get("advertiser_mu") or 0.0)
                if "mediaCostInvestment" in item:
                    media_cost_investment = float(item.get("mediaCostInvestment") or 0.0)
                else:
                    media_cost_investment = advertiser_mu
                item["grossProfit"] = round(media_cost_investment - profit, 2) if media_cost_investment > 0 else 0.0
                item["mediaCostRate"] = round((profit / media_cost_investment * 100.0) if media_cost_investment > 0 else 0.0, 4)
                item["ctr"] = round((click / impress * 100.0) if impress > 0 else 0.0, 4)
                item["dspEcpm"] = round((advertiser_mu / impress * 1000.0) if impress > 0 else 0.0, 4)
                item["dspEcpc"] = round((advertiser_mu / click) if click > 0 else 0.0, 4)
                item["dailyInvestment"] = round(advertiser_mu / days, 2)
                item["dailyRequest"] = round(request / days, 2)
                item["dailyImpress"] = round(impress / days, 2)
                item["dailyClick"] = round(click / days, 2)
                return item

            days = int(monthly["days"])
            monthly = enrich(monthly, days)
            child_network = enrich(child_network, days)
            creative_traffic = enrich(creative_traffic, days)
            tw_country = enrich(tw_country, days)
            tw_child_country = enrich(tw_child_country, days)
            creative_traffic["trafficFormat"] = "創意型流量池"
            main_network = empty_month(month_text)
            for key in ("request", "impress", "click", "profit", "advertiser_mu"):
                main_network[key] = max(0.0, float(monthly.get(key) or 0.0) - float(child_network.get(key) or 0.0))
            main_network = enrich(main_network, days)
            country_source = any(str(row.get("country_scope") or "") == "total" for row in country_rows)
            if country_source:
                tw_main_country = empty_month(month_text)
                for key in ("request", "impress"):
                    tw_main_country[key] = max(0.0, float(tw_country.get(key) or 0.0) - float(tw_child_country.get(key) or 0.0))
                network_total = tw_country
                network_child = tw_child_country
                network_main = enrich(tw_main_country, days)
            else:
                network_total = monthly
                network_child = child_network
                network_main = main_network
            network_usage_rows.append(
                {
                    "month": month_text,
                    "groupId": int(child_group.get("group_id") or 117),
                    "groupName": child_group_name,
                    "total": network_total,
                    "child": network_child,
                    "main": network_main,
                    "tw": tw_country,
                    "countrySource": country_source,
                    "childRequestShare": round(
                        (float(network_child.get("request") or 0.0) / float(network_total.get("request") or 0.0) * 100.0)
                        if float(network_total.get("request") or 0.0) > 0
                        else 0.0,
                        4,
                    ),
                    "childImpressShare": round(
                        (float(network_child.get("impress") or 0.0) / float(network_total.get("impress") or 0.0) * 100.0)
                        if float(network_total.get("impress") or 0.0) > 0
                        else 0.0,
                        4,
                    ),
                    "childInvestmentShare": round(
                        (
                            float(child_network.get("advertiser_mu") or 0.0)
                            / float(monthly.get("advertiser_mu") or 0.0)
                            * 100.0
                        )
                        if float(monthly.get("advertiser_mu") or 0.0) > 0
                        else 0.0,
                        4,
                    ),
                }
            )
            monthly_rows.append(monthly)
            format_rows = [enrich(item, days) for item in by_format.values()]
            format_rows.sort(key=lambda item: float(item.get("advertiser_mu") or 0.0), reverse=True)
            ad_formats_by_month[month_text] = format_rows
            top_zones_by_month[month_text] = [
                enrich(item, days)
                for item in sorted(by_zone.values(), key=lambda item: float(item.get("advertiser_mu") or 0.0), reverse=True)[:top_limit]
            ]
            top_campaigns_by_month[month_text] = [
                enrich(item, days)
                for item in sorted(by_campaign.values(), key=lambda item: float(item.get("advertiser_mu") or 0.0), reverse=True)[:top_limit]
            ]
            creative = next((item for item in format_rows if str(item.get("adFormat") or "") == "創意廣告"), None)
            if creative:
                creative_daily_rows.append(dict(creative))
            creative_traffic_rows.append(dict(creative_traffic))

        format_names = sorted({str(item.get("adFormat") or "") for rows in ad_formats_by_month.values() for item in rows})
        return {
            "source": "monthly_report.sqlite",
            "months": month_keys,
            "monthly": monthly_rows,
            "adFormats": {
                "names": format_names,
                "byMonth": ad_formats_by_month,
            },
            "creativeDaily": creative_daily_rows,
            "trafficDaily": {
                "creative": creative_traffic_rows,
            },
            "networkUsage": network_usage_rows,
            "networkGroup": {
                "groupId": int(child_group.get("group_id") or 117),
                "groupName": child_group_name,
                "zoneCount": len(child_zone_ids),
                "updatedAt": str(child_group.get("updated_at") or ""),
            },
            "topZonesByMonth": top_zones_by_month,
            "topCampaignsByMonth": top_campaigns_by_month,
            "notes": [
                "目前可由 SSP 月報資料直接組出：媒體成本、整體流量、主/子聯播網、廣告形式、創意型、Top 版位、Top 訂單。",
                "TW 流量使用 SSP country daily report；子聯播網依 monthly_zone_groups group_id=117 對照。",
            ],
        }

    def fetch_ssp_ad_group_api(
        self,
        *,
        zone_group_id: int,
        start_day: str,
        end_day: str,
        template_version: str,
        rule_version: str,
        email: str | None = None,
        password: str | None = None,
        scope_check_url: str | None = None,
        api_base_url: str | None = None,
        auth_decrypt_key: str | None = None,
        service_id: int | None = None,
        source_name: str | None = None,
        timeout_seconds: int | None = None,
    ) -> dict:
        settings = resolve_ssp_api_settings(
            email=email,
            password=password,
            scope_check_url=scope_check_url,
            api_base_url=api_base_url,
            auth_decrypt_key=auth_decrypt_key,
            service_id=service_id,
            source_name=source_name,
            timeout_seconds=timeout_seconds,
        )
        catalog_item = SSP_AD_GROUP_CATALOG_BY_ID.get(int(zone_group_id), {})
        bundle = SspApiClient(settings).fetch_ad_group_report_bundle(
            start_day=start_day,
            end_day=end_day,
            zone_group_id=zone_group_id,
            zone_group_name=str(catalog_item.get("name") or ""),
        )
        rows = normalize_ssp_ad_group_report_rows(
            [row for row in bundle["rows"] if isinstance(row, dict)],
            zone_group_id=zone_group_id,
            source_name=settings.source_name,
        )
        for row in rows:
            row["zone_group_name"] = str(catalog_item.get("name") or f"zone_group {zone_group_id}")
            row["ad_format"] = str(catalog_item.get("format") or "")
            row["price_tier"] = str(catalog_item.get("tier") or "")
        auth = bundle.get("auth") if isinstance(bundle.get("auth"), dict) else {}
        auth_user = auth.get("user") if isinstance(auth.get("user"), dict) else {}
        login = bundle.get("login") if isinstance(bundle.get("login"), dict) else {}
        login_user_id = int(auth_user.get("id") or login.get("id") or 0)
        login_email = str(auth_user.get("email") or login.get("email") or "")
        report_condition = bundle.get("report_condition") if isinstance(bundle.get("report_condition"), dict) else {}
        report_result = bundle.get("report_result") if isinstance(bundle.get("report_result"), dict) else {}

        with self.repo.connect() as conn:
            self.repo.resolve_trace_binding(conn, "ssp", template_version, rule_version)
            trace = self.repo.build_trace_meta(conn, "ssp", template_version, rule_version)
            detail = {
                "zone_group_id": int(zone_group_id),
                "start_day": start_day,
                "end_day": end_day,
                "row_count": len(rows),
                "records_total": int(bundle.get("records_total") or 0),
                "report_id": int(bundle.get("report_id") or 0),
                "report_ids": list(bundle.get("report_ids") or []),
                "chunk_mode": str(bundle.get("chunk_mode") or "single"),
                "chunk_days": int(bundle.get("chunk_days") or 1),
                "service_id": int(auth.get("service_id") or 0),
                "source_name": settings.source_name,
                "login_user_id": login_user_id,
                "login_email": login_email,
                "sum_row": bundle.get("sum_row") if isinstance(bundle.get("sum_row"), dict) else {},
            }
            run_id = self.repo.insert_run_log(
                conn,
                run_type="fetch_ssp_ad_group_api",
                workflow="ssp",
                status="ok",
                trace=trace,
                detail=detail,
            )
            changed = self.repo.save_ssp_ad_group_report(
                conn,
                run_id=run_id,
                zone_group_id=zone_group_id,
                start_day=start_day,
                end_day=end_day,
                report_id=int(bundle.get("report_id") or 0),
                records_total=int(bundle.get("records_total") or 0),
                source=settings.source_name,
                request_payload=report_condition,
                response_payload=report_result,
                rows=rows,
            )
            self.repo.append_audit_event(
                conn,
                event_type="fetch_ssp_ad_group_api",
                scope="service",
                status="ok",
                payload={"run_id": run_id, **detail, "row_count": changed},
            )
            conn.commit()

        return {
            "status": "ok",
            "workflow": "ssp",
            "run_id": run_id,
            "zone_group_id": int(zone_group_id),
            "start_day": start_day,
            "end_day": end_day,
            "row_count": changed,
            "records_total": int(bundle.get("records_total") or 0),
            "report_id": int(bundle.get("report_id") or 0),
            "report_ids": list(bundle.get("report_ids") or []),
            "service_id": int(auth.get("service_id") or 0),
            "login_user_id": login_user_id,
            "login_email": login_email,
            "source_name": settings.source_name,
            "sum_row": bundle.get("sum_row") if isinstance(bundle.get("sum_row"), dict) else {},
        }

    def fetch_all_ssp_ad_group_api(
        self,
        *,
        start_day: str,
        end_day: str,
        template_version: str,
        rule_version: str,
        email: str | None = None,
        password: str | None = None,
        scope_check_url: str | None = None,
        api_base_url: str | None = None,
        auth_decrypt_key: str | None = None,
        service_id: int | None = None,
        source_name: str | None = None,
        timeout_seconds: int | None = None,
    ) -> dict:
        groups: list[dict[str, object]] = []
        row_count = 0
        records_total = 0
        for item in SSP_AD_GROUP_CATALOG:
            result = self.fetch_ssp_ad_group_api(
                zone_group_id=int(item["id"]),
                start_day=start_day,
                end_day=end_day,
                template_version=template_version,
                rule_version=rule_version,
                email=email,
                password=password,
                scope_check_url=scope_check_url,
                api_base_url=api_base_url,
                auth_decrypt_key=auth_decrypt_key,
                service_id=service_id,
                source_name=source_name,
                timeout_seconds=timeout_seconds,
            )
            groups.append(result)
            row_count += int(result.get("row_count") or 0)
            records_total += int(result.get("records_total") or 0)
        return {
            "status": "ok",
            "workflow": "ssp",
            "start_day": start_day,
            "end_day": end_day,
            "group_count": len(groups),
            "row_count": row_count,
            "records_total": records_total,
            "groups": groups,
        }

    def build_ssp_ad_group_monitor_snapshot(
        self,
        *,
        start_day: str,
        end_day: str,
    ) -> dict[str, object]:
        catalog = [dict(item) for item in SSP_AD_GROUP_CATALOG]
        rows = self.repo.read_ssp_ad_group_metrics_for_groups(
            zone_group_ids=[int(item["id"]) for item in catalog],
            start_day=start_day,
            end_day=end_day,
        )
        latest_runs = self.repo.read_latest_ssp_ad_group_runs(
            zone_group_ids=[int(item["id"]) for item in catalog],
        )
        latest_by_group = {int(item.get("zone_group_id") or 0): item for item in latest_runs}
        metric_keys = ["request", "impress", "click", "profit", "advertiser_mu"]

        def empty_group(item: dict[str, object]) -> dict[str, object]:
            return {
                "zone_group_id": int(item["id"]),
                "zone_group_name": str(item["name"]),
                "ad_format": str(item["format"]),
                "price_tier": str(item["tier"]),
                **{key: 0.0 for key in metric_keys},
                "invalid_impress": 0.0,
                "invalid_click": 0.0,
            }

        total = {key: 0.0 for key in metric_keys}
        by_group: dict[int, dict[str, object]] = {int(item["id"]): empty_group(item) for item in catalog}
        by_format: dict[str, dict[str, object]] = {}
        placements_by_group: dict[int, dict[int, dict[str, object]]] = {}
        all_dates = sorted({str(row.get("date") or "") for row in rows if str(row.get("date") or "")})
        group_daily: dict[int, dict[str, dict[str, object]]] = {}
        format_daily: dict[str, dict[str, dict[str, object]]] = {}
        placement_daily: dict[int, dict[int, dict[str, dict[str, object]]]] = {}
        for row in rows:
            group_id = int(row.get("zone_group_id") or 0)
            date_key = str(row.get("date") or "")
            catalog_item = SSP_AD_GROUP_CATALOG_BY_ID.get(group_id, {})
            group = by_group.setdefault(
                group_id,
                {
                    "zone_group_id": group_id,
                    "zone_group_name": str(row.get("zone_group_name") or catalog_item.get("name") or ""),
                    "ad_format": str(row.get("ad_format") or catalog_item.get("format") or ""),
                    "price_tier": str(row.get("price_tier") or catalog_item.get("tier") or ""),
                    **{key: 0.0 for key in metric_keys},
                    "invalid_impress": 0.0,
                    "invalid_click": 0.0,
                },
            )
            ad_format = str(group.get("ad_format") or "")
            fmt = by_format.setdefault(
                ad_format,
                {
                    "ad_format": ad_format,
                    **{key: 0.0 for key in metric_keys},
                    "invalid_impress": 0.0,
                    "invalid_click": 0.0,
                },
            )
            zone_id = int(row.get("zone_id") or 0)
            daily_group = group_daily.setdefault(group_id, {}).setdefault(
                date_key,
                {
                    "date": date_key,
                    "zone_group_id": group_id,
                    "zone_group_name": str(group.get("zone_group_name") or ""),
                    "ad_format": ad_format,
                    "price_tier": str(group.get("price_tier") or ""),
                    **{key: 0.0 for key in metric_keys},
                    "invalid_impress": 0.0,
                    "invalid_click": 0.0,
                },
            )
            daily_format = format_daily.setdefault(ad_format, {}).setdefault(
                date_key,
                {
                    "date": date_key,
                    "ad_format": ad_format,
                    **{key: 0.0 for key in metric_keys},
                    "invalid_impress": 0.0,
                    "invalid_click": 0.0,
                },
            )
            placement = placements_by_group.setdefault(group_id, {}).setdefault(
                zone_id,
                {
                    "zone_id": zone_id,
                    "zone_name": str(row.get("zone_name") or ""),
                    "zone_group_id": group_id,
                    "zone_group_name": str(group.get("zone_group_name") or ""),
                    "ad_format": ad_format,
                    "price_tier": str(group.get("price_tier") or ""),
                    **{key: 0.0 for key in metric_keys},
                    "invalid_impress": 0.0,
                    "invalid_click": 0.0,
                },
            )
            daily_placement = placement_daily.setdefault(group_id, {}).setdefault(zone_id, {}).setdefault(
                date_key,
                {
                    "date": date_key,
                    "zone_id": zone_id,
                    "zone_name": str(row.get("zone_name") or ""),
                    "zone_group_id": group_id,
                    "zone_group_name": str(group.get("zone_group_name") or ""),
                    "ad_format": ad_format,
                    "price_tier": str(group.get("price_tier") or ""),
                    **{key: 0.0 for key in metric_keys},
                    "invalid_impress": 0.0,
                    "invalid_click": 0.0,
                },
            )
            for key in metric_keys:
                value = float(row.get(key) or 0.0)
                total[key] += value
                group[key] = float(group.get(key) or 0.0) + value
                fmt[key] = float(fmt.get(key) or 0.0) + value
                placement[key] = float(placement.get(key) or 0.0) + value
                daily_group[key] = float(daily_group.get(key) or 0.0) + value
                daily_format[key] = float(daily_format.get(key) or 0.0) + value
                daily_placement[key] = float(daily_placement.get(key) or 0.0) + value
            for key in ("invalid_impress", "invalid_click"):
                value = float(row.get(key) or 0.0)
                group[key] = float(group.get(key) or 0.0) + value
                fmt[key] = float(fmt.get(key) or 0.0) + value
                placement[key] = float(placement.get(key) or 0.0) + value
                daily_group[key] = float(daily_group.get(key) or 0.0) + value
                daily_format[key] = float(daily_format.get(key) or 0.0) + value
                daily_placement[key] = float(daily_placement.get(key) or 0.0) + value

        def enrich(item: dict[str, object]) -> dict[str, object]:
            impress = float(item.get("impress") or 0.0)
            click = float(item.get("click") or 0.0)
            profit = float(item.get("profit") or 0.0)
            advertiser_mu = float(item.get("advertiser_mu") or 0.0)
            item["ctr"] = round((click / impress) * 100, 6) if impress else 0.0
            item["ecpm"] = round((advertiser_mu / impress) * 1000, 6) if impress else 0.0
            item["ecpc"] = round(advertiser_mu / click, 6) if click else 0.0
            item["dsp_cpm"] = round((advertiser_mu / impress) * 1000, 6) if impress else 0.0
            item["dsp_cpc"] = round(advertiser_mu / click, 6) if click else 0.0
            return item

        def compact_daily(item: dict[str, object]) -> dict[str, object]:
            enriched = enrich(dict(item))
            return {key: enriched.get(key, 0.0) for key in SSP_AD_GROUP_METRICS}

        def avg_metrics(daily_items: dict[str, dict[str, object]], recent_dates: list[str]) -> dict[str, object]:
            base = {key: 0.0 for key in metric_keys}
            active_dates = 0
            for date_key in recent_dates:
                daily_item = daily_items.get(date_key)
                if not daily_item:
                    continue
                active_dates += 1
                for key in metric_keys:
                    base[key] += float(daily_item.get(key) or 0.0)
            divisor = max(1, active_dates)
            for key in metric_keys:
                base[key] = base[key] / divisor
            return compact_daily(base)

        def has_metric_anomaly(
            daily_metrics: dict[str, object] | None,
            average_metrics: dict[str, object] | None,
        ) -> bool:
            if not daily_metrics or not average_metrics:
                return False
            for metric in ("ctr", "ecpm", "ecpc"):
                baseline = float(average_metrics.get(metric) or 0.0)
                value = float(daily_metrics.get(metric) or 0.0)
                if baseline <= 0:
                    continue
                delta = (value - baseline) / baseline
                if abs(delta) <= 0.05:
                    continue
                lower_is_better = metric in {"ecpm", "ecpc"}
                if (lower_is_better and delta > 0) or (not lower_is_better and delta < 0):
                    return True
            return False

        summary = enrich({"label": "全部", **total})
        group_rows = [enrich(item) for item in by_group.values()]
        format_rows = [enrich(item) for item in by_format.values()]
        recent_dates = list(reversed(all_dates[-7:]))
        compare_date = recent_dates[0] if recent_dates else ""

        for item in format_rows:
            ad_format = str(item.get("ad_format") or "")
            item["zone_group_name"] = ad_format
            item["price_tier"] = "全部"
            item["daily_metrics"] = {
                date_key: compact_daily(daily_item)
                for date_key, daily_item in format_daily.get(ad_format, {}).items()
            }
            item["avg_metrics"] = avg_metrics(format_daily.get(ad_format, {}), recent_dates)
            item["status"] = "alert" if has_metric_anomaly(
                item["daily_metrics"].get(compare_date) if isinstance(item.get("daily_metrics"), dict) else None,
                item.get("avg_metrics") if isinstance(item.get("avg_metrics"), dict) else None,
            ) else "ok"
            item["reasons"] = []

        for item in group_rows:
            group_id = int(item.get("zone_group_id") or 0)
            item["daily_metrics"] = {
                date_key: compact_daily(daily_item)
                for date_key, daily_item in group_daily.get(group_id, {}).items()
            }
            item["avg_metrics"] = avg_metrics(group_daily.get(group_id, {}), recent_dates)
            item["status"] = "alert" if has_metric_anomaly(
                item["daily_metrics"].get(compare_date) if isinstance(item.get("daily_metrics"), dict) else None,
                item.get("avg_metrics") if isinstance(item.get("avg_metrics"), dict) else None,
            ) else "ok"
            item["reasons"] = []
            item["latest_run"] = latest_by_group.get(int(item.get("zone_group_id") or 0))

        group_rows.sort(
            key=lambda item: (
                str(item.get("ad_format") or ""),
                {"高": 0, "中": 1, "低": 2}.get(str(item.get("price_tier") or ""), 9),
            )
        )
        format_rows.sort(key=lambda item: str(item.get("ad_format") or ""))

        placement_groups: dict[str, list[dict[str, object]]] = {}
        for group_id, placements in placements_by_group.items():
            group_key = str(group_id)
            enriched = [enrich(item) for item in placements.values()]
            enriched.sort(key=lambda item: float(item.get("request") or 0.0), reverse=True)
            for placement in enriched:
                zone_id = int(placement.get("zone_id") or 0)
                placement["daily_metrics"] = {
                    date_key: compact_daily(daily_item)
                    for date_key, daily_item in placement_daily.get(group_id, {}).get(zone_id, {}).items()
                }
                placement["avg_metrics"] = avg_metrics(
                    placement_daily.get(group_id, {}).get(zone_id, {}),
                    recent_dates,
                )
                placement["status"] = "alert" if has_metric_anomaly(
                    placement["daily_metrics"].get(compare_date) if isinstance(placement.get("daily_metrics"), dict) else None,
                    placement.get("avg_metrics") if isinstance(placement.get("avg_metrics"), dict) else None,
                ) else "ok"
                placement["reasons"] = []
            important_alerts = [
                item
                for item in enriched
                if item.get("status") == "alert" and float(item.get("request") or 0.0) >= 100.0
            ][:10]
            limited_by_zone: dict[int, dict[str, object]] = {}
            for item in [*enriched[:20], *important_alerts]:
                limited_by_zone[int(item.get("zone_id") or 0)] = item
            placement_groups[group_key] = sorted(
                limited_by_zone.values(),
                key=lambda item: float(item.get("request") or 0.0),
                reverse=True,
            )

        return {
            "start_day": start_day,
            "end_day": end_day,
            "catalog": catalog,
            "formats": sorted({str(item["format"]) for item in catalog}),
            "metrics": list(SSP_AD_GROUP_METRICS),
            "default_metric": "ecpc",
            "summary": summary,
            "groups": group_rows,
            "format_summary": format_rows,
            "placements_by_group": placement_groups,
            "date_keys_desc": list(reversed(all_dates[-30:])),
            "latest_runs": latest_runs,
            "row_count": len(rows),
            "group_count": len(group_rows),
        }

    def fetch_dsp_api(
        self,
        *,
        start_day: str,
        end_day: str,
        template_version: str,
        rule_version: str,
        email: str | None = None,
        password: str | None = None,
        scope_check_url: str | None = None,
        api_base_url: str | None = None,
        auth_decrypt_key: str | None = None,
        service_id: int | None = None,
        source_name: str | None = None,
        timeout_seconds: int | None = None,
    ) -> dict:
        settings = resolve_dsp_api_settings(
            email=email,
            password=password,
            scope_check_url=scope_check_url,
            api_base_url=api_base_url,
            auth_decrypt_key=auth_decrypt_key,
            service_id=service_id,
            source_name=source_name,
            timeout_seconds=timeout_seconds,
        )
        bundle = DspApiClient(settings).fetch_report_bundle(start_day=start_day, end_day=end_day)
        rows = normalize_dsp_report_rows(
            [row for row in bundle["rows"] if isinstance(row, dict)],
            source_name=settings.source_name,
        )
        normalized_rows = self._field_contract.validate_and_normalize_save_rows(rows)

        with self.repo.connect() as conn:
            self.repo.resolve_trace_binding(conn, "dsp", template_version, rule_version)
            existing_rows = self.repo.read_canonical_rows_in_tx(conn, "dsp")
            requested_days = _inclusive_day_texts(start_day, end_day)
            preserved_rows = [row for row in existing_rows if _canonical_day_text(row) not in requested_days]
            merged_rows = [*preserved_rows, *normalized_rows]
            merged_rows.sort(
                key=lambda row: (
                    str(row.get("日期時間") or ""),
                    str(row.get("經銷商") or ""),
                    str(row.get("訂單") or ""),
                    str(row.get("素材") or ""),
                )
            )
            written = self.repo.save_canonical_rows(conn, "dsp", merged_rows)
            trace = self.repo.build_trace_meta(conn, "dsp", template_version, rule_version)
            model = bundle.get("model") if isinstance(bundle.get("model"), dict) else {}
            detail = {
                "start_day": start_day,
                "end_day": end_day,
                "row_count": len(normalized_rows),
                "total_row_count": written,
                "fetched_row_count": len(normalized_rows),
                "retained_row_count": len(preserved_rows),
                "replaced_day_count": len(requested_days),
                "records_total": int(bundle.get("records_total") or 0),
                "job_id": str(bundle.get("job_id") or ""),
                "job_ids": list(bundle.get("job_ids") or []),
                "chunk_mode": str(bundle.get("chunk_mode") or "single"),
                "chunk_days": int(bundle.get("chunk_days") or 1),
                "service_id": int((bundle.get("auth") or {}).get("service_id") or 0),
                "source_name": settings.source_name,
                "login_user_id": int((((bundle.get("auth") or {}).get("user") or {}) if isinstance((bundle.get("auth") or {}).get("user"), dict) else {}).get("id") or 0),
                "login_email": str((((bundle.get("auth") or {}).get("user") or {}) if isinstance((bundle.get("auth") or {}).get("user"), dict) else {}).get("email") or ""),
                "job_status": int(model.get("status") or 0),
            }
            run_id = self.repo.insert_run_log(
                conn,
                run_type="fetch_dsp_api",
                workflow="dsp",
                status="ok",
                trace=trace,
                detail=detail,
            )
            self.repo.append_audit_event(
                conn,
                event_type="fetch_dsp_api",
                scope="service",
                status="ok",
                payload={"run_id": run_id, **detail},
            )
            conn.commit()

        return {
            "status": "ok",
            "workflow": "dsp",
            "run_id": run_id,
            "start_day": start_day,
            "end_day": end_day,
            "row_count": len(normalized_rows),
            "total_row_count": written,
            "fetched_row_count": len(normalized_rows),
            "retained_row_count": len(preserved_rows),
            "replaced_day_count": len(requested_days),
            "records_total": int(bundle.get("records_total") or 0),
            "job_id": str(bundle.get("job_id") or ""),
            "job_ids": list(bundle.get("job_ids") or []),
            "chunk_mode": str(bundle.get("chunk_mode") or "single"),
            "chunk_days": int(bundle.get("chunk_days") or 1),
            "service_id": int((bundle.get("auth") or {}).get("service_id") or 0),
            "login_user_id": int((((bundle.get("auth") or {}).get("user") or {}) if isinstance((bundle.get("auth") or {}).get("user"), dict) else {}).get("id") or 0),
            "login_email": str((((bundle.get("auth") or {}).get("user") or {}) if isinstance((bundle.get("auth") or {}).get("user"), dict) else {}).get("email") or ""),
            "source_name": settings.source_name,
        }

    def mark_tab4_delivery(
        self,
        *,
        workflow: str,
        main_tab: str,
        sub_tab: str,
        template_version: str,
        rule_version: str,
        week_start: str | None = None,
        week_end: str | None = None,
    ) -> dict:
        if workflow != "dsp":
            raise ValueError("tab4_delivery only supports dsp workflow")
        if main_tab != "dsp_tab3" or sub_tab != "pivot":
            raise ValueError("tab4_delivery must be triggered from dsp_tab3/pivot")
        resolved_week_start, resolved_week_end = self._resolve_export_period(
            week_start=week_start,
            week_end=week_end,
        )
        with self.repo.connect() as conn:
            self.repo.resolve_trace_binding(conn, workflow, template_version, rule_version)
            rows = self.repo.read_canonical_rows_in_tx(conn, workflow)
            trace = self.repo.build_trace_meta(conn, workflow, template_version, rule_version)
            run_id = self.repo.insert_run_log(
                conn,
                run_type="tab4_delivery",
                workflow=workflow,
                status="ok",
                trace=trace,
                detail={
                    "source": "pivot_handoff",
                    "main_tab": main_tab,
                    "sub_tab": sub_tab,
                    "row_count": len(rows),
                    "delivery_snapshot_token": trace.canonical_token,
                    "delivery_source_db_hash": trace.source_db_hash,
                    "week_start": resolved_week_start,
                    "week_end": resolved_week_end,
                },
            )
            self.repo.append_audit_event(
                conn,
                event_type="tab4_delivery",
                scope="service",
                status="ok",
                payload={
                    "workflow": workflow,
                    "run_id": run_id,
                    "template_version": template_version,
                    "rule_version": rule_version,
                    "canonical_token": trace.canonical_token,
                    "row_count": len(rows),
                    "main_tab": main_tab,
                    "sub_tab": sub_tab,
                    "week_start": resolved_week_start,
                    "week_end": resolved_week_end,
                },
            )
            state = self.repo.get_tab4_delivery_state(conn, workflow)
        out = {
            "run_id": run_id,
            "ready": bool(state.get("ready")),
            "reason": str(state.get("reason") or ""),
            "updated_at": str(state.get("updated_at") or ""),
            "delivery_snapshot_token": str(state.get("delivery_snapshot_token") or ""),
            "delivery_row_count": int(state.get("delivery_row_count") or 0),
            "week_start": str(state.get("delivery_week_start") or ""),
            "week_end": str(state.get("delivery_week_end") or ""),
        }
        if self._feature_flags.get("enable_test_hooks", False):
            out["test_hooks_enabled"] = True
        return out

    def validate_dsp_export_request(
        self,
        *,
        workflow: str,
        main_tab: str,
        sub_tab: str,
        template_version: str,
        rule_version: str,
        week_start: str | None = None,
        week_end: str | None = None,
    ) -> dict[str, str]:
        if workflow != "dsp":
            raise ValueError("dsp export gate only supports dsp workflow")
        if main_tab != "dsp_tab4":
            raise PermissionError("dsp export must be triggered from dsp_tab4")
        if sub_tab not in {"overview"}:
            raise PermissionError("dsp export sub_tab out of scope")
        with self.repo.connect() as conn:
            self.repo.resolve_trace_binding(conn, workflow, template_version, rule_version)
            delivery_state = self.repo.assert_tab4_delivery_ready(conn, workflow)
            trace = self.repo.build_trace_meta(conn, workflow, template_version, rule_version)
        resolved_week_start, resolved_week_end = self._resolve_export_period(
            week_start=week_start,
            week_end=week_end,
        )
        delivery_snapshot_token = str(delivery_state.get("delivery_snapshot_token") or "")
        if not delivery_snapshot_token:
            raise PermissionError("tab4 delivery snapshot token missing")
        if delivery_snapshot_token != trace.canonical_token:
            raise PermissionError("tab4 delivery snapshot mismatch with canonical")
        delivery_week_start = str(delivery_state.get("delivery_week_start") or "")
        delivery_week_end = str(delivery_state.get("delivery_week_end") or "")
        if delivery_week_start != resolved_week_start or delivery_week_end != resolved_week_end:
            raise PermissionError("tab4 delivery period mismatch with export period")
        return {
            "delivery_snapshot_token": delivery_snapshot_token,
            "delivery_run_id": str(delivery_state.get("last_delivery_run_id") or ""),
        }

    def export(
        self,
        *,
        workflow: str,
        artifact_root: Path,
        template_version: str,
        rule_version: str,
        main_tab: str | None = None,
        sub_tab: str | None = None,
        week_start: str | None = None,
        week_end: str | None = None,
        delivery_snapshot_token: str | None = None,
        delivery_run_id: str | None = None,
    ) -> dict:
        artifact_root.mkdir(parents=True, exist_ok=True)
        resolved_week_start, resolved_week_end = self._resolve_export_period(
            week_start=week_start,
            week_end=week_end,
        )
        if workflow == "dsp":
            artifact_name = self._build_dsp_export_filename(resolved_week_start, resolved_week_end)
        else:
            artifact_name = f"{workflow}_export.xlsx"
        artifact_path = artifact_root / artifact_name
        with self.repo.connect() as conn:
            self.repo.resolve_trace_binding(conn, workflow, template_version, rule_version)
            export_rows: list[dict]
            hydrate_rows: list[dict]
            export_columns: list[str]
            if workflow == "ssp":
                snapshot = self._resolve_ssp_effective_snapshot_in_tx(conn)
                export_rows = list(snapshot["rows"])
                hydrate_rows = export_rows
                export_columns = list(snapshot["field_names"])
            else:
                workflow_rows = self.repo.read_canonical_rows_in_tx(conn, workflow)
                export_rows = workflow_rows
                hydrate_rows = workflow_rows
                if workflow == "dsp":
                    export_rows = self._filter_rows_by_period(
                        workflow_rows,
                        week_start=resolved_week_start,
                        week_end=resolved_week_end,
                    )
                export_columns = list(self.repo.canonical_columns)
            trace = self.repo.build_trace_meta(conn, workflow, template_version, rule_version)
            export_delivery_snapshot_token = str(delivery_snapshot_token or "")
            export_delivery_run_id = str(delivery_run_id or "")
            try:
                if workflow == "dsp":
                    template_path = self._resolve_dsp_export_template_path(
                        week_start=resolved_week_start,
                        week_end=resolved_week_end,
                    )
                    self._hydrate_dsp_template_workbook(
                        template_path=template_path,
                        artifact_path=artifact_path,
                        rows=hydrate_rows,
                        week_start=resolved_week_start,
                        week_end=resolved_week_end,
                    )
                else:
                    wb = Workbook(write_only=True)
                    try:
                        ws_data = wb.create_sheet("canonical_data")
                        ws_data.append(export_columns)
                        for row in export_rows:
                            ws_data.append([_sanitize_workbook_cell_value(row.get(col, "")) for col in export_columns])

                        ws_meta = wb.create_sheet("metadata")
                        ws_meta.append(["key", "value"])
                        ws_meta.append(["workflow", workflow])
                        ws_meta.append(["template_version", template_version])
                        ws_meta.append(["rule_version", rule_version])
                        ws_meta.append(["source_db_hash", trace.source_db_hash])
                        ws_meta.append(["canonical_token", trace.canonical_token])
                        ws_meta.append(["week_start", resolved_week_start])
                        ws_meta.append(["week_end", resolved_week_end])
                        wb.save(artifact_path)
                    finally:
                        wb.close()
                # 讀回一次，確認檔案可開啟，避免留半壞檔案。
                verify_wb = load_workbook(artifact_path, read_only=True, data_only=True)
                verify_wb.close()
                checksum = hashlib.sha256(artifact_path.read_bytes()).hexdigest()
                trace = replace(trace, artifact_checksum=checksum)
                run_id = self.repo.insert_run_log(
                    conn,
                    run_type="export",
                    workflow=workflow,
                    status="ok",
                    trace=trace,
                    detail={
                        "artifact_path": str(artifact_path),
                        "row_count": len(export_rows),
                        "week_start": resolved_week_start,
                        "week_end": resolved_week_end,
                        "delivery_snapshot_token": export_delivery_snapshot_token,
                        "delivery_run_id": export_delivery_run_id,
                    },
                )
                self.repo.insert_publish_run(
                    conn,
                    run_id,
                    workflow,
                    artifact_path,
                    trace,
                    status="ok",
                    week_start=resolved_week_start,
                    week_end=resolved_week_end,
                )
                self.repo.insert_evidence(conn, run_id, artifact_path, checksum, status="ok")
                marker = self._trace_marker(workflow=workflow, run_type="export", run_id=run_id)
                audit_payload = {
                    "workflow": workflow,
                    "run_id": run_id,
                    "template_version": template_version,
                    "rule_version": rule_version,
                    "canonical_token": trace.canonical_token,
                    "artifact_path": str(artifact_path),
                    "artifact_checksum": checksum,
                    "row_count": len(export_rows),
                    "week_start": resolved_week_start,
                    "week_end": resolved_week_end,
                    "delivery_snapshot_token": export_delivery_snapshot_token,
                    "delivery_run_id": export_delivery_run_id,
                    **self._extra_debug_payload(),
                }
                if marker:
                    audit_payload["trace_marker"] = marker
                self.repo.append_audit_event(
                    conn,
                    event_type="export",
                    scope="service",
                    status="ok",
                    payload=audit_payload,
                )
            except Exception:
                if artifact_path.exists():
                    artifact_path.unlink()
                raise

        out = {
            "run_id": run_id,
            "artifact_path": str(artifact_path),
            "artifact_checksum": checksum,
            "row_count": len(export_rows),
            "week_start": resolved_week_start,
            "week_end": resolved_week_end,
        }
        if workflow == "dsp":
            out["delivery_snapshot_token"] = export_delivery_snapshot_token
            out["delivery_run_id"] = export_delivery_run_id
        if marker:
            out["trace_marker"] = marker
        if self._feature_flags.get("enable_test_hooks", False):
            out["test_hooks_enabled"] = True
        return out
