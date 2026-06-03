from __future__ import annotations

import io
import json
import hashlib
import sqlite3
import tempfile
import threading
import unittest
from datetime import date, timedelta
from unittest.mock import patch
from urllib.error import HTTPError
from urllib.request import Request, urlopen
from urllib.parse import parse_qs, urlencode, urlparse
from http.server import ThreadingHTTPServer
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

import app.ui_shell as ui_shell_module
from app.main import run_cli
from app.ui_shell import UiContext, UiRequestHandler, collect_runtime_status, collect_workflow_frame, dispatch_action
from infra.dsp_api import DspApiSettings
from infra.ssp_api import SspApiSettings
from infra.sqlite.bootstrap import AcceptanceGateError, build_config
from infra.sqlite.repository import SQLiteRepository


class UiShellTests(unittest.TestCase):
    @staticmethod
    def _same_cell_value(left: object, right: object, *, tol: float = 1e-6) -> bool:
        if isinstance(left, (int, float)) and isinstance(right, (int, float)):
            lf = float(left)
            rf = float(right)
            scale = max(1.0, abs(lf), abs(rf))
            return abs(lf - rf) <= tol * scale
        return left == right

    def _fmt_num(self, value: int | float) -> str:
        return f"{float(value):,.2f}"

    def _week_start(self, current: date | None = None) -> date:
        current = current or date.today()
        return current - timedelta(days=current.weekday())

    def _dsp_bucket_date(self, *, weeks_ago: int = 2, day_offset: int = 4) -> str:
        return (self._week_start() - timedelta(days=weeks_ago * 7) + timedelta(days=day_offset)).isoformat()

    def _dsp_bucket_datetime(self, *, weeks_ago: int = 2, day_offset: int = 4) -> str:
        return f"{self._dsp_bucket_date(weeks_ago=weeks_ago, day_offset=day_offset)} 00:00:00"

    def _full_row(self, **overrides: object) -> dict:
        row = {
            "日期時間": "2026-05-01 00:00:00",
            "經銷商": "A",
            "訂單": "O1",
            "素材": "C1",
            "廣告形式": "Banner",
            "尺寸": "300x250",
            "素材樣板": "tplA",
            "執行金額": 10.0,
            "系統營收": 12.5,
            "媒體費用": 8.0,
            "原始經銷商": "A",
            "原始廣告形式": "Banner",
            "最終經銷商": "A1",
            "規則命中_經銷商": "r1",
            "最終來源_經銷商": "rule",
            "分類層級B": "B1",
            "分類層級C": "C1",
            "分類層級D": "D1",
            "最終廣告形式": "Banner",
            "規則命中_廣告形式": "r2",
            "最終來源_廣告形式": "rule",
        }
        row.update(overrides)
        return row

    def _dsp_mutable_cells(self) -> dict[str, set[str]]:
        mutable: dict[str, set[str]] = {
            "mF投資量_總表": {"A1"},
            "各經銷商明細": {f"A{row_idx}" for row_idx in (5, 24, 44, 63, 82)},
            "北流進單追蹤": {"A1"},
        }
        detail_cells = mutable["各經銷商明細"]
        month_cols = [5 + idx * 2 for idx in range(12)]
        for row_idx in (
            7, 8, 9, 10, 11, 12, 13,
            26, 27, 28, 29, 30, 31, 32,
            46, 47, 48, 49, 50, 51, 52,
            65, 66, 67, 68, 69, 70, 71,
            84, 85, 86, 87, 88, 89, 90,
        ):
            for col_idx in month_cols:
                detail_cells.add(f"{get_column_letter(col_idx)}{row_idx}")
        return mutable

    def _assert_dsp_export_template_parity(self, template_path: Path, export_path: Path) -> None:
        template_wb = load_workbook(template_path, data_only=False)
        export_wb = load_workbook(export_path, data_only=False)
        mutable_cells = self._dsp_mutable_cells()
        try:
            self.assertEqual(template_wb.sheetnames, export_wb.sheetnames)
            for sheet_name in template_wb.sheetnames:
                ws_t = template_wb[sheet_name]
                ws_e = export_wb[sheet_name]
                self.assertEqual(str(ws_t.sheet_state), str(ws_e.sheet_state), sheet_name)
                self.assertEqual(str(ws_t.freeze_panes or ""), str(ws_e.freeze_panes or ""), sheet_name)
                self.assertEqual(
                    sorted(str(item) for item in ws_t.merged_cells.ranges),
                    sorted(str(item) for item in ws_e.merged_cells.ranges),
                    sheet_name,
                )
                self.assertEqual(repr(ws_t.sheet_properties.tabColor), repr(ws_e.sheet_properties.tabColor), sheet_name)
                self.assertEqual(
                    sorted(idx for idx, dim in ws_t.row_dimensions.items() if bool(getattr(dim, "hidden", False))),
                    sorted(idx for idx, dim in ws_e.row_dimensions.items() if bool(getattr(dim, "hidden", False))),
                    sheet_name,
                )
                self.assertEqual(
                    sorted(name for name, dim in ws_t.column_dimensions.items() if bool(getattr(dim, "hidden", False))),
                    sorted(name for name, dim in ws_e.column_dimensions.items() if bool(getattr(dim, "hidden", False))),
                    sheet_name,
                )
                self.assertEqual(
                    {name: float(dim.width) for name, dim in ws_t.column_dimensions.items() if dim.width is not None},
                    {name: float(dim.width) for name, dim in ws_e.column_dimensions.items() if dim.width is not None},
                    sheet_name,
                )

                t_cells = getattr(ws_t, "_cells", {})
                e_cells = getattr(ws_e, "_cells", {})
                for coord in sorted(set(t_cells.keys()) | set(e_cells.keys())):
                    cell_t = ws_t.cell(row=coord[0], column=coord[1])
                    cell_e = ws_e.cell(row=coord[0], column=coord[1])
                    coord_name = cell_t.coordinate
                    self.assertEqual(repr(cell_t.font), repr(cell_e.font), f"{sheet_name}!{coord_name} font")
                    self.assertEqual(repr(cell_t.fill), repr(cell_e.fill), f"{sheet_name}!{coord_name} fill")
                    self.assertEqual(repr(cell_t.border), repr(cell_e.border), f"{sheet_name}!{coord_name} border")
                    self.assertEqual(repr(cell_t.alignment), repr(cell_e.alignment), f"{sheet_name}!{coord_name} alignment")
                    self.assertEqual(repr(cell_t.protection), repr(cell_e.protection), f"{sheet_name}!{coord_name} protection")
                    self.assertEqual(
                        str(cell_t.number_format or ""),
                        str(cell_e.number_format or ""),
                        f"{sheet_name}!{coord_name} number_format",
                    )
                    t_formula = isinstance(cell_t.value, str) and cell_t.value.startswith("=")
                    e_formula = isinstance(cell_e.value, str) and cell_e.value.startswith("=")
                    self.assertEqual(t_formula, e_formula, f"{sheet_name}!{coord_name} formula marker")
                    if t_formula:
                        self.assertEqual(str(cell_t.value), str(cell_e.value), f"{sheet_name}!{coord_name} formula text")
                    if coord_name in mutable_cells.get(sheet_name, set()):
                        continue
                    self.assertTrue(
                        self._same_cell_value(cell_t.value, cell_e.value),
                        f"{sheet_name}!{coord_name} static cell",
                    )
        finally:
            export_wb.close()
            template_wb.close()

    def _make_project(self, root: Path, *, include_ssp_template: bool = True) -> None:
        src = Path(__file__).resolve().parents[1]
        (root / "migrations").mkdir(parents=True, exist_ok=True)
        (root / "templates").mkdir(parents=True, exist_ok=True)
        (root / "contracts").mkdir(parents=True, exist_ok=True)
        (root / "data_seed" / "templates_rules_mapping").mkdir(parents=True, exist_ok=True)
        (root / "data_seed_test" / "templates_rules_mapping").mkdir(parents=True, exist_ok=True)
        (root / "migrations" / "0001_initial.sql").write_text((src / "migrations" / "0001_initial.sql").read_text(encoding="utf-8"), encoding="utf-8")
        (root / "templates" / "template_registry.seed.json").write_text((src / "templates" / "template_registry.seed.json").read_text(encoding="utf-8"), encoding="utf-8")
        (root / "templates" / "ruleset.seed.json").write_text((src / "templates" / "ruleset.seed.json").read_text(encoding="utf-8"), encoding="utf-8")
        self._write_dsp_tab4_template(root / "templates" / "dsp_tab4_template.xlsx")
        if include_ssp_template:
            self._write_ssp_media_template(root / "templates" / "ssp_template.xlsx")
        sidecar_src = src / "templates" / "dsp_tab4_template.xlsx.period.json"
        if sidecar_src.exists():
            (root / "templates" / "dsp_tab4_template.xlsx.period.json").write_text(
                sidecar_src.read_text(encoding="utf-8"),
                encoding="utf-8",
            )
        (root / "contracts" / "fields_contract.json").write_text((src / "contracts" / "fields_contract.json").read_text(encoding="utf-8"), encoding="utf-8")
        (root / "bootstrap.manifest.json").write_text((src / "bootstrap.manifest.json").read_text(encoding="utf-8"), encoding="utf-8")
        (root / "bootstrap.test.manifest.json").write_text((src / "bootstrap.test.manifest.json").read_text(encoding="utf-8"), encoding="utf-8")
        default_group_overrides = {
            "蓋板": [{"placement_id": 8435, "placement_name": "MW_蓋版_COOL", "remark": "prod", "media_target": 1000}],
            "置底": [{"placement_id": 17236, "placement_name": "MW_置底創意_UDN聯合新聞", "remark": "", "media_target": 1200}],
            "置底展開": [{"placement_id": 22218, "placement_name": "MW_置底(展開)_上報", "remark": "", "media_target": 800}],
            "文中300x250": [{"placement_id": 16980, "placement_name": "MW_文中創意300x250_大人物", "remark": "", "media_target": 900}],
            "文中320x480": [{"placement_id": 16977, "placement_name": "MW_文中創意320x480_大人物", "remark": "", "media_target": 700}],
        }
        for rel_path in (
            root / "data_seed" / "templates_rules_mapping" / "group_overrides.json",
            root / "data_seed_test" / "templates_rules_mapping" / "group_overrides.json",
        ):
            rel_path.write_text(json.dumps(default_group_overrides, ensure_ascii=False), encoding="utf-8")

    def _write_dsp_tab4_template(self, path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        try:
            ws_hidden_a = wb.active
            ws_hidden_a.title = "2025年_MF_合作績效統計總表"
            ws_hidden_a.sheet_state = "hidden"
            ws_hidden_a.freeze_panes = "C1"
            ws_hidden_a["A1"] = 2025

            ws_hidden_b = wb.create_sheet("2025_外部+行政_合作績效統計總表 ")
            ws_hidden_b.sheet_state = "hidden"
            ws_hidden_b.freeze_panes = "C1"
            ws_hidden_b["A1"] = "外部經銷商"

            ws_summary = wb.create_sheet("mF投資量_總表")
            ws_summary.freeze_panes = "M1"
            ws_summary.merge_cells("A1:D1")
            ws_summary["A1"] = 2026
            ws_summary["A2"] = "DSP投資額 總計"
            ws_summary["E2"] = "=SUM(E3:E8)"
            ws_summary["F2"] = "=SUM(F3:F8)"
            ws_summary["M2"] = "=SUM(M3:M8)"
            ws_summary["N2"] = "=SUM(N3:N8)"
            ws_summary["AC2"] = "=SUM(AC3:AC8)"
            ws_summary["AD2"] = "=SUM(AD3:AD8)"
            ws_summary["E3"] = "=各經銷商明細!E$6"
            ws_summary["F3"] = "=E3/E$2"
            ws_summary["M3"] = "=各經銷商明細!M$6"
            ws_summary["N3"] = "=M3/M$2"
            ws_summary["AC3"] = "=E3+G3+I3+K3+M3+O3+Q3+S3+U3+W3+Y3+AA3"
            ws_summary["AD3"] = "=AC3/AC$2"
            for idx in range(16, 24):
                ws_summary.row_dimensions[idx].hidden = True

            ws_detail = wb.create_sheet("各經銷商明細")
            ws_detail.freeze_panes = "U1"
            ws_detail.merge_cells("A2:D2")
            ws_detail["A2"] = "全體經銷 總投資量目標 & 達成率 (含北流)"
            ws_detail["A5"] = 2026
            for row_idx in (5, 24, 44, 63, 82):
                ws_detail[f"A{row_idx}"] = 2026
            for row_idx in (6, 25, 45, 64, 83):
                ws_detail[f"E{row_idx}"] = f"=SUM(E{row_idx + 1}:E{row_idx + 7})"
                ws_detail[f"F{row_idx}"] = f"=SUM(F{row_idx + 1}:F{row_idx + 7})"
                ws_detail[f"M{row_idx}"] = f"=SUM(M{row_idx + 1}:M{row_idx + 7})"
                ws_detail[f"N{row_idx}"] = f"=SUM(N{row_idx + 1}:N{row_idx + 7})"
                ws_detail[f"AC{row_idx}"] = f"=SUM(AC{row_idx + 1}:AC{row_idx + 7})"
                ws_detail[f"AD{row_idx}"] = f"=SUM(AD{row_idx + 1}:AD{row_idx + 7})"
            for row_idx in (7, 26, 46, 65, 84):
                ws_detail[f"F{row_idx}"] = f"=E{row_idx}/E$6"
                ws_detail[f"N{row_idx}"] = f"=M{row_idx}/M$6"
                ws_detail[f"AC{row_idx}"] = f"=E{row_idx}+G{row_idx}+I{row_idx}+K{row_idx}+M{row_idx}+O{row_idx}+Q{row_idx}+S{row_idx}+U{row_idx}+W{row_idx}+Y{row_idx}+AA{row_idx}"
                ws_detail[f"AD{row_idx}"] = f"=AC{row_idx}/AC$6"
            ws_detail["AB2"] = "=SUM(AA6,AA25,AA45,AA64)/AA2"
            ws_detail["AB3"] = "=AA6/AA3"

            ws_tracking = wb.create_sheet("北流進單追蹤")
            ws_tracking["A1"] = "2026年5月份_北流進單狀態"
            ws_tracking["K2"] = None
            ws_tracking.column_dimensions["I"].hidden = True
            ws_tracking.column_dimensions["J"].hidden = True

            wb.save(path)
        finally:
            wb.close()

    def _write_ssp_media_template(self, path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        template_rows = [
            ("蓋板", 8435, "MW_蓋版_COOL", "頭部媒體", True, "60-80%", 1000, "prod"),
            ("置底", 17236, "MW_置底創意_UDN聯合新聞", "新聞站", False, "55-75%", 1200, ""),
            ("置底展開", 22218, "MW_置底(展開)_上報", "討論站", True, "50-70%", 800, ""),
            ("文中300x250", 16980, "MW_文中創意300x250_大人物", "內容站", False, "45-65%", 900, ""),
            ("文中320x480", 16977, "MW_文中創意320x480_大人物", "內容站", True, "40-60%", 700, ""),
        ]
        try:
            ws_master = wb.active
            ws_master.title = "版位編號"
            ws_master["A1"] = "版位"
            for row_idx, (_, placement_id, _, _, _, _, _, _) in enumerate(template_rows, start=2):
                ws_master[f"A{row_idx}"] = placement_id

            for category, placement_id, placement_name, media_quality, need_call, target_fr, media_target, remark in template_rows:
                ws = wb.create_sheet(category)
                ws["A1"] = "Remark"
                ws["B1"] = "版位"
                ws["C1"] = "版位名稱"
                ws["D1"] = "媒體質量"
                ws["E1"] = "需喊量"
                ws["F1"] = "目標FR"
                ws["G1"] = "預估量(7-22點)"
                ws["A2"] = remark
                ws["B2"] = placement_id
                ws["C2"] = placement_name
                ws["D2"] = media_quality
                ws["E2"] = need_call
                ws["F2"] = target_fr
                ws["G2"] = media_target
            wb.save(path)
        finally:
            wb.close()

    def _ctx(self, root: Path, workflow: str = "dsp") -> UiContext:
        return UiContext(
            root=root,
            runtime_env="prod",
            manifest_rel="bootstrap.manifest.json",
            workflow=workflow,
            template_version="v1",
            rule_version="v1",
            artifact_root=(root / "artifacts").resolve(),
        )

    def _run_cli_json(self, argv: list[str]) -> tuple[int, dict]:
        stdout = io.StringIO()
        with patch("sys.stdout", new=stdout):
            code = run_cli(argv)
        payload = json.loads(stdout.getvalue() or "{}")
        return code, payload

    def _mock_ssp_fetch_bundle(self) -> dict[str, object]:
        return {
            "auth": {
                "service_id": 14,
                "token": "ssp-service-token",
                "user": {"id": 2072, "email": "matt@clickforce.com.tw"},
            },
            "login": {"id": 2072, "email": "matt@clickforce.com.tw"},
            "report_id": 174425,
            "records_total": 1,
            "sum_row": {"request": 2885, "impress": 1386, "profit": 2.08},
            "rows": [
                {
                    "data_time": "2026-05-11 00:00:00",
                    "supplier_id": "1",
                    "supplierName": "域動測試",
                    "site_id": "784",
                    "siteName": "DEMO link",
                    "zone_id": "10230",
                    "zoneName": "DEMO LINK 專用",
                    "request": "2885",
                    "impress": "1386",
                    "click": "0",
                    "profit": "2.08",
                    "advertiser_mu": "8.32",
                }
            ],
        }

    def _mock_ssp_ad_group_fetch_bundle(self) -> dict[str, object]:
        bundle = self._mock_ssp_fetch_bundle()
        bundle["report_id"] = 274425
        bundle["records_total"] = 1
        bundle["sum_row"] = {"request": 2885, "impress": 1386, "profit": 2.08}
        return bundle

    def _mock_ssp_fetch_bundle_multi_day(self) -> dict[str, object]:
        return {
            "auth": {
                "service_id": 14,
                "token": "ssp-service-token",
                "user": {"id": 2072, "email": "matt@clickforce.com.tw"},
            },
            "login": {"id": 2072, "email": "matt@clickforce.com.tw"},
            "report_id": 174426,
            "report_ids": [174425, 174426],
            "records_total": 3,
            "chunk_mode": "daily",
            "chunk_days": 2,
            "sum_row": {"request": 3000, "impress": 1500, "profit": 4.0},
            "rows": [
                {
                    "data_time": "2026-05-10 00:00:00",
                    "supplier_id": "1",
                    "supplierName": "域動測試",
                    "site_id": "784",
                    "siteName": "DEMO link",
                    "zone_id": "10230",
                    "zoneName": "DEMO LINK 專用",
                    "request": "1000",
                    "impress": "500",
                    "click": "0",
                    "profit": "1.50",
                    "advertiser_mu": "4.10",
                },
                {
                    "data_time": "2026-05-10 01:00:00",
                    "supplier_id": "1",
                    "supplierName": "域動測試",
                    "site_id": "784",
                    "siteName": "DEMO link",
                    "zone_id": "10230",
                    "zoneName": "DEMO LINK 專用",
                    "request": "500",
                    "impress": "250",
                    "click": "0",
                    "profit": "0.50",
                    "advertiser_mu": "1.20",
                },
                {
                    "data_time": "2026-05-11 00:00:00",
                    "supplier_id": "1",
                    "supplierName": "域動測試",
                    "site_id": "784",
                    "siteName": "DEMO link",
                    "zone_id": "10230",
                    "zoneName": "DEMO LINK 專用",
                    "request": "1500",
                    "impress": "750",
                    "click": "0",
                    "profit": "2.00",
                    "advertiser_mu": "3.02",
                },
            ],
        }

    def _mock_dsp_fetch_bundle(self) -> dict[str, object]:
        return {
            "auth": {
                "service_id": 10,
                "token": "dsp-service-token",
                "user": {"id": 2072, "email": "matt@clickforce.com.tw"},
            },
            "job_id": "35cffe17660dad7fbdfb7080ffa2f1a6",
            "records_total": 1,
            "model": {"status": 1},
            "rows": [
                {
                    "data_time": "2026-05-10",
                    "distributor_id": "[台灣]域動行銷股份有限公司",
                    "campaign_id": "(42031)活動",
                    "creative_id": "(314928)0422_純蓋板",
                    "size_id": "純蓋板",
                    "content_type": "HTML/JS",
                    "campaign_mu": 10934.99,
                    "distributor_mu": 9000.11,
                    "advertiser_mu": 8000.22,
                }
            ],
        }

    def _make_seed_canonical_dsp_db(self, root: Path, *, seed_root: str = "data_seed") -> Path:
        seed_db = root / seed_root / "canonical" / "mdreport_dsp.sqlite"
        seed_db.parent.mkdir(parents=True, exist_ok=True)
        conn = sqlite3.connect(str(seed_db))
        try:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS dsp_rawdata (
                  row_order INTEGER PRIMARY KEY,
                  日期時間 TEXT NOT NULL DEFAULT '',
                  經銷商 TEXT NOT NULL DEFAULT '',
                  訂單 TEXT NOT NULL DEFAULT '',
                  素材 TEXT NOT NULL DEFAULT '',
                  廣告形式 TEXT NOT NULL DEFAULT '',
                  尺寸 TEXT NOT NULL DEFAULT '',
                  素材樣板 TEXT NOT NULL DEFAULT '',
                  執行金額 REAL NOT NULL DEFAULT 0.0,
                  系統營收 REAL NOT NULL DEFAULT 0.0,
                  媒體費用 REAL NOT NULL DEFAULT 0.0,
                  原始經銷商 TEXT NOT NULL DEFAULT '',
                  原始廣告形式 TEXT NOT NULL DEFAULT '',
                  最終經銷商 TEXT NOT NULL DEFAULT '',
                  規則命中_經銷商 TEXT NOT NULL DEFAULT '',
                  最終來源_經銷商 TEXT NOT NULL DEFAULT '',
                  分類層級B TEXT NOT NULL DEFAULT '',
                  分類層級C TEXT NOT NULL DEFAULT '',
                  分類層級D TEXT NOT NULL DEFAULT '',
                  最終廣告形式 TEXT NOT NULL DEFAULT '',
                  規則命中_廣告形式 TEXT NOT NULL DEFAULT '',
                  最終來源_廣告形式 TEXT NOT NULL DEFAULT '',
                  updated_at TEXT NOT NULL DEFAULT ''
                )
                """
            )
            conn.execute(
                """
                INSERT INTO dsp_rawdata(
                  row_order, 日期時間, 經銷商, 訂單, 素材, 廣告形式, 尺寸, 素材樣板, 執行金額, 系統營收, 媒體費用,
                  原始經銷商, 原始廣告形式, 最終經銷商, 規則命中_經銷商, 最終來源_經銷商,
                  分類層級B, 分類層級C, 分類層級D, 最終廣告形式, 規則命中_廣告形式, 最終來源_廣告形式, updated_at
                ) VALUES
                (0, '2026-05-01 00:00:00', 'A', 'O1', 'C1', 'Banner', '300x250', 'tplA', 10.0, 12.5, 8.0,
                 'A', 'Banner', 'API_PROMOTED_A1', 'r1', 'api', 'B1', 'C1', 'D1', 'Banner', 'r2', 'api', '2026-05-09T00:00:00')
                """
            )
            conn.commit()
        finally:
            conn.close()
        return seed_db

    def _make_seed_canonical_ssp_truth_db(self, root: Path, *, seed_root: str = "data_seed") -> Path:
        seed_db = root / seed_root / "canonical" / "mdreport.sqlite"
        seed_db.parent.mkdir(parents=True, exist_ok=True)
        conn = sqlite3.connect(str(seed_db))
        try:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS raw (
                  source TEXT,
                  ts TEXT,
                  date TEXT,
                  hour INTEGER,
                  placement_id INTEGER,
                  placement_name TEXT,
                  request INTEGER,
                  impression INTEGER,
                  clicks INTEGER,
                  revenue REAL,
                  dsp_amount REAL,
                  order_id TEXT,
                  order_name TEXT,
                  supplier_id INTEGER,
                  supplier_name TEXT,
                  site_id INTEGER,
                  site_name TEXT
                )
                """
            )
            conn.execute(
                """
                INSERT INTO raw(
                  source, ts, date, hour, placement_id, placement_name, request, impression, clicks,
                  revenue, dsp_amount, order_id, order_name, supplier_id, supplier_name, site_id, site_name
                ) VALUES
                ('times_api', '2026-05-05 10:00:00', '2026-05-05', 10, 2001, '版位A_300x250', 320, 4500, 11, 500.0, 420.0, 'O-100', '訂單A', 7, '時報供應商A', 701, 'times-site-A'),
                ('times_api', '2026-05-05 11:00:00', '2026-05-05', 11, 2002, '版位B_320x480', 280, 3900, 12, 450.0, 360.0, 'O-101', '訂單B', 8, '時報供應商B', 702, 'times-site-B')
                """
            )
            conn.commit()
        finally:
            conn.close()
        return seed_db

    def test_ui_shell_runtime_actions_and_status(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            ctx = self._ctx(root)

            before = collect_runtime_status(ctx)
            self.assertEqual(before["health"]["status"], "fail")

            boot = dispatch_action(ctx, {"action": "bootstrap"})
            self.assertEqual(boot["status"], "ok")

            health = dispatch_action(ctx, {"action": "health"})
            self.assertEqual(health["status"], "ok")

            save_out = dispatch_action(ctx, {"action": "save", "rows": [self._full_row()]})
            self.assertTrue(str(save_out["run_id"]).startswith("run-"))

            modify_out = dispatch_action(
                ctx,
                {
                    "action": "modify",
                    "updates": [{"row_order": 0, "column": "最終經銷商", "value": "A2"}],
                },
            )
            self.assertEqual(modify_out["changed_count"], 1)
            self.assertEqual(modify_out["adjustment_count"], 1)

            dispatch_action(
                ctx,
                {
                    "action": "tab4_delivery",
                    "main_tab": "dsp_tab3",
                    "sub_tab": "pivot",
                    "period_week_start": "2026-04-27",
                    "period_week_end": "2026-05-03",
                },
            )
            export_out = dispatch_action(
                ctx,
                {
                    "action": "export",
                    "main_tab": "dsp_tab4",
                    "sub_tab": "overview",
                    "period_week_start": "2026-04-27",
                    "period_week_end": "2026-05-03",
                },
            )
            self.assertTrue(Path(str(export_out["artifact_path"])).exists())

            after = collect_runtime_status(ctx)
            self.assertEqual(after["canonical_source"], "sqlite")
            self.assertEqual(after["health"]["status"], "ok")
            self.assertGreaterEqual(len(after["recent"]["run_log"]), 3)
            self.assertGreaterEqual(len(after["recent"]["audit_log"]), 1)

    def test_seed_rebuild_action_restores_rows_from_seed_manifest(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            ctx = self._ctx(root)

            dispatch_action(ctx, {"action": "bootstrap"})
            seed_row = self._full_row(最終經銷商="RAW_API_A1", 最終來源_經銷商="api")
            raw_root = root / "raw-inbox"
            raw_root.mkdir(parents=True, exist_ok=True)
            (raw_root / "dsp_rawdata_20260506_150144_recalc.json").write_text(
                json.dumps([seed_row], ensure_ascii=False),
                encoding="utf-8",
            )
            code, payload = self._run_cli_json(
                [
                    "--root",
                    str(root),
                    "seed-bootstrap",
                    "--raw-source",
                    "raw-inbox",
                ]
            )
            self.assertEqual(code, 0)
            self.assertEqual(payload["status"], "ok")

            dispatch_action(
                ctx,
                {
                    "action": "save",
                    "rows": [self._full_row(最終經銷商="LIVE_DIRTY_A9", 最終來源_經銷商="manual")],
                },
            )

            rebuild_out = dispatch_action(
                ctx,
                {
                    "action": "seed_rebuild",
                    "workflow": "dsp",
                },
            )
            self.assertEqual(rebuild_out["status"], "ok")
            self.assertEqual(int(rebuild_out["files_used"]), 1)
            self.assertEqual(int(rebuild_out["workflows"]["dsp"]["row_count"]), 1)

            conn = sqlite3.connect(str(root / "data" / "mdrep.sqlite"))
            try:
                row = conn.execute(
                    """
                    SELECT 最終經銷商, 最終來源_經銷商
                    FROM canonical_raw
                    WHERE workflow='dsp'
                    ORDER BY row_order ASC
                    LIMIT 1
                    """
                ).fetchone()
            finally:
                conn.close()
            self.assertEqual(str(row[0]), "RAW_API_A1")
            self.assertEqual(str(row[1]), "api")

    def test_seed_promote_live_action_rebuilds_prod_and_test_runtime_dbs(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            self._make_seed_canonical_dsp_db(root, seed_root="data_seed")
            self._make_seed_canonical_ssp_truth_db(root, seed_root="data_seed")
            self._make_seed_canonical_dsp_db(root, seed_root="data_seed_test")
            self._make_seed_canonical_ssp_truth_db(root, seed_root="data_seed_test")

            prod_dsp_ctx = self._ctx(root, workflow="dsp")
            prod_ssp_ctx = self._ctx(root, workflow="ssp")
            test_dsp_ctx = UiContext(
                root=root,
                runtime_env="test",
                manifest_rel="bootstrap.test.manifest.json",
                workflow="dsp",
                template_version="v1",
                rule_version="v1",
                artifact_root=(root / "artifacts_test").resolve(),
            )
            test_ssp_ctx = UiContext(
                root=root,
                runtime_env="test",
                manifest_rel="bootstrap.test.manifest.json",
                workflow="ssp",
                template_version="v1",
                rule_version="v1",
                artifact_root=(root / "artifacts_test").resolve(),
            )

            prod_dsp_out = dispatch_action(prod_dsp_ctx, {"action": "seed_promote_live", "workflow": "dsp"})
            prod_ssp_out = dispatch_action(prod_ssp_ctx, {"action": "seed_promote_live", "workflow": "ssp"})
            test_dsp_out = dispatch_action(test_dsp_ctx, {"action": "seed_promote_live", "workflow": "dsp"})
            test_ssp_out = dispatch_action(test_ssp_ctx, {"action": "seed_promote_live", "workflow": "ssp"})

            self.assertEqual(prod_dsp_out["status"], "ok")
            self.assertTrue(str(prod_dsp_out["source_db"]).endswith("/data_seed/canonical/mdreport_dsp.sqlite"))
            self.assertEqual(prod_ssp_out["status"], "ok")
            self.assertTrue(str(prod_ssp_out["source_db"]).endswith("/data_seed/canonical/mdreport.sqlite"))
            self.assertEqual(test_dsp_out["status"], "ok")
            self.assertTrue(str(test_dsp_out["source_db"]).endswith("/data_seed_test/canonical/mdreport_dsp.sqlite"))
            self.assertEqual(test_ssp_out["status"], "ok")
            self.assertTrue(str(test_ssp_out["source_db"]).endswith("/data_seed_test/canonical/mdreport.sqlite"))

            prod_conn = sqlite3.connect(str(root / "data" / "mdrep.sqlite"))
            try:
                prod_dsp = prod_conn.execute(
                    "SELECT COUNT(1), MIN(最終經銷商) FROM canonical_raw WHERE workflow='dsp'"
                ).fetchone()
                prod_ssp = prod_conn.execute(
                    "SELECT COUNT(1), MIN(supplier_name) FROM ssp_raw"
                ).fetchone()
                prod_ssp_canonical = prod_conn.execute(
                    "SELECT COUNT(1) FROM canonical_raw WHERE workflow='ssp'"
                ).fetchone()
            finally:
                prod_conn.close()
            self.assertEqual(int(prod_dsp[0] or 0), 1)
            self.assertEqual(str(prod_dsp[1]), "API_PROMOTED_A1")
            self.assertEqual(int(prod_ssp[0] or 0), 2)
            self.assertEqual(str(prod_ssp[1]), "時報供應商A")
            self.assertEqual(int(prod_ssp_canonical[0] or 0), 0)

            test_conn = sqlite3.connect(str(root / "data_test" / "mdrep.test.sqlite"))
            try:
                test_dsp = test_conn.execute(
                    "SELECT COUNT(1), MIN(最終經銷商) FROM canonical_raw WHERE workflow='dsp'"
                ).fetchone()
                test_ssp = test_conn.execute(
                    "SELECT COUNT(1), MIN(supplier_name) FROM ssp_raw"
                ).fetchone()
                test_ssp_canonical = test_conn.execute(
                    "SELECT COUNT(1) FROM canonical_raw WHERE workflow='ssp'"
                ).fetchone()
            finally:
                test_conn.close()
            self.assertEqual(int(test_dsp[0] or 0), 1)
            self.assertEqual(str(test_dsp[1]), "API_PROMOTED_A1")
            self.assertEqual(int(test_ssp[0] or 0), 2)
            self.assertEqual(str(test_ssp[1]), "時報供應商A")
            self.assertEqual(int(test_ssp_canonical[0] or 0), 0)

            dispatch_action(
                prod_dsp_ctx,
                {"action": "tab4_delivery", "main_tab": "dsp_tab3", "sub_tab": "pivot"},
            )
            prod_dsp_export = dispatch_action(
                prod_dsp_ctx,
                {"action": "export", "main_tab": "dsp_tab4", "sub_tab": "overview"},
            )
            prod_ssp_export = dispatch_action(prod_ssp_ctx, {"action": "export"})
            dispatch_action(
                test_dsp_ctx,
                {"action": "tab4_delivery", "main_tab": "dsp_tab3", "sub_tab": "pivot"},
            )
            test_dsp_export = dispatch_action(
                test_dsp_ctx,
                {"action": "export", "main_tab": "dsp_tab4", "sub_tab": "overview"},
            )
            test_ssp_export = dispatch_action(test_ssp_ctx, {"action": "export"})

            self.assertTrue(Path(str(prod_dsp_export["artifact_path"])).exists())
            self.assertTrue(Path(str(prod_ssp_export["artifact_path"])).exists())
            self.assertTrue(Path(str(test_dsp_export["artifact_path"])).exists())
            self.assertTrue(Path(str(test_ssp_export["artifact_path"])).exists())
            self.assertIn("artifacts_test", str(test_dsp_export["artifact_path"]))
            self.assertIn("artifacts_test", str(test_ssp_export["artifact_path"]))

    def test_dispatch_action_fetch_ssp_api_writes_ssp_raw_and_clears_canonical_pollution(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            ctx = self._ctx(root, workflow="ssp")

            dispatch_action(ctx, {"action": "bootstrap"})
            dispatch_action(
                ctx,
                {
                    "action": "save",
                    "workflow": "ssp",
                    "rows": [self._full_row(最終經銷商="DIRTY_SSP_ROW", 最終來源_經銷商="manual")],
                },
            )

            with patch("domain.services.resolve_ssp_api_settings") as mock_settings, patch("domain.services.SspApiClient") as mock_client:
                mock_settings.return_value = SspApiSettings(email="matt@clickforce.com.tw", password="24450379")
                mock_client.return_value.fetch_report_bundle.return_value = self._mock_ssp_fetch_bundle()

                out = dispatch_action(
                    ctx,
                    {
                        "action": "fetch_ssp_api",
                        "workflow": "ssp",
                        "date": "2026-05-11",
                    },
                )

            self.assertEqual(out["status"], "ok")
            self.assertEqual(int(out["service_id"]), 14)
            self.assertEqual(int(out["row_count"]), 1)
            self.assertEqual(int(out["records_total"]), 1)
            self.assertEqual(int(out["login_user_id"]), 2072)
            self.assertEqual(str(out["login_email"]), "matt@clickforce.com.tw")
            self.assertEqual(out["sum_row"], {"request": 2885, "impress": 1386, "profit": 2.08})

            conn = sqlite3.connect(str(root / "data" / "mdrep.sqlite"))
            try:
                ssp_row = conn.execute(
                    """
                    SELECT source, ts, supplier_name, site_name, placement_name, request, impression, revenue, dsp_amount
                    FROM ssp_raw
                    ORDER BY row_order ASC
                    LIMIT 1
                    """
                ).fetchone()
                canonical_count = conn.execute(
                    "SELECT COUNT(1) FROM canonical_raw WHERE workflow='ssp'"
                ).fetchone()
                run_row = conn.execute(
                    """
                    SELECT run_type, workflow, status
                    FROM run_log
                    WHERE workflow='ssp'
                    ORDER BY rowid DESC
                    LIMIT 1
                    """
                ).fetchone()
            finally:
                conn.close()

            self.assertEqual(str(ssp_row[0]), "ssp3_api")
            self.assertEqual(str(ssp_row[1]), "2026-05-11 00:00:00")
            self.assertEqual(str(ssp_row[2]), "域動測試")
            self.assertEqual(str(ssp_row[3]), "DEMO link")
            self.assertEqual(str(ssp_row[4]), "DEMO LINK 專用")
            self.assertEqual(float(ssp_row[5]), 2885.0)
            self.assertEqual(float(ssp_row[6]), 1386.0)
            self.assertEqual(float(ssp_row[7]), 2.08)
            self.assertEqual(float(ssp_row[8]), 8.32)
            self.assertEqual(int(canonical_count[0] or 0), 0)
            self.assertEqual(tuple(str(v) for v in run_row), ("fetch_ssp_api", "ssp", "ok"))

    def test_dispatch_action_fetch_ssp_api_preserves_other_days_when_refreshing_single_day(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            ctx = self._ctx(root, workflow="ssp")
            dispatch_action(ctx, {"action": "bootstrap"})

            cfg = build_config(root, ctx.manifest_rel, ctx.runtime_env)
            repo = SQLiteRepository(cfg.db_path, project_root=root)
            with repo.connect() as conn:
                repo.save_ssp_raw_rows(
                    conn,
                    [
                        {
                            "source": "ssp3_api",
                            "ts": "2026-05-10 00:00:00",
                            "date": "2026-05-10",
                            "hour": 0,
                            "placement_id": 10229,
                            "placement_name": "OLD_PLACEMENT",
                            "request": 100.0,
                            "impression": 50.0,
                            "clicks": 1.0,
                            "revenue": 2.0,
                            "dsp_amount": 4.0,
                            "supplier_id": 1,
                            "supplier_name": "OLD_SUPPLIER",
                            "site_id": 784,
                            "site_name": "OLD_SITE",
                        }
                    ],
                )
                conn.commit()

            with patch("domain.services.resolve_ssp_api_settings") as mock_settings, patch("domain.services.SspApiClient") as mock_client:
                mock_settings.return_value = SspApiSettings(email="matt@clickforce.com.tw", password="24450379")
                mock_client.return_value.fetch_report_bundle.return_value = self._mock_ssp_fetch_bundle()

                out = dispatch_action(
                    ctx,
                    {
                        "action": "fetch_ssp_api",
                        "workflow": "ssp",
                        "date": "2026-05-11",
                    },
                )

            self.assertEqual(out["status"], "ok")
            self.assertEqual(int(out["fetched_row_count"]), 1)
            self.assertEqual(int(out["retained_row_count"]), 1)
            self.assertEqual(int(out["replaced_day_count"]), 1)
            self.assertEqual(int(out["row_count"]), 1)
            self.assertEqual(int(out["total_row_count"]), 2)

            conn = sqlite3.connect(str(root / "data" / "mdrep.sqlite"))
            try:
                rows = conn.execute(
                    """
                    SELECT date, placement_name, request
                    FROM ssp_raw
                    ORDER BY date ASC, row_order ASC
                    """
                ).fetchall()
            finally:
                conn.close()

            self.assertEqual(rows, [("2026-05-10", "OLD_PLACEMENT", 100.0), ("2026-05-11", "DEMO LINK 專用", 2885.0)])

    def test_fetch_ssp_api_cli_uses_runtime_command_contract(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            with patch("domain.services.resolve_ssp_api_settings") as mock_settings, patch("domain.services.SspApiClient") as mock_client:
                mock_settings.return_value = SspApiSettings(email="matt@clickforce.com.tw", password="24450379")
                mock_client.return_value.fetch_report_bundle.return_value = self._mock_ssp_fetch_bundle()

                code, payload = self._run_cli_json(
                    [
                        "--root",
                        str(root),
                        "fetch-ssp-api",
                        "--date",
                        "2026-05-11",
                    ]
                )

            self.assertEqual(code, 0)
            self.assertEqual(payload["status"], "ok")
            result = payload["result"]
            self.assertEqual(result["workflow"], "ssp")
            self.assertEqual(result["start_day"], "2026-05-11")
            self.assertEqual(result["end_day"], "2026-05-11")
            self.assertEqual(int(result["service_id"]), 14)
            self.assertEqual(int(result["row_count"]), 1)
            self.assertEqual(int(result["report_id"]), 174425)
            self.assertEqual(result["sum_row"], {"request": 2885, "impress": 1386, "profit": 2.08})

    def test_fetch_ssp_ad_group_api_cli_uses_runtime_command_contract(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            with patch("domain.services.resolve_ssp_api_settings") as mock_settings, patch("domain.services.SspApiClient") as mock_client:
                mock_settings.return_value = SspApiSettings(email="matt@clickforce.com.tw", password="24450379")
                mock_client.return_value.fetch_ad_group_report_bundle.return_value = self._mock_ssp_ad_group_fetch_bundle()

                code, payload = self._run_cli_json(
                    [
                        "--root",
                        str(root),
                        "fetch-ssp-ad-group-api",
                        "--date",
                        "2026-05-11",
                        "--zone-group-id",
                        "335",
                    ]
                )

            self.assertEqual(code, 0)
            self.assertEqual(payload["status"], "ok")
            result = payload["result"]
            self.assertEqual(result["workflow"], "ssp")
            self.assertEqual(result["start_day"], "2026-05-11")
            self.assertEqual(result["end_day"], "2026-05-11")
            self.assertEqual(int(result["group_count"]), 1)
            self.assertEqual(int(result["row_count"]), 1)
            self.assertEqual(int(result["records_total"]), 1)
            group = result["groups"][0]
            self.assertEqual(int(group["zone_group_id"]), 335)
            self.assertEqual(int(group["service_id"]), 14)
            self.assertEqual(int(group["report_id"]), 274425)

            conn = sqlite3.connect(root / "data" / "mdrep.sqlite")
            try:
                metric_row = conn.execute(
                    """
                    SELECT source, zone_group_id, zone_group_name, ad_format, price_tier, date, request, impress, profit
                    FROM ssp_ad_group_daily_metrics
                    LIMIT 1
                    """
                ).fetchone()
            finally:
                conn.close()

            self.assertEqual(str(metric_row[0]), "ssp3_api")
            self.assertEqual(int(metric_row[1]), 335)
            self.assertEqual(str(metric_row[2]), "知名媒體 高價版位 BN")
            self.assertEqual(str(metric_row[3]), "知名媒體 BN")
            self.assertEqual(str(metric_row[4]), "高")
            self.assertEqual(str(metric_row[5]), "2026-05-11")
            self.assertEqual(float(metric_row[6]), 2885.0)
            self.assertEqual(float(metric_row[7]), 1386.0)
            self.assertEqual(float(metric_row[8]), 2.08)

    def test_fetch_ssp_api_cli_multi_day_sum_row_is_aggregated(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            with patch("domain.services.resolve_ssp_api_settings") as mock_settings, patch("domain.services.SspApiClient") as mock_client:
                mock_settings.return_value = SspApiSettings(email="matt@clickforce.com.tw", password="24450379")
                mock_client.return_value.fetch_report_bundle.return_value = self._mock_ssp_fetch_bundle_multi_day()

                code, payload = self._run_cli_json(
                    [
                        "--root",
                        str(root),
                        "fetch-ssp-api",
                        "--start-day",
                        "2026-05-10",
                        "--end-day",
                        "2026-05-11",
                    ]
                )

            self.assertEqual(code, 0)
            self.assertEqual(payload["status"], "ok")
            result = payload["result"]
            self.assertEqual(result["workflow"], "ssp")
            self.assertEqual(result["start_day"], "2026-05-10")
            self.assertEqual(result["end_day"], "2026-05-11")
            self.assertEqual(int(result["records_total"]), 3)
            self.assertEqual(int(result["report_id"]), 174426)
            self.assertEqual(list(result["report_ids"]), [174425, 174426])
            self.assertEqual(result["sum_row"], {"request": 3000, "impress": 1500, "profit": 4.0})

    def test_dispatch_action_fetch_ssp_api_accepts_camelcase_auth_contract(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            ctx = self._ctx(root, workflow="ssp")
            bundle = self._mock_ssp_fetch_bundle()
            bundle["login"] = {}
            bundle["sum_row"] = {"request": 5770, "impress": 2772, "profit": 4.16}
            bundle["auth"] = {
                "service_id": 14,
                "token": "ssp-service-token",
                "user": {"id": 3001, "email": "ssp-auth-user@clickforce.com.tw"},
            }

            def _resolve_no_legacy(
                *,
                email: str | None = None,
                password: str | None = None,
                scope_check_url: str | None = None,
                api_base_url: str | None = None,
                auth_decrypt_key: str | None = None,
                service_id: int | None = None,
                source_name: str | None = None,
                timeout_seconds: int | None = None,
            ) -> SspApiSettings:
                self.assertEqual(email, "matt@clickforce.com.tw")
                self.assertEqual(password, "24450379")
                self.assertEqual(scope_check_url, "https://example.com/ssp-scope-check")
                self.assertEqual(api_base_url, "https://example.com/ssp-api")
                self.assertEqual(auth_decrypt_key, "ssp-auth-key")
                self.assertEqual(service_id, 14)
                self.assertEqual(source_name, "ssp3_api")
                self.assertEqual(timeout_seconds, 15)
                return SspApiSettings(email="matt@clickforce.com.tw", password="24450379")

            with patch("domain.services.resolve_ssp_api_settings", side_effect=_resolve_no_legacy), patch("domain.services.SspApiClient") as mock_client:
                mock_client.return_value.fetch_report_bundle.return_value = bundle
                out = dispatch_action(
                    ctx,
                    {
                        "action": "fetch_ssp_api",
                        "workflow": "ssp",
                        "startDay": "2026-05-10",
                        "endDay": "2026-05-11",
                        "email": "matt@clickforce.com.tw",
                        "password": "24450379",
                        "scopeCheckUrl": "https://example.com/ssp-scope-check",
                        "apiBaseUrl": "https://example.com/ssp-api",
                        "authDecryptKey": "ssp-auth-key",
                        "serviceId": 14,
                        "sourceName": "ssp3_api",
                        "timeoutSeconds": 15,
                    },
                )

            self.assertEqual(out["status"], "ok")
            self.assertEqual(out["start_day"], "2026-05-10")
            self.assertEqual(out["end_day"], "2026-05-11")
            self.assertEqual(out["sum_row"], {"request": 5770, "impress": 2772, "profit": 4.16})
            self.assertEqual(int(out["login_user_id"]), 3001)
            self.assertEqual(str(out["login_email"]), "ssp-auth-user@clickforce.com.tw")

    def test_fetch_dsp_api_cli_accepts_runtime_auth_contract(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)

            def _resolve_no_legacy(
                *,
                email: str | None = None,
                password: str | None = None,
                scope_check_url: str | None = None,
                api_base_url: str | None = None,
                auth_decrypt_key: str | None = None,
                service_id: int | None = None,
                source_name: str | None = None,
                timeout_seconds: int | None = None,
            ) -> DspApiSettings:
                self.assertEqual(email, "matt@clickforce.com.tw")
                self.assertEqual(password, "24450379")
                self.assertEqual(scope_check_url, "https://example.com/dsp-scope-check")
                self.assertEqual(api_base_url, "https://example.com/dsp-api")
                self.assertEqual(auth_decrypt_key, "dsp-auth-key")
                self.assertEqual(service_id, 10)
                self.assertEqual(source_name, "dsp3_api")
                self.assertEqual(timeout_seconds, 20)
                return DspApiSettings(email="matt@clickforce.com.tw", password="24450379")

            with patch("domain.services.resolve_dsp_api_settings", side_effect=_resolve_no_legacy), patch("domain.services.DspApiClient") as mock_client:
                mock_client.return_value.fetch_report_bundle.return_value = self._mock_dsp_fetch_bundle()
                code, payload = self._run_cli_json(
                    [
                        "--root",
                        str(root),
                        "fetch-dsp-api",
                        "--start-day",
                        "2026-05-10",
                        "--end-day",
                        "2026-05-11",
                        "--email",
                        "matt@clickforce.com.tw",
                        "--password",
                        "24450379",
                        "--scope-check-url",
                        "https://example.com/dsp-scope-check",
                        "--api-base-url",
                        "https://example.com/dsp-api",
                        "--auth-decrypt-key",
                        "dsp-auth-key",
                        "--service-id",
                        "10",
                        "--source-name",
                        "dsp3_api",
                        "--timeout-seconds",
                        "20",
                    ]
                )

            self.assertEqual(code, 0)
            self.assertEqual(payload["status"], "ok")
            result = payload["result"]
            self.assertEqual(result["workflow"], "dsp")
            self.assertEqual(result["start_day"], "2026-05-10")
            self.assertEqual(result["end_day"], "2026-05-11")
            self.assertEqual(int(result["service_id"]), 10)

    def test_fetch_api_commands_reject_removed_mdreport_config_surface(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)

            for command in ("fetch-ssp-api", "fetch-ssp-ad-group-api", "fetch-dsp-api"):
                code, payload = self._run_cli_json(
                    [
                        "--root",
                        str(root),
                        command,
                        "--date",
                        "2026-05-10",
                        "--mdreport-config",
                        "/tmp/legacy-api-config.py",
                    ]
                )
                self.assertNotEqual(code, 0)
                self.assertEqual(payload["error_code"], "CLI_USAGE_ERROR")

    def test_dispatch_action_fetch_ssp_api_rejects_removed_mdreport_config_surface(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            ctx = self._ctx(root, workflow="ssp")

            with self.assertRaises(ValueError) as exc_ctx:
                dispatch_action(
                    ctx,
                    {
                        "action": "fetch_ssp_api",
                        "workflow": "ssp",
                        "date": "2026-05-10",
                        "mdreportConfig": "/tmp/legacy-api-config.py",
                    },
                )

            self.assertIn("mdreport-config 已移除", str(exc_ctx.exception))

    def test_dispatch_action_fetch_dsp_api_rejects_removed_mdreport_config_surface(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            ctx = self._ctx(root, workflow="dsp")

            with self.assertRaises(ValueError) as exc_ctx:
                dispatch_action(
                    ctx,
                    {
                        "action": "fetch_dsp_api",
                        "workflow": "dsp",
                        "date": "2026-05-10",
                        "mdreportConfig": "/tmp/legacy-api-config.py",
                    },
                )

            self.assertIn("mdreport-config 已移除", str(exc_ctx.exception))

    def test_dispatch_action_fetch_dsp_api_writes_canonical_rows_from_regular_api(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            ctx = self._ctx(root, workflow="dsp")

            with patch("domain.services.resolve_dsp_api_settings") as mock_settings, patch("domain.services.DspApiClient") as mock_client:
                mock_settings.return_value = DspApiSettings(email="matt@clickforce.com.tw", password="24450379")
                mock_client.return_value.fetch_report_bundle.return_value = self._mock_dsp_fetch_bundle()

                out = dispatch_action(
                    ctx,
                    {
                        "action": "fetch_dsp_api",
                        "workflow": "dsp",
                        "date": "2026-05-10",
                    },
                )

            self.assertEqual(out["status"], "ok")
            self.assertEqual(int(out["service_id"]), 10)
            self.assertEqual(int(out["row_count"]), 1)
            self.assertEqual(int(out["total_row_count"]), 1)
            self.assertEqual(str(out["job_id"]), "35cffe17660dad7fbdfb7080ffa2f1a6")
            self.assertEqual(str(out["source_name"]), "dsp3_api")

            conn = sqlite3.connect(str(root / "data" / "mdrep.sqlite"))
            try:
                row = conn.execute(
                    """
                    SELECT 日期時間, 經銷商, 訂單, 素材, 廣告形式, 尺寸, 素材樣板, 執行金額, 系統營收, 媒體費用,
                           原始經銷商, 最終經銷商, 最終來源_經銷商
                    FROM canonical_raw
                    WHERE workflow='dsp'
                    ORDER BY row_order ASC
                    LIMIT 1
                    """
                ).fetchone()
                run_row = conn.execute(
                    """
                    SELECT run_type, workflow, status
                    FROM run_log
                    WHERE workflow='dsp'
                    ORDER BY rowid DESC
                    LIMIT 1
                    """
                ).fetchone()
            finally:
                conn.close()

            self.assertEqual(str(row[0]), "2026-05-10")
            self.assertEqual(str(row[1]), "[台灣]域動行銷股份有限公司")
            self.assertEqual(str(row[2]), "(42031)活動")
            self.assertEqual(str(row[3]), "(314928)0422_純蓋板")
            self.assertEqual(str(row[4]), "純蓋板")
            self.assertEqual(str(row[5]), "純蓋板")
            self.assertEqual(str(row[6]), "HTML/JS")
            self.assertEqual(float(row[7]), 10934.99)
            self.assertEqual(float(row[8]), 9000.11)
            self.assertEqual(float(row[9]), 8000.22)
            self.assertEqual(str(row[10]), "[台灣]域動行銷股份有限公司")
            self.assertEqual(str(row[11]), "[台灣]域動行銷股份有限公司")
            self.assertEqual(str(row[12]), "raw")
            self.assertEqual(tuple(str(v) for v in run_row), ("fetch_dsp_api", "dsp", "ok"))

    def test_dispatch_action_fetch_dsp_api_preserves_other_days_when_refreshing_single_day(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            ctx = self._ctx(root, workflow="dsp")
            dispatch_action(ctx, {"action": "bootstrap"})

            dispatch_action(
                ctx,
                {
                    "action": "save",
                    "workflow": "dsp",
                    "rows": [
                        self._full_row(日期時間="2026-05-09", 訂單="OLD-0509", 執行金額=900.0),
                        self._full_row(日期時間="2026-05-10", 訂單="OLD-0510", 執行金額=1000.0),
                    ],
                },
            )

            with patch("domain.services.resolve_dsp_api_settings") as mock_settings, patch("domain.services.DspApiClient") as mock_client:
                mock_settings.return_value = DspApiSettings(email="matt@clickforce.com.tw", password="24450379")
                mock_client.return_value.fetch_report_bundle.return_value = self._mock_dsp_fetch_bundle()

                out = dispatch_action(
                    ctx,
                    {
                        "action": "fetch_dsp_api",
                        "workflow": "dsp",
                        "date": "2026-05-10",
                    },
                )

            self.assertEqual(out["status"], "ok")
            self.assertEqual(int(out["fetched_row_count"]), 1)
            self.assertEqual(int(out["retained_row_count"]), 1)
            self.assertEqual(int(out["replaced_day_count"]), 1)
            self.assertEqual(int(out["row_count"]), 1)
            self.assertEqual(int(out["total_row_count"]), 2)

            conn = sqlite3.connect(str(root / "data" / "mdrep.sqlite"))
            try:
                rows = conn.execute(
                    """
                    SELECT 日期時間, 訂單, 執行金額
                    FROM canonical_raw
                    WHERE workflow='dsp'
                    ORDER BY 日期時間 ASC, row_order ASC
                    """
                ).fetchall()
            finally:
                conn.close()

            self.assertEqual(len(rows), 2)
            self.assertEqual(tuple(str(v) for v in rows[0]), ("2026-05-09", "OLD-0509", "900.0"))
            self.assertEqual(str(rows[1][0]), "2026-05-10")
            self.assertEqual(str(rows[1][1]), "(42031)活動")
            self.assertEqual(float(rows[1][2]), 10934.99)

    def test_dispatch_action_fetch_dsp_api_clears_requested_day_when_api_returns_no_rows(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            ctx = self._ctx(root, workflow="dsp")
            dispatch_action(ctx, {"action": "bootstrap"})

            dispatch_action(
                ctx,
                {
                    "action": "save",
                    "workflow": "dsp",
                    "rows": [
                        self._full_row(日期時間="2026-05-09", 訂單="OLD-0509", 執行金額=900.0),
                        self._full_row(日期時間="2026-05-10", 訂單="OLD-0510", 執行金額=1000.0),
                    ],
                },
            )

            empty_bundle = {**self._mock_dsp_fetch_bundle(), "records_total": 0, "rows": []}
            with patch("domain.services.resolve_dsp_api_settings") as mock_settings, patch("domain.services.DspApiClient") as mock_client:
                mock_settings.return_value = DspApiSettings(email="matt@clickforce.com.tw", password="24450379")
                mock_client.return_value.fetch_report_bundle.return_value = empty_bundle

                out = dispatch_action(
                    ctx,
                    {
                        "action": "fetch_dsp_api",
                        "workflow": "dsp",
                        "date": "2026-05-10",
                    },
                )

            self.assertEqual(out["status"], "ok")
            self.assertEqual(int(out["fetched_row_count"]), 0)
            self.assertEqual(int(out["retained_row_count"]), 1)
            self.assertEqual(int(out["replaced_day_count"]), 1)
            self.assertEqual(int(out["row_count"]), 0)
            self.assertEqual(int(out["total_row_count"]), 1)

            conn = sqlite3.connect(str(root / "data" / "mdrep.sqlite"))
            try:
                rows = conn.execute(
                    """
                    SELECT 日期時間, 訂單, 執行金額
                    FROM canonical_raw
                    WHERE workflow='dsp'
                    ORDER BY 日期時間 ASC, row_order ASC
                    """
                ).fetchall()
            finally:
                conn.close()

            self.assertEqual(len(rows), 1)
            self.assertEqual(tuple(str(v) for v in rows[0]), ("2026-05-09", "OLD-0509", "900.0"))

    def test_fetch_dsp_api_cli_uses_runtime_command_contract(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            with patch("domain.services.resolve_dsp_api_settings") as mock_settings, patch("domain.services.DspApiClient") as mock_client:
                mock_settings.return_value = DspApiSettings(email="matt@clickforce.com.tw", password="24450379")
                mock_client.return_value.fetch_report_bundle.return_value = self._mock_dsp_fetch_bundle()

                code, payload = self._run_cli_json(
                    [
                        "--root",
                        str(root),
                        "fetch-dsp-api",
                        "--date",
                        "2026-05-10",
                    ]
                )

            self.assertEqual(code, 0)
            self.assertEqual(payload["status"], "ok")
            result = payload["result"]
            self.assertEqual(result["workflow"], "dsp")
            self.assertEqual(result["start_day"], "2026-05-10")
            self.assertEqual(result["end_day"], "2026-05-10")
            self.assertEqual(int(result["service_id"]), 10)
            self.assertEqual(int(result["row_count"]), 1)
            self.assertEqual(str(result["job_id"]), "35cffe17660dad7fbdfb7080ffa2f1a6")

    def test_resolve_ui_context_switches_to_test_env_defaults(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)

            ctx = ui_shell_module._resolve_ui_context(
                root=root,
                runtime_env_raw="test",
                manifest_raw="",
                artifact_root_raw="",
                workflow="dsp",
                template_version="v1",
                rule_version="v1",
            )

            self.assertEqual(ctx.runtime_env, "test")
            self.assertEqual(ctx.manifest_rel, "bootstrap.test.manifest.json")
            self.assertTrue(str(ctx.artifact_root).endswith("artifacts_test"))

    def test_resolve_ui_context_rejects_artifact_root_outside_scope(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            outside_root = Path(td).parent / "escaped-artifacts"

            with self.assertRaises(PermissionError):
                ui_shell_module._resolve_ui_context(
                    root=root,
                    runtime_env_raw="",
                    manifest_raw="",
                    artifact_root_raw=str(outside_root),
                    workflow="dsp",
                    template_version="v1",
                    rule_version="v1",
                )

    def test_sandbox_context_isolates_db_and_reset_restores_baseline(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            prod_ctx = self._ctx(root)
            dispatch_action(prod_ctx, {"action": "bootstrap"})
            dispatch_action(prod_ctx, {"action": "save", "rows": [self._full_row(最終經銷商="BASELINE")]})
            prepare_out = dispatch_action(prod_ctx, {"action": "sandbox_prepare"})
            self.assertEqual(prepare_out["status"], "ok")

            sandbox_ctx = ui_shell_module._resolve_ui_context(
                root=root,
                runtime_env_raw="",
                manifest_raw="",
                artifact_root_raw="",
                sandbox_raw="case-151",
                workflow="dsp",
                template_version="v1",
                rule_version="v1",
            )

            self.assertEqual(sandbox_ctx.sandbox_id, "case-151")
            self.assertEqual(sandbox_ctx.db_path, (root / "data_sandbox" / "case-151" / "mdrep.sqlite").resolve())
            self.assertEqual(sandbox_ctx.artifact_root, (root / "artifacts_sandbox" / "case-151").resolve())
            self.assertTrue((sandbox_ctx.db_path or Path("")).exists())

            dispatch_action(sandbox_ctx, {"action": "save", "rows": [self._full_row(最終經銷商="SANDBOX")]})
            sandbox_ctx.artifact_root.mkdir(parents=True, exist_ok=True)
            (sandbox_ctx.artifact_root / "old.txt").write_text("stale", encoding="utf-8")

            prod_frame = collect_workflow_frame(prod_ctx)
            sandbox_frame = collect_workflow_frame(sandbox_ctx)
            self.assertEqual(prod_frame["rows"][0]["最終經銷商"], "BASELINE")
            self.assertEqual(sandbox_frame["rows"][0]["最終經銷商"], "SANDBOX")

            reset_out = dispatch_action(sandbox_ctx, {"action": "sandbox_reset"})
            self.assertEqual(reset_out["status"], "ok")
            self.assertEqual(reset_out["sandbox"], "case-151")
            self.assertGreaterEqual(int(reset_out["removed_artifact_entries"]), 1)
            self.assertFalse((sandbox_ctx.artifact_root / "old.txt").exists())

            reset_frame = collect_workflow_frame(sandbox_ctx)
            self.assertEqual(reset_frame["rows"][0]["最終經銷商"], "BASELINE")

    def test_sandbox_reset_uses_immutable_baseline_snapshot(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            prod_ctx = self._ctx(root)
            dispatch_action(prod_ctx, {"action": "bootstrap"})
            dispatch_action(prod_ctx, {"action": "save", "rows": [self._full_row(最終經銷商="BASELINE")]})
            dispatch_action(prod_ctx, {"action": "sandbox_prepare"})

            sandbox_ctx = ui_shell_module._resolve_ui_context(
                root=root,
                runtime_env_raw="",
                manifest_raw="",
                artifact_root_raw="",
                sandbox_raw="fresh",
                workflow="dsp",
                template_version="v1",
                rule_version="v1",
            )
            dispatch_action(sandbox_ctx, {"action": "save", "rows": [self._full_row(最終經銷商="DIRTY_SANDBOX")]})
            dispatch_action(prod_ctx, {"action": "save", "rows": [self._full_row(最終經銷商="POLLUTED_BASELINE")]})
            dispatch_action(sandbox_ctx, {"action": "sandbox_reset"})
            reset_frame = collect_workflow_frame(sandbox_ctx)
            prod_frame = collect_workflow_frame(prod_ctx)
            self.assertEqual(reset_frame["rows"][0]["最終經銷商"], "BASELINE")
            self.assertEqual(prod_frame["rows"][0]["最終經銷商"], "POLLUTED_BASELINE")

    def test_sandbox_prepare_refuses_to_overwrite_existing_snapshot_without_force(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            prod_ctx = self._ctx(root)
            dispatch_action(prod_ctx, {"action": "bootstrap"})
            dispatch_action(prod_ctx, {"action": "save", "rows": [self._full_row(最終經銷商="BASELINE")]})
            dispatch_action(prod_ctx, {"action": "sandbox_prepare"})
            dispatch_action(prod_ctx, {"action": "save", "rows": [self._full_row(最終經銷商="POLLUTED")]})

            with self.assertRaises(FileExistsError):
                dispatch_action(prod_ctx, {"action": "sandbox_prepare"})

            with self.assertRaises(FileExistsError):
                dispatch_action(prod_ctx, {"action": "sandbox_prepare", "force": "false"})

            with self.assertRaises(FileExistsError):
                dispatch_action(prod_ctx, {"action": "sandbox_prepare", "force": "0"})

            forced = dispatch_action(prod_ctx, {"action": "sandbox_prepare", "force": True})
            self.assertEqual(forced["status"], "ok")
            self.assertTrue(bool(forced["force"]))

            sandbox_ctx = ui_shell_module._resolve_ui_context(
                root=root,
                runtime_env_raw="",
                manifest_raw="",
                artifact_root_raw="",
                sandbox_raw="force-check",
                workflow="dsp",
                template_version="v1",
                rule_version="v1",
            )
            frame = collect_workflow_frame(sandbox_ctx)
            self.assertEqual(frame["rows"][0]["最終經銷商"], "POLLUTED")

    def test_sandbox_context_requires_prepared_non_empty_baseline_snapshot(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            prod_ctx = self._ctx(root)

            with self.assertRaises(FileNotFoundError):
                ui_shell_module._resolve_ui_context(
                    root=root,
                    runtime_env_raw="",
                    manifest_raw="",
                    artifact_root_raw="",
                    sandbox_raw="fresh",
                    workflow="dsp",
                    template_version="v1",
                    rule_version="v1",
                )

            dispatch_action(prod_ctx, {"action": "bootstrap"})
            with self.assertRaises(ValueError):
                dispatch_action(prod_ctx, {"action": "sandbox_prepare"})

    def test_invalid_sandbox_id_is_rejected_instead_of_falling_back_to_baseline(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            prod_ctx = self._ctx(root)
            dispatch_action(prod_ctx, {"action": "bootstrap"})
            dispatch_action(prod_ctx, {"action": "save", "rows": [self._full_row(最終經銷商="BASELINE")]})

            with self.assertRaises(ValueError):
                ui_shell_module._resolve_ui_context(
                    root=root,
                    runtime_env_raw="",
                    manifest_raw="",
                    artifact_root_raw="",
                    sandbox_raw="qa/01",
                    workflow="dsp",
                    template_version="v1",
                    rule_version="v1",
                )

    def test_sandbox_reset_without_sandbox_is_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            ctx = self._ctx(root)
            dispatch_action(ctx, {"action": "bootstrap"})
            with self.assertRaises(ValueError):
                dispatch_action(ctx, {"action": "sandbox_reset"})

    def test_multiple_sandbox_ids_do_not_pollute_each_other(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            prod_ctx = self._ctx(root)
            dispatch_action(prod_ctx, {"action": "bootstrap"})
            dispatch_action(prod_ctx, {"action": "save", "rows": [self._full_row(最終經銷商="BASELINE")]})
            dispatch_action(prod_ctx, {"action": "sandbox_prepare"})
            sandbox_a = ui_shell_module._resolve_ui_context(
                root=root,
                runtime_env_raw="",
                manifest_raw="",
                artifact_root_raw="",
                sandbox_raw="qa-a",
                workflow="dsp",
                template_version="v1",
                rule_version="v1",
            )
            sandbox_b = ui_shell_module._resolve_ui_context(
                root=root,
                runtime_env_raw="",
                manifest_raw="",
                artifact_root_raw="",
                sandbox_raw="qa-b",
                workflow="dsp",
                template_version="v1",
                rule_version="v1",
            )
            dispatch_action(sandbox_a, {"action": "save", "rows": [self._full_row(最終經銷商="A_DIRTY")]})
            dispatch_action(sandbox_b, {"action": "save", "rows": [self._full_row(最終經銷商="B_DIRTY")]})
            dispatch_action(sandbox_a, {"action": "sandbox_reset"})

            frame_a = collect_workflow_frame(sandbox_a)
            frame_b = collect_workflow_frame(sandbox_b)
            self.assertEqual(frame_a["rows"][0]["最終經銷商"], "BASELINE")
            self.assertEqual(frame_b["rows"][0]["最終經銷商"], "B_DIRTY")

    def test_dsp_export_fills_template_inputs_without_overwriting_formulas(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            ctx = self._ctx(root)
            dispatch_action(ctx, {"action": "bootstrap"})
            dispatch_action(
                ctx,
                {
                    "action": "save",
                    "rows": [
                        self._full_row(
                            日期時間="2026-05-01 00:00:00",
                            分類層級B="內部經銷商",
                            分類層級C="營銷事業處",
                            分類層級D="一般廣告",
                            最終廣告形式="一般廣告",
                            執行金額=123.0,
                        ),
                        self._full_row(
                            日期時間="2026-05-02 00:00:00",
                            分類層級B="內部經銷商",
                            分類層級C="策略部",
                            分類層級D="蓋板/置底(展開&不展)/文中",
                            最終廣告形式="創意",
                            執行金額=456.0,
                        ),
                        self._full_row(
                            日期時間="2026-05-03 00:00:00",
                            分類層級B="外部經銷商",
                            分類層級C="經銷推廣",
                            分類層級D="玩藝國際股份有限公司",
                            最終廣告形式="一般廣告",
                            執行金額=111.0,
                        ),
                        self._full_row(
                            日期時間="2026-05-03 00:00:00",
                            分類層級B="外部經銷商",
                            分類層級C="IO委刊",
                            分類層級D="momo",
                            最終廣告形式="一般廣告",
                            執行金額=222.0,
                        ),
                        self._full_row(
                            日期時間="2026-05-03 00:00:00",
                            分類層級B="HB串接",
                            分類層級C="MD",
                            分類層級D="appier",
                            最終廣告形式="一般廣告",
                            執行金額=333.0,
                        ),
                    ],
                },
            )

            dispatch_action(
                ctx,
                {
                    "action": "tab4_delivery",
                    "main_tab": "dsp_tab3",
                    "sub_tab": "pivot",
                    "period_week_start": "2026-04-27",
                    "period_week_end": "2026-05-03",
                },
            )
            export_out = dispatch_action(
                ctx,
                {
                    "action": "export",
                    "main_tab": "dsp_tab4",
                    "sub_tab": "overview",
                    "period_week_start": "2026-04-27",
                    "period_week_end": "2026-05-03",
                },
            )
            export_path = Path(str(export_out["artifact_path"]))
            wb = load_workbook(export_path, data_only=False)
            try:
                ws_summary = wb["mF投資量_總表"]
                ws_detail = wb["各經銷商明細"]
                ws_tracking = wb["北流進單追蹤"]

                self.assertEqual(ws_summary["M2"].value, "=SUM(M3:M8)")
                self.assertEqual(ws_summary["N3"].value, "=M3/M$2")
                self.assertEqual(ws_summary["AC3"].value, "=E3+G3+I3+K3+M3+O3+Q3+S3+U3+W3+Y3+AA3")
                self.assertEqual(ws_detail["N7"].value, "=M7/M$6")
                self.assertEqual(ws_detail["AB2"].value, "=SUM(AA6,AA25,AA45,AA64)/AA2")
                self.assertIsNone(ws_tracking["K2"].value)

                self.assertEqual(ws_detail["M7"].value, 123.0)
                self.assertEqual(ws_detail["M27"].value, 456.0)
                self.assertEqual(ws_detail["M26"].value, 0.0)
                self.assertEqual(ws_detail["M46"].value, 111.0)
                self.assertEqual(ws_detail["M65"].value, 222.0)
                self.assertEqual(ws_detail["M84"].value, 333.0)
                self.assertEqual(ws_detail["A24"].value, 2026)
            finally:
                wb.close()

            self._assert_dsp_export_template_parity(
                root / "templates" / "dsp_tab4_template.xlsx",
                export_path,
            )

    def test_dsp_export_download_endpoint_returns_artifact_attachment(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            ctx = self._ctx(root)
            dispatch_action(ctx, {"action": "bootstrap"})
            dispatch_action(ctx, {"action": "save", "rows": [self._full_row()]})
            dispatch_action(
                ctx,
                {
                    "action": "tab4_delivery",
                    "main_tab": "dsp_tab3",
                    "sub_tab": "pivot",
                    "period_week_start": "2026-04-27",
                    "period_week_end": "2026-05-03",
                },
            )
            export_out = dispatch_action(
                ctx,
                {
                    "action": "export",
                    "main_tab": "dsp_tab4",
                    "sub_tab": "overview",
                    "period_week_start": "2026-04-27",
                    "period_week_end": "2026-05-03",
                },
            )
            artifact_path = Path(str(export_out.get("artifact_path") or ""))
            self.assertTrue(artifact_path.exists())

            try:
                server = ThreadingHTTPServer(("127.0.0.1", 0), UiRequestHandler)
            except PermissionError:
                self.skipTest("sandbox 禁止本地 socket bind，略過 download endpoint smoke")
            thread = threading.Thread(target=server.serve_forever, daemon=True)
            thread.start()
            try:
                host, port = server.server_address
                query = urlencode(
                    {
                        "root": str(root),
                        "manifest": "bootstrap.manifest.json",
                        "workflow": "dsp",
                        "template_version": "v1",
                        "rule_version": "v1",
                        "artifact_root": "artifacts",
                        "artifact_path": str(artifact_path),
                        "main_tab": "dsp_tab4",
                        "sub_tab": "overview",
                    }
                )
                with urlopen(Request(f"http://{host}:{port}/api/export/download?{query}")) as resp:
                    body = resp.read()
                    self.assertEqual(resp.status, 200)
                    self.assertIn("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", resp.headers.get("Content-Type", ""))
                    disposition = resp.headers.get("Content-Disposition", "")
                    self.assertIn("attachment;", disposition)
                    self.assertIn("filename*=", disposition)
                    self.assertIn("2026%20DSP", disposition)
                    self.assertEqual(
                        hashlib.sha256(body).hexdigest(),
                        hashlib.sha256(artifact_path.read_bytes()).hexdigest(),
                    )
            finally:
                server.shutdown()
                server.server_close()
                thread.join(timeout=2.0)

    def test_dsp_export_download_endpoint_rejects_artifact_root_outside_scope(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            ctx = self._ctx(root)
            dispatch_action(ctx, {"action": "bootstrap"})
            dispatch_action(ctx, {"action": "save", "rows": [self._full_row(最終經銷商="OUTSIDE_SCOPE_A1")]})
            dispatch_action(
                ctx,
                {
                    "action": "tab4_delivery",
                    "main_tab": "dsp_tab3",
                    "sub_tab": "pivot",
                },
            )
            export_out = dispatch_action(
                ctx,
                {
                    "action": "export",
                    "main_tab": "dsp_tab4",
                    "sub_tab": "overview",
                },
            )
            artifact_path = Path(str(export_out.get("artifact_path") or ""))
            self.assertTrue(artifact_path.exists())

            try:
                server = ThreadingHTTPServer(("127.0.0.1", 0), UiRequestHandler)
            except PermissionError:
                self.skipTest("sandbox 禁止本地 socket bind，略過 download endpoint scope guard")
            thread = threading.Thread(target=server.serve_forever, daemon=True)
            thread.start()
            try:
                host, port = server.server_address
                query = urlencode(
                    {
                        "root": str(root),
                        "manifest": "bootstrap.manifest.json",
                        "workflow": "dsp",
                        "template_version": "v1",
                        "rule_version": "v1",
                        "artifact_root": str(root.parent / "escaped-artifacts"),
                        "artifact_path": str(artifact_path),
                    }
                )
                with self.assertRaises(HTTPError) as exc_ctx:
                    urlopen(Request(f"http://{host}:{port}/api/export/download?{query}"))
                self.assertEqual(exc_ctx.exception.code, 400)
                error_payload = json.loads(exc_ctx.exception.read().decode("utf-8"))
                self.assertEqual(error_payload.get("error_code"), "DOWNLOAD_FAILED")
                self.assertIn("artifact_root out of allowed artifact scope", str(error_payload.get("message") or ""))
            finally:
                server.shutdown()
                server.server_close()
                thread.join(timeout=2.0)

    def test_api_action_rejects_artifact_root_outside_scope(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)

            try:
                server = ThreadingHTTPServer(("127.0.0.1", 0), UiRequestHandler)
            except PermissionError:
                self.skipTest("sandbox 禁止本地 socket bind，略過 api/action scope guard")
            thread = threading.Thread(target=server.serve_forever, daemon=True)
            thread.start()
            try:
                host, port = server.server_address
                req = Request(
                    f"http://{host}:{port}/api/action",
                    data=json.dumps(
                        {
                            "action": "health",
                            "root": str(root),
                            "manifest": "bootstrap.manifest.json",
                            "workflow": "dsp",
                            "template_version": "v1",
                            "rule_version": "v1",
                            "artifact_root": str(root.parent / "escaped-artifacts"),
                        }
                    ).encode("utf-8"),
                    headers={"Content-Type": "application/json"},
                    method="POST",
                )
                with self.assertRaises(HTTPError) as exc_ctx:
                    urlopen(req)
                self.assertEqual(exc_ctx.exception.code, 400)
                error_payload = json.loads(exc_ctx.exception.read().decode("utf-8"))
                self.assertEqual(error_payload.get("error_code"), "UI_ACTION_FAILED")
                self.assertIn("artifact_root out of allowed artifact scope", str(error_payload.get("message") or ""))
            finally:
                server.shutdown()
                server.server_close()
                thread.join(timeout=2.0)

    def test_api_action_sandbox_prepare_works_before_sandbox_snapshot_exists(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            prod_ctx = self._ctx(root)
            dispatch_action(prod_ctx, {"action": "bootstrap"})
            dispatch_action(prod_ctx, {"action": "save", "rows": [self._full_row(最終經銷商="BASELINE")]})

            try:
                server = ThreadingHTTPServer(("127.0.0.1", 0), UiRequestHandler)
            except PermissionError:
                self.skipTest("sandbox 禁止本地 socket bind，略過 sandbox_prepare endpoint")
            thread = threading.Thread(target=server.serve_forever, daemon=True)
            thread.start()
            try:
                host, port = server.server_address
                req = Request(
                    f"http://{host}:{port}/api/action",
                    data=json.dumps(
                        {
                            "action": "sandbox_prepare",
                            "root": str(root),
                            "manifest": "bootstrap.manifest.json",
                            "workflow": "dsp",
                            "template_version": "v1",
                            "rule_version": "v1",
                            "artifact_root": "artifacts_sandbox/fresh-link",
                            "sandbox": "fresh-link",
                        }
                    ).encode("utf-8"),
                    headers={"Content-Type": "application/json"},
                    method="POST",
                )
                with urlopen(req) as resp:
                    response_payload = json.loads(resp.read().decode("utf-8"))
                self.assertEqual(response_payload.get("status"), "ok")
                self.assertEqual(response_payload.get("result", {}).get("status"), "ok")

                sandbox_ctx = ui_shell_module._resolve_ui_context(
                    root=root,
                    runtime_env_raw="",
                    manifest_raw="",
                    artifact_root_raw="",
                    sandbox_raw="fresh-link",
                    workflow="dsp",
                    template_version="v1",
                    rule_version="v1",
                )
                self.assertTrue((sandbox_ctx.db_path or Path("")).exists())
            finally:
                server.shutdown()
                server.server_close()
                thread.join(timeout=2.0)

    def test_ssp_media_demand_endpoint_wraps_sandbox_errors_as_json(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            try:
                server = ThreadingHTTPServer(("127.0.0.1", 0), UiRequestHandler)
            except PermissionError:
                self.skipTest("sandbox 禁止本地 socket bind，略過 media-demand error wrapper")
            thread = threading.Thread(target=server.serve_forever, daemon=True)
            thread.start()
            try:
                host, port = server.server_address
                query = urlencode(
                    {
                        "root": str(root),
                        "manifest": "bootstrap.manifest.json",
                        "workflow": "ssp",
                        "template_version": "v1",
                        "rule_version": "v1",
                        "artifact_root": "artifacts",
                        "sandbox": "bad/id",
                    }
                )
                with self.assertRaises(HTTPError) as exc_ctx:
                    urlopen(f"http://{host}:{port}/api/ssp/media-demand?{query}")
                self.assertEqual(exc_ctx.exception.code, 400)
                payload = json.loads(exc_ctx.exception.read().decode("utf-8"))
                self.assertEqual(payload.get("status"), "error")
                self.assertEqual(payload.get("error_code"), "UI_MEDIA_DEMAND_FAILED")
                self.assertIn("sandbox", str(payload.get("message") or ""))
            finally:
                server.shutdown()
                server.server_close()
                thread.join(timeout=2.0)

    def test_ui_shell_strict_gate_respected(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            ctx = self._ctx(root)
            dispatch_action(ctx, {"action": "bootstrap"})

            conn = sqlite3.connect(str(root / "data" / "mdrep.sqlite"))
            try:
                conn.execute("DELETE FROM rule_bindings WHERE workflow='ssp'")
                conn.commit()
            finally:
                conn.close()

            with self.assertRaises(AcceptanceGateError):
                dispatch_action(ctx, {"action": "save", "rows": [self._full_row()]})

    def test_ui_shell_status_has_machine_readable_summaries(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            ctx = self._ctx(root)
            dispatch_action(ctx, {"action": "bootstrap"})
            status = collect_runtime_status(ctx)
            self.assertIn("health", status)
            self.assertIn("recent", status)
            self.assertIn("run_log", status["recent"])
            self.assertIn("audit_log", status["recent"])
            # JSON serializable snapshot for UI rendering.
            json.dumps(status, ensure_ascii=False)

    def test_ssp_status_recent_entries_do_not_mix_dsp_runs(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            dsp_ctx = self._ctx(root)
            ssp_ctx = self._ctx(root, workflow="ssp")
            dispatch_action(dsp_ctx, {"action": "bootstrap"})
            dispatch_action(dsp_ctx, {"action": "save", "rows": [self._full_row(經銷商="DSP_ONLY")]})
            dispatch_action(dsp_ctx, {"action": "tab4_delivery", "main_tab": "dsp_tab3", "sub_tab": "pivot"})
            dispatch_action(
                dsp_ctx,
                {
                    "action": "export",
                    "main_tab": "dsp_tab4",
                    "sub_tab": "overview",
                },
            )
            dispatch_action(ssp_ctx, {"action": "save", "workflow": "ssp", "rows": [self._full_row(經銷商="SSP_ONLY")]})

            ssp_status = collect_runtime_status(ssp_ctx)
            run_log = ssp_status["recent"]["run_log"]
            self.assertGreaterEqual(len(run_log), 1)
            self.assertTrue(all(str(item.get("workflow") or "") == "ssp" for item in run_log), run_log)
            audit_log = ssp_status["recent"]["audit_log"]
            self.assertGreaterEqual(len(audit_log), 1)
            self.assertTrue(
                all(
                    str(item.get("scope") or "").startswith("ssp:")
                    or str(item.get("workflow") or "") == "ssp"
                    for item in audit_log
                ),
                audit_log,
            )
            self.assertFalse(any(str(item.get("event_type") or "") == "tab4_delivery" for item in audit_log), audit_log)
            self.assertEqual(ssp_status["recent"]["publish_runs"], [])
            self.assertEqual(ssp_status["recent"]["evidence_index"], [])
            self.assertFalse(bool(ssp_status["tab4_delivery"].get("ready")))
            self.assertEqual(str(ssp_status["tab4_delivery"].get("last_delivery_run_id") or ""), "")

    def test_ssp_frame_falls_back_to_canonical_rows_when_ssp_raw_empty(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            ctx = self._ctx(root, workflow="ssp")
            dispatch_action(ctx, {"action": "bootstrap"})
            dispatch_action(
                ctx,
                {
                    "action": "save",
                    "workflow": "ssp",
                    "rows": [self._full_row(經銷商="SSP_FRAME", 最終經銷商="SSP_FRAME_CANONICAL")],
                },
            )

            frame = collect_workflow_frame(ctx)
            self.assertEqual(frame.get("source_table"), "canonical_raw")
            self.assertEqual(frame.get("row_count"), 1)
            self.assertIn("經銷商", frame.get("field_names") or [])
            self.assertIn("最終經銷商", frame.get("manual_fields") or [])
            self.assertEqual((frame.get("rows") or [{}])[0].get("經銷商"), "SSP_FRAME")
            self.assertEqual((frame.get("rows") or [{}])[0].get("最終經銷商"), "SSP_FRAME_CANONICAL")

    def test_ssp_media_demand_config_loads_defaults_and_saves_per_env(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            prod_ctx = self._ctx(root, workflow="ssp")
            test_ctx = UiContext(
                root=root,
                runtime_env="test",
                manifest_rel="bootstrap.test.manifest.json",
                workflow="ssp",
                template_version="v1",
                rule_version="v1",
                artifact_root=(root / "artifacts_test").resolve(),
            )
            dispatch_action(prod_ctx, {"action": "bootstrap"})
            dispatch_action(test_ctx, {"action": "bootstrap"})

            prod_frame = collect_workflow_frame(prod_ctx)
            prod_media = prod_frame.get("ssp_media_demand") or {}
            self.assertEqual(prod_media.get("storage_source"), "defaults")
            self.assertEqual(prod_media.get("defaults_source"), "template+json")
            self.assertEqual(prod_media.get("template_path"), str((root / "templates" / "ssp_template.xlsx").resolve()))
            self.assertEqual(
                prod_media.get("group_overrides_path"),
                str((root / "data_seed" / "templates_rules_mapping" / "group_overrides.json").resolve()),
            )
            self.assertEqual(len(prod_media.get("categories") or []), 5)
            self.assertGreaterEqual(len(prod_media.get("slots") or []), 5)
            self.assertTrue(any(str(slot.get("media_quality") or "").strip() for slot in (prod_media.get("slots") or [])))
            self.assertTrue(any(str(slot.get("target_fr") or "").strip() for slot in (prod_media.get("slots") or [])))

            save_out = dispatch_action(
                test_ctx,
                {
                    "action": "ssp_media_save",
                    "ssp_media_slots": [
                        {
                            "category": "蓋板",
                            "slot_order": 0,
                            "placement_id": "99999",
                            "placement_name": "TEST_SLOT",
                            "media_quality": "知名媒體",
                            "need_call": True,
                            "target_fr": "60-80%",
                            "remark": "test env only",
                            "media_target": 3210,
                            "is_active": True,
                        }
                    ],
                },
            )
            self.assertEqual(save_out.get("status"), "ok")
            self.assertEqual(int(save_out.get("row_count") or 0), 1)
            self.assertTrue(str(save_out.get("run_id") or "").startswith("run-"))

            cfg = build_config(root, test_ctx.manifest_rel, test_ctx.runtime_env)
            conn = sqlite3.connect(str(cfg.db_path))
            try:
                run_row = conn.execute(
                    """
                    SELECT run_type, workflow, template_version, rule_version, detail_json
                    FROM run_log
                    WHERE run_id = ?
                    """,
                    (str(save_out.get("run_id") or ""),),
                ).fetchone()
                self.assertIsNotNone(run_row)
                assert run_row is not None
                self.assertEqual(str(run_row[0]), "ssp_media_save")
                self.assertEqual(str(run_row[1]), "ssp")
                self.assertEqual(str(run_row[2]), "v1")
                self.assertEqual(str(run_row[3]), "v1")
                run_detail = json.loads(str(run_row[4]))
                self.assertEqual(str(run_detail.get("runtime_env") or ""), "test")
                self.assertEqual(int(run_detail.get("row_count") or 0), 1)
                self.assertTrue(str(run_detail.get("template_id") or ""))
                self.assertTrue(str(run_detail.get("mapping_version") or ""))
                self.assertTrue(str(run_detail.get("rule_hash") or ""))

                audit_row = conn.execute(
                    """
                    SELECT event_type, scope, status, payload_json
                    FROM audit_log
                    WHERE event_type = 'ssp_media_save'
                    ORDER BY id DESC
                    LIMIT 1
                    """
                ).fetchone()
                self.assertIsNotNone(audit_row)
                assert audit_row is not None
                self.assertEqual(str(audit_row[0]), "ssp_media_save")
                self.assertEqual(str(audit_row[1]), "service")
                self.assertEqual(str(audit_row[2]), "ok")
                audit_payload = json.loads(str(audit_row[3]))
                self.assertEqual(str(audit_payload.get("workflow") or ""), "ssp")
                self.assertEqual(str(audit_payload.get("run_id") or ""), str(save_out.get("run_id") or ""))
                self.assertEqual(str(audit_payload.get("template_version") or ""), "v1")
                self.assertEqual(str(audit_payload.get("rule_version") or ""), "v1")
                self.assertTrue(str(audit_payload.get("canonical_token") or ""))
                self.assertEqual(str(audit_payload.get("runtime_env") or ""), "test")
                self.assertEqual(int(audit_payload.get("row_count") or 0), 1)
            finally:
                conn.close()

            refreshed_test = collect_workflow_frame(test_ctx).get("ssp_media_demand") or {}
            refreshed_prod = collect_workflow_frame(prod_ctx).get("ssp_media_demand") or {}
            self.assertEqual(refreshed_test.get("storage_source"), "db")
            self.assertEqual((refreshed_test.get("slots") or [{}])[0].get("placement_id"), "99999")
            self.assertEqual((refreshed_test.get("slots") or [{}])[0].get("media_quality"), "知名媒體")
            self.assertEqual((refreshed_test.get("slots") or [{}])[0].get("need_call"), True)
            self.assertEqual((refreshed_test.get("slots") or [{}])[0].get("target_fr"), "60-80%")
            self.assertEqual((refreshed_test.get("slots") or [{}])[0].get("media_target"), 3210.0)
            self.assertEqual(refreshed_prod.get("storage_source"), "defaults")
            self.assertNotEqual((refreshed_prod.get("slots") or [{}])[0].get("placement_id"), "99999")

    def test_ssp_media_demand_config_uses_json_only_when_local_template_missing(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root, include_ssp_template=False)
            ctx = self._ctx(root, workflow="ssp")
            dispatch_action(ctx, {"action": "bootstrap"})

            media = (collect_workflow_frame(ctx).get("ssp_media_demand") or {})
            slots = media.get("slots") or []
            self.assertEqual(media.get("storage_source"), "defaults")
            self.assertEqual(media.get("defaults_source"), "json")
            self.assertEqual(media.get("template_path"), "")
            self.assertEqual(
                media.get("group_overrides_path"),
                str((root / "data_seed" / "templates_rules_mapping" / "group_overrides.json").resolve()),
            )
            self.assertEqual((slots[0] if slots else {}).get("placement_id"), "8435")
            self.assertEqual((slots[0] if slots else {}).get("placement_name"), "MW_蓋版_COOL")
            self.assertEqual((slots[0] if slots else {}).get("media_quality"), "")
            self.assertEqual((slots[0] if slots else {}).get("target_fr"), "")

    def test_ssp_media_demand_view_honors_scope_and_source_contract(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            ctx = self._ctx(root, workflow="ssp")
            dispatch_action(ctx, {"action": "bootstrap"})
            dispatch_action(
                ctx,
                {
                    "action": "ssp_media_save",
                    "ssp_media_slots": [
                        {"category": "蓋板", "slot_order": 0, "placement_id": "111", "placement_name": "A", "remark": "", "media_target": 100, "is_active": True},
                        {"category": "蓋板", "slot_order": 1, "placement_id": "222", "placement_name": "B", "remark": "", "media_target": 100, "is_active": True},
                        {"category": "蓋板", "slot_order": 2, "placement_id": "333", "placement_name": "C", "remark": "", "media_target": 100, "is_active": True},
                    ],
                },
            )

            cfg = build_config(root, "bootstrap.manifest.json", "prod")
            repo = SQLiteRepository(cfg.db_path, project_root=root)
            repo.replace_ssp_raw_rows(
                [
                    {
                        "source": "times_api",
                        "ts": "2026-05-10 23:00:00",
                        "date": "2026-05-10",
                        "hour": 23,
                        "placement_id": 111,
                        "placement_name": "A",
                        "request": 200,
                        "impression": 80,
                        "clicks": 4,
                        "revenue": 50,
                        "dsp_amount": 30,
                    },
                    {
                        "source": "times_api",
                        "ts": "2026-05-10 11:00:00",
                        "date": "2026-05-10",
                        "hour": 11,
                        "placement_id": 222,
                        "placement_name": "B",
                        "request": 100,
                        "impression": 60,
                        "clicks": 3,
                        "revenue": 40,
                        "dsp_amount": 20,
                    },
                    {
                        "source": "backup_api",
                        "ts": "2026-05-10 12:00:00",
                        "date": "2026-05-10",
                        "hour": 12,
                        "placement_id": 222,
                        "placement_name": "B",
                        "request": 250,
                        "impression": 90,
                        "clicks": 5,
                        "revenue": 55,
                        "dsp_amount": 35,
                    },
                    {
                        "source": "times_api",
                        "ts": "2026-05-09 11:00:00",
                        "date": "2026-05-09",
                        "hour": 11,
                        "placement_id": 333,
                        "placement_name": "C",
                        "request": 999,
                        "impression": 300,
                        "clicks": 10,
                        "revenue": 70,
                        "dsp_amount": 50,
                    },
                ]
            )

            view = repo.resolve_ssp_media_demand_view(
                runtime_env="prod",
                data_seed_root=cfg.data_seed_root,
                category="蓋板",
                source="__all__",
                start_date="2026-05-01",
                end_date="2026-05-10",
                scope_mode="all",
                day_limit=7,
                threshold=60,
                only_unmet=False,
            )
            rows = view.get("rows") or []
            self.assertEqual([row["slot"]["placement_id"] for row in rows], ["222", "111", "333"])
            self.assertEqual(view.get("source"), "__all__")
            self.assertAlmostEqual(float(rows[0]["latest_compliance_rate"]), 350.0)
            self.assertAlmostEqual(float(rows[0]["metrics_by_date"]["2026-05-10"]["all"]["complianceRate"]), 350.0)
            self.assertAlmostEqual(float(rows[1]["latest_compliance_rate"]), 200.0)
            self.assertAlmostEqual(float(rows[0]["latest_request"]), 350.0)

            unmet_view = repo.resolve_ssp_media_demand_view(
                runtime_env="prod",
                data_seed_root=cfg.data_seed_root,
                category="蓋板",
                source="__all__",
                start_date="2026-05-01",
                end_date="2026-05-10",
                scope_mode="all",
                day_limit=7,
                threshold=250,
                only_unmet=True,
            )
            unmet_rows = unmet_view.get("rows") or []
            self.assertEqual([row["slot"]["placement_id"] for row in unmet_rows], ["111"])

            scoped_view = repo.resolve_ssp_media_demand_view(
                runtime_env="prod",
                data_seed_root=cfg.data_seed_root,
                category="蓋板",
                source="times_api",
                start_date="2026-05-01",
                end_date="2026-05-10",
                scope_mode="07-22",
                day_limit=7,
                threshold=150,
                only_unmet=True,
            )
            scoped_rows = scoped_view.get("rows") or []
            self.assertEqual(scoped_view.get("source"), "times_api")
            self.assertEqual(scoped_view.get("scope_mode"), "07-22")
            self.assertEqual([row["slot"]["placement_id"] for row in scoped_rows], ["222"])
            self.assertAlmostEqual(float(scoped_rows[0]["latest_request"]), 100.0)
            self.assertAlmostEqual(float(scoped_rows[0]["latest_compliance_rate"]), 100.0)

    def test_tab4_delivery_state_relocks_after_rawdata_save(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            ctx = self._ctx(root)
            dispatch_action(ctx, {"action": "bootstrap"})
            dispatch_action(ctx, {"action": "save", "rows": [self._full_row()]})

            delivered = dispatch_action(
                ctx,
                {
                    "action": "tab4_delivery",
                    "main_tab": "dsp_tab3",
                    "sub_tab": "pivot",
                },
            )
            self.assertTrue(bool(delivered.get("ready")))

            ready_status = collect_runtime_status(ctx)
            self.assertTrue(bool(ready_status.get("tab4_delivery", {}).get("ready")))
            self.assertEqual(
                str(ready_status.get("tab4_delivery", {}).get("reason") or ""),
                "pivot_handoff",
            )

            dispatch_action(ctx, {"action": "save", "rows": [self._full_row(最終經銷商="A2")]})
            relocked_status = collect_runtime_status(ctx)
            self.assertFalse(bool(relocked_status.get("tab4_delivery", {}).get("ready")))
            self.assertEqual(
                str(relocked_status.get("tab4_delivery", {}).get("reason") or ""),
                "rawdata_saved",
            )

    def test_dsp_export_requires_tab4_route_and_delivery_snapshot_identity(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            ctx = self._ctx(root)
            dispatch_action(ctx, {"action": "bootstrap"})
            dispatch_action(ctx, {"action": "save", "rows": [self._full_row()]})

            with self.assertRaises(PermissionError):
                dispatch_action(
                    ctx,
                    {
                        "action": "export",
                        "main_tab": "dsp_tab3",
                        "sub_tab": "rawdata",
                    },
                )

            delivered = dispatch_action(
                ctx,
                {
                    "action": "tab4_delivery",
                    "main_tab": "dsp_tab3",
                    "sub_tab": "pivot",
                },
            )
            self.assertTrue(bool(delivered.get("ready")))
            delivery_token = str(delivered.get("delivery_snapshot_token") or "")
            self.assertTrue(delivery_token)

            status = collect_runtime_status(ctx)
            status_delivery = status.get("tab4_delivery", {})
            self.assertEqual(str(status_delivery.get("delivery_snapshot_token") or ""), delivery_token)
            self.assertEqual(str(status_delivery.get("last_delivery_run_id") or ""), str(delivered.get("run_id") or ""))

            frame = collect_workflow_frame(ctx)
            frame_snapshot = frame.get("tab4_delivery_snapshot", {})
            self.assertEqual(str(frame_snapshot.get("delivery_snapshot_token") or ""), delivery_token)
            self.assertEqual(str(frame_snapshot.get("delivery_run_id") or ""), str(delivered.get("run_id") or ""))
            preview_contract = frame.get("tab4_preview_contract", {})
            self.assertEqual(str(preview_contract.get("kind") or ""), "template_preview")
            summary_preview = frame.get("tab4_preview_template_summary", {})
            detail_preview = frame.get("tab4_preview_template_detail", {})
            self.assertIn("tab4_preview_template_summary", frame)
            self.assertIn("tab4_preview_template_detail", frame)
            self.assertGreater(float((summary_preview.get("monthTotals") or [0.0] * 12)[4]), 0.0)
            self.assertGreater(float((summary_preview.get("rows") or [{}])[0].get("monthlyAmounts", [0.0] * 12)[4]), 0.0)
            self.assertGreater(float((detail_preview.get("kpiRows") or [{}])[0].get("monthlyAmounts", [0.0] * 12)[4]), 0.0)
            first_section = (detail_preview.get("sections") or [{}])[0]
            self.assertGreater(float(((first_section.get("rows") or [{}])[0]).get("monthlyAmounts", [0.0] * 12)[4]), 0.0)
            self.assertNotIn("tab4_template_summary", frame)
            self.assertNotIn("tab4_template_detail", frame)

            with self.assertRaises(PermissionError):
                dispatch_action(
                    ctx,
                    {
                        "action": "export",
                        "main_tab": "dsp_tab4",
                        "sub_tab": "pivot",
                    },
                )

            conn = sqlite3.connect(str(root / "data" / "mdrep.sqlite"))
            try:
                conn.execute(
                    "UPDATE run_log SET canonical_token = '' WHERE run_id = ?",
                    (str(delivered.get("run_id") or ""),),
                )
                conn.commit()
            finally:
                conn.close()

            with self.assertRaises(PermissionError):
                dispatch_action(
                    ctx,
                    {
                        "action": "export",
                        "main_tab": "dsp_tab4",
                        "sub_tab": "overview",
                    },
                )
            status_after_token_removed = collect_runtime_status(ctx)
            self.assertFalse(bool(status_after_token_removed.get("tab4_delivery", {}).get("ready")))
            self.assertEqual(
                str(status_after_token_removed.get("tab4_delivery", {}).get("reason") or ""),
                "missing_snapshot_token",
            )

            delivered = dispatch_action(
                ctx,
                {
                    "action": "tab4_delivery",
                    "main_tab": "dsp_tab3",
                    "sub_tab": "pivot",
                },
            )
            delivery_token = str(delivered.get("delivery_snapshot_token") or "")
            self.assertTrue(delivery_token)

            export_out = dispatch_action(
                ctx,
                {
                    "action": "export",
                    "main_tab": "dsp_tab4",
                    "sub_tab": "overview",
                },
            )
            self.assertEqual(str(export_out.get("delivery_snapshot_token") or ""), delivery_token)
            self.assertEqual(str(export_out.get("delivery_run_id") or ""), str(delivered.get("run_id") or ""))

    def test_dsp_export_requires_explicit_tab4_route_in_api(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            ctx = self._ctx(root)
            dispatch_action(ctx, {"action": "bootstrap"})
            dispatch_action(ctx, {"action": "save", "rows": [self._full_row()]})
            dispatch_action(
                ctx,
                {
                    "action": "tab4_delivery",
                    "main_tab": "dsp_tab3",
                    "sub_tab": "pivot",
                },
            )
            with self.assertRaisesRegex(PermissionError, "dsp export must be triggered from dsp_tab4"):
                dispatch_action(
                    ctx,
                    {
                        "action": "export",
                    },
                )

    def test_cli_dsp_export_defaults_route_to_tab4_overview(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            rows_json_path = root / "rows.json"
            rows_json_path.write_text(
                json.dumps([self._full_row()], ensure_ascii=False),
                encoding="utf-8",
            )

            code, boot = self._run_cli_json(["--root", str(root), "bootstrap"])
            self.assertEqual(code, 0)
            self.assertEqual(boot.get("status"), "ok")

            code, save_out = self._run_cli_json(
                [
                    "--root",
                    str(root),
                    "save",
                    "--workflow",
                    "dsp",
                    "--template-version",
                    "v1",
                    "--rule-version",
                    "v1",
                    "--rows-json",
                    str(rows_json_path),
                ]
            )
            self.assertEqual(code, 0)
            self.assertEqual(save_out.get("status"), "ok")

            ctx = self._ctx(root)
            delivered = dispatch_action(
                ctx,
                {
                    "action": "tab4_delivery",
                    "main_tab": "dsp_tab3",
                    "sub_tab": "pivot",
                },
            )
            delivery_token = str(delivered.get("delivery_snapshot_token") or "")
            self.assertTrue(delivery_token)

            code, export_out = self._run_cli_json(
                [
                    "--root",
                    str(root),
                    "export",
                    "--workflow",
                    "dsp",
                    "--template-version",
                    "v1",
                    "--rule-version",
                    "v1",
                ]
            )
            self.assertEqual(code, 0)
            self.assertEqual(export_out.get("status"), "ok")
            result = export_out.get("result") if isinstance(export_out, dict) else {}
            self.assertIsInstance(result, dict)
            artifact_path = Path(str((result or {}).get("artifact_path") or ""))
            self.assertTrue(artifact_path.exists())
            self.assertRegex(artifact_path.name, r"^2026 DSP投資量報表_\d{4}-\d{4}\.xlsx$")
            self.assertEqual(str((result or {}).get("delivery_snapshot_token") or ""), delivery_token)

    def test_dsp_period_bound_template_requires_matching_period_for_save_and_export(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            (root / "templates" / "dsp_tab4_template.xlsx.period.json").write_text(
                json.dumps(
                    {
                        "week_start": "2026-01-01",
                        "week_end": "2026-05-03",
                        "note": "test narrow period window",
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            ctx = self._ctx(root)
            dispatch_action(ctx, {"action": "bootstrap"})

            save_out = dispatch_action(
                ctx,
                {
                    "action": "save",
                    "rows": [self._full_row()],
                    "period_week_start": "2026-04-27",
                    "period_week_end": "2026-05-03",
                },
            )
            self.assertTrue(str(save_out.get("run_id") or ""))

            dispatch_action(
                ctx,
                {
                    "action": "tab4_delivery",
                    "main_tab": "dsp_tab3",
                    "sub_tab": "pivot",
                    "period_week_start": "2026-04-27",
                    "period_week_end": "2026-05-03",
                },
            )

            export_out = dispatch_action(
                ctx,
                {
                    "action": "export",
                    "main_tab": "dsp_tab4",
                    "sub_tab": "overview",
                    "period_week_start": "2026-04-27",
                    "period_week_end": "2026-05-03",
                },
            )
            self.assertTrue(Path(str(export_out.get("artifact_path") or "")).exists())

            with self.assertRaisesRegex(ValueError, "dsp period has no matching base template"):
                dispatch_action(
                    ctx,
                    {
                        "action": "save",
                        "rows": [self._full_row()],
                        "period_week_start": "2026-05-04",
                        "period_week_end": "2026-05-10",
                    },
                )
            with self.assertRaisesRegex(PermissionError, "period mismatch"):
                dispatch_action(
                    ctx,
                    {
                        "action": "export",
                        "main_tab": "dsp_tab4",
                        "sub_tab": "overview",
                        "period_week_start": "2026-05-04",
                        "period_week_end": "2026-05-10",
                    },
                )

    def test_ui_shell_root_returns_503_when_frontend_dist_missing(self) -> None:
        try:
            server = ThreadingHTTPServer(("127.0.0.1", 0), UiRequestHandler)
        except PermissionError:
            self.skipTest("sandbox 禁止本地 socket bind，略過 root entry smoke")

        thread = threading.Thread(target=server.serve_forever, daemon=True)
        thread.start()
        try:
            with tempfile.TemporaryDirectory() as missing_dist:
                with patch.object(ui_shell_module, "FRONTEND_DIST_DIR", Path(missing_dist)):
                    host, port = server.server_address
                    req = Request(f"http://{host}:{port}/")
                    try:
                        with urlopen(req) as resp:
                            body = resp.read().decode("utf-8")
                            status_code = resp.status
                            content_type = resp.headers.get("Content-Type", "")
                    except HTTPError as exc:
                        body = exc.read().decode("utf-8")
                        status_code = exc.code
                        content_type = exc.headers.get("Content-Type", "")
                    self.assertEqual(status_code, 503)
                    self.assertIn("text/html", content_type)
                    self.assertIn("React frontend artifact not found", body)
        finally:
            server.shutdown()
            server.server_close()
            thread.join(timeout=2.0)

    def test_ui_shell_root_serves_frontend_entry_when_dist_exists(self) -> None:
        frontend_index = ui_shell_module.FRONTEND_DIST_DIR / "index.html"
        if not frontend_index.exists():
            self.skipTest("frontend dist 不存在，略過 frontend entry smoke（請先 pnpm build）")

        try:
            server = ThreadingHTTPServer(("127.0.0.1", 0), UiRequestHandler)
        except PermissionError:
            self.skipTest("sandbox 禁止本地 socket bind，略過 root entry smoke")

        thread = threading.Thread(target=server.serve_forever, daemon=True)
        thread.start()
        try:
            host, port = server.server_address
            with urlopen(Request(f"http://{host}:{port}/")) as resp:
                body = resp.read().decode("utf-8")
                self.assertEqual(resp.status, 200)
                self.assertIn("text/html", resp.headers.get("Content-Type", ""))
                self.assertIn("<div id=\"root\"></div>", body)
                self.assertNotIn("React frontend artifact not found", body)
        finally:
            server.shutdown()
            server.server_close()
            thread.join(timeout=2.0)

    def test_ui_browser_acceptance_smoke(self) -> None:
        frontend_index = ui_shell_module.FRONTEND_DIST_DIR / "index.html"
        if not frontend_index.exists():
            self.skipTest("frontend dist 不存在，無法做真 browser acceptance（請先 pnpm build）")

        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            ctx = self._ctx(root, workflow="dsp")
            browser_acceptance_datetime = self._dsp_bucket_datetime(weeks_ago=2)
            browser_acceptance_date = browser_acceptance_datetime[:10]
            dispatch_action(ctx, {"action": "bootstrap"})
            dispatch_action(
                ctx,
                {
                    "action": "save",
                    "rows": [
                        self._full_row(日期時間=browser_acceptance_datetime),
                        self._full_row(
                            日期時間=browser_acceptance_datetime,
                            訂單="O2",
                            素材="C2",
                            廣告形式="Video",
                            最終廣告形式="Video",
                            執行金額=20.0,
                            媒體費用=16.0,
                        ),
                    ],
                },
            )

            try:
                server = ThreadingHTTPServer(("127.0.0.1", 0), UiRequestHandler)
            except PermissionError:
                self.skipTest("sandbox 禁止本地 socket bind，略過 browser acceptance smoke")

            try:
                from playwright.sync_api import Error as PlaywrightError
                from playwright.sync_api import sync_playwright
            except Exception:
                self.skipTest("缺少 playwright 依賴，請先安裝 playwright 並執行 playwright install chromium")

            thread = threading.Thread(target=server.serve_forever, daemon=True)
            thread.start()
            try:
                host, port = server.server_address
                base_url = f"http://{host}:{port}"
                with sync_playwright() as p:
                    try:
                        browser = p.chromium.launch(headless=True)
                    except PlaywrightError as exc:
                        self.skipTest(f"playwright browser context 無法啟動，請先 playwright install chromium: {exc}")
                    context = browser.new_context(viewport={"width": 1680, "height": 1400}, accept_downloads=True)
                    page = context.new_page()
                    try:
                        query = urlencode(
                            {
                                "root": str(root),
                                "manifest": "bootstrap.manifest.json",
                                "workflow": "dsp",
                                "template_version": "v1",
                                "rule_version": "v1",
                                "artifact_root": "artifacts",
                            }
                        )
                        page.goto(f"{base_url}/?{query}", wait_until="domcontentloaded")
                        layout_box = page.locator(".layout").bounding_box()
                        self.assertIsNotNone(layout_box)
                        self.assertGreaterEqual(int(layout_box["width"] if layout_box else 0), 1660)
                        self.assertEqual(int(layout_box["x"] if layout_box else -1), 0)
                        self.assertFalse(page.is_visible("text=React frontend artifact not found"))
                        self.assertTrue(page.is_visible("text=MDREP Frontend Shell"))
                        self.assertTrue(page.is_visible("text=Use DSP"))
                        self.assertTrue(page.is_visible("text=Use SSP"))
                        self.assertTrue(page.is_visible("text=Runtime Context"))
                        self.assertTrue(page.is_visible("text=Service Input"))

                        # Main/Sub Tabs aria-selected：預設 DSP tab3 + rawdata。
                        main_tab3 = page.locator("[data-testid='main-tab-dsp-tab3']")
                        main_tab4 = page.locator("[data-testid='main-tab-dsp-tab4']")
                        sub_rawdata = page.locator("[data-testid='sub-tab-rawdata']")
                        sub_pivot = page.locator("[data-testid='sub-tab-pivot']")
                        sub_result = page.locator("[data-testid='sub-tab-result']")
                        main_tab3.click()
                        self.assertEqual(main_tab3.get_attribute("aria-selected"), "true")
                        self.assertEqual(main_tab4.get_attribute("aria-selected"), "false")
                        sub_rawdata.click()
                        self.assertEqual(sub_rawdata.get_attribute("aria-selected"), "true")
                        self.assertEqual(sub_result.get_attribute("aria-selected"), "false")
                        rawdata_workspace = page.locator("[data-testid='section-rawdata']")
                        self.assertTrue(rawdata_workspace.is_visible())
                        self.assertTrue(rawdata_workspace.get_by_text("編修工作量").first.is_visible())
                        self.assertTrue(rawdata_workspace.get_by_text("提交判定").first.is_visible())
                        self.assertEqual(rawdata_workspace.locator("[data-testid='action-export']").count(), 0)
                        self.assertEqual(rawdata_workspace.get_by_text("Keyword Filter").count(), 0)
                        self.assertEqual(page.locator("[data-testid='dsp-rawdata-view-mode']").count(), 1)
                        view_user = page.locator("[data-testid='dsp-rawdata-view-user']")
                        view_verify = page.locator("[data-testid='dsp-rawdata-view-verify']")
                        view_pm = page.locator("[data-testid='dsp-rawdata-view-pm']")
                        self.assertEqual(view_user.get_attribute("aria-selected"), "true")
                        self.assertEqual(view_verify.get_attribute("aria-selected"), "false")
                        self.assertEqual(view_pm.get_attribute("aria-selected"), "false")
                        self.assertTrue(page.locator("[data-testid='dsp-rawdata-date-bucket']").is_visible())
                        self.assertTrue(page.locator("[data-testid='dsp-rawdata-row-limit']").is_visible())
                        self.assertTrue(page.locator("[data-testid='dsp-rawdata-distributor']").is_visible())
                        self.assertTrue(page.locator("[data-testid='dsp-rawdata-ad-format']").is_visible())
                        self.assertTrue(page.locator("[data-testid='dsp-rawdata-size']").is_visible())
                        self.assertTrue(page.locator("[data-testid='dsp-rawdata-template']").is_visible())
                        self.assertEqual(page.locator("[data-testid='dsp-rawdata-date-bucket'] option").count(), 4)
                        main_rawdata_table = rawdata_workspace.locator("table", has=page.locator("thead tr th", has_text="最終經銷商")).first
                        rawdata_row_count = main_rawdata_table.locator("tbody tr").count()
                        expected_user_headers = [
                            "日期時間", "最終經銷商", "訂單", "素材", "最終廣告形式", "尺寸", "素材樣板", "執行金額", "媒體費用",
                        ]
                        user_headers = [
                            main_rawdata_table.locator("thead tr th").nth(i).inner_text().strip()
                            for i in range(main_rawdata_table.locator("thead tr th").count())
                        ]
                        self.assertEqual(user_headers, expected_user_headers)
                        rawdata_table_wrap = page.locator("[data-testid='rawdata-table-wrap']").first
                        user_wrap_metrics = rawdata_table_wrap.evaluate(
                            """el => ({
                              overflowX: window.getComputedStyle(el).overflowX,
                              scrollWidth: el.scrollWidth,
                              clientWidth: el.clientWidth
                            })"""
                        )
                        self.assertIn(str(user_wrap_metrics.get("overflowX", "")), {"hidden", "clip"})
                        self.assertLessEqual(
                            int(user_wrap_metrics.get("scrollWidth", 0)),
                            int(user_wrap_metrics.get("clientWidth", 0)) + 1,
                        )
                        first_header_cell = main_rawdata_table.locator("thead tr th").first
                        first_header_width_before = int(
                            first_header_cell.evaluate("el => Math.round(el.getBoundingClientRect().width)")
                        )
                        first_resizer = page.locator("[data-testid='rawdata-col-resizer-0']").first
                        first_resizer_box = first_resizer.bounding_box()
                        self.assertIsNotNone(first_resizer_box)
                        if first_resizer_box:
                            page.mouse.move(
                                first_resizer_box["x"] + (first_resizer_box["width"] / 2),
                                first_resizer_box["y"] + (first_resizer_box["height"] / 2),
                            )
                            page.mouse.down()
                            page.mouse.move(
                                first_resizer_box["x"] + (first_resizer_box["width"] / 2) - 80,
                                first_resizer_box["y"] + (first_resizer_box["height"] / 2),
                            )
                            page.mouse.up()
                        page.wait_for_timeout(80)
                        first_header_width_after_resize = int(
                            first_header_cell.evaluate("el => Math.round(el.getBoundingClientRect().width)")
                        )
                        self.assertGreaterEqual(first_header_width_after_resize, first_header_width_before)
                        view_verify.click()
                        self.assertEqual(view_verify.get_attribute("aria-selected"), "true")
                        self.assertEqual(main_rawdata_table.locator("tbody tr").count(), rawdata_row_count)
                        expected_verify_headers = [
                            "日期時間", "最終經銷商", "訂單", "素材", "最終廣告形式", "尺寸", "素材樣板", "執行金額", "媒體費用",
                            "原始經銷商", "原始廣告形式", "最終來源_經銷商", "規則命中_經銷商", "最終來源_廣告形式", "規則命中_廣告形式",
                        ]
                        verify_headers = [
                            main_rawdata_table.locator("thead tr th").nth(i).inner_text().strip()
                            for i in range(main_rawdata_table.locator("thead tr th").count())
                        ]
                        self.assertEqual(verify_headers, expected_verify_headers)
                        verify_wrap_metrics = rawdata_table_wrap.evaluate(
                            """el => ({
                              overflowX: window.getComputedStyle(el).overflowX,
                              scrollWidth: el.scrollWidth,
                              clientWidth: el.clientWidth
                            })"""
                        )
                        self.assertIn(str(verify_wrap_metrics.get("overflowX", "")), {"auto", "scroll"})
                        self.assertGreater(
                            int(verify_wrap_metrics.get("scrollWidth", 0)),
                            int(verify_wrap_metrics.get("clientWidth", 0)),
                        )
                        view_pm.click()
                        self.assertEqual(view_pm.get_attribute("aria-selected"), "true")
                        self.assertEqual(main_rawdata_table.locator("tbody tr").count(), rawdata_row_count)
                        expected_pm_headers = [
                            "日期時間", "最終經銷商", "訂單", "素材", "最終廣告形式", "尺寸", "素材樣板", "執行金額", "媒體費用",
                            "原始經銷商", "原始廣告形式", "最終來源_經銷商", "規則命中_經銷商", "最終來源_廣告形式", "規則命中_廣告形式",
                            "經銷商", "分類層級B", "分類層級C", "分類層級D", "廣告形式", "系統營收",
                        ]
                        pm_headers = [
                            main_rawdata_table.locator("thead tr th").nth(i).inner_text().strip()
                            for i in range(main_rawdata_table.locator("thead tr th").count())
                        ]
                        self.assertEqual(pm_headers, expected_pm_headers)
                        view_user.click()
                        self.assertEqual(view_user.get_attribute("aria-selected"), "true")
                        first_header_width_after_switch = int(
                            main_rawdata_table.locator("thead tr th").first.evaluate(
                                "el => Math.round(el.getBoundingClientRect().width)"
                            )
                        )
                        self.assertGreaterEqual(first_header_width_after_switch, first_header_width_after_resize - 4)
                        date_bucket = page.locator("[data-testid='dsp-rawdata-date-bucket']")
                        page.wait_for_function(
                            """() => {
                              const el = document.querySelector("[data-testid='dsp-rawdata-date-bucket']");
                              return !!el && el.value === "two_weeks_ago";
                            }"""
                        )
                        self.assertEqual(date_bucket.input_value(), "two_weeks_ago")
                        date_bucket.select_option("last_week")
                        page.wait_for_timeout(120)
                        page.get_by_role("button", name="展開詳細").click()
                        with page.expect_response(lambda resp: resp.request.method == "GET" and "/api/frame" in resp.url):
                            page.get_by_role("button", name="Refresh Frame").click()
                        page.wait_for_function(
                            """() => {
                              const el = document.querySelector("[data-testid='dsp-rawdata-date-bucket']");
                              return !!el && el.value === "two_weeks_ago";
                            }"""
                        )
                        self.assertEqual(date_bucket.input_value(), "two_weeks_ago")
                        self.assertTrue(main_rawdata_table.get_by_text(browser_acceptance_date).first.is_visible())
                        page.reload(wait_until="domcontentloaded")
                        page.locator("[data-testid='main-tab-dsp-tab3']").click()
                        page.locator("[data-testid='sub-tab-rawdata']").click()
                        page.locator("[data-testid='dsp-rawdata-view-user']").click()
                        reloaded_table = page.locator("[data-testid='section-rawdata']").locator(
                            "table", has=page.locator("thead tr th", has_text="最終經銷商")
                        ).first
                        first_header_width_after_reload = int(
                            reloaded_table.locator("thead tr th").first.evaluate(
                                "el => Math.round(el.getBoundingClientRect().width)"
                            )
                        )
                        self.assertGreaterEqual(first_header_width_after_reload, first_header_width_after_resize - 4)
                        self.assertEqual(main_rawdata_table.locator("thead tr th", has_text="變更狀態").count(), 0)
                        self.assertEqual(main_rawdata_table.locator("thead tr th", has_text="列操作").count(), 0)
                        rawdata_workspace = page.locator("[data-testid='section-rawdata']")
                        self.assertEqual(rawdata_workspace.get_by_role("button", name="Revert").count(), 0)
                        main_rawdata_table = rawdata_workspace.locator(
                            "table", has=page.locator("thead tr th", has_text="最終經銷商")
                        ).first
                        header_cells = main_rawdata_table.locator("thead tr th")
                        distributor_col_index = -1
                        for i in range(header_cells.count()):
                            if header_cells.nth(i).inner_text().strip() == "最終經銷商":
                                distributor_col_index = i
                                break
                        self.assertGreaterEqual(distributor_col_index, 0)
                        target_input = main_rawdata_table.locator("tbody tr").first.locator("td").nth(distributor_col_index).locator("input")
                        self.assertEqual(target_input.count(), 1)
                        before_value = target_input.input_value()
                        after_value = f"{before_value}_mdrep"
                        target_input.fill(after_value)
                        with page.expect_response(lambda resp: resp.request.method == "POST" and "/api/action" in resp.url) as modify_resp:
                            page.locator("[data-testid='action-modify']").click()
                        modify_payload = modify_resp.value.request.post_data_json
                        if callable(modify_payload):
                            modify_payload = modify_payload()
                        self.assertIsInstance(modify_payload, dict)
                        self.assertEqual(modify_payload.get("action"), "modify")
                        updates = modify_payload.get("updates") or []
                        self.assertGreaterEqual(len(updates), 1)
                        modify_result = modify_resp.value.json()
                        self.assertEqual(modify_result.get("status"), "ok")
                        modify_result_payload = modify_result.get("result") if isinstance(modify_result, dict) else {}
                        self.assertIsInstance(modify_result_payload, dict)
                        modify_run_id = str((modify_result_payload or {}).get("run_id") or "")
                        self.assertTrue(modify_run_id)
                        first_update = updates[0]
                        expected_row_order = str(first_update.get("row_order"))
                        expected_column = str(first_update.get("column"))
                        expected_value = str(first_update.get("value"))
                        frame_payload = page.evaluate(
                            """async ({root, manifest}) => {
                              const q = new URLSearchParams({
                                root, manifest, workflow: "dsp", template_version: "v1", rule_version: "v1", artifact_root: "artifacts"
                              });
                              const resp = await fetch(`/api/frame?${q.toString()}`);
                              return await resp.json();
                            }""",
                            {"root": str(root), "manifest": "bootstrap.manifest.json"},
                        )
                        self.assertEqual(frame_payload.get("status"), "ok")
                        frame_rows = ((frame_payload.get("result") or {}).get("rows") or [])
                        self.assertGreaterEqual(len(frame_rows), 1)
                        self.assertTrue(
                            any(
                                str(row.get("row_order")) == expected_row_order
                                and str(row.get(expected_column, "")) == expected_value
                                for row in frame_rows
                            ),
                            f"frame 未回讀到更新值: row_order={expected_row_order}, column={expected_column}, value={expected_value}",
                        )
                        page.locator("[data-testid='sub-tab-result']").click()
                        result_workspace = page.locator("[data-testid='section-result']")
                        self.assertTrue(result_workspace.is_visible())
                        self.assertIn("last_action: modify", result_workspace.inner_text())
                        self.assertIn("result_status: ok", result_workspace.inner_text())
                        self.assertIn(f"run_id: {modify_run_id}", result_workspace.inner_text())
                        result_text = result_workspace.inner_text()
                        self.assertTrue(
                            f"rows: {len(frame_rows)}" in result_text or f"rows: {self._fmt_num(len(frame_rows))}" in result_text,
                            result_text,
                        )
                        modify_status_payload = page.evaluate(
                            """async ({root, manifest}) => {
                              const q = new URLSearchParams({
                                root, manifest, workflow: "dsp", template_version: "v1", rule_version: "v1", artifact_root: "artifacts"
                              });
                              const resp = await fetch(`/api/status?${q.toString()}`);
                              return await resp.json();
                            }""",
                            {"root": str(root), "manifest": "bootstrap.manifest.json"},
                        )
                        self.assertEqual(modify_status_payload.get("status"), "ok")
                        modify_run_log = (((modify_status_payload.get("result") or {}).get("recent") or {}).get("run_log") or [])
                        self.assertGreaterEqual(len(modify_run_log), 1)
                        self.assertEqual(str(modify_run_log[0].get("run_type", "")).lower(), "modify")
                        self.assertEqual(str(modify_run_log[0].get("run_id", "")), modify_run_id)

                        # Runtime strip 預設收合，只顯示單行摘要。
                        self.assertFalse(page.is_visible("text=Runtime Actions"))
                        page.get_by_role("button", name="展開詳細").click()
                        self.assertTrue(page.is_visible("text=Runtime Actions"))
                        with page.expect_response(lambda resp: resp.request.method == "GET" and "/api/status?" in resp.url) as refresh_status_resp:
                            page.get_by_role("button", name="Refresh Status").click()
                        refresh_status_payload = refresh_status_resp.value.json()
                        self.assertEqual(refresh_status_payload.get("status"), "ok")
                        self.assertIn("result", refresh_status_payload)
                        self.assertIn("recent", refresh_status_payload["result"])
                        with page.expect_response(lambda resp: resp.request.method == "GET" and "/api/frame?" in resp.url) as refresh_frame_resp:
                            page.get_by_role("button", name="Refresh Frame").click()
                        refresh_frame_payload = refresh_frame_resp.value.json()
                        self.assertEqual(refresh_frame_payload.get("status"), "ok")
                        self.assertIn("result", refresh_frame_payload)
                        self.assertIn("row_count", refresh_frame_payload["result"])
                        self.assertGreaterEqual(int(refresh_frame_payload["result"]["row_count"]), 1)
                        page.get_by_role("button", name="收合詳細").click()
                        self.assertFalse(page.is_visible("text=Runtime Actions"))

                        # Tab4 交付卡控：未交付前先鎖住。
                        page.locator("[data-testid='main-tab-dsp-tab4']").click()
                        self.assertEqual(main_tab4.get_attribute("aria-selected"), "true")
                        self.assertEqual(main_tab3.get_attribute("aria-selected"), "false")
                        page.locator("[data-testid='sub-tab-overview']").click()
                        tab4_workspace = page.locator("section.panel", has_text="DSP Tab4 Workspace").first
                        tab4_workspace.wait_for(state="visible")
                        self.assertTrue(tab4_workspace.is_visible())
                        self.assertTrue(page.locator("[data-testid='tab4-delivery-locked']").is_visible())
                        self.assertTrue(page.locator("[data-testid='action-return-pivot']").is_visible())
                        self.assertTrue(page.locator("[data-testid='action-export']").is_disabled())
                        locked_export_result = page.evaluate(
                            """async ({root, manifest}) => {
                              const body = {
                                root,
                                manifest,
                                workflow: "dsp",
                                template_version: "v1",
                                rule_version: "v1",
                                artifact_root: "artifacts",
                                main_tab: "dsp_tab4",
                                sub_tab: "overview",
                                action: "export"
                              };
                              const resp = await fetch('/api/action', {
                                method: 'POST',
                                headers: { 'Content-Type': 'application/json' },
                                body: JSON.stringify(body),
                              });
                              return { status: resp.status, payload: await resp.json() };
                            }""",
                            {"root": str(root), "manifest": "bootstrap.manifest.json"},
                        )
                        self.assertEqual(locked_export_result["status"], 400)
                        self.assertEqual(locked_export_result["payload"].get("status"), "error")
                        self.assertIn("tab4 delivery required", locked_export_result["payload"].get("message", ""))
                        page.locator("[data-testid='action-return-pivot']").click()
                        self.assertEqual(main_tab3.get_attribute("aria-selected"), "true")
                        self.assertEqual(page.locator("[data-testid='sub-tab-pivot']").get_attribute("aria-selected"), "true")

                        # 樞紐 workspace：insight panel 與樞紐內容需共存。
                        sub_pivot.click()
                        self.assertEqual(sub_pivot.get_attribute("aria-selected"), "true")
                        pivot_workspace = page.locator("[data-testid='section-pivot']")
                        self.assertTrue(pivot_workspace.is_visible())
                        self.assertTrue(pivot_workspace.get_by_text("樞紐核對資訊").first.is_visible())
                        self.assertFalse(pivot_workspace.get_by_text("Latest Publish").first.is_visible())
                        self.assertFalse(pivot_workspace.get_by_text("Latest Evidence").first.is_visible())
                        self.assertFalse(pivot_workspace.get_by_text("樞紐只做核對").first.is_visible())
                        self.assertFalse(pivot_workspace.get_by_text("樞紐節奏").first.is_visible())
                        self.assertFalse(pivot_workspace.get_by_text("當前焦點").first.is_visible())
                        pivot_matrix_table = pivot_workspace.locator(".table-wrap table").first
                        pivot_headers = [
                            pivot_matrix_table.locator("thead tr th").nth(i).inner_text().strip()
                            for i in range(pivot_matrix_table.locator("thead tr th").count())
                        ]
                        self.assertEqual(pivot_headers[0], "經銷商")
                        self.assertIn("Banner", pivot_headers)
                        self.assertIn("Video", pivot_headers)
                        self.assertEqual(pivot_headers[-1], "總計")
                        self.assertGreaterEqual(pivot_matrix_table.locator("tbody tr").count(), 1)
                        first_pivot_row_distributor = pivot_matrix_table.locator("tbody tr").first.locator("td").first.inner_text().strip()
                        self.assertTrue(first_pivot_row_distributor)
                        with page.expect_response(lambda resp: resp.request.method == "POST" and "/api/action" in resp.url) as delivery_resp:
                            page.locator("[data-testid='action-send-tab4']").click()
                        delivery_payload = delivery_resp.value.request.post_data_json
                        if callable(delivery_payload):
                            delivery_payload = delivery_payload()
                        self.assertIsInstance(delivery_payload, dict)
                        self.assertEqual(delivery_payload.get("action"), "tab4_delivery")
                        delivery_result = delivery_resp.value.json()
                        self.assertEqual(delivery_result.get("status"), "ok")
                        self.assertTrue(bool(delivery_result.get("result", {}).get("ready")))
                        delivery_snapshot_token = str(delivery_result.get("result", {}).get("delivery_snapshot_token") or "")
                        self.assertTrue(delivery_snapshot_token)
                        delivery_run_id = str(delivery_result.get("result", {}).get("run_id") or "")
                        self.assertTrue(delivery_run_id)
                        delivery_status_payload = page.evaluate(
                            """async ({root, manifest}) => {
                              const params = new URLSearchParams({
                                root,
                                manifest,
                                workflow: "dsp",
                                template_version: "v1",
                                rule_version: "v1",
                                artifact_root: "artifacts",
                              });
                              const resp = await fetch(`/api/status?${params.toString()}`);
                              return await resp.json();
                            }""",
                            {"root": str(root), "manifest": "bootstrap.manifest.json"},
                        )
                        self.assertEqual(delivery_status_payload.get("status"), "ok")
                        self.assertTrue(bool(delivery_status_payload.get("result", {}).get("tab4_delivery", {}).get("ready")))
                        self.assertEqual(
                            str(delivery_status_payload.get("result", {}).get("tab4_delivery", {}).get("delivery_snapshot_token") or ""),
                            delivery_snapshot_token,
                        )
                        delivery_frame_payload = page.evaluate(
                            """async ({root, manifest}) => {
                              const params = new URLSearchParams({
                                root,
                                manifest,
                                workflow: "dsp",
                                template_version: "v1",
                                rule_version: "v1",
                                artifact_root: "artifacts",
                              });
                              const resp = await fetch(`/api/frame?${params.toString()}`);
                              return await resp.json();
                            }""",
                            {"root": str(root), "manifest": "bootstrap.manifest.json"},
                        )
                        self.assertEqual(delivery_frame_payload.get("status"), "ok")
                        self.assertEqual(
                            str((delivery_frame_payload.get("result", {}).get("tab4_delivery_snapshot") or {}).get("delivery_snapshot_token") or ""),
                            delivery_snapshot_token,
                        )
                        frame_summary = delivery_frame_payload.get("result", {}).get("tab4_preview_template_summary") or {}
                        self.assertGreater(float((frame_summary.get("monthTotals") or [0.0] * 12)[4]), 0.0)
                        page.wait_for_function(
                            """() => {
                              const tab4 = document.querySelector("[data-testid='main-tab-dsp-tab4']");
                              return tab4 && tab4.getAttribute("aria-selected") === "true";
                            }"""
                        )
                        page.locator("[data-testid='tab4-delivery-locked']").wait_for(state="detached")
                        self.assertEqual(main_tab4.get_attribute("aria-selected"), "true")
                        self.assertEqual(main_tab3.get_attribute("aria-selected"), "false")
                        tab4_workspace = page.locator("section.panel", has_text="DSP Tab4 Workspace").first
                        self.assertTrue(tab4_workspace.is_visible())
                        self.assertEqual(page.locator("[data-testid='tab4-delivery-locked']").count(), 0)
                        self.assertIn("交付身份：", tab4_workspace.inner_text())
                        self.assertIn(delivery_snapshot_token, tab4_workspace.inner_text())
                        sub_overview = page.locator("[data-testid='sub-tab-overview']")
                        self.assertEqual(sub_overview.get_attribute("aria-selected"), "true")

                        # DSP Tab4 workspace tabs：summary/detail/tracking 切換不應回退成大表堆疊。
                        page.locator("[data-testid='sub-tab-overview']").click()
                        self.assertEqual(sub_overview.get_attribute("aria-selected"), "true")
                        mf_matrix = page.locator("[data-testid='tab4-mf-summary-matrix']")
                        self.assertTrue(mf_matrix.is_visible())
                        self.assertIn("DSP投資額 總計", mf_matrix.inner_text())
                        self.assertIn("1月", mf_matrix.inner_text())
                        self.assertIn("年度(總)", mf_matrix.inner_text())
                        self.assertIn("全體經銷商", mf_matrix.inner_text())
                        self.assertEqual(mf_matrix.locator("tbody tr").count(), 13)
                        tab4_detail = page.locator("[data-testid='tab4-detail']")
                        tab4_tracking = page.locator("[data-testid='tab4-tracking']")
                        tab4_detail.click()
                        self.assertIn("btn-primary", tab4_detail.get_attribute("class") or "")
                        self.assertEqual(tab4_detail.get_attribute("aria-selected"), "true")
                        detail_matrix = page.locator("[data-testid='tab4-mf-detail-matrix']")
                        self.assertTrue(detail_matrix.is_visible())
                        self.assertIn("各經銷商明細", detail_matrix.inner_text())
                        self.assertIn("各經銷商明細分區", detail_matrix.inner_text())
                        self.assertIn("全體經銷 總投資量目標", detail_matrix.inner_text())
                        self.assertIn("年度(總)", detail_matrix.inner_text())
                        tab4_tracking.click()
                        self.assertIn("btn-primary", tab4_tracking.get_attribute("class") or "")
                        self.assertEqual(tab4_tracking.get_attribute("aria-selected"), "true")
                        self.assertEqual(tab4_detail.get_attribute("aria-selected"), "false")
                        self.assertTrue(page.is_visible("text=北流進單追蹤"))
                        self.assertTrue(tab4_workspace.is_visible())
                        self.assertTrue(tab4_workspace.locator("[data-testid='action-export']").is_visible())
                        with page.expect_download() as download_info:
                            with page.expect_response(lambda resp: resp.request.method == "POST" and "/api/action" in resp.url) as export_resp:
                                tab4_workspace.locator("[data-testid='action-export']").click()
                        export_payload = export_resp.value.request.post_data_json
                        if callable(export_payload):
                            export_payload = export_payload()
                        self.assertIsInstance(export_payload, dict)
                        self.assertEqual(export_payload.get("action"), "export")
                        self.assertEqual(export_payload.get("main_tab"), "dsp_tab4")
                        self.assertEqual(export_payload.get("sub_tab"), "overview")
                        export_result = export_resp.value.json()
                        self.assertEqual(export_result.get("status"), "ok")
                        export_result_payload = export_result.get("result") if isinstance(export_result, dict) else {}
                        self.assertIsInstance(export_result_payload, dict)
                        export_run_id = str((export_result_payload or {}).get("run_id") or "")
                        export_artifact_checksum = str((export_result_payload or {}).get("artifact_checksum") or "")
                        export_artifact_path = ""
                        if isinstance(export_result_payload, dict):
                            export_artifact_path = str(export_result_payload.get("artifact_path") or "")
                        self.assertEqual(str(export_result_payload.get("delivery_snapshot_token") or ""), delivery_snapshot_token)
                        self.assertEqual(str(export_result_payload.get("delivery_run_id") or ""), delivery_run_id)
                        self.assertTrue(export_run_id)
                        self.assertTrue(export_artifact_checksum)
                        self.assertTrue(export_artifact_path)
                        self.assertTrue(Path(export_artifact_path).exists(), f"export artifact not found: {export_artifact_path}")
                        self.assertRegex(Path(export_artifact_path).name, r"^2026 DSP投資量報表_\d{4}-\d{4}\.xlsx$")
                        export_actual_checksum = hashlib.sha256(Path(export_artifact_path).read_bytes()).hexdigest()
                        self.assertEqual(export_artifact_checksum, export_actual_checksum)
                        download = download_info.value
                        self.assertRegex(download.suggested_filename, r"^2026 DSP投資量報表_\d{4}-\d{4}\.xlsx$")
                        download_query = parse_qs(urlparse(download.url).query)
                        self.assertEqual(download_query.get("main_tab", [""])[0], "dsp_tab4")
                        self.assertEqual(download_query.get("sub_tab", [""])[0], "overview")
                        download_temp_path = download.path()
                        self.assertIsNotNone(download_temp_path)
                        if download_temp_path is not None:
                            self.assertEqual(
                                hashlib.sha256(Path(download_temp_path).read_bytes()).hexdigest(),
                                export_artifact_checksum,
                            )
                        export_wb = load_workbook(Path(export_artifact_path), data_only=True)
                        try:
                            self.assertEqual(
                                export_wb.sheetnames,
                                [
                                    "2025年_MF_合作績效統計總表",
                                    "2025_外部+行政_合作績效統計總表 ",
                                    "mF投資量_總表",
                                    "各經銷商明細",
                                    "北流進單追蹤",
                                ],
                            )
                            self.assertEqual(export_wb["2025年_MF_合作績效統計總表"].sheet_state, "hidden")
                            self.assertEqual(export_wb["2025_外部+行政_合作績效統計總表 "].sheet_state, "hidden")
                            self.assertEqual(export_wb["mF投資量_總表"].freeze_panes, "M1")
                            self.assertEqual(export_wb["各經銷商明細"].freeze_panes, "U1")
                        finally:
                            export_wb.close()
                        page.locator("[data-testid='sub-tab-result']").click()
                        export_result_workspace = page.locator("[data-testid='section-result']")
                        self.assertTrue(export_result_workspace.is_visible())
                        self.assertIn("last_action: export", export_result_workspace.inner_text())
                        self.assertIn("result_status: ok", export_result_workspace.inner_text())
                        self.assertIn(f"run_id: {export_run_id}", export_result_workspace.inner_text())
                        export_frame_payload = page.evaluate(
                            """async ({root, manifest}) => {
                              const q = new URLSearchParams({
                                root, manifest, workflow: "dsp", template_version: "v1", rule_version: "v1", artifact_root: "artifacts"
                              });
                              const resp = await fetch(`/api/frame?${q.toString()}`);
                              return await resp.json();
                            }""",
                            {"root": str(root), "manifest": "bootstrap.manifest.json"},
                        )
                        self.assertEqual(export_frame_payload.get("status"), "ok")
                        export_row_count = int(((export_frame_payload.get("result") or {}).get("row_count") or 0))
                        self.assertGreaterEqual(export_row_count, 1)
                        export_result_text = export_result_workspace.inner_text()
                        self.assertTrue(
                            f"rows: {export_row_count}" in export_result_text
                            or f"rows: {self._fmt_num(export_row_count)}" in export_result_text,
                            export_result_text,
                        )
                        page.locator("[data-testid='result-segment-action']").click()
                        self.assertIn(export_artifact_path, page.locator("pre.json-view").inner_text())
                        page.locator("[data-testid='result-segment-state']").click()
                        self.assertIn('"lastAction": "export"', page.locator("pre.json-view").inner_text())
                        page.locator("[data-testid='sub-tab-overview']").click()
                        export_blocked_requests: list[dict] = []

                        def _capture_export_blocked_request(req: object) -> None:
                            request_method = getattr(req, "method", "")
                            if callable(request_method):
                                request_method = request_method()
                            request_url = getattr(req, "url", "")
                            if callable(request_url):
                                request_url = request_url()
                            if request_method == "POST" and "/api/action" in request_url:
                                body = getattr(req, "post_data_json", None)
                                if callable(body):
                                    body = body()
                                if isinstance(body, dict):
                                    export_blocked_requests.append(body)

                        page.on("request", _capture_export_blocked_request)
                        pre_publish_count = len(export_blocked_requests)
                        tab4_workspace.locator("[data-testid='action-publish']").evaluate("el => el.click()")
                        page.wait_for_timeout(250)
                        post_publish_count = len(export_blocked_requests)
                        self.assertEqual(post_publish_count, pre_publish_count)
                        page.remove_listener("request", _capture_export_blocked_request)

                        # Rawdata save 後，前一次 Tab4 交付必須失效（重新鎖住）。
                        page.locator("[data-testid='main-tab-dsp-tab3']").click()
                        page.locator("[data-testid='sub-tab-rawdata']").click()
                        with page.expect_response(lambda resp: resp.request.method == "POST" and "/api/action" in resp.url) as save_resp:
                            page.locator("[data-testid='action-save']").click()
                        save_payload = save_resp.value.request.post_data_json
                        if callable(save_payload):
                            save_payload = save_payload()
                        self.assertIsInstance(save_payload, dict)
                        self.assertEqual(save_payload.get("action"), "save")
                        save_result = save_resp.value.json()
                        self.assertEqual(str(save_result.get("status")), "ok")
                        page.locator("[data-testid='main-tab-dsp-tab4']").click()
                        page.wait_for_function(
                            """() => {
                              const tab4 = document.querySelector("[data-testid='main-tab-dsp-tab4']");
                              return tab4 && tab4.getAttribute("aria-selected") === "true";
                            }"""
                        )
                        page.locator("[data-testid='sub-tab-overview']").wait_for(state="visible")
                        relocked_status_payload = page.evaluate(
                            """async ({root, manifest}) => {
                              const params = new URLSearchParams({
                                root,
                                manifest,
                                workflow: "dsp",
                                template_version: "v1",
                                rule_version: "v1",
                                artifact_root: "artifacts",
                              });
                              const resp = await fetch(`/api/status?${params.toString()}`);
                              return await resp.json();
                            }""",
                            {"root": str(root), "manifest": "bootstrap.manifest.json"},
                        )
                        self.assertEqual(relocked_status_payload.get("status"), "ok")
                        self.assertFalse(bool(relocked_status_payload.get("result", {}).get("tab4_delivery", {}).get("ready")))
                        self.assertEqual(
                            str(relocked_status_payload.get("result", {}).get("tab4_delivery", {}).get("reason") or ""),
                            "rawdata_saved",
                        )

                        # SSP parity：成效救火與媒體要量都必須可切換且可見。
                        page.get_by_role("button", name="Use SSP").click()
                        main_ssp_anomaly = page.locator("[data-testid='main-tab-ssp-anomaly']")
                        main_ssp_media_demand = page.locator("[data-testid='main-tab-ssp-media-demand']")
                        main_ssp_anomaly.click()
                        self.assertEqual(main_ssp_anomaly.get_attribute("aria-selected"), "true")
                        self.assertEqual(page.locator("[data-testid='main-tab-ssp-volume']").count(), 0)
                        self.assertEqual(main_ssp_media_demand.count(), 1)
                        self.assertEqual(page.locator("[data-testid='sub-tabs']").count(), 0)
                        self.assertEqual(page.locator("[data-testid='sub-tab-overview']").count(), 0)
                        self.assertEqual(page.locator("[data-testid='sub-tab-rawdata']").count(), 0)
                        self.assertEqual(page.locator("[data-testid='sub-tab-pivot']").count(), 0)
                        self.assertEqual(page.locator("[data-testid='sub-tab-result']").count(), 0)
                        ssp_anomaly_workspace = page.locator("section.panel", has_text="SSP 成效異常 Workspace").first
                        self.assertTrue(ssp_anomaly_workspace.is_visible())
                        self.assertTrue(ssp_anomaly_workspace.get_by_text("控制列").first.is_visible())
                        self.assertTrue(ssp_anomaly_workspace.get_by_text("每日總表").first.is_visible())
                        self.assertTrue(ssp_anomaly_workspace.get_by_text("異常供應商收合區").first.is_visible())
                        self.assertTrue(page.locator("[data-testid='ssp-anomaly-visibility-mode']").is_visible())
                        self.assertTrue(page.locator("[data-testid='ssp-anomaly-dod-threshold']").is_visible())
                        self.assertEqual(page.locator("[data-testid='ssp-anomaly-supplier-filter']").count(), 0)
                        self.assertEqual(page.locator("[data-testid='ssp-anomaly-only-toggle']").count(), 0)
                        self.assertEqual(page.locator("[data-testid='ssp-anomaly-min-request']").count(), 0)
                        daily_summary = page.locator("[data-testid='ssp-anomaly-daily-summary']")
                        self.assertEqual(daily_summary.locator("thead tr").count(), 2)
                        self.assertLessEqual(daily_summary.locator("tbody tr").count(), 13)
                        self.assertEqual(daily_summary.locator("thead tr:first-child th", has_text="DoD 變動(萬)").count(), 0)
                        self.assertEqual(daily_summary.locator("thead tr:first-child th", has_text="網站異常數").count(), 0)
                        self.assertTrue(
                            page.locator("[data-testid='ssp-anomaly-suppliers-accordion']").is_visible()
                            or ssp_anomaly_workspace.get_by_text("目前沒有異常供應商").first.is_visible()
                        )
                        if page.locator("[data-testid='ssp-anomaly-suppliers-accordion']").is_visible():
                            suppliers_head = page.locator("[data-testid='ssp-anomaly-suppliers-head']")
                            self.assertTrue(suppliers_head.is_visible())
                            self.assertIn("供應商", suppliers_head.inner_text())
                            self.assertIn("DoD 變動(萬)", suppliers_head.inner_text())
                            self.assertIn("網站異常數", suppliers_head.inner_text())
                            first_supplier_detail = page.locator("[data-testid='ssp-anomaly-suppliers-accordion'] details").first
                            if first_supplier_detail.count() > 0:
                                self.assertFalse(first_supplier_detail.evaluate("el => el.hasAttribute('open')"))

                        main_ssp_media_demand.click()
                        self.assertEqual(main_ssp_media_demand.get_attribute("aria-selected"), "true")
                        self.assertEqual(page.locator("[data-testid='sub-tabs']").count(), 0)
                        ssp_media_workspace = page.locator("section.panel", has_text="SSP 媒體要量 Workspace").first
                        self.assertTrue(ssp_media_workspace.is_visible())
                        self.assertTrue(page.locator("[data-testid='ssp-media-workbench']").is_visible())
                        page.locator("[data-testid='ssp-media-toggle-summary']").click()
                        page.wait_for_function(
                            """() => !!document.querySelector("[data-testid='ssp-media-kpi']")"""
                        )
                        self.assertTrue(page.locator("[data-testid='ssp-media-kpi']").is_visible())
                        self.assertTrue(page.locator("[data-testid='ssp-media-category-tabs']").is_visible())
                        self.assertTrue(page.locator("[data-testid='ssp-media-threshold']").is_visible())
                        self.assertTrue(page.locator("[data-testid='ssp-media-only-unmet-toggle']").is_visible())
                        self.assertEqual(page.locator("[data-testid='ssp-media-toggle-slot-editor']").count(), 1)
                        self.assertEqual(page.locator("[data-testid='ssp-media-slot-editor']").count(), 0)
                        self.assertTrue(page.locator("[data-testid='ssp-media-demand-table']").is_visible())
                        self.assertEqual(page.locator("[data-testid='ssp-media-category-蓋板']").count(), 1)
                        self.assertEqual(page.locator("[data-testid='ssp-media-category-置底']").count(), 1)
                        self.assertEqual(page.locator("[data-testid='ssp-media-category-置底展開']").count(), 1)
                        self.assertEqual(page.locator("[data-testid='ssp-media-category-文中300x250']").count(), 1)
                        self.assertEqual(page.locator("[data-testid='ssp-media-category-文中320x480']").count(), 1)
                        self.assertEqual(page.locator("[data-testid='ssp-anomaly-visibility-mode']").count(), 0)

                        page.wait_for_function("() => window.location.search.includes('main_tab=ssp_media_demand')")
                        self.assertIn("main_tab=ssp_media_demand", page.url)
                        self.assertNotIn("sub_tab=", page.url)
                        runtime_strip_text = page.locator(".workbench-runtime-strip").inner_text()
                        self.assertIn("main: ssp_media_demand", runtime_strip_text)
                        self.assertNotIn("rawdata", runtime_strip_text.lower())
                        self.assertTrue(page.locator("[data-testid='ssp-media-kpi']").is_visible())

                        # 最小 happy path：SSP 直接改日期區間後觸發 action，確認 payload 帶最新週期。
                        captured_actions: list[dict] = []

                        def _capture_action_request(req: object) -> None:
                            request_method = getattr(req, "method", "")
                            if callable(request_method):
                                request_method = request_method()
                            request_url = getattr(req, "url", "")
                            if callable(request_url):
                                request_url = request_url()
                            if request_method == "POST" and "/api/action" in request_url:
                                body = getattr(req, "post_data_json", None)
                                if callable(body):
                                    body = body()
                                if isinstance(body, dict):
                                    captured_actions.append(body)

                        page.on("request", _capture_action_request)
                        page.locator("[data-testid='period-range-toggle']").click()
                        page.locator("[data-testid='period-range-day-2026-05-05']").click()
                        page.locator("[data-testid='period-range-day-2026-05-11']").click()
                        page.get_by_role("button", name="展開詳細").click()
                        with page.expect_response(lambda resp: resp.request.method == "POST" and "/api/action" in resp.url):
                            page.get_by_role("button", name="Health").click()
                        self.assertGreaterEqual(len(captured_actions), 1)
                        last_action_payload = captured_actions[-1]
                        self.assertEqual(last_action_payload.get("action"), "health")
                        self.assertEqual(last_action_payload.get("workflow"), "ssp")
                        self.assertEqual(last_action_payload.get("period_preset"), "custom")
                        self.assertEqual(last_action_payload.get("period_week_start"), "2026-05-05")
                        self.assertEqual(last_action_payload.get("period_week_end"), "2026-05-11")
                        page.get_by_role("button", name="收合詳細").click()

                        # 切回 DSP Result，驗證 Result segmented view 仍是摘要優先。
                        page.get_by_role("button", name="Use DSP").click()
                        page.locator("[data-testid='main-tab-dsp-tab3']").click()
                        page.locator("[data-testid='sub-tab-result']").click()
                        self.assertTrue(page.is_visible("text=Execution Summary"))
                        self.assertFalse(page.is_visible("text=Action Payload JSON"))
                        self.assertFalse(page.is_visible("text=Result State JSON"))
                        self.assertEqual(page.locator("[data-testid='result-segment-summary']").get_attribute("aria-selected"), "true")

                        page.locator("[data-testid='result-segment-action']").click()
                        self.assertTrue(page.is_visible("text=Action Payload JSON"))
                        self.assertFalse(page.is_visible("text=Execution Summary"))
                        self.assertEqual(page.locator("[data-testid='result-segment-action']").get_attribute("aria-selected"), "true")

                        page.locator("[data-testid='result-segment-state']").click()
                        self.assertTrue(page.is_visible("text=Result State JSON"))
                        self.assertFalse(page.is_visible("text=Action Payload JSON"))
                        self.assertEqual(page.locator("[data-testid='result-segment-state']").get_attribute("aria-selected"), "true")

                        status_payload = page.evaluate(
                            """async ({root, manifest}) => {
                              const q = new URLSearchParams({
                                root, manifest, workflow: "dsp", template_version: "v1", rule_version: "v1", artifact_root: "artifacts"
                              });
                              const resp = await fetch(`/api/status?${q.toString()}`);
                              return await resp.json();
                            }""",
                            {"root": str(root), "manifest": "bootstrap.manifest.json"},
                        )
                        self.assertEqual(status_payload["status"], "ok")
                        self.assertIn("result", status_payload)
                        run_log = status_payload["result"]["recent"]["run_log"]
                        self.assertGreaterEqual(len(run_log), 3)
                        export_entries = [item for item in run_log if str(item.get("run_type", "")).lower() == "export"]
                        modify_entries = [item for item in run_log if str(item.get("run_type", "")).lower() == "modify"]
                        self.assertGreaterEqual(len(export_entries), 1)
                        self.assertGreaterEqual(len(modify_entries), 1)
                        self.assertEqual(str(export_entries[0].get("run_id", "")), export_run_id)
                        self.assertEqual(str(modify_entries[0].get("run_id", "")), modify_run_id)
                        publish_runs = status_payload["result"]["recent"]["publish_runs"]
                        self.assertGreaterEqual(len(publish_runs), 1)
                        self.assertEqual(str(publish_runs[0].get("run_id", "")), export_run_id)
                    finally:
                        context.close()
                        browser.close()
            finally:
                server.shutdown()
                server.server_close()
                thread.join(timeout=2.0)

    def test_dsp_tab4_browser_export_keeps_weekly_period_contract(self) -> None:
        frontend_index = ui_shell_module.FRONTEND_DIST_DIR / "index.html"
        if not frontend_index.exists():
            self.skipTest("frontend dist 不存在，無法做真 browser acceptance（請先 pnpm build）")

        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            baseline_dir = root / "data_seed_test" / "dsp_weekly_baselines"
            baseline_dir.mkdir(parents=True, exist_ok=True)
            baseline_name = "2026 DSP投資量報表_0101-0503.xlsx"
            self._write_dsp_tab4_template(baseline_dir / baseline_name)
            (baseline_dir / "manifest.json").write_text(
                json.dumps(
                    {
                        "workflow": "dsp",
                        "baselines": [
                            {
                                "week_end": "2026-05-03",
                                "file": baseline_name,
                            }
                        ],
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            ctx = UiContext(
                root=root,
                runtime_env="test",
                manifest_rel="bootstrap.test.manifest.json",
                workflow="dsp",
                template_version="v1",
                rule_version="v1",
                artifact_root=(root / "artifacts_test").resolve(),
            )
            dispatch_action(ctx, {"action": "bootstrap"})
            dispatch_action(
                ctx,
                {
                    "action": "save",
                    "rows": [
                        self._full_row(日期時間="2026-05-04 00:00:00", 執行金額=125.0),
                        self._full_row(日期時間="2026-05-03 23:59:59", 訂單="OUTSIDE", 執行金額=999.0),
                    ],
                },
            )
            delivery = dispatch_action(
                ctx,
                {
                    "action": "tab4_delivery",
                    "main_tab": "dsp_tab3",
                    "sub_tab": "pivot",
                    "period_week_start": "2026-05-04",
                    "period_week_end": "2026-05-10",
                },
            )
            self.assertTrue(bool(delivery.get("ready")))

            try:
                server = ThreadingHTTPServer(("127.0.0.1", 0), UiRequestHandler)
            except PermissionError:
                self.skipTest("sandbox 禁止本地 socket bind，略過 browser acceptance")

            try:
                from playwright.sync_api import Error as PlaywrightError
                from playwright.sync_api import sync_playwright
            except Exception:
                self.skipTest("缺少 playwright 依賴，請先安裝 playwright 並執行 playwright install chromium")

            thread = threading.Thread(target=server.serve_forever, daemon=True)
            thread.start()
            try:
                host, port = server.server_address
                base_url = f"http://{host}:{port}"
                query = urlencode(
                    {
                        "root": str(root),
                        "env": "test",
                        "manifest": "bootstrap.test.manifest.json",
                        "workflow": "dsp",
                        "template_version": "v1",
                        "rule_version": "v1",
                        "artifact_root": "artifacts_test",
                        "main_tab": "dsp_tab4",
                        "sub_tab": "overview",
                        "period_preset": "last_week",
                        "period_week_start": "2026-05-04",
                        "period_week_end": "2026-05-10",
                    }
                )
                with sync_playwright() as p:
                    try:
                        browser = p.chromium.launch(headless=True)
                    except PlaywrightError as exc:
                        self.skipTest(f"playwright browser context 無法啟動，請先 playwright install chromium: {exc}")
                    context = browser.new_context(viewport={"width": 1440, "height": 1000}, accept_downloads=True)
                    page = context.new_page()
                    try:
                        with page.expect_response(lambda resp: resp.request.method == "GET" and "/api/frame?" in resp.url) as frame_resp:
                            page.goto(f"{base_url}/?{query}", wait_until="domcontentloaded")
                        frame_url_query = parse_qs(urlparse(frame_resp.value.url).query)
                        self.assertEqual(frame_url_query.get("period_week_start", [""])[0], "2026-05-04")
                        self.assertEqual(frame_url_query.get("period_week_end", [""])[0], "2026-05-10")
                        frame_payload = frame_resp.value.json()
                        self.assertEqual(frame_payload.get("status"), "ok")
                        page.locator("[data-testid='action-export']").wait_for(state="visible")

                        with page.expect_download() as download_info:
                            with page.expect_response(lambda resp: resp.request.method == "POST" and "/api/action" in resp.url) as export_resp:
                                page.locator("[data-testid='action-export']").click()
                        export_payload = export_resp.value.request.post_data_json
                        if callable(export_payload):
                            export_payload = export_payload()
                        self.assertIsInstance(export_payload, dict)
                        self.assertEqual(export_payload.get("action"), "export")
                        self.assertEqual(export_payload.get("main_tab"), "dsp_tab4")
                        self.assertEqual(export_payload.get("sub_tab"), "overview")
                        self.assertEqual(export_payload.get("period_week_start"), "2026-05-04")
                        self.assertEqual(export_payload.get("period_week_end"), "2026-05-10")
                        export_result = export_resp.value.json()
                        self.assertEqual(export_result.get("status"), "ok")
                        result = export_result.get("result") if isinstance(export_result, dict) else {}
                        self.assertIsInstance(result, dict)
                        self.assertEqual(int(result.get("row_count") or 0), 1)
                        self.assertEqual(str(result.get("week_start") or ""), "2026-05-04")
                        self.assertEqual(str(result.get("week_end") or ""), "2026-05-10")
                        artifact_path = Path(str(result.get("artifact_path") or ""))
                        artifact_checksum = str(result.get("artifact_checksum") or "")
                        self.assertEqual(artifact_path.name, "2026 DSP投資量報表_0504-0510.xlsx")
                        self.assertTrue(artifact_path.exists())
                        self.assertEqual(hashlib.sha256(artifact_path.read_bytes()).hexdigest(), artifact_checksum)
                        download = download_info.value
                        self.assertEqual(download.suggested_filename, "2026 DSP投資量報表_0504-0510.xlsx")
                        download_query = parse_qs(urlparse(download.url).query)
                        self.assertEqual(download_query.get("main_tab", [""])[0], "dsp_tab4")
                        self.assertEqual(download_query.get("sub_tab", [""])[0], "overview")
                        downloaded_path = download.path()
                        self.assertIsNotNone(downloaded_path)
                        if downloaded_path is not None:
                            self.assertEqual(hashlib.sha256(Path(downloaded_path).read_bytes()).hexdigest(), artifact_checksum)
                    finally:
                        context.close()
                        browser.close()
            finally:
                server.shutdown()
                server.server_close()
                thread.join(timeout=2.0)

    def test_ui_browser_modify_failure_keeps_local_edits(self) -> None:
        frontend_index = ui_shell_module.FRONTEND_DIST_DIR / "index.html"
        if not frontend_index.exists():
            self.skipTest("frontend dist 不存在，無法做真 browser mutation smoke（請先 pnpm build）")

        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            ctx = self._ctx(root, workflow="dsp")
            dispatch_action(ctx, {"action": "bootstrap"})
            dispatch_action(
                ctx,
                {"action": "save", "rows": [self._full_row(日期時間=self._dsp_bucket_datetime(weeks_ago=2))]},
            )

            try:
                server = ThreadingHTTPServer(("127.0.0.1", 0), UiRequestHandler)
            except PermissionError:
                self.skipTest("sandbox 禁止本地 socket bind，略過 browser mutation smoke")

            try:
                from playwright.sync_api import Error as PlaywrightError
                from playwright.sync_api import sync_playwright
            except Exception:
                self.skipTest("缺少 playwright 依賴，請先安裝 playwright 並執行 playwright install chromium")

            original_dispatch_action = ui_shell_module.dispatch_action

            def _failing_dispatch_action(local_ctx: UiContext, payload: dict[str, object]) -> dict[str, object]:
                if str(payload.get("action") or "").strip() == "modify":
                    raise ValueError("simulated modify failure")
                return original_dispatch_action(local_ctx, payload)

            thread = threading.Thread(target=server.serve_forever, daemon=True)
            thread.start()
            try:
                host, port = server.server_address
                base_url = f"http://{host}:{port}"
                with patch.object(ui_shell_module, "dispatch_action", side_effect=_failing_dispatch_action):
                    with sync_playwright() as p:
                        try:
                            browser = p.chromium.launch(headless=True)
                        except PlaywrightError as exc:
                            self.skipTest(f"playwright browser context 無法啟動，請先 playwright install chromium: {exc}")
                        page = browser.new_page()
                        try:
                            query = urlencode(
                                {
                                    "root": str(root),
                                    "manifest": "bootstrap.manifest.json",
                                    "workflow": "dsp",
                                    "template_version": "v1",
                                    "rule_version": "v1",
                                    "artifact_root": "artifacts",
                                }
                            )
                            page.goto(f"{base_url}/?{query}", wait_until="domcontentloaded")
                            page.locator("[data-testid='main-tab-dsp-tab3']").click()
                            page.locator("[data-testid='sub-tab-rawdata']").click()
                            rawdata_workspace = page.locator("[data-testid='section-rawdata']")
                            editable_inputs = rawdata_workspace.locator("tbody tr input")
                            self.assertGreaterEqual(editable_inputs.count(), 1)
                            target_input = editable_inputs.first
                            before_value = target_input.input_value()
                            after_value = f"{before_value}_failcase"
                            target_input.fill(after_value)
                            with page.expect_response(lambda resp: resp.request.method == "POST" and "/api/action" in resp.url) as modify_resp:
                                page.locator("[data-testid='action-modify']").click()
                            modify_payload = modify_resp.value.request.post_data_json
                            if callable(modify_payload):
                                modify_payload = modify_payload()
                            self.assertIsInstance(modify_payload, dict)
                            self.assertEqual(modify_payload.get("action"), "modify")
                            self.assertGreaterEqual(len(modify_payload.get("updates") or []), 1)
                            modify_result = modify_resp.value.json()
                            self.assertEqual(modify_result.get("status"), "error")
                            self.assertEqual(modify_result.get("error_code"), "UI_ACTION_FAILED")
                            page.wait_for_timeout(120)
                            self.assertEqual(target_input.input_value(), after_value)
                            self.assertIn("dirty_rows: 1", rawdata_workspace.inner_text())
                        finally:
                            browser.close()
            finally:
                server.shutdown()
                server.server_close()
                thread.join(timeout=2.0)

    def _seed_monthly_chart_rows(self, root: Path) -> None:
        cfg = build_config(root, "bootstrap.manifest.json", "prod")
        repo = SQLiteRepository(cfg.db_path, project_root=root)
        with repo.connect_monthly_report() as conn:
            repo.save_monthly_report_rows(
                conn,
                run_id="run-2026-04",
                report_kind="ssp_regular_monthly_zone_campaign_size",
                start_day="2026-04-01",
                end_day="2026-04-30",
                report_id=1,
                records_total=2,
                source="test",
                pb=0,
                request_payload={},
                response_payload={},
                sum_row={},
                rows=[
                    {
                        "month": "2026-04",
                        "date": "2026-04-01",
                        "zone_id": 1,
                        "zone_name": "測試版位",
                        "campaign_id": 10,
                        "campaign_name": "測試訂單",
                        "creative_size_id": "300x250",
                        "ad_format": "Banner",
                        "request": 1000,
                        "impress": 800,
                        "click": 40,
                        "profit": 100,
                        "advertiser_mu": 200,
                    },
                    {
                        "month": "2026-04",
                        "date": "2026-04-02",
                        "zone_id": 1,
                        "zone_name": "測試版位",
                        "campaign_id": 10,
                        "campaign_name": "測試訂單",
                        "creative_size_id": "320x480",
                        "ad_format": "Video",
                        "request": 1200,
                        "impress": 900,
                        "click": 45,
                        "profit": 150,
                        "advertiser_mu": 260,
                    },
                ],
            )

    def test_ui_browser_monthly_charts_tab_refreshes_frame(self) -> None:
        frontend_index = ui_shell_module.FRONTEND_DIST_DIR / "index.html"
        if not frontend_index.exists():
            self.skipTest("frontend dist 不存在，無法做 monthly charts browser smoke（請先 pnpm build）")

        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            ctx = self._ctx(root, workflow="monthly")
            dispatch_action(ctx, {"action": "bootstrap"})
            self._seed_monthly_chart_rows(root)

            try:
                server = ThreadingHTTPServer(("127.0.0.1", 0), UiRequestHandler)
            except PermissionError:
                self.skipTest("sandbox 禁止本地 socket bind，略過 monthly charts browser smoke")

            try:
                from playwright.sync_api import Error as PlaywrightError
                from playwright.sync_api import sync_playwright
            except Exception:
                self.skipTest("缺少 playwright 依賴，請先安裝 playwright 並執行 playwright install chromium")

            thread = threading.Thread(target=server.serve_forever, daemon=True)
            thread.start()
            try:
                host, port = server.server_address
                base_url = f"http://{host}:{port}"
                query = urlencode(
                    {
                        "root": str(root),
                        "manifest": "bootstrap.manifest.json",
                        "workflow": "monthly",
                        "template_version": "v1",
                        "rule_version": "v1",
                        "artifact_root": "artifacts",
                        "main_tab": "monthly_p4",
                    }
                )
                with sync_playwright() as p:
                    try:
                        browser = p.chromium.launch(headless=True)
                    except PlaywrightError as exc:
                        self.skipTest(f"playwright browser context 無法啟動，請先 playwright install chromium: {exc}")
                    context = browser.new_context(viewport={"width": 1440, "height": 1000})
                    page = context.new_page()
                    try:
                        page.goto(f"{base_url}/?{query}", wait_until="domcontentloaded")
                        page.locator("[data-testid='main-tab-monthly-p4']").wait_for(state="visible")
                        with page.expect_response(lambda resp: resp.request.method == "GET" and "/api/frame?" in resp.url and "main_tab=monthly_charts" in resp.url) as frame_resp:
                            page.locator("[data-testid='main-tab-monthly-charts']").click()
                        frame_payload = frame_resp.value.json()
                        self.assertEqual(frame_payload.get("status"), "ok")
                        self.assertIn("monthly_charts", frame_payload.get("result") or {})
                        page.get_by_role("heading", name="月報簡報素材").wait_for(state="visible")
                        self.assertGreater(page.locator(".monthly-copy-button").count(), 0)
                        self.assertEqual(page.locator(".state-block.empty").count(), 0)
                    finally:
                        context.close()
                        browser.close()
            finally:
                server.shutdown()
                server.server_close()
                thread.join(timeout=2.0)

    def test_ui_browser_monthly_charts_copy_failure_does_not_claim_image_success(self) -> None:
        frontend_index = ui_shell_module.FRONTEND_DIST_DIR / "index.html"
        if not frontend_index.exists():
            self.skipTest("frontend dist 不存在，無法做 monthly charts copy smoke（請先 pnpm build）")

        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            ctx = self._ctx(root, workflow="monthly")
            dispatch_action(ctx, {"action": "bootstrap"})
            self._seed_monthly_chart_rows(root)

            try:
                server = ThreadingHTTPServer(("127.0.0.1", 0), UiRequestHandler)
            except PermissionError:
                self.skipTest("sandbox 禁止本地 socket bind，略過 monthly charts copy smoke")

            try:
                from playwright.sync_api import Error as PlaywrightError
                from playwright.sync_api import sync_playwright
            except Exception:
                self.skipTest("缺少 playwright 依賴，請先安裝 playwright 並執行 playwright install chromium")

            thread = threading.Thread(target=server.serve_forever, daemon=True)
            thread.start()
            try:
                host, port = server.server_address
                base_url = f"http://{host}:{port}"
                query = urlencode(
                    {
                        "root": str(root),
                        "manifest": "bootstrap.manifest.json",
                        "workflow": "monthly",
                        "template_version": "v1",
                        "rule_version": "v1",
                        "artifact_root": "artifacts",
                        "main_tab": "monthly_charts",
                    }
                )
                with sync_playwright() as p:
                    try:
                        browser = p.chromium.launch(headless=True)
                    except PlaywrightError as exc:
                        self.skipTest(f"playwright browser context 無法啟動，請先 playwright install chromium: {exc}")
                    context = browser.new_context(viewport={"width": 1440, "height": 1000})
                    page = context.new_page()
                    page.add_init_script(
                        """
                        Object.defineProperty(navigator, 'clipboard', {
                          configurable: true,
                          value: {
                            write: async (items) => {
                              const item = items && items[0];
                              const payload = item && (item.items || item);
                              if (payload && payload['image/png']) {
                                throw new Error('simulated image clipboard failure');
                              }
                              window.__monthlyClipboardWrite = true;
                            },
                            writeText: async (text) => { window.__monthlyClipboardText = String(text || ''); },
                          },
                        });
                        Object.defineProperty(window, 'ClipboardItem', {
                          configurable: true,
                          value: function ClipboardItem(items) { this.items = items; },
                        });
                        """
                    )
                    try:
                        page.goto(f"{base_url}/?{query}", wait_until="domcontentloaded")
                        page.get_by_role("heading", name="月報簡報素材").wait_for(state="visible")
                        page.locator(".monthly-copy-button").first.click()
                        page.get_by_text("圖片複製失敗，已改複製文字/HTML。").wait_for(state="visible")
                        self.assertEqual(page.get_by_text("已複製圖片").count(), 0)
                        self.assertTrue(page.evaluate("() => Boolean(window.__monthlyClipboardWrite || window.__monthlyClipboardText)"))
                    finally:
                        context.close()
                        browser.close()
            finally:
                server.shutdown()
                server.server_close()
                thread.join(timeout=2.0)


if __name__ == "__main__":
    unittest.main()
