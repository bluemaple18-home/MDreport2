from __future__ import annotations

import json
import sqlite3
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from domain.services import CanonicalService
from infra.sqlite.bootstrap import bootstrap_init
from infra.sqlite.repository import SQLiteRepository
from infra.ssp_api import SspApiSettings


class Phase2ServicesTests(unittest.TestCase):
    def _write_dsp_tab4_template(self, path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        try:
            ws_hidden_a = wb.active
            ws_hidden_a.title = "2025年_MF_合作績效統計總表"
            ws_hidden_a.sheet_state = "hidden"
            ws_hidden_a.freeze_panes = "C1"
            ws_hidden_a["A1"] = 2025

            ws_hidden_b = wb.create_sheet("2025_外部+行政_合作績效統計總表 ")
            ws_hidden_b.sheet_state = "hidden"
            ws_hidden_b.freeze_panes = "C1"
            ws_hidden_b["A1"] = "外部經銷商"

            ws_summary = wb.create_sheet("mF投資量_總表")
            ws_summary.freeze_panes = "M1"
            ws_summary.merge_cells("A1:D1")
            ws_summary["A1"] = 2026
            ws_summary["A2"] = "DSP投資額 總計"
            for idx in range(16, 24):
                ws_summary.row_dimensions[idx].hidden = True

            ws_detail = wb.create_sheet("各經銷商明細")
            ws_detail.freeze_panes = "U1"
            ws_detail.merge_cells("A2:D2")
            ws_detail["A2"] = "全體經銷 總投資量目標 & 達成率 (含北流)"
            ws_detail["A5"] = 2026

            ws_tracking = wb.create_sheet("北流進單追蹤")
            ws_tracking["A1"] = "2026年5月份_北流進單狀態"
            ws_tracking.column_dimensions["I"].hidden = True
            ws_tracking.column_dimensions["J"].hidden = True

            wb.save(path)
        finally:
            wb.close()

    def _full_row(self, **overrides: object) -> dict:
        row = {
            "日期時間": "2026-05-01 00:00:00",
            "經銷商": "A",
            "訂單": "O1",
            "素材": "C1",
            "廣告形式": "Banner",
            "尺寸": "300x250",
            "素材樣板": "tplA",
            "執行金額": 10.0,
            "系統營收": 12.5,
            "媒體費用": 8.0,
            "原始經銷商": "A",
            "原始廣告形式": "Banner",
            "最終經銷商": "A1",
            "規則命中_經銷商": "r1",
            "最終來源_經銷商": "rule",
            "分類層級B": "B1",
            "分類層級C": "C1",
            "分類層級D": "D1",
            "最終廣告形式": "Banner",
            "規則命中_廣告形式": "r2",
            "最終來源_廣告形式": "rule",
        }
        row.update(overrides)
        return row

    def _setup_project(self, root: Path) -> SQLiteRepository:
        (root / "migrations").mkdir(parents=True, exist_ok=True)
        (root / "templates").mkdir(parents=True, exist_ok=True)
        (root / "contracts").mkdir(parents=True, exist_ok=True)
        src = Path(__file__).resolve().parents[1]
        (root / "migrations" / "0001_initial.sql").write_text((src / "migrations" / "0001_initial.sql").read_text(encoding="utf-8"), encoding="utf-8")
        (root / "templates" / "template_registry.seed.json").write_text((src / "templates" / "template_registry.seed.json").read_text(encoding="utf-8"), encoding="utf-8")
        (root / "templates" / "ruleset.seed.json").write_text((src / "templates" / "ruleset.seed.json").read_text(encoding="utf-8"), encoding="utf-8")
        self._write_dsp_tab4_template(root / "templates" / "dsp_tab4_template.xlsx")
        (root / "contracts" / "fields_contract.json").write_text((src / "contracts" / "fields_contract.json").read_text(encoding="utf-8"), encoding="utf-8")
        (root / "bootstrap.manifest.json").write_text((src / "bootstrap.manifest.json").read_text(encoding="utf-8"), encoding="utf-8")

        result = bootstrap_init(root)
        return SQLiteRepository(Path(result["db_path"]), project_root=root)

    def test_save_modify_export_and_traceability(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            repo = self._setup_project(Path(td))
            svc = CanonicalService(repo)

            with repo.connect() as conn:
                self.assertEqual(
                    conn.execute("SELECT COUNT(1) FROM rule_bindings WHERE workflow='dsp'").fetchone()[0],
                    1,
                )

            save_out = svc.save(
                workflow="dsp",
                rows=[self._full_row()],
                template_version="v1",
                rule_version="v1",
            )
            self.assertTrue(save_out["run_id"].startswith("run-"))

            mod_out = svc.modify(
                workflow="dsp",
                updates=[{"row_order": 0, "column": "最終經銷商", "value": "A2"}],
                template_version="v1",
                rule_version="v1",
            )
            self.assertEqual(mod_out["changed_count"], 1)
            self.assertEqual(mod_out["adjustment_count"], 1)

            artifact_root = Path(td) / "artifacts"
            export_out = svc.export(
                workflow="dsp",
                artifact_root=artifact_root,
                template_version="v1",
                rule_version="v1",
            )
            self.assertTrue(Path(export_out["artifact_path"]).exists())
            self.assertTrue(str(export_out["artifact_path"]).endswith(".xlsx"))
            self.assertTrue(
                Path(export_out["artifact_path"]).name.startswith("2026 DSP投資量報表_")
            )
            wb = load_workbook(Path(export_out["artifact_path"]), data_only=True)
            expected_sheetnames = [
                "2025年_MF_合作績效統計總表",
                "2025_外部+行政_合作績效統計總表 ",
                "mF投資量_總表",
                "各經銷商明細",
                "北流進單追蹤",
            ]
            self.assertEqual(wb.sheetnames, expected_sheetnames)
            self.assertEqual(wb["2025年_MF_合作績效統計總表"].sheet_state, "hidden")
            self.assertEqual(wb["2025_外部+行政_合作績效統計總表 "].sheet_state, "hidden")
            self.assertEqual(wb["mF投資量_總表"].sheet_state, "visible")
            self.assertEqual(wb["各經銷商明細"].sheet_state, "visible")
            self.assertEqual(wb["北流進單追蹤"].sheet_state, "visible")
            self.assertEqual(wb["mF投資量_總表"].freeze_panes, "M1")
            self.assertEqual(wb["各經銷商明細"].freeze_panes, "U1")
            self.assertTrue(bool(wb["北流進單追蹤"].column_dimensions["I"].hidden))
            self.assertTrue(bool(wb["北流進單追蹤"].column_dimensions["J"].hidden))
            wb.close()

            conn = sqlite3.connect(str(repo.db_path))
            try:
                rows = conn.execute(
                    "SELECT run_type, source_db_hash, canonical_token, template_version, rule_version, artifact_checksum, detail_json FROM run_log ORDER BY created_at ASC"
                ).fetchall()
                self.assertEqual(len(rows), 3)
                export_detail: dict[str, object] | None = None
                for run_type, source_db_hash, canonical_token, template_version, rule_version, artifact_checksum, detail_json in rows:
                    self.assertIn(run_type, {"save", "modify", "export"})
                    self.assertTrue(str(source_db_hash))
                    self.assertTrue(str(canonical_token))
                    self.assertEqual(template_version, "v1")
                    self.assertEqual(rule_version, "v1")
                    detail = json.loads(detail_json)
                    self.assertEqual(detail["template_id"], "dsp-default")
                    self.assertEqual(detail["mapping_version"], "v1")
                    self.assertEqual(detail["rule_hash"], "bootstrap-v1")
                    if run_type == "export":
                        self.assertTrue(str(artifact_checksum))
                        self.assertTrue(str(detail["week_start"]))
                        self.assertTrue(str(detail["week_end"]))
                        export_detail = detail

                publish_rows = conn.execute(
                    "SELECT template_id, template_version, week_start, week_end, detail_json FROM publish_runs ORDER BY created_at ASC"
                ).fetchall()
                self.assertEqual(len(publish_rows), 1)
                self.assertEqual(publish_rows[0][0], "dsp-default")
                self.assertEqual(publish_rows[0][1], "v1")
                publish_week_start = str(publish_rows[0][2])
                publish_week_end = str(publish_rows[0][3])
                self.assertTrue(publish_week_start)
                self.assertTrue(publish_week_end)
                publish_detail = json.loads(publish_rows[0][4])
                self.assertEqual(publish_detail["template_id"], "dsp-default")
                self.assertEqual(publish_detail["mapping_version"], "v1")
                self.assertEqual(publish_detail["week_start"], publish_week_start)
                self.assertEqual(publish_detail["week_end"], publish_week_end)
                self.assertIsNotNone(export_detail)
                assert export_detail is not None
                self.assertEqual(str(export_detail["week_start"]), publish_week_start)
                self.assertEqual(str(export_detail["week_end"]), publish_week_end)

                adjustments = conn.execute(
                    "SELECT workflow, target_type, target_key, override_value, detail_json FROM overrides_adjustments ORDER BY id ASC"
                ).fetchall()
                self.assertEqual(len(adjustments), 1)
                adj_workflow, target_type, target_key, override_value, detail_json = adjustments[0]
                self.assertEqual(adj_workflow, "dsp")
                self.assertEqual(target_type, "manual_field")
                self.assertEqual(target_key, "row:0:最終經銷商")
                self.assertEqual(override_value, "A2")
                adj_detail = json.loads(detail_json)
                self.assertEqual(adj_detail["source"], "modify")
                self.assertEqual(adj_detail["row_order"], 0)
                self.assertEqual(adj_detail["column"], "最終經銷商")
                self.assertEqual(adj_detail["template_version"], "v1")
                self.assertEqual(adj_detail["rule_version"], "v1")
                self.assertEqual(adj_detail["run_id"], mod_out["run_id"])

                canonical_before = conn.execute("SELECT COUNT(1) FROM canonical_raw WHERE workflow='dsp'").fetchone()[0]

                audit_rows = conn.execute(
                    "SELECT event_type, scope, status, payload_json FROM audit_log ORDER BY id ASC"
                ).fetchall()
                service_events = [row for row in audit_rows if row[1] == "service"]
                self.assertGreaterEqual(len(service_events), 3)
                event_types = [row[0] for row in service_events]
                self.assertIn("save", event_types)
                self.assertIn("modify", event_types)
                self.assertIn("export", event_types)
                for event_type, scope, status, payload_json in service_events:
                    self.assertEqual(scope, "service")
                    self.assertEqual(status, "ok")
                    payload = json.loads(payload_json)
                    self.assertEqual(payload["workflow"], "dsp")
                    self.assertEqual(payload["template_version"], "v1")
                    self.assertEqual(payload["rule_version"], "v1")
                    self.assertTrue(str(payload["canonical_token"]))
                    if event_type == "export":
                        self.assertTrue(str(payload["artifact_checksum"]))
                        self.assertEqual(str(payload["week_start"]), publish_week_start)
                        self.assertEqual(str(payload["week_end"]), publish_week_end)
            finally:
                conn.close()

            # 刪除 artifact 不應影響 canonical。
            Path(export_out["artifact_path"]).unlink()
            conn = sqlite3.connect(str(repo.db_path))
            try:
                canonical_after = conn.execute("SELECT COUNT(1) FROM canonical_raw WHERE workflow='dsp'").fetchone()[0]
                self.assertEqual(canonical_before, canonical_after)
            finally:
                conn.close()

    def test_modify_rejects_uncontrolled_column(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            repo = self._setup_project(Path(td))
            svc = CanonicalService(repo)
            svc.save(
                workflow="dsp",
                rows=[self._full_row()],
                template_version="v1",
                rule_version="v1",
            )
            with self.assertRaises(ValueError):
                svc.modify(
                    workflow="dsp",
                    updates=[{"row_order": 0, "column": "經銷商", "value": "B"}],
                    template_version="v1",
                    rule_version="v1",
                )

    def test_modify_missing_row_raises_and_rolls_back(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            repo = self._setup_project(Path(td))
            svc = CanonicalService(repo)
            svc.save(
                workflow="dsp",
                rows=[self._full_row()],
                template_version="v1",
                rule_version="v1",
            )
            with self.assertRaises(LookupError):
                svc.modify(
                    workflow="dsp",
                    updates=[{"row_order": 9, "column": "最終經銷商", "value": "B"}],
                    template_version="v1",
                    rule_version="v1",
                )
            conn = sqlite3.connect(str(repo.db_path))
            try:
                row = conn.execute(
                    "SELECT 最終經銷商 FROM canonical_raw WHERE workflow='dsp' AND row_order=0"
                ).fetchone()
                self.assertEqual(row[0], "A1")
                run_count = conn.execute("SELECT COUNT(1) FROM run_log WHERE run_type='modify'").fetchone()[0]
                self.assertEqual(run_count, 0)
                adjustment_count = conn.execute("SELECT COUNT(1) FROM overrides_adjustments").fetchone()[0]
                self.assertEqual(adjustment_count, 0)
            finally:
                conn.close()

    def test_modify_rolls_back_when_adjustment_write_fails(self) -> None:
        class BrokenRepo(SQLiteRepository):
            def insert_override_adjustments(self, *args, **kwargs):  # type: ignore[override]
                raise RuntimeError("adjustment down")

        with tempfile.TemporaryDirectory() as td:
            repo = self._setup_project(Path(td))
            svc = CanonicalService(repo)
            svc.save(
                workflow="dsp",
                rows=[self._full_row()],
                template_version="v1",
                rule_version="v1",
            )
            broken = BrokenRepo(repo.db_path, project_root=Path(td))
            broken_svc = CanonicalService(broken)
            with self.assertRaises(RuntimeError):
                broken_svc.modify(
                    workflow="dsp",
                    updates=[{"row_order": 0, "column": "最終經銷商", "value": "A2"}],
                    template_version="v1",
                    rule_version="v1",
                )
            conn = sqlite3.connect(str(repo.db_path))
            try:
                row = conn.execute(
                    "SELECT 最終經銷商 FROM canonical_raw WHERE workflow='dsp' AND row_order=0"
                ).fetchone()
                self.assertEqual(row[0], "A1")
                run_count = conn.execute("SELECT COUNT(1) FROM run_log WHERE run_type='modify'").fetchone()[0]
                self.assertEqual(run_count, 0)
                adjustment_count = conn.execute("SELECT COUNT(1) FROM overrides_adjustments").fetchone()[0]
                self.assertEqual(adjustment_count, 0)
            finally:
                conn.close()

    def test_canonical_token_changes_on_raw_state_change(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            repo = self._setup_project(Path(td))
            svc = CanonicalService(repo)
            svc.save(
                workflow="dsp",
                rows=[self._full_row()],
                template_version="v1",
                rule_version="v1",
            )
            with repo.connect() as conn:
                before = repo.canonical_token(conn, "dsp")
            svc.save(
                workflow="dsp",
                rows=[self._full_row(經銷商="B")],
                template_version="v1",
                rule_version="v1",
            )
            with repo.connect() as conn:
                after = repo.canonical_token(conn, "dsp")
            self.assertNotEqual(before, after)

    def test_save_rolls_back_when_run_log_fails(self) -> None:
        class BrokenRepo(SQLiteRepository):
            def insert_run_log(self, *args, **kwargs):  # type: ignore[override]
                raise RuntimeError("boom")

        with tempfile.TemporaryDirectory() as td:
            repo = self._setup_project(Path(td))
            broken = BrokenRepo(repo.db_path, project_root=Path(td))
            svc = CanonicalService(broken)
            with self.assertRaises(RuntimeError):
                svc.save(
                    workflow="dsp",
                    rows=[self._full_row()],
                    template_version="v1",
                    rule_version="v1",
                )
            conn = sqlite3.connect(str(repo.db_path))
            try:
                count = conn.execute("SELECT COUNT(1) FROM canonical_raw WHERE workflow='dsp'").fetchone()[0]
                self.assertEqual(count, 0)
            finally:
                conn.close()

    def test_export_cleans_up_orphan_artifact_on_log_failure(self) -> None:
        class BrokenRepo(SQLiteRepository):
            def insert_run_log(self, *args, **kwargs):  # type: ignore[override]
                raise RuntimeError("boom")

        with tempfile.TemporaryDirectory() as td:
            repo = self._setup_project(Path(td))
            svc = CanonicalService(repo)
            svc.save(
                workflow="dsp",
                rows=[self._full_row()],
                template_version="v1",
                rule_version="v1",
            )

            broken = BrokenRepo(repo.db_path, project_root=Path(td))
            broken_svc = CanonicalService(broken)
            artifact_root = Path(td) / "artifacts"
            with self.assertRaises(RuntimeError):
                broken_svc.export(
                    workflow="dsp",
                    artifact_root=artifact_root,
                    template_version="v1",
                    rule_version="v1",
                )
            self.assertEqual(len(list(artifact_root.glob("*.xlsx"))), 0)

    def test_export_supports_explicit_week_period_and_validates_period(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            repo = self._setup_project(Path(td))
            svc = CanonicalService(repo)
            svc.save(
                workflow="dsp",
                rows=[self._full_row()],
                template_version="v1",
                rule_version="v1",
            )

            export_out = svc.export(
                workflow="dsp",
                artifact_root=Path(td) / "artifacts",
                template_version="v1",
                rule_version="v1",
                week_start="2026-04-27",
                week_end="2026-05-03",
            )
            self.assertEqual(export_out["week_start"], "2026-04-27")
            self.assertEqual(export_out["week_end"], "2026-05-03")
            self.assertEqual(
                Path(export_out["artifact_path"]).name,
                "2026 DSP投資量報表_0427-0503.xlsx",
            )

            conn = sqlite3.connect(str(repo.db_path))
            try:
                publish = conn.execute(
                    "SELECT week_start, week_end FROM publish_runs WHERE run_id = ?",
                    (export_out["run_id"],),
                ).fetchone()
                self.assertIsNotNone(publish)
                assert publish is not None
                self.assertEqual(str(publish[0]), "2026-04-27")
                self.assertEqual(str(publish[1]), "2026-05-03")
            finally:
                conn.close()

            with self.assertRaises(ValueError):
                svc.export(
                    workflow="dsp",
                    artifact_root=Path(td) / "artifacts",
                    template_version="v1",
                    rule_version="v1",
                    week_start="2026-04-27",
                )
            with self.assertRaises(ValueError):
                svc.export(
                    workflow="dsp",
                    artifact_root=Path(td) / "artifacts",
                    template_version="v1",
                    rule_version="v1",
                    week_start="2026/04/27",
                    week_end="2026-05-03",
                )
            with self.assertRaises(ValueError):
                svc.export(
                    workflow="dsp",
                    artifact_root=Path(td) / "artifacts",
                    template_version="v1",
                    rule_version="v1",
                    week_start="2026-05-10",
                    week_end="2026-05-03",
                )

    def test_dsp_export_adds_only_period_rows_to_previous_week_baseline(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            repo = self._setup_project(root)
            baseline_root = root / "data_seed" / "dsp_weekly_baselines"
            baseline_root.mkdir(parents=True, exist_ok=True)
            baseline_path = baseline_root / "2026 DSP投資量報表_0101-0503.xlsx"

            wb = load_workbook(root / "templates" / "dsp_tab4_template.xlsx")
            try:
                ws_detail = wb["各經銷商明細"]
                ws_detail["E7"] = 50
                ws_detail["M7"] = 100
                wb.save(baseline_path)
            finally:
                wb.close()
            (baseline_root / "manifest.json").write_text(
                json.dumps(
                    {
                        "kind": "dsp_weekly_baseline_workbooks",
                        "env": "test",
                        "files": [
                            {
                                "week_end": "2026-05-03",
                                "path": baseline_path.name,
                            }
                        ],
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            svc = CanonicalService(repo)
            svc.save(
                workflow="dsp",
                rows=[
                    self._full_row(日期時間="2026-05-05 00:00:00", 執行金額=25.0),
                    self._full_row(日期時間="2026-04-20 00:00:00", 執行金額=999.0),
                ],
                template_version="v1",
                rule_version="v1",
            )

            export_out = svc.export(
                workflow="dsp",
                artifact_root=root / "artifacts",
                template_version="v1",
                rule_version="v1",
                week_start="2026-05-04",
                week_end="2026-05-10",
            )
            self.assertEqual(int(export_out["row_count"]), 1)

            export_wb = load_workbook(Path(export_out["artifact_path"]), data_only=False)
            try:
                ws_detail = export_wb["各經銷商明細"]
                self.assertEqual(float(ws_detail["E7"].value), 50.0)
                self.assertEqual(float(ws_detail["M7"].value), 125.0)
            finally:
                export_wb.close()

    def test_dsp_export_supports_baselines_manifest_shape(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            repo = self._setup_project(root)
            baseline_root = root / "data_seed" / "dsp_weekly_baselines"
            baseline_root.mkdir(parents=True, exist_ok=True)
            baseline_path = baseline_root / "2026 DSP投資量報表_0101-0503.xlsx"

            wb = load_workbook(root / "templates" / "dsp_tab4_template.xlsx")
            try:
                ws_detail = wb["各經銷商明細"]
                ws_detail["M7"] = 100
                wb.save(baseline_path)
            finally:
                wb.close()
            (baseline_root / "manifest.json").write_text(
                json.dumps(
                    {
                        "workflow": "dsp",
                        "baselines": [
                            {
                                "week_end": "2026-05-03",
                                "file": baseline_path.name,
                            }
                        ],
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            svc = CanonicalService(repo)
            svc.save(
                workflow="dsp",
                rows=[self._full_row(日期時間="2026-05-05 00:00:00", 執行金額=25.0)],
                template_version="v1",
                rule_version="v1",
            )

            export_out = svc.export(
                workflow="dsp",
                artifact_root=root / "artifacts",
                template_version="v1",
                rule_version="v1",
                week_start="2026-05-04",
                week_end="2026-05-10",
            )

            export_wb = load_workbook(Path(export_out["artifact_path"]), data_only=False)
            try:
                ws_detail = export_wb["各經銷商明細"]
                self.assertEqual(float(ws_detail["M7"].value), 125.0)
            finally:
                export_wb.close()

    def test_dsp_export_fails_closed_when_only_test_weekly_baselines_exist(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            repo = self._setup_project(root)
            baseline_root = root / "data_seed_test" / "dsp_weekly_baselines"
            baseline_root.mkdir(parents=True, exist_ok=True)
            (baseline_root / "manifest.json").write_text(
                json.dumps(
                    {
                        "kind": "dsp_weekly_baseline_workbooks",
                        "env": "test",
                        "files": [
                            {
                                "week_end": "2026-05-03",
                                "path": "2026 DSP投資量報表_0101-0503.xlsx",
                            }
                        ],
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            svc = CanonicalService(repo)
            svc.save(
                workflow="dsp",
                rows=[self._full_row(日期時間="2026-05-05 00:00:00", 執行金額=25.0)],
                template_version="v1",
                rule_version="v1",
            )

            with self.assertRaisesRegex(FileNotFoundError, "找不到 DSP 週報基底 workbook"):
                svc.export(
                    workflow="dsp",
                    artifact_root=root / "artifacts",
                    template_version="v1",
                    rule_version="v1",
                    week_start="2026-05-04",
                    week_end="2026-05-10",
                )

    def test_validate_dsp_export_request_only_gates_ui_flow(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            repo = self._setup_project(Path(td))
            svc = CanonicalService(repo)
            svc.save(
                workflow="dsp",
                rows=[self._full_row()],
                template_version="v1",
                rule_version="v1",
            )

            artifact_root = Path(td) / "artifacts"
            export_out = svc.export(
                workflow="dsp",
                artifact_root=artifact_root,
                template_version="v1",
                rule_version="v1",
            )
            self.assertTrue(Path(str(export_out["artifact_path"])).exists())
            self.assertEqual(str(export_out.get("delivery_snapshot_token") or ""), "")
            self.assertEqual(str(export_out.get("delivery_run_id") or ""), "")

            with self.assertRaisesRegex(PermissionError, "tab4 delivery required"):
                svc.validate_dsp_export_request(
                    workflow="dsp",
                    main_tab="dsp_tab4",
                    sub_tab="overview",
                    template_version="v1",
                    rule_version="v1",
                )

            delivered = svc.mark_tab4_delivery(
                workflow="dsp",
                main_tab="dsp_tab3",
                sub_tab="pivot",
                template_version="v1",
                rule_version="v1",
                week_start="2026-04-27",
                week_end="2026-05-03",
            )
            delivery_meta = svc.validate_dsp_export_request(
                workflow="dsp",
                main_tab="dsp_tab4",
                sub_tab="overview",
                template_version="v1",
                rule_version="v1",
                week_start="2026-04-27",
                week_end="2026-05-03",
            )
            self.assertEqual(
                delivery_meta["delivery_snapshot_token"],
                str(delivered["delivery_snapshot_token"]),
            )
            self.assertEqual(
                delivery_meta["delivery_run_id"],
                str(delivered["run_id"]),
            )
            with self.assertRaisesRegex(PermissionError, "period mismatch"):
                svc.validate_dsp_export_request(
                    workflow="dsp",
                    main_tab="dsp_tab4",
                    sub_tab="overview",
                    template_version="v1",
                    rule_version="v1",
                    week_start="2026-05-04",
                    week_end="2026-05-10",
                )

    def test_save_rejects_contract_unknown_fields(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            repo = self._setup_project(Path(td))
            svc = CanonicalService(repo)
            bad_row = self._full_row(不存在欄位="x")
            with self.assertRaises(ValueError):
                svc.save(
                    workflow="dsp",
                    rows=[bad_row],
                    template_version="v1",
                    rule_version="v1",
                )

    def test_save_rejects_contract_type_mismatch(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            repo = self._setup_project(Path(td))
            svc = CanonicalService(repo)
            with self.assertRaises(ValueError):
                svc.save(
                    workflow="dsp",
                    rows=[self._full_row(執行金額="oops")],
                    template_version="v1",
                    rule_version="v1",
                )

    def test_export_fails_when_binding_missing(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            repo = self._setup_project(Path(td))
            svc = CanonicalService(repo)
            svc.save(
                workflow="dsp",
                rows=[self._full_row()],
                template_version="v1",
                rule_version="v1",
            )
            conn = sqlite3.connect(str(repo.db_path))
            try:
                conn.execute("DELETE FROM rule_bindings WHERE workflow='dsp'")
                conn.commit()
            finally:
                conn.close()
            with self.assertRaises(LookupError):
                svc.export(
                    workflow="dsp",
                    artifact_root=Path(td) / "artifacts",
                    template_version="v1",
                    rule_version="v1",
                )

    def test_save_applies_contract_defaults_for_missing_fields(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            repo = self._setup_project(Path(td))
            svc = CanonicalService(repo)
            # 只提供部分欄位，其餘應由 fields_contract default 補齊。
            minimal = {
                "日期時間": "2026-05-01 00:00:00",
                "經銷商": "A",
                "訂單": "O1",
                "素材": "C1",
                "廣告形式": "Banner",
                "尺寸": "300x250",
                "素材樣板": "tplA",
                "執行金額": 10.0,
                "系統營收": 12.5,
                "媒體費用": 8.0,
            }
            out = svc.save(workflow="dsp", rows=[minimal], template_version="v1", rule_version="v1")
            self.assertTrue(out["run_id"].startswith("run-"))
            conn = sqlite3.connect(str(repo.db_path))
            try:
                row = conn.execute(
                    "SELECT 原始經銷商, 最終經銷商, 規則命中_經銷商 FROM canonical_raw WHERE workflow='dsp' AND row_order=0"
                ).fetchone()
                self.assertEqual(row[0], "")
                self.assertEqual(row[1], "")
                self.assertEqual(row[2], "")
            finally:
                conn.close()

    def test_contract_loads_with_nonstandard_db_path(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            repo = self._setup_project(root)
            conn = sqlite3.connect(str(repo.db_path))
            conn.close()
            custom_db = root / "runtime" / "sqlite" / "prod.sqlite"
            custom_db.parent.mkdir(parents=True, exist_ok=True)
            src_conn = sqlite3.connect(str(repo.db_path))
            dst_conn = sqlite3.connect(str(custom_db))
            try:
                src_conn.backup(dst_conn)
            finally:
                src_conn.close()
                dst_conn.close()
            custom_repo = SQLiteRepository(custom_db, project_root=root)
            svc = CanonicalService(custom_repo)
            out = svc.save(workflow="dsp", rows=[self._full_row()], template_version="v1", rule_version="v1")
            self.assertTrue(out["run_id"].startswith("run-"))

    def test_repository_requires_explicit_contract_source(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            repo = self._setup_project(Path(td))
            with self.assertRaises(ValueError):
                SQLiteRepository(repo.db_path)

    def test_canonical_schema_matches_field_contract(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            repo = self._setup_project(Path(td))
            conn = sqlite3.connect(str(repo.db_path))
            try:
                schema_columns = [
                    str(row[1])
                    for row in conn.execute("PRAGMA table_info(canonical_raw)").fetchall()
                ]
            finally:
                conn.close()
            canonical_columns = [
                col for col in schema_columns if col not in {"id", "workflow", "row_order", "updated_at"}
            ]
            self.assertEqual(canonical_columns, repo.field_contract.field_names)

    def test_ssp_save_modify_export_and_traceability_parity(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            repo = self._setup_project(Path(td))
            svc = CanonicalService(repo)

            save_out = svc.save(
                workflow="ssp",
                rows=[self._full_row(經銷商="SSP_A", 最終經銷商="SSP_A1")],
                template_version="v1",
                rule_version="v1",
            )
            self.assertTrue(save_out["run_id"].startswith("run-"))

            mod_out = svc.modify(
                workflow="ssp",
                updates=[{"row_order": 0, "column": "最終經銷商", "value": "SSP_A2"}],
                template_version="v1",
                rule_version="v1",
            )
            self.assertEqual(mod_out["changed_count"], 1)
            self.assertEqual(mod_out["adjustment_count"], 1)

            artifact_root = Path(td) / "artifacts"
            export_out = svc.export(
                workflow="ssp",
                artifact_root=artifact_root,
                template_version="v1",
                rule_version="v1",
            )
            self.assertTrue(Path(export_out["artifact_path"]).exists())
            self.assertTrue(str(export_out["artifact_path"]).endswith(".xlsx"))
            wb = load_workbook(Path(export_out["artifact_path"]), data_only=True)
            try:
                ws_data = wb["canonical_data"]
                headers = [cell.value for cell in ws_data[1]]
                first_row = dict(zip(headers, [cell.value for cell in ws_data[2]]))
                self.assertEqual(first_row["最終經銷商"], "SSP_A2")
                ws_meta = wb["metadata"]
                meta = {str(r[0]): str(r[1]) for r in ws_meta.iter_rows(min_row=2, values_only=True)}
                self.assertEqual(meta["workflow"], "ssp")
                self.assertEqual(meta["template_version"], "v1")
                self.assertEqual(meta["rule_version"], "v1")
            finally:
                wb.close()

            conn = sqlite3.connect(str(repo.db_path))
            try:
                run_rows = conn.execute(
                    "SELECT run_type, workflow, template_version, rule_version FROM run_log WHERE workflow='ssp' ORDER BY created_at ASC"
                ).fetchall()
                self.assertEqual(len(run_rows), 3)
                for run_type, workflow, template_version, rule_version in run_rows:
                    self.assertIn(run_type, {"save", "modify", "export"})
                    self.assertEqual(workflow, "ssp")
                    self.assertEqual(template_version, "v1")
                    self.assertEqual(rule_version, "v1")

                publish = conn.execute(
                    "SELECT template_id, template_version FROM publish_runs WHERE run_id = ?",
                    (export_out["run_id"],),
                ).fetchone()
                self.assertIsNotNone(publish)
                self.assertEqual(publish[0], "ssp-default")
                self.assertEqual(publish[1], "v1")

                evidence_count = conn.execute(
                    "SELECT COUNT(1) FROM evidence_index WHERE run_id = ?",
                    (export_out["run_id"],),
                ).fetchone()[0]
                self.assertEqual(evidence_count, 1)

                adjustment = conn.execute(
                    "SELECT workflow, target_key, override_value FROM overrides_adjustments WHERE workflow='ssp' ORDER BY id DESC LIMIT 1"
                ).fetchone()
                self.assertIsNotNone(adjustment)
                self.assertEqual(adjustment[0], "ssp")
                self.assertEqual(adjustment[1], "row:0:最終經銷商")
                self.assertEqual(adjustment[2], "SSP_A2")
            finally:
                conn.close()

    def test_ssp_media_save_writes_run_and_audit_log(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            repo = self._setup_project(Path(td))
            svc = CanonicalService(repo)

            out = svc.save_ssp_media_slots(
                runtime_env="test",
                slots=[
                    {
                        "category": "蓋板",
                        "slot_order": 0,
                        "placement_id": "M100",
                        "placement_name": "Media Slot",
                        "media_target": 1000,
                        "is_active": True,
                    }
                ],
                template_version="v1",
                rule_version="v1",
            )
            self.assertEqual(out["status"], "ok")
            self.assertEqual(out["runtime_env"], "test")
            self.assertEqual(out["row_count"], 1)
            self.assertTrue(str(out["run_id"]).startswith("run-"))

            conn = sqlite3.connect(str(repo.db_path))
            try:
                run = conn.execute(
                    """
                    SELECT run_type, workflow, template_version, rule_version, detail_json
                    FROM run_log
                    WHERE run_id = ?
                    """,
                    (out["run_id"],),
                ).fetchone()
                self.assertIsNotNone(run)
                assert run is not None
                self.assertEqual(run[0], "ssp_media_save")
                self.assertEqual(run[1], "ssp")
                self.assertEqual(run[2], "v1")
                self.assertEqual(run[3], "v1")
                detail = json.loads(str(run[4]))
                self.assertEqual(str(detail.get("runtime_env") or ""), "test")
                self.assertEqual(int(detail.get("row_count") or 0), 1)
                self.assertTrue(str(detail.get("template_id") or ""))
                self.assertTrue(str(detail.get("mapping_version") or ""))
                self.assertTrue(str(detail.get("rule_hash") or ""))

                audit = conn.execute(
                    """
                    SELECT event_type, scope, status, payload_json
                    FROM audit_log
                    WHERE event_type = 'ssp_media_save'
                    ORDER BY id DESC
                    LIMIT 1
                    """
                ).fetchone()
                self.assertIsNotNone(audit)
                assert audit is not None
                self.assertEqual(audit[0], "ssp_media_save")
                self.assertEqual(audit[1], "service")
                self.assertEqual(audit[2], "ok")
                payload = json.loads(str(audit[3]))
                self.assertEqual(str(payload.get("workflow") or ""), "ssp")
                self.assertEqual(str(payload.get("run_id") or ""), out["run_id"])
                self.assertEqual(str(payload.get("template_version") or ""), "v1")
                self.assertEqual(str(payload.get("rule_version") or ""), "v1")
                self.assertEqual(str(payload.get("runtime_env") or ""), "test")
                self.assertEqual(int(payload.get("row_count") or 0), 1)
                self.assertTrue(str(payload.get("canonical_token") or ""))
            finally:
                conn.close()

    def test_ssp_export_falls_back_to_canonical_rows_when_ssp_raw_empty(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            repo = self._setup_project(Path(td))
            svc = CanonicalService(repo)

            svc.save(
                workflow="ssp",
                rows=[self._full_row(經銷商="SSP_FALLBACK", 最終經銷商="SSP_CANONICAL")],
                template_version="v1",
                rule_version="v1",
            )

            snapshot = svc.resolve_ssp_effective_snapshot()
            self.assertEqual(snapshot["source"], "canonical_raw")
            self.assertEqual(set(snapshot["manual_fields"]), set(repo.modify_allowed_columns))

            export_out = svc.export(
                workflow="ssp",
                artifact_root=Path(td) / "artifacts",
                template_version="v1",
                rule_version="v1",
            )

            wb = load_workbook(Path(export_out["artifact_path"]), data_only=True)
            try:
                ws_data = wb["canonical_data"]
                headers = [cell.value for cell in ws_data[1]]
                first_row = dict(zip(headers, [cell.value for cell in ws_data[2]]))
                self.assertEqual(headers, repo.canonical_columns)
                self.assertEqual(first_row["經銷商"], "SSP_FALLBACK")
                self.assertEqual(first_row["最終經銷商"], "SSP_CANONICAL")
            finally:
                wb.close()

    def test_fetch_ssp_api_replaces_only_ssp_runtime_rows_and_preserves_dsp(self) -> None:
        class _FakeSspApiClient:
            def __init__(self, settings: SspApiSettings) -> None:
                self.settings = settings

            def fetch_report_bundle(self, *, start_day: str, end_day: str) -> dict[str, object]:
                if start_day != "2026-05-10" or end_day != "2026-05-11":
                    raise AssertionError("SSP API fetch range should be preserved")
                return {
                    "auth": {
                        "service_id": self.settings.service_id,
                        "user": {"id": 2072, "email": "ssp@example.com"},
                    },
                    "login": {"id": 2072, "email": "ssp@example.com"},
                    "report_id": 102,
                    "report_ids": [101, 102],
                    "records_total": 2,
                    "chunk_mode": "daily",
                    "chunk_days": 2,
                    "daily": [
                        {"date": "2026-05-10", "report_id": 101, "row_count": 1, "records_total": 1},
                        {"date": "2026-05-11", "report_id": 102, "row_count": 1, "records_total": 1},
                    ],
                    "sum_row": {"request": 30, "impress": 300, "click": 20},
                    "rows": [
                        {
                            "data_time": "2026-05-10 00:00:00",
                            "zone_id": "1001",
                            "zoneName": "SSP Slot A",
                            "supplier_id": "11",
                            "supplierName": "Supplier A",
                            "site_id": "21",
                            "siteName": "Site A",
                            "request": "10",
                            "impress": "100",
                            "click": "5",
                            "profit": "1.5",
                            "advertiser_mu": "3.0",
                        },
                        {
                            "data_time": "2026-05-11 00:00:00",
                            "zone_id": "1002",
                            "zoneName": "SSP Slot B",
                            "supplier_id": "12",
                            "supplierName": "Supplier B",
                            "site_id": "22",
                            "siteName": "Site B",
                            "request": "20",
                            "impress": "200",
                            "click": "15",
                            "profit": "2.5",
                            "advertiser_mu": "5.0",
                        },
                    ],
                }

        with tempfile.TemporaryDirectory() as td:
            repo = self._setup_project(Path(td))
            svc = CanonicalService(repo)
            svc.save(
                workflow="dsp",
                rows=[self._full_row(經銷商="DSP_KEEP", 最終經銷商="DSP_KEEP")],
                template_version="v1",
                rule_version="v1",
            )
            svc.save(
                workflow="ssp",
                rows=[self._full_row(經銷商="SSP_OLD", 最終經銷商="SSP_OLD")],
                template_version="v1",
                rule_version="v1",
            )

            with repo.connect() as conn:
                dsp_before = conn.execute("SELECT COUNT(1) FROM canonical_raw WHERE workflow='dsp'").fetchone()[0]
                ssp_before = conn.execute("SELECT COUNT(1) FROM canonical_raw WHERE workflow='ssp'").fetchone()[0]
            self.assertEqual(dsp_before, 1)
            self.assertEqual(ssp_before, 1)

            with patch("domain.services.SspApiClient", _FakeSspApiClient):
                out = svc.fetch_ssp_api(
                    start_day="2026-05-10",
                    end_day="2026-05-11",
                    template_version="v1",
                    rule_version="v1",
                    email="ssp@example.com",
                    password="secret",
                )

            self.assertEqual(out["status"], "ok")
            self.assertEqual(out["row_count"], 2)
            self.assertEqual(out["records_total"], 2)
            self.assertEqual(out["report_ids"], [101, 102])
            self.assertEqual(out["chunk_mode"], "daily")
            self.assertEqual(out["chunk_days"], 2)
            self.assertEqual(len(out["daily"]), 2)

            with repo.connect() as conn:
                dsp_after = conn.execute("SELECT COUNT(1) FROM canonical_raw WHERE workflow='dsp'").fetchone()[0]
                ssp_after = conn.execute("SELECT COUNT(1) FROM canonical_raw WHERE workflow='ssp'").fetchone()[0]
                ssp_raw_summary = conn.execute(
                    "SELECT COUNT(1), MIN(date), MAX(date), MIN(source), MAX(source) FROM ssp_raw"
                ).fetchone()
                dsp_label = conn.execute(
                    "SELECT 最終經銷商 FROM canonical_raw WHERE workflow='dsp' AND row_order=0"
                ).fetchone()[0]
                detail_json = conn.execute(
                    "SELECT detail_json FROM run_log WHERE run_type='fetch_ssp_api' AND workflow='ssp' ORDER BY created_at DESC LIMIT 1"
                ).fetchone()[0]

            self.assertEqual(dsp_after, dsp_before)
            self.assertEqual(dsp_label, "DSP_KEEP")
            self.assertEqual(ssp_after, 0)
            self.assertEqual(tuple(ssp_raw_summary), (2, "2026-05-10", "2026-05-11", "ssp3_api", "ssp3_api"))
            detail = json.loads(str(detail_json))
            self.assertEqual(detail["daily"], out["daily"])
            self.assertEqual(detail["chunk_days"], 2)


if __name__ == "__main__":
    unittest.main()
