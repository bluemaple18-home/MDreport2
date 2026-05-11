from __future__ import annotations

import json
import shutil
import subprocess
import tempfile
import unittest
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook


class AppShellCliTests(unittest.TestCase):
    def _write_dsp_tab4_template(self, path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        try:
            ws_hidden_a = wb.active
            ws_hidden_a.title = "2025年_MF_合作績效統計總表"
            ws_hidden_a.sheet_state = "hidden"
            ws_hidden_a.freeze_panes = "C1"
            ws_hidden_a["A1"] = 2025

            ws_hidden_b = wb.create_sheet("2025_外部+行政_合作績效統計總表 ")
            ws_hidden_b.sheet_state = "hidden"
            ws_hidden_b.freeze_panes = "C1"
            ws_hidden_b["A1"] = "外部經銷商"

            ws_summary = wb.create_sheet("mF投資量_總表")
            ws_summary.freeze_panes = "M1"
            ws_summary.merge_cells("A1:D1")
            ws_summary["A1"] = 2026
            ws_summary["A2"] = "DSP投資額 總計"
            for idx in range(16, 24):
                ws_summary.row_dimensions[idx].hidden = True

            ws_detail = wb.create_sheet("各經銷商明細")
            ws_detail.freeze_panes = "U1"
            ws_detail.merge_cells("A2:D2")
            ws_detail["A2"] = "全體經銷 總投資量目標 & 達成率 (含北流)"
            ws_detail["A5"] = 2026

            ws_tracking = wb.create_sheet("北流進單追蹤")
            ws_tracking["A1"] = "2026年5月份_北流進單狀態"
            ws_tracking.column_dimensions["I"].hidden = True
            ws_tracking.column_dimensions["J"].hidden = True

            wb.save(path)
        finally:
            wb.close()

    def _make_project(self, root: Path) -> None:
        src = Path(__file__).resolve().parents[1]
        (root / "migrations").mkdir(parents=True, exist_ok=True)
        (root / "templates").mkdir(parents=True, exist_ok=True)
        (root / "contracts").mkdir(parents=True, exist_ok=True)
        (root / "data_seed" / "templates_rules_mapping").mkdir(parents=True, exist_ok=True)
        (root / "data_seed_test" / "templates_rules_mapping").mkdir(parents=True, exist_ok=True)
        (root / "migrations" / "0001_initial.sql").write_text((src / "migrations" / "0001_initial.sql").read_text(encoding="utf-8"), encoding="utf-8")
        (root / "templates" / "template_registry.seed.json").write_text((src / "templates" / "template_registry.seed.json").read_text(encoding="utf-8"), encoding="utf-8")
        (root / "templates" / "ruleset.seed.json").write_text((src / "templates" / "ruleset.seed.json").read_text(encoding="utf-8"), encoding="utf-8")
        self._write_dsp_tab4_template(root / "templates" / "dsp_tab4_template.xlsx")
        (root / "contracts" / "fields_contract.json").write_text((src / "contracts" / "fields_contract.json").read_text(encoding="utf-8"), encoding="utf-8")
        (root / "bootstrap.manifest.json").write_text((src / "bootstrap.manifest.json").read_text(encoding="utf-8"), encoding="utf-8")
        (root / "bootstrap.test.manifest.json").write_text((src / "bootstrap.test.manifest.json").read_text(encoding="utf-8"), encoding="utf-8")
        default_group_overrides = {
            "蓋板": [{"placement_id": 8435, "placement_name": "MW_蓋版_COOL", "remark": "", "media_target": 1000}],
            "置底": [{"placement_id": 17236, "placement_name": "MW_置底創意_UDN聯合新聞", "remark": "", "media_target": 1200}],
            "置底展開": [{"placement_id": 22218, "placement_name": "MW_置底(展開)_上報", "remark": "", "media_target": 800}],
            "文中300x250": [{"placement_id": 16980, "placement_name": "MW_文中創意300x250_大人物", "remark": "", "media_target": 900}],
            "文中320x480": [{"placement_id": 16977, "placement_name": "MW_文中創意320x480_大人物", "remark": "", "media_target": 700}],
        }
        for rel_path in (
            root / "data_seed" / "templates_rules_mapping" / "group_overrides.json",
            root / "data_seed_test" / "templates_rules_mapping" / "group_overrides.json",
        ):
            rel_path.write_text(json.dumps(default_group_overrides, ensure_ascii=False), encoding="utf-8")

    def _run(self, root: Path, *args: str) -> subprocess.CompletedProcess[str]:
        cmd = [
            "python3",
            str(Path(__file__).resolve().parents[1] / "app" / "main.py"),
            "--root",
            str(root),
            *args,
        ]
        return subprocess.run(cmd, capture_output=True, text=True, check=False)

    def _make_legacy_mdreport_seed_source(self, root: Path) -> Path:
        legacy = root / "legacy-mdreport"
        (legacy / "data").mkdir(parents=True, exist_ok=True)
        (legacy / "artifacts").mkdir(parents=True, exist_ok=True)

        import sqlite3

        for db_name in ("mdreport.sqlite", "mdreport_dsp.sqlite", "anomaly.sqlite", "volume.sqlite"):
            db_path = legacy / "data" / db_name
            conn = sqlite3.connect(str(db_path))
            try:
                conn.execute("CREATE TABLE IF NOT EXISTS probe(id INTEGER PRIMARY KEY, v TEXT)")
                conn.execute("INSERT INTO probe(v) VALUES('ok')")
                conn.commit()
            finally:
                conn.close()
        (legacy / "data" / "group_overrides.json").write_text(
            json.dumps({"A": "A1"}, ensure_ascii=False),
            encoding="utf-8",
        )

        (legacy / "artifacts" / "dsp_rawdata_20260506_150144_recalc.json").write_text(
            json.dumps([{"日期時間": "2026-05-06 00:00:00", "經銷商": "A", "訂單": "O1"}], ensure_ascii=False),
            encoding="utf-8",
        )
        (legacy / "artifacts" / "dsp_weekly_accum_20260507_155315_weekly_meta.json").write_text(
            json.dumps({"week_start": "2026-05-05", "week_end": "2026-05-11"}, ensure_ascii=False),
            encoding="utf-8",
        )
        (legacy / "artifacts" / "dsp_dataset_manifest.json").write_text(
            json.dumps([{"batch": "b1"}], ensure_ascii=False),
            encoding="utf-8",
        )
        (legacy / "artifacts" / "dsp_rawdata_20260506_150144_check.json").write_text(
            json.dumps({"skip": True}, ensure_ascii=False),
            encoding="utf-8",
        )
        (legacy / "artifacts" / "CARD-T34-245_probe.json").write_text(
            json.dumps({"noise": True}, ensure_ascii=False),
            encoding="utf-8",
        )
        return legacy

    def _make_seed_canonical_dsp_db(self, root: Path) -> Path:
        seed_db = root / "data_seed" / "canonical" / "mdreport_dsp.sqlite"
        seed_db.parent.mkdir(parents=True, exist_ok=True)
        import sqlite3
        conn = sqlite3.connect(str(seed_db))
        try:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS dsp_rawdata (
                  row_order INTEGER PRIMARY KEY,
                  日期時間 TEXT NOT NULL DEFAULT '',
                  經銷商 TEXT NOT NULL DEFAULT '',
                  訂單 TEXT NOT NULL DEFAULT '',
                  素材 TEXT NOT NULL DEFAULT '',
                  廣告形式 TEXT NOT NULL DEFAULT '',
                  尺寸 TEXT NOT NULL DEFAULT '',
                  素材樣板 TEXT NOT NULL DEFAULT '',
                  執行金額 REAL NOT NULL DEFAULT 0.0,
                  系統營收 REAL NOT NULL DEFAULT 0.0,
                  媒體費用 REAL NOT NULL DEFAULT 0.0,
                  原始經銷商 TEXT NOT NULL DEFAULT '',
                  原始廣告形式 TEXT NOT NULL DEFAULT '',
                  最終經銷商 TEXT NOT NULL DEFAULT '',
                  規則命中_經銷商 TEXT NOT NULL DEFAULT '',
                  最終來源_經銷商 TEXT NOT NULL DEFAULT '',
                  分類層級B TEXT NOT NULL DEFAULT '',
                  分類層級C TEXT NOT NULL DEFAULT '',
                  分類層級D TEXT NOT NULL DEFAULT '',
                  最終廣告形式 TEXT NOT NULL DEFAULT '',
                  規則命中_廣告形式 TEXT NOT NULL DEFAULT '',
                  最終來源_廣告形式 TEXT NOT NULL DEFAULT '',
                  updated_at TEXT NOT NULL DEFAULT ''
                )
                """
            )
            conn.execute(
                """
                INSERT INTO dsp_rawdata(
                  row_order, 日期時間, 經銷商, 訂單, 素材, 廣告形式, 尺寸, 素材樣板, 執行金額, 系統營收, 媒體費用,
                  原始經銷商, 原始廣告形式, 最終經銷商, 規則命中_經銷商, 最終來源_經銷商,
                  分類層級B, 分類層級C, 分類層級D, 最終廣告形式, 規則命中_廣告形式, 最終來源_廣告形式, updated_at
                ) VALUES
                (0, '2026-05-01 00:00:00', 'A', 'O1', 'C1', 'Banner', '300x250', 'tplA', 10.0, 12.5, 8.0, 'A', 'Banner', 'PROMOTED_A1', 'r1', 'rule', 'B1', 'C1', 'D1', 'Banner', 'r2', 'rule', '2026-05-09T00:00:00'),
                (1, '2026-05-01 01:00:00', 'B', 'O2', 'C2', 'Banner', '300x250', 'tplA', 20.0, 25.0, 16.0, 'B', 'Banner', 'PROMOTED_B1', 'r1', 'rule', 'B2', 'C2', 'D2', 'Banner', 'r2', 'rule', '2026-05-09T00:00:00')
                """
            )
            conn.commit()
        finally:
            conn.close()
        return seed_db

    def _make_seed_canonical_ssp_db(self, root: Path) -> Path:
        seed_db = root / "data_seed" / "canonical" / "volume.sqlite"
        seed_db.parent.mkdir(parents=True, exist_ok=True)
        import sqlite3
        conn = sqlite3.connect(str(seed_db))
        try:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS raw (
                  source TEXT,
                  ts TEXT,
                  date TEXT,
                  hour INTEGER,
                  placement_id INTEGER,
                  placement_name TEXT,
                  request INTEGER,
                  impression INTEGER,
                  clicks INTEGER,
                  revenue REAL,
                  dsp_amount REAL,
                  order_id TEXT,
                  order_name TEXT,
                  supplier_id INTEGER,
                  supplier_name TEXT,
                  site_id INTEGER,
                  site_name TEXT
                )
                """
            )
            conn.execute(
                """
                INSERT INTO raw(
                  source, ts, date, hour, placement_id, placement_name, request, impression, clicks,
                  revenue, dsp_amount, order_id, order_name, supplier_id, supplier_name, site_id, site_name
                ) VALUES
                ('clickforce_api', '2026-05-04 00:00:00', '2026-05-04', 0, 1001, '母檔_純蓋板_300x250', 150, 1200, 12, 500.0, 420.0, 'O-100', '訂單A', 7, '供應商A', 501, 'site-A'),
                ('clickforce_api', '2026-05-04 01:00:00', '2026-05-04', 1, 1002, '母檔_置底展開_320x480', 200, 2100, 15, 700.0, 610.0, 'O-101', '訂單B', 8, '供應商B', 502, 'site-B')
                """
            )
            conn.commit()
        finally:
            conn.close()
        return seed_db

    def _make_seed_canonical_ssp_mdreport_db(self, root: Path) -> Path:
        seed_db = root / "data_seed" / "canonical" / "mdreport.sqlite"
        seed_db.parent.mkdir(parents=True, exist_ok=True)
        import sqlite3
        conn = sqlite3.connect(str(seed_db))
        try:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS raw (
                  source TEXT,
                  ts TEXT,
                  date TEXT,
                  hour INTEGER,
                  placement_id INTEGER,
                  placement_name TEXT,
                  request INTEGER,
                  impression INTEGER,
                  supplier_id INTEGER,
                  supplier_name TEXT,
                  site_id INTEGER,
                  site_name TEXT
                )
                """
            )
            conn.execute(
                """
                INSERT INTO raw(
                  source, ts, date, hour, placement_id, placement_name, request, impression,
                  supplier_id, supplier_name, site_id, site_name
                ) VALUES
                ('times_api', '2026-05-05 10:00:00', '2026-05-05', 10, 2001, '版位A_300x250', 320, 4500, 11, '時報供應商A', 701, 'times-site-A'),
                ('times_api', '2026-05-05 11:00:00', '2026-05-05', 11, 2002, '版位B_320x480', 280, 3900, 12, '時報供應商B', 702, 'times-site-B')
                """
            )
            conn.commit()
        finally:
            conn.close()
        return seed_db

    def _set_feature_flags(self, root: Path, **flags: bool) -> None:
        manifest_path = root / "bootstrap.manifest.json"
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        feature_flags = manifest.get("feature_flags", {})
        if not isinstance(feature_flags, dict):
            feature_flags = {}
        feature_flags.update(flags)
        manifest["feature_flags"] = feature_flags
        manifest_path.write_text(json.dumps(manifest, ensure_ascii=False), encoding="utf-8")

    def _run_bootstrap_wrapper(self, root: Path, *args: str) -> subprocess.CompletedProcess[str]:
        cmd = [
            "python3",
            str(Path(__file__).resolve().parents[1] / "app" / "bootstrap_init.py"),
            "--root",
            str(root),
            *args,
        ]
        return subprocess.run(cmd, capture_output=True, text=True, check=False)

    def test_help(self) -> None:
        proc = subprocess.run(
            ["python3", str(Path(__file__).resolve().parents[1] / "app" / "main.py"), "--help"],
            capture_output=True,
            text=True,
            check=False,
        )
        self.assertEqual(proc.returncode, 0)
        self.assertIn("bootstrap", proc.stdout)
        self.assertIn("health", proc.stdout)

    def test_bootstrap_then_health(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)

            p1 = self._run(root, "bootstrap")
            self.assertEqual(p1.returncode, 0)
            j1 = json.loads(p1.stdout)
            self.assertEqual(j1["status"], "ok")

            p2 = self._run(root, "health")
            self.assertEqual(p2.returncode, 0)
            j2 = json.loads(p2.stdout)
            self.assertEqual(j2["status"], "ok")

    def test_cli_usage_errors_are_json(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)

            p1 = self._run(root)
            self.assertNotEqual(p1.returncode, 0)
            j1 = json.loads(p1.stdout)
            self.assertEqual(j1["status"], "error")
            self.assertEqual(j1["error_code"], "CLI_USAGE_ERROR")

            p2 = self._run(root, "health", "--unknown")
            self.assertNotEqual(p2.returncode, 0)
            j2 = json.loads(p2.stdout)
            self.assertEqual(j2["error_code"], "CLI_USAGE_ERROR")

    def test_health_fail_fast_without_db(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)

            p = self._run(root, "health")
            self.assertNotEqual(p.returncode, 0)
            j = json.loads(p.stdout)
            self.assertEqual(j["status"], "error")
            self.assertEqual(j["error_code"], "DB_NOT_FOUND")

    def test_health_fail_fast_missing_migration(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            self._run(root, "bootstrap")
            (root / "migrations" / "0001_initial.sql").unlink()

            p = self._run(root, "health")
            self.assertNotEqual(p.returncode, 0)
            j = json.loads(p.stdout)
            self.assertEqual(j["error_code"], "MIGRATION_NOT_FOUND")

    def test_health_fail_fast_missing_seed(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            self._run(root, "bootstrap")
            (root / "templates" / "ruleset.seed.json").unlink()

            p = self._run(root, "health")
            self.assertNotEqual(p.returncode, 0)
            j = json.loads(p.stdout)
            self.assertEqual(j["error_code"], "RULE_SEED_NOT_FOUND")

    def test_health_fail_fast_binding_incomplete(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            self._run(root, "bootstrap")
            db = root / "data" / "mdrep.sqlite"
            import sqlite3
            conn = sqlite3.connect(str(db))
            try:
                conn.execute("DELETE FROM rule_bindings WHERE workflow='ssp'")
                conn.commit()
            finally:
                conn.close()

            p = self._run(root, "health")
            self.assertNotEqual(p.returncode, 0)
            j = json.loads(p.stdout)
            self.assertEqual(j["error_code"], "RULE_BINDING_INCOMPLETE")

    def test_bootstrap_fails_without_manifest(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            p = self._run(root, "bootstrap")
            self.assertNotEqual(p.returncode, 0)
            j = json.loads(p.stdout)
            self.assertEqual(j["status"], "error")
            self.assertEqual(j["error_code"], "FILE_NOT_FOUND")

    def test_service_invalid_json_and_missing_db(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            bad_rows = root / "rows.json"
            bad_rows.write_text("{bad", encoding="utf-8")

            p1 = self._run(root, "save", "--workflow", "dsp", "--template-version", "v1", "--rule-version", "v1", "--rows-json", str(bad_rows))
            self.assertNotEqual(p1.returncode, 0)
            j1 = json.loads(p1.stdout)
            self.assertEqual(j1["error_code"], "INVALID_ROWS_JSON")

            good_rows = root / "rows-ok.json"
            good_rows.write_text(json.dumps([{
                "日期時間": "2026-05-01 00:00:00",
                "經銷商": "A",
                "訂單": "O1",
                "素材": "C1",
                "廣告形式": "Banner",
                "尺寸": "300x250",
                "素材樣板": "tplA",
                "執行金額": 10.0,
                "系統營收": 12.5,
                "媒體費用": 8.0,
                "原始經銷商": "A",
                "原始廣告形式": "Banner",
                "最終經銷商": "A1",
                "規則命中_經銷商": "r1",
                "最終來源_經銷商": "rule",
                "分類層級B": "B1",
                "分類層級C": "C1",
                "分類層級D": "D1",
                "最終廣告形式": "Banner",
                "規則命中_廣告形式": "r2",
                "最終來源_廣告形式": "rule"
            }], ensure_ascii=False), encoding="utf-8")
            p2 = self._run(root, "save", "--workflow", "dsp", "--template-version", "v1", "--rule-version", "v1", "--rows-json", str(good_rows))
            self.assertNotEqual(p2.returncode, 0)
            j2 = json.loads(p2.stdout)
            self.assertEqual(j2["error_code"], "STRICT_ACCEPTANCE_GATE_FAILED")

    def test_health_malformed_manifest_reason_code(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            (root / "bootstrap.manifest.json").write_text("{bad", encoding="utf-8")

            p = self._run(root, "health")
            self.assertNotEqual(p.returncode, 0)
            j = json.loads(p.stdout)
            self.assertEqual(j["error_code"], "MANIFEST_JSON_INVALID")

    def test_bootstrap_wrapper_does_not_misread_option_values_as_command(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            # `bootstrap` 作為 --manifest 的值，不應被當成 command，wrapper 仍需補上 bootstrap 命令。
            p = self._run_bootstrap_wrapper(root, "--manifest", "bootstrap")
            self.assertNotEqual(p.returncode, 0)
            j = json.loads(p.stdout)
            self.assertEqual(j["status"], "error")
            self.assertIn(j["error_code"], {"FILE_NOT_FOUND", "MANIFEST_NOT_FOUND"})

    def test_bootstrap_wrapper_keeps_explicit_command(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            p = self._run_bootstrap_wrapper(root, "health")
            self.assertNotEqual(p.returncode, 0)
            j = json.loads(p.stdout)
            self.assertEqual(j["error_code"], "DB_NOT_FOUND")

    def test_export_cli_outputs_workbook_artifact(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            pb = self._run(root, "bootstrap")
            self.assertEqual(pb.returncode, 0)
            row = {
                "日期時間": "2026-05-01 00:00:00",
                "經銷商": "A",
                "訂單": "O1",
                "素材": "C1",
                "廣告形式": "Banner",
                "尺寸": "300x250",
                "素材樣板": "tplA",
                "執行金額": 10.0,
                "系統營收": 12.5,
                "媒體費用": 8.0,
                "原始經銷商": "A",
                "原始廣告形式": "Banner",
                "最終經銷商": "A1",
                "規則命中_經銷商": "r1",
                "最終來源_經銷商": "rule",
                "分類層級B": "B1",
                "分類層級C": "C1",
                "分類層級D": "D1",
                "最終廣告形式": "Banner",
                "規則命中_廣告形式": "r2",
                "最終來源_廣告形式": "rule",
            }
            rows_path = root / "rows.json"
            rows_path.write_text(json.dumps([row], ensure_ascii=False), encoding="utf-8")
            ps = self._run(
                root,
                "save",
                "--workflow",
                "dsp",
                "--template-version",
                "v1",
                "--rule-version",
                "v1",
                "--rows-json",
                str(rows_path),
            )
            self.assertEqual(ps.returncode, 0)

            pe = self._run(
                root,
                "export",
                "--workflow",
                "dsp",
                "--template-version",
                "v1",
                "--rule-version",
                "v1",
            )
            self.assertEqual(pe.returncode, 0)
            je = json.loads(pe.stdout)
            artifact_path = Path(je["result"]["artifact_path"])
            self.assertTrue(artifact_path.exists())
            self.assertEqual(artifact_path.suffix, ".xlsx")
            today = date.today()
            this_week_start = today - timedelta(days=today.weekday())
            prev_week_start = this_week_start - timedelta(days=7)
            prev_week_end = this_week_start - timedelta(days=1)
            self.assertEqual(
                artifact_path.name,
                f"{prev_week_end.year} DSP投資量報表_{prev_week_start.strftime('%m%d')}-{prev_week_end.strftime('%m%d')}.xlsx",
            )
            wb = load_workbook(artifact_path, data_only=True)
            try:
                self.assertEqual(
                    wb.sheetnames,
                    [
                        "2025年_MF_合作績效統計總表",
                        "2025_外部+行政_合作績效統計總表 ",
                        "mF投資量_總表",
                        "各經銷商明細",
                        "北流進單追蹤",
                    ],
                )
            finally:
                wb.close()

    def test_export_cli_accepts_explicit_week_period(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            pb = self._run(root, "bootstrap")
            self.assertEqual(pb.returncode, 0)
            row = {
                "日期時間": "2026-05-01 00:00:00",
                "經銷商": "A",
                "訂單": "O1",
                "素材": "C1",
                "廣告形式": "Banner",
                "尺寸": "300x250",
                "素材樣板": "tplA",
                "執行金額": 10.0,
                "系統營收": 12.5,
                "媒體費用": 8.0,
                "原始經銷商": "A",
                "原始廣告形式": "Banner",
                "最終經銷商": "A1",
                "規則命中_經銷商": "r1",
                "最終來源_經銷商": "rule",
                "分類層級B": "B1",
                "分類層級C": "C1",
                "分類層級D": "D1",
                "最終廣告形式": "Banner",
                "規則命中_廣告形式": "r2",
                "最終來源_廣告形式": "rule",
            }
            rows_path = root / "rows.json"
            rows_path.write_text(json.dumps([row], ensure_ascii=False), encoding="utf-8")
            ps = self._run(
                root,
                "save",
                "--workflow",
                "dsp",
                "--template-version",
                "v1",
                "--rule-version",
                "v1",
                "--rows-json",
                str(rows_path),
            )
            self.assertEqual(ps.returncode, 0)

            pe = self._run(
                root,
                "export",
                "--workflow",
                "dsp",
                "--template-version",
                "v1",
                "--rule-version",
                "v1",
                "--week-start",
                "2026-04-27",
                "--week-end",
                "2026-05-03",
            )
            self.assertEqual(pe.returncode, 0)
            je = json.loads(pe.stdout)
            self.assertEqual(je["result"]["week_start"], "2026-04-27")
            self.assertEqual(je["result"]["week_end"], "2026-05-03")
            self.assertEqual(
                Path(str(je["result"]["artifact_path"])).name,
                "2026 DSP投資量報表_0427-0503.xlsx",
            )

            import sqlite3
            db = root / "data" / "mdrep.sqlite"
            conn = sqlite3.connect(str(db))
            try:
                publish = conn.execute(
                    "SELECT week_start, week_end FROM publish_runs ORDER BY created_at DESC LIMIT 1"
                ).fetchone()
            finally:
                conn.close()
            self.assertIsNotNone(publish)
            assert publish is not None
            self.assertEqual(str(publish[0]), "2026-04-27")
            self.assertEqual(str(publish[1]), "2026-05-03")

    def test_export_cli_rejects_invalid_week_period(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            pb = self._run(root, "bootstrap")
            self.assertEqual(pb.returncode, 0)
            row = {
                "日期時間": "2026-05-01 00:00:00",
                "經銷商": "A",
                "訂單": "O1",
                "素材": "C1",
                "廣告形式": "Banner",
                "尺寸": "300x250",
                "素材樣板": "tplA",
                "執行金額": 10.0,
                "系統營收": 12.5,
                "媒體費用": 8.0,
                "原始經銷商": "A",
                "原始廣告形式": "Banner",
                "最終經銷商": "A1",
                "規則命中_經銷商": "r1",
                "最終來源_經銷商": "rule",
                "分類層級B": "B1",
                "分類層級C": "C1",
                "分類層級D": "D1",
                "最終廣告形式": "Banner",
                "規則命中_廣告形式": "r2",
                "最終來源_廣告形式": "rule",
            }
            rows_path = root / "rows.json"
            rows_path.write_text(json.dumps([row], ensure_ascii=False), encoding="utf-8")
            ps = self._run(
                root,
                "save",
                "--workflow",
                "dsp",
                "--template-version",
                "v1",
                "--rule-version",
                "v1",
                "--rows-json",
                str(rows_path),
            )
            self.assertEqual(ps.returncode, 0)

            pe = self._run(
                root,
                "export",
                "--workflow",
                "dsp",
                "--template-version",
                "v1",
                "--rule-version",
                "v1",
                "--week-start",
                "2026-04-27",
            )
            self.assertNotEqual(pe.returncode, 0)
            je = json.loads(pe.stdout)
            self.assertEqual(je["error_code"], "VALIDATION_ERROR")

    def test_ssp_cli_save_modify_export_parity(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            pb = self._run(root, "bootstrap")
            self.assertEqual(pb.returncode, 0)

            row = {
                "日期時間": "2026-05-01 00:00:00",
                "經銷商": "SSP_A",
                "訂單": "O1",
                "素材": "C1",
                "廣告形式": "Banner",
                "尺寸": "300x250",
                "素材樣板": "tplA",
                "執行金額": 10.0,
                "系統營收": 12.5,
                "媒體費用": 8.0,
                "原始經銷商": "SSP_A",
                "原始廣告形式": "Banner",
                "最終經銷商": "SSP_A1",
                "規則命中_經銷商": "r1",
                "最終來源_經銷商": "rule",
                "分類層級B": "B1",
                "分類層級C": "C1",
                "分類層級D": "D1",
                "最終廣告形式": "Banner",
                "規則命中_廣告形式": "r2",
                "最終來源_廣告形式": "rule",
            }
            rows_path = root / "ssp-rows.json"
            rows_path.write_text(json.dumps([row], ensure_ascii=False), encoding="utf-8")
            ps = self._run(
                root,
                "save",
                "--workflow",
                "ssp",
                "--template-version",
                "v1",
                "--rule-version",
                "v1",
                "--rows-json",
                str(rows_path),
            )
            self.assertEqual(ps.returncode, 0)

            updates_path = root / "ssp-updates.json"
            updates_path.write_text(
                json.dumps([{"row_order": 0, "column": "最終經銷商", "value": "SSP_A2"}], ensure_ascii=False),
                encoding="utf-8",
            )
            pm = self._run(
                root,
                "modify",
                "--workflow",
                "ssp",
                "--template-version",
                "v1",
                "--rule-version",
                "v1",
                "--updates-json",
                str(updates_path),
            )
            self.assertEqual(pm.returncode, 0)

            pe = self._run(
                root,
                "export",
                "--workflow",
                "ssp",
                "--template-version",
                "v1",
                "--rule-version",
                "v1",
            )
            self.assertEqual(pe.returncode, 0)
            je = json.loads(pe.stdout)
            artifact_path = Path(je["result"]["artifact_path"])
            self.assertTrue(artifact_path.exists())
            self.assertEqual(artifact_path.suffix, ".xlsx")
            wb = load_workbook(artifact_path, data_only=True)
            try:
                ws_data = wb["canonical_data"]
                headers = [cell.value for cell in ws_data[1]]
                first_row = dict(zip(headers, [cell.value for cell in ws_data[2]]))
                self.assertEqual(first_row["最終經銷商"], "SSP_A2")
                ws_meta = wb["metadata"]
                meta = {str(r[0]): str(r[1]) for r in ws_meta.iter_rows(min_row=2, values_only=True)}
                self.assertEqual(meta["workflow"], "ssp")
            finally:
                wb.close()

    def test_strict_acceptance_gate_blocks_service_when_binding_broken(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            pb = self._run(root, "bootstrap")
            self.assertEqual(pb.returncode, 0)
            db = root / "data" / "mdrep.sqlite"
            import sqlite3
            conn = sqlite3.connect(str(db))
            try:
                conn.execute("DELETE FROM rule_bindings WHERE workflow='ssp'")
                conn.commit()
            finally:
                conn.close()

            rows_path = root / "rows.json"
            rows_path.write_text(json.dumps([], ensure_ascii=False), encoding="utf-8")
            ps = self._run(
                root,
                "save",
                "--workflow",
                "dsp",
                "--template-version",
                "v1",
                "--rule-version",
                "v1",
                "--rows-json",
                str(rows_path),
            )
            self.assertNotEqual(ps.returncode, 0)
            js = json.loads(ps.stdout)
            self.assertEqual(js["error_code"], "STRICT_ACCEPTANCE_GATE_FAILED")
            self.assertIn("details", js)

    def test_non_strict_acceptance_gate_allows_service_when_binding_broken(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            self._set_feature_flags(root, strict_acceptance_gate=False)
            pb = self._run(root, "bootstrap")
            self.assertEqual(pb.returncode, 0)
            db = root / "data" / "mdrep.sqlite"
            import sqlite3
            conn = sqlite3.connect(str(db))
            try:
                conn.execute("DELETE FROM rule_bindings WHERE workflow='ssp'")
                conn.commit()
            finally:
                conn.close()

            row = {
                "日期時間": "2026-05-01 00:00:00",
                "經銷商": "A",
                "訂單": "O1",
                "素材": "C1",
                "廣告形式": "Banner",
                "尺寸": "300x250",
                "素材樣板": "tplA",
                "執行金額": 10.0,
                "系統營收": 12.5,
                "媒體費用": 8.0,
                "原始經銷商": "A",
                "原始廣告形式": "Banner",
                "最終經銷商": "A1",
                "規則命中_經銷商": "r1",
                "最終來源_經銷商": "rule",
                "分類層級B": "B1",
                "分類層級C": "C1",
                "分類層級D": "D1",
                "最終廣告形式": "Banner",
                "規則命中_廣告形式": "r2",
                "最終來源_廣告形式": "rule",
            }
            rows_path = root / "rows-ok.json"
            rows_path.write_text(json.dumps([row], ensure_ascii=False), encoding="utf-8")
            ps = self._run(
                root,
                "save",
                "--workflow",
                "dsp",
                "--template-version",
                "v1",
                "--rule-version",
                "v1",
                "--rows-json",
                str(rows_path),
            )
            self.assertEqual(ps.returncode, 0)

    def test_feature_flags_enable_trace_and_test_hooks(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            self._set_feature_flags(
                root,
                strict_acceptance_gate=False,
                enable_test_hooks=True,
                enable_trace_markers=True,
            )
            pb = self._run(root, "bootstrap")
            self.assertEqual(pb.returncode, 0)
            jb = json.loads(pb.stdout)
            self.assertEqual(jb["result"]["feature_flags"]["enable_test_hooks"], True)
            self.assertEqual(jb["result"]["feature_flags"]["enable_trace_markers"], True)

            row = {
                "日期時間": "2026-05-01 00:00:00",
                "經銷商": "A",
                "訂單": "O1",
                "素材": "C1",
                "廣告形式": "Banner",
                "尺寸": "300x250",
                "素材樣板": "tplA",
                "執行金額": 10.0,
                "系統營收": 12.5,
                "媒體費用": 8.0,
                "原始經銷商": "A",
                "原始廣告形式": "Banner",
                "最終經銷商": "A1",
                "規則命中_經銷商": "r1",
                "最終來源_經銷商": "rule",
                "分類層級B": "B1",
                "分類層級C": "C1",
                "分類層級D": "D1",
                "最終廣告形式": "Banner",
                "規則命中_廣告形式": "r2",
                "最終來源_廣告形式": "rule",
            }
            rows_path = root / "rows.json"
            rows_path.write_text(json.dumps([row], ensure_ascii=False), encoding="utf-8")
            ps = self._run(
                root,
                "save",
                "--workflow",
                "dsp",
                "--template-version",
                "v1",
                "--rule-version",
                "v1",
                "--rows-json",
                str(rows_path),
            )
            self.assertEqual(ps.returncode, 0)
            js = json.loads(ps.stdout)
            self.assertEqual(js["result"]["test_hooks_enabled"], True)
            self.assertTrue(str(js["result"]["trace_marker"]))

    def test_health_reports_both_workflow_bindings(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            pb = self._run(root, "bootstrap")
            self.assertEqual(pb.returncode, 0)
            ph = self._run(root, "health")
            self.assertEqual(ph.returncode, 0)
            jh = json.loads(ph.stdout)
            self.assertEqual(jh["status"], "ok")

            import sqlite3
            db = root / "data" / "mdrep.sqlite"
            conn = sqlite3.connect(str(db))
            try:
                rows = conn.execute(
                    "SELECT workflow, template_id FROM rule_bindings ORDER BY workflow, template_id"
                ).fetchall()
            finally:
                conn.close()
            workflows = {str(r[0]) for r in rows}
            self.assertIn("dsp", workflows)
            self.assertIn("ssp", workflows)

    def test_health_fails_when_only_ssp_binding_missing(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            pb = self._run(root, "bootstrap")
            self.assertEqual(pb.returncode, 0)

            import sqlite3
            db = root / "data" / "mdrep.sqlite"
            conn = sqlite3.connect(str(db))
            try:
                conn.execute("DELETE FROM rule_bindings WHERE workflow='ssp'")
                conn.commit()
            finally:
                conn.close()

            ph = self._run(root, "health")
            self.assertNotEqual(ph.returncode, 0)
            jh = json.loads(ph.stdout)
            self.assertEqual(jh["status"], "error")
            self.assertEqual(jh["error_code"], "RULE_BINDING_INCOMPLETE")
            self.assertIn("missing_bindings", jh["details"]["health"]["checks"])
            self.assertTrue(
                any(
                    item["workflow"] == "ssp" and item["template_id"] == "ssp-default"
                    for item in jh["details"]["health"]["checks"]["missing_bindings"]
                )
            )

    def test_seed_bootstrap_builds_old_data_seed_scaffold(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            pb = self._run(root, "bootstrap")
            self.assertEqual(pb.returncode, 0)

            row = {
                "日期時間": "2026-05-01 00:00:00",
                "經銷商": "A",
                "訂單": "O1",
                "素材": "C1",
                "廣告形式": "Banner",
                "尺寸": "300x250",
                "素材樣板": "tplA",
                "執行金額": 10.0,
                "系統營收": 12.5,
                "媒體費用": 8.0,
                "原始經銷商": "A",
                "原始廣告形式": "Banner",
                "最終經銷商": "A1",
                "規則命中_經銷商": "r1",
                "最終來源_經銷商": "rule",
                "分類層級B": "B1",
                "分類層級C": "C1",
                "分類層級D": "D1",
                "最終廣告形式": "Banner",
                "規則命中_廣告形式": "r2",
                "最終來源_廣告形式": "rule",
            }
            rows_path = root / "rows.json"
            rows_path.write_text(json.dumps([row], ensure_ascii=False), encoding="utf-8")
            ps = self._run(
                root,
                "save",
                "--workflow",
                "dsp",
                "--template-version",
                "v1",
                "--rule-version",
                "v1",
                "--rows-json",
                str(rows_path),
            )
            self.assertEqual(ps.returncode, 0)
            save_run_id = str(json.loads(ps.stdout)["result"]["run_id"])
            self.assertTrue(save_run_id.startswith("run-"))

            raw_inbox = root / "raw-inbox"
            raw_inbox.mkdir(parents=True, exist_ok=True)
            raw_file = raw_inbox / "dsp_raw_20260501.json"
            raw_file.write_text(json.dumps([{"x": 1}], ensure_ascii=False), encoding="utf-8")
            noise_file = raw_inbox / "debug_probe_20260501.json"
            noise_file.write_text("{}", encoding="utf-8")

            pseed = self._run(root, "seed-bootstrap", "--raw-source", "raw-inbox")
            self.assertEqual(pseed.returncode, 0)
            jseed = json.loads(pseed.stdout)
            self.assertEqual(jseed["status"], "ok")
            self.assertEqual(int(jseed["result"]["raw_entry_count"]), 1)
            manifest_path = Path(str(jseed["result"]["manifest_path"]))
            self.assertTrue(manifest_path.exists())

            seed_manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            self.assertEqual(seed_manifest["seed_manifest_version"], "v1")
            self.assertEqual(seed_manifest["layers"]["raw_seed"]["entry_count"], 1)
            self.assertEqual(len(seed_manifest["entries"]), 1)
            entry = seed_manifest["entries"][0]
            self.assertEqual(entry["workflow"], "dsp")
            self.assertEqual(entry["source_date"], "2026-05-01")
            self.assertTrue(str(entry["checksum_sha256"]))
            self.assertEqual(str(entry["import_run_id"]), save_run_id)

            seed_root = Path(str(jseed["result"]["seed_root"]))
            self.assertTrue((seed_root / "canonical" / "mdrep.sqlite").exists())
            self.assertTrue((seed_root / "logs" / "run_log.json").exists())
            self.assertTrue((seed_root / "templates_rules_mapping" / "fields_contract.json").exists())

    def test_seed_rebuild_restores_canonical_from_raw_seed(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            pb = self._run(root, "bootstrap")
            self.assertEqual(pb.returncode, 0)

            seed_row = {
                "日期時間": "2026-05-01 00:00:00",
                "經銷商": "A",
                "訂單": "O1",
                "素材": "C1",
                "廣告形式": "Banner",
                "尺寸": "300x250",
                "素材樣板": "tplA",
                "執行金額": 10.0,
                "系統營收": 12.5,
                "媒體費用": 8.0,
                "原始經銷商": "A",
                "原始廣告形式": "Banner",
                "最終經銷商": "SEED_A1",
                "規則命中_經銷商": "r1",
                "最終來源_經銷商": "rule",
                "分類層級B": "B1",
                "分類層級C": "C1",
                "分類層級D": "D1",
                "最終廣告形式": "Banner",
                "規則命中_廣告形式": "r2",
                "最終來源_廣告形式": "rule",
            }
            raw_inbox = root / "raw-inbox"
            raw_inbox.mkdir(parents=True, exist_ok=True)
            (raw_inbox / "dsp_raw_20260501.json").write_text(
                json.dumps([seed_row], ensure_ascii=False),
                encoding="utf-8",
            )
            pseed = self._run(root, "seed-bootstrap", "--raw-source", "raw-inbox")
            self.assertEqual(pseed.returncode, 0)

            current_row = dict(seed_row)
            current_row["最終經銷商"] = "LIVE_A9"
            rows_path = root / "rows-now.json"
            rows_path.write_text(json.dumps([current_row], ensure_ascii=False), encoding="utf-8")
            ps = self._run(
                root,
                "save",
                "--workflow",
                "dsp",
                "--template-version",
                "v1",
                "--rule-version",
                "v1",
                "--rows-json",
                str(rows_path),
            )
            self.assertEqual(ps.returncode, 0)

            pr = self._run(
                root,
                "seed-rebuild",
                "--workflow",
                "dsp",
                "--template-version",
                "v1",
                "--rule-version",
                "v1",
            )
            self.assertEqual(pr.returncode, 0)
            jr = json.loads(pr.stdout)
            self.assertEqual(jr["status"], "ok")
            self.assertEqual(int(jr["result"]["files_used"]), 1)
            self.assertEqual(int(jr["result"]["workflows"]["dsp"]["row_count"]), 1)

            pe = self._run(
                root,
                "export",
                "--workflow",
                "dsp",
                "--template-version",
                "v1",
                "--rule-version",
                "v1",
            )
            self.assertEqual(pe.returncode, 0)
            je = json.loads(pe.stdout)
            artifact_path = Path(je["result"]["artifact_path"])
            self.assertTrue(artifact_path.exists())
            wb = load_workbook(artifact_path, data_only=True)
            try:
                self.assertEqual(
                    wb.sheetnames,
                    [
                        "2025年_MF_合作績效統計總表",
                        "2025_外部+行政_合作績效統計總表 ",
                        "mF投資量_總表",
                        "各經銷商明細",
                        "北流進單追蹤",
                    ],
                )
            finally:
                wb.close()

    def test_seed_rebuild_restores_ssp_truth_table_from_raw_seed(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            pb = self._run(root, "bootstrap")
            self.assertEqual(pb.returncode, 0)

            seed_rows = [
                {
                    "source": "ssp_seed",
                    "ts": "2026-05-06 08:00:00",
                    "date": "2026-05-06",
                    "hour": 8,
                    "placement_id": 9101,
                    "placement_name": "SEED_SSP_SLOT_300x250",
                    "request": 321.0,
                    "impression": 4321.0,
                    "clicks": 21.0,
                    "revenue": 654.5,
                    "dsp_amount": 543.2,
                    "order_id": "SSP-SEED-001",
                    "order_name": "Seed Order",
                    "supplier_id": 77,
                    "supplier_name": "Seed Supplier",
                    "site_id": 8801,
                    "site_name": "seed-site",
                }
            ]
            raw_inbox = root / "raw-inbox"
            raw_inbox.mkdir(parents=True, exist_ok=True)
            (raw_inbox / "ssp_raw_20260506.json").write_text(
                json.dumps(seed_rows, ensure_ascii=False),
                encoding="utf-8",
            )
            pseed = self._run(root, "seed-bootstrap", "--raw-source", "raw-inbox")
            self.assertEqual(pseed.returncode, 0)

            self._make_seed_canonical_ssp_db(root)
            pp = self._run(
                root,
                "seed-promote-live",
                "--workflow",
                "ssp",
                "--source-db-rel",
                "canonical/volume.sqlite",
                "--template-version",
                "v1",
                "--rule-version",
                "v1",
            )
            self.assertEqual(pp.returncode, 0)

            pr = self._run(
                root,
                "seed-rebuild",
                "--workflow",
                "ssp",
                "--template-version",
                "v1",
                "--rule-version",
                "v1",
            )
            self.assertEqual(pr.returncode, 0)
            jr = json.loads(pr.stdout)
            self.assertEqual(jr["status"], "ok")
            self.assertEqual(int(jr["result"]["files_used"]), 1)
            self.assertEqual(int(jr["result"]["workflows"]["ssp"]["row_count"]), 1)

            import sqlite3

            db = root / "data" / "mdrep.sqlite"
            conn = sqlite3.connect(str(db))
            try:
                ssp_count = conn.execute("SELECT COUNT(1) FROM ssp_raw").fetchone()
                canonical_count = conn.execute("SELECT COUNT(1) FROM canonical_raw WHERE workflow='ssp'").fetchone()
                row = conn.execute(
                    """
                    SELECT source, placement_name, supplier_name, request, impression
                    FROM ssp_raw
                    ORDER BY row_order ASC
                    LIMIT 1
                    """
                ).fetchone()
            finally:
                conn.close()

            self.assertEqual(int(ssp_count[0] or 0), 1)
            self.assertEqual(int(canonical_count[0] or 0), 0)
            self.assertIsNotNone(row)
            assert row is not None
            self.assertEqual(str(row[0]), "ssp_seed")
            self.assertEqual(str(row[1]), "SEED_SSP_SLOT_300x250")
            self.assertEqual(str(row[2]), "Seed Supplier")
            self.assertEqual(float(row[3]), 321.0)
            self.assertEqual(float(row[4]), 4321.0)

    def test_seed_import_mdreport_layers_raw_and_canonical(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            legacy_root = self._make_legacy_mdreport_seed_source(root)

            pi = self._run(
                root,
                "seed-import-mdreport",
                "--mdreport-root",
                str(legacy_root),
            )
            self.assertEqual(pi.returncode, 0)
            ji = json.loads(pi.stdout)
            self.assertEqual(ji["status"], "ok")
            self.assertEqual(int(ji["result"]["canonical_seed_count"]), 4)
            self.assertEqual(int(ji["result"]["raw_entry_count"]), 3)

            seed_root = Path(str(ji["result"]["seed_root"]))
            self.assertTrue((seed_root / "canonical" / "mdreport.sqlite").exists())
            self.assertTrue((seed_root / "canonical" / "mdreport_dsp.sqlite").exists())
            self.assertTrue((seed_root / "templates_rules_mapping" / "group_overrides.json").exists())

            manifest_path = Path(str(ji["result"]["manifest_path"]))
            self.assertTrue(manifest_path.exists())
            payload = json.loads(manifest_path.read_text(encoding="utf-8"))
            self.assertEqual(Path(str(payload["source_project_root"])).resolve(), legacy_root.resolve())
            self.assertEqual(payload["layers"]["raw_seed"]["entry_count"], 3)
            names = {Path(item["raw_file_rel_path"]).name for item in payload["entries"]}
            self.assertIn("dsp_rawdata_20260506_150144_recalc.json", names)
            self.assertIn("dsp_weekly_accum_20260507_155315_weekly_meta.json", names)
            self.assertIn("dsp_dataset_manifest.json", names)
            self.assertNotIn("dsp_rawdata_20260506_150144_check.json", names)

    def test_seed_promote_live_converts_legacy_canonical_to_runtime_db(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            self._make_seed_canonical_dsp_db(root)

            pp = self._run(
                root,
                "seed-promote-live",
                "--workflow",
                "dsp",
                "--template-version",
                "v1",
                "--rule-version",
                "v1",
            )
            self.assertEqual(pp.returncode, 0)
            jp = json.loads(pp.stdout)
            self.assertEqual(jp["status"], "ok")
            self.assertEqual(int(jp["result"]["row_count"]), 2)
            self.assertTrue(str(jp["result"]["run_id"]).startswith("run-"))

            import sqlite3
            db = root / "data" / "mdrep.sqlite"
            conn = sqlite3.connect(str(db))
            try:
                count = conn.execute("SELECT COUNT(1) FROM canonical_raw WHERE workflow='dsp'").fetchone()
                top = conn.execute(
                    "SELECT 最終經銷商 FROM canonical_raw WHERE workflow='dsp' ORDER BY row_order ASC LIMIT 1"
                ).fetchone()
            finally:
                conn.close()
            self.assertEqual(int(count[0] or 0), 2)
            self.assertEqual(str(top[0]), "PROMOTED_A1")

    def test_seed_promote_live_supports_ssp_raw_table(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            self._make_seed_canonical_ssp_db(root)

            pp = self._run(
                root,
                "seed-promote-live",
                "--workflow",
                "ssp",
                "--source-db-rel",
                "canonical/volume.sqlite",
                "--template-version",
                "v1",
                "--rule-version",
                "v1",
            )
            self.assertEqual(pp.returncode, 0)
            jp = json.loads(pp.stdout)
            self.assertEqual(jp["status"], "ok")
            self.assertEqual(int(jp["result"]["row_count"]), 2)
            self.assertEqual(str(jp["result"]["workflow"]), "ssp")

            import sqlite3
            db = root / "data" / "mdrep.sqlite"
            conn = sqlite3.connect(str(db))
            try:
                count = conn.execute("SELECT COUNT(1) FROM ssp_raw").fetchone()
                canonical_count = conn.execute("SELECT COUNT(1) FROM canonical_raw WHERE workflow='ssp'").fetchone()
                rows = conn.execute(
                    """
                    SELECT supplier_name, placement_name, request, impression
                    FROM ssp_raw
                    ORDER BY row_order ASC
                    """
                ).fetchall()
            finally:
                conn.close()

            self.assertEqual(int(count[0] or 0), 2)
            self.assertEqual(int(canonical_count[0] or 0), 0)
            self.assertEqual(str(rows[0][0]), "供應商A")
            self.assertEqual(str(rows[0][1]), "母檔_純蓋板_300x250")
            self.assertEqual(float(rows[0][2]), 150.0)
            self.assertEqual(float(rows[0][3]), 1200.0)

            pe = self._run(
                root,
                "export",
                "--workflow",
                "ssp",
                "--template-version",
                "v1",
                "--rule-version",
                "v1",
            )
            self.assertEqual(pe.returncode, 0)
            je = json.loads(pe.stdout)
            self.assertEqual(int(je["result"]["row_count"]), 2)
            artifact_path = Path(str(je["result"]["artifact_path"]))
            self.assertTrue(artifact_path.exists())
            wb = load_workbook(artifact_path, data_only=True)
            try:
                ws_data = wb["canonical_data"]
                headers = [cell.value for cell in ws_data[1]]
                first_row = dict(zip(headers, [cell.value for cell in ws_data[2]]))
                self.assertEqual(first_row["supplier_name"], "供應商A")
                self.assertEqual(first_row["placement_name"], "母檔_純蓋板_300x250")
            finally:
                wb.close()

    def test_seed_promote_live_ssp_uses_mdreport_default_truth_db(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            source_seed_db = self._make_seed_canonical_ssp_mdreport_db(root)
            test_seed_db = root / "data_seed_test" / "canonical" / "mdreport.sqlite"
            test_seed_db.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source_seed_db, test_seed_db)

            pp = self._run(
                root,
                "seed-promote-live",
                "--workflow",
                "ssp",
                "--template-version",
                "v1",
                "--rule-version",
                "v1",
            )
            self.assertEqual(pp.returncode, 0)
            jp = json.loads(pp.stdout)
            self.assertEqual(jp["status"], "ok")
            self.assertTrue(str(jp["result"]["source_db"]).endswith("/data_seed/canonical/mdreport.sqlite"))
            self.assertEqual(int(jp["result"]["row_count"]), 2)

            import sqlite3
            db = root / "data" / "mdrep.sqlite"
            conn = sqlite3.connect(str(db))
            try:
                count = conn.execute("SELECT COUNT(1) FROM ssp_raw").fetchone()
                canonical_count = conn.execute("SELECT COUNT(1) FROM canonical_raw WHERE workflow='ssp'").fetchone()
                first = conn.execute(
                    """
                    SELECT supplier_name, placement_name, request, impression
                    FROM ssp_raw
                    ORDER BY row_order ASC
                    LIMIT 1
                    """
                ).fetchone()
            finally:
                conn.close()

            self.assertEqual(int(count[0] or 0), 2)
            self.assertEqual(int(canonical_count[0] or 0), 0)
            self.assertEqual(str(first[0]), "時報供應商A")
            self.assertEqual(str(first[1]), "版位A_300x250")
            self.assertEqual(float(first[2]), 320.0)
            self.assertEqual(float(first[3]), 4500.0)

    def test_cli_env_test_routes_db_seed_and_artifact_paths(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            self._make_project(root)
            source_seed_db = self._make_seed_canonical_ssp_mdreport_db(root)
            test_seed_db = root / "data_seed_test" / "canonical" / "mdreport.sqlite"
            test_seed_db.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source_seed_db, test_seed_db)

            pb = self._run(root, "--env", "test", "bootstrap")
            self.assertEqual(pb.returncode, 0)
            jb = json.loads(pb.stdout)
            self.assertEqual(jb["status"], "ok")
            self.assertEqual(jb["result"]["runtime_env"], "test")
            self.assertTrue(str(jb["result"]["db_path"]).endswith("data_test/mdrep.test.sqlite"))
            self.assertTrue(str(jb["result"]["artifact_root"]).endswith("artifacts_test"))
            self.assertTrue(str(jb["result"]["data_seed_root"]).endswith("data_seed_test"))

            row = {
                "日期時間": "2026-05-01 00:00:00",
                "經銷商": "A",
                "訂單": "O1",
                "素材": "C1",
                "廣告形式": "Banner",
                "尺寸": "300x250",
                "素材樣板": "tplA",
                "執行金額": 10.0,
                "系統營收": 12.5,
                "媒體費用": 8.0,
                "原始經銷商": "A",
                "原始廣告形式": "Banner",
                "最終經銷商": "A1",
                "規則命中_經銷商": "r1",
                "最終來源_經銷商": "rule",
                "分類層級B": "B1",
                "分類層級C": "C1",
                "分類層級D": "D1",
                "最終廣告形式": "Banner",
                "規則命中_廣告形式": "r2",
                "最終來源_廣告形式": "rule",
            }
            rows_path = root / "rows.json"
            rows_path.write_text(json.dumps([row], ensure_ascii=False), encoding="utf-8")

            ps = self._run(
                root,
                "--env",
                "test",
                "save",
                "--workflow",
                "dsp",
                "--template-version",
                "v1",
                "--rule-version",
                "v1",
                "--rows-json",
                str(rows_path),
            )
            self.assertEqual(ps.returncode, 0)

            pp = self._run(
                root,
                "--env",
                "test",
                "seed-promote-live",
                "--workflow",
                "ssp",
                "--template-version",
                "v1",
                "--rule-version",
                "v1",
            )
            self.assertEqual(pp.returncode, 0)

            pe = self._run(
                root,
                "--env",
                "test",
                "export",
                "--workflow",
                "ssp",
                "--template-version",
                "v1",
                "--rule-version",
                "v1",
            )
            self.assertEqual(pe.returncode, 0)
            je = json.loads(pe.stdout)
            artifact_path = Path(str(je["result"]["artifact_path"]))
            self.assertTrue(artifact_path.exists())
            self.assertIn("artifacts_test", str(artifact_path))


if __name__ == "__main__":
    unittest.main()
