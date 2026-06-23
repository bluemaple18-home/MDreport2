from __future__ import annotations

import tempfile
import unittest
import base64
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook

from domain.services import CanonicalService
from infra.sqlite.bootstrap import bootstrap_init
from infra.sqlite.repository import SQLiteRepository


class MonthlyP4Tests(unittest.TestCase):
    def _write_dsp_tab4_template(self, path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        try:
            ws = wb.active
            ws.title = "2025年_MF_合作績效統計總表"
            ws.sheet_state = "hidden"
            wb.create_sheet("2025_外部+行政_合作績效統計總表 ")
            wb.create_sheet("mF投資量_總表")
            ws_detail = wb.create_sheet("各經銷商明細")
            ws_detail["A5"] = 2026
            wb.create_sheet("北流進單追蹤")
            wb.save(path)
        finally:
            wb.close()

    def _setup_project(self, root: Path) -> SQLiteRepository:
        src = Path(__file__).resolve().parents[1]
        (root / "migrations").mkdir(parents=True, exist_ok=True)
        (root / "templates").mkdir(parents=True, exist_ok=True)
        (root / "contracts").mkdir(parents=True, exist_ok=True)
        (root / "migrations" / "0001_initial.sql").write_text((src / "migrations" / "0001_initial.sql").read_text(encoding="utf-8"), encoding="utf-8")
        (root / "templates" / "template_registry.seed.json").write_text((src / "templates" / "template_registry.seed.json").read_text(encoding="utf-8"), encoding="utf-8")
        (root / "templates" / "ruleset.seed.json").write_text((src / "templates" / "ruleset.seed.json").read_text(encoding="utf-8"), encoding="utf-8")
        (root / "contracts" / "fields_contract.json").write_text((src / "contracts" / "fields_contract.json").read_text(encoding="utf-8"), encoding="utf-8")
        (root / "bootstrap.manifest.json").write_text((src / "bootstrap.manifest.json").read_text(encoding="utf-8"), encoding="utf-8")
        self._write_dsp_tab4_template(root / "templates" / "dsp_tab4_template.xlsx")
        result = bootstrap_init(root)
        return SQLiteRepository(Path(result["db_path"]), project_root=root)

    def _row(self, **overrides: object) -> dict[str, object]:
        row: dict[str, object] = {
            "日期時間": "2026-04-01 00:00:00",
            "經銷商": "外部",
            "訂單": "O1",
            "素材": "C1",
            "廣告形式": "Banner",
            "尺寸": "300x250",
            "素材樣板": "tpl",
            "執行金額": 100.0,
            "系統營收": 100.0,
            "媒體費用": 100.0,
            "原始經銷商": "外部",
            "原始廣告形式": "Banner",
            "最終經銷商": "外部經銷商",
            "規則命中_經銷商": "external",
            "最終來源_經銷商": "rule",
            "分類層級B": "外部經銷商",
            "分類層級C": "經銷推廣",
            "分類層級D": "玩藝",
            "最終廣告形式": "一般廣告",
            "規則命中_廣告形式": "rule",
            "最終來源_廣告形式": "rule",
        }
        row.update(overrides)
        return row

    def test_monthly_p4_snapshot_seeds_targets_and_uses_manual_inputs(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            repo = self._setup_project(Path(td))
            svc = CanonicalService(repo)
            svc.save(
                workflow="dsp",
                rows=[
                    self._row(
                        經銷商="[台灣]域動行銷股份有限公司",
                        最終經銷商="[台灣]域動行銷股份有限公司",
                        分類層級B="內部經銷商",
                        分類層級C="營銷事業處",
                        分類層級D="營銷事業處",
                        執行金額=500.0,
                    ),
                    self._row(
                        經銷商="域動行銷-MD",
                        訂單="直播 IO 專案",
                        素材="直播素材",
                        最終經銷商="IO委刊",
                        分類層級B="外部經銷商",
                        分類層級C="IO委刊",
                        分類層級D="直播IO",
                        執行金額=6000.0,
                    ),
                    self._row(
                        經銷商="域動行銷-MD",
                        訂單="momo直播 專案",
                        素材="momo素材",
                        最終經銷商="[台灣]域動行銷股份有限公司",
                        規則命中_經銷商="momo_marketing",
                        分類層級B="內部經銷商",
                        分類層級C="營銷事業處",
                        分類層級D="momo",
                        執行金額=7000.0,
                    ),
                    self._row(
                        經銷商="玩藝國際股份有限公司",
                        訂單="直播 專案",
                        素材="直播素材",
                        最終經銷商="外部經銷商",
                        分類層級B="外部經銷商",
                        分類層級C="經銷推廣",
                        分類層級D="玩藝國際股份有限公司",
                        執行金額=8000.0,
                    ),
                    self._row(
                        經銷商="QA經銷商",
                        訂單="測試訂單",
                        最終經銷商="QA經銷商",
                        分類層級B="內部經銷商",
                        分類層級C="營銷事業處",
                        分類層級D="QA經銷商",
                        執行金額=9000.0,
                    ),
                ],
                template_version="v1",
                rule_version="v1",
            )

            before = svc.build_monthly_p4_snapshot(week_start="2026-04-01", week_end="2026-04-30")
            april = next(item for item in before["monthPayloads"] if item["month"] == "2026-04")
            self.assertEqual(before["anchorMonth"], "2026-04")
            self.assertEqual(before["availableMonths"], [f"2026-{month:02d}" for month in range(1, 13)])
            self.assertEqual(float(april["targets"]["external_total"]), 1350000.0)

            svc.save_monthly_p4_manual_inputs(
                month="2026-04",
                inputs={
                    "hb_revenue": 84728,
                    "remaining_traffic_revenue": 141009,
                    "data_monetization_adjustment": 72000,
                },
                template_version="v1",
                rule_version="v1",
            )
            after = svc.build_monthly_p4_snapshot(week_start="2026-04-01", week_end="2026-04-30")
            april_after = next(item for item in after["monthPayloads"] if item["month"] == "2026-04")
            self.assertEqual(float(april_after["actuals"]["external_total"]), 8000.0)
            self.assertEqual(float(april_after["actuals"]["mf_marketing"]), 13500.0)
            self.assertEqual(float(april_after["computed"]["mf_marketing"]), 7500.0)
            self.assertEqual(float(april_after["computed"]["external_io_live_auto"]), 6000.0)
            self.assertGreater(float(april_after["actuals"]["data_fee"]), 72000.0)

    def test_monthly_p4_snapshot_keeps_same_month_different_years_separate(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            repo = self._setup_project(Path(td))
            svc = CanonicalService(repo)
            svc.save(
                workflow="dsp",
                rows=[
                    self._row(日期時間="2025-04-01 00:00:00", 訂單="O-2025", 執行金額=900.0),
                    self._row(日期時間="2026-04-01 00:00:00", 訂單="O-2026", 執行金額=100.0),
                ],
                template_version="v1",
                rule_version="v1",
            )

            snapshot = svc.build_monthly_p4_snapshot(week_start="2026-04-01", week_end="2026-04-30")
            april = next(item for item in snapshot["monthPayloads"] if item["month"] == "2026-04")

            self.assertEqual(snapshot["anchorMonth"], "2026-04")
            self.assertEqual(float(april["computed"]["external_self_operated"]), 100.0)
            self.assertEqual(float(april["actuals"]["external_total"]), 100.0)

    def test_monthly_p4_snapshot_normalizes_slash_date_period(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            repo = self._setup_project(Path(td))
            svc = CanonicalService(repo)

            snapshot = svc.build_monthly_p4_snapshot(week_start="2026/04/01", week_end="2026/04/30")

            self.assertEqual(snapshot["anchorMonth"], "2026-04")
            self.assertEqual(snapshot["months"], ["2026-02", "2026-03", "2026-04"])

    def test_archive_dsp_month_replaces_raw_rows_with_monthly_summary_rows(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            repo = self._setup_project(Path(td))
            svc = CanonicalService(repo)
            svc.save(
                workflow="dsp",
                rows=[
                    self._row(日期時間="2026-04-01", 訂單="raw-1", 執行金額=100.0),
                    self._row(日期時間="2026-04-02", 訂單="raw-2", 執行金額=50.0),
                    self._row(日期時間="2026-05-01", 訂單="raw-3", 執行金額=25.0),
                ],
                template_version="v1",
                rule_version="v1",
            )
            before = svc.build_monthly_p4_snapshot(week_start="2026-04-01", week_end="2026-04-30")
            before_april = next(item for item in before["monthPayloads"] if item["month"] == "2026-04")

            result = svc.archive_dsp_month(month="2026-04")
            after = svc.build_monthly_p4_snapshot(week_start="2026-04-01", week_end="2026-04-30")
            after_april = next(item for item in after["monthPayloads"] if item["month"] == "2026-04")

            self.assertEqual(result["status"], "ok")
            self.assertEqual(result["source_row_count"], 2)
            self.assertEqual(result["archive_row_count"], 1)
            self.assertEqual(float(before_april["actuals"]["external_total"]), 150.0)
            self.assertEqual(float(after_april["actuals"]["external_total"]), 150.0)
            with repo.connect() as conn:
                april_rows = conn.execute(
                    "SELECT COUNT(*) FROM canonical_raw WHERE workflow = 'dsp' AND substr(\"日期時間\", 1, 7) = '2026-04'"
                ).fetchone()[0]
                may_rows = conn.execute(
                    "SELECT COUNT(*) FROM canonical_raw WHERE workflow = 'dsp' AND substr(\"日期時間\", 1, 7) = '2026-05'"
                ).fetchone()[0]
                archive = conn.execute(
                    "SELECT marker, source_row_count, archive_row_count FROM monthly_dsp_archives WHERE month = '2026-04'"
                ).fetchone()
            self.assertEqual(april_rows, 1)
            self.assertEqual(may_rows, 1)
            self.assertEqual(archive[0], "MONTHLY_ARCHIVE_DSP_2026-04")
            self.assertEqual(archive[1], 2)
            self.assertEqual(archive[2], 1)

    def test_monthly_p4_test_inputs_do_not_modify_formal_inputs(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            repo = self._setup_project(Path(td))
            test_db = Path(td) / "data" / "monthly_p4_test.sqlite"
            test_db.touch()
            test_repo = SQLiteRepository(test_db, project_root=Path(td))
            svc = CanonicalService(repo, monthly_test_repo=test_repo)

            svc.save_monthly_p4_manual_inputs(
                month="2026-03",
                inputs={"hb_revenue": 100},
                template_version="v1",
                rule_version="v1",
            )
            svc.save_monthly_p4_test_inputs(
                month="2026-03",
                inputs={"hb_revenue": 999},
                template_version="v1",
                rule_version="v1",
            )

            formal = svc.build_monthly_p4_snapshot(week_start="2026-03-01", week_end="2026-03-31")
            test = svc.build_monthly_p4_snapshot(
                week_start="2026-03-01",
                week_end="2026-03-31",
                manual_source="test",
            )
            formal_march = next(item for item in formal["monthPayloads"] if item["month"] == "2026-03")
            test_march = next(item for item in test["monthPayloads"] if item["month"] == "2026-03")

            self.assertEqual(float(formal_march["manualInputs"]["hb_revenue"]), 100.0)
            self.assertEqual(float(test_march["manualInputs"]["hb_revenue"]), 999.0)
            self.assertEqual(test["testDbPath"], str(test_db.resolve()))
            with repo.connect() as conn:
                formal_count = conn.execute("SELECT COUNT(*) FROM monthly_p4_manual_inputs").fetchone()[0]
                formal_has_test_table = conn.execute(
                    "SELECT COUNT(*) FROM sqlite_master WHERE type = 'table' AND name = 'monthly_p4_test_inputs'"
                ).fetchone()[0]
            with test_repo.connect() as conn:
                test_count = conn.execute("SELECT COUNT(*) FROM monthly_p4_test_inputs").fetchone()[0]
            self.assertEqual(formal_count, 1)
            self.assertEqual(formal_has_test_table, 0)
            self.assertEqual(test_count, 1)

    def test_monthly_p4_test_template_uploads_are_separate(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            repo = self._setup_project(Path(td))
            test_db = Path(td) / "data" / "monthly_p4_test.sqlite"
            test_db.touch()
            test_repo = SQLiteRepository(test_db, project_root=Path(td))
            svc = CanonicalService(repo, monthly_test_repo=test_repo)
            wb = Workbook()
            try:
                wb.active.title = "績效追蹤 p4 5 (j)"
                ws = wb.active
                ws["A1"] = "產品處 2026績效 (不含電商)"
                ws["F3"] = "Jan"
                ws["G3"] = "Feb"
                ws["H3"] = "Mar"
                ws["A2"] = "產品處 廣告總營收"
                ws["B2"] = "目標"
                ws["F4"] = 8847250
                ws["G4"] = 6927250
                ws["H4"] = 6546750
                ws["B3"] = "實績"
                ws["F5"] = 123
                ws["G5"] = 456
                ws["H5"] = 789
                ws["B4"] = "月 達成率"
                ws["F6"] = "1%"
                ws["G6"] = "2%"
                ws["H6"] = "3%"
                wb.create_sheet("檢核")
                buffer = BytesIO()
                wb.save(buffer)
            finally:
                wb.close()
            payload = base64.b64encode(buffer.getvalue()).decode("ascii")

            base_result = svc.save_monthly_p4_test_template(
                template_kind="base",
                filename="月報_基礎模板.xlsx",
                content_base64=payload,
                template_version="v1",
                rule_version="v1",
            )
            check_result = svc.save_monthly_p4_test_template(
                template_kind="check",
                filename="月報_檢核模板.xlsx",
                content_base64=payload,
                template_version="v1",
                rule_version="v1",
            )
            snapshot = svc.build_monthly_p4_snapshot(
                week_start="2026-03-01",
                week_end="2026-03-31",
                manual_source="test",
            )

            self.assertEqual(base_result["template_kind"], "base")
            self.assertEqual(check_result["template_kind"], "check")
            self.assertEqual(snapshot["testTemplates"]["base"]["filename"], "月報_基礎模板.xlsx")
            self.assertEqual(snapshot["testTemplates"]["check"]["filename"], "月報_檢核模板.xlsx")
            self.assertEqual(snapshot["testTemplates"]["base"]["sheetNames"], ["績效追蹤 p4 5 (j)", "檢核"])
            self.assertGreater(snapshot["testTemplates"]["check"]["snapshot"]["entryCount"], 0)
            self.assertEqual(snapshot["diff"]["status"], "mismatch")
            self.assertGreater(snapshot["diff"]["diffCount"], 0)
            self.assertTrue(
                any(item["reason"] == "missing_in_check_template" for item in snapshot["diff"]["diffs"])
            )

    def test_monthly_p4_template_parser_prefers_p4_tracking_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            repo = self._setup_project(Path(td))
            svc = CanonicalService(repo)
            wb = Workbook()
            try:
                target_ws = wb.active
                target_ws.title = "2026各營收目標明細"
                target_ws["F5"] = "Jan"
                target_ws["G5"] = "Feb"
                target_ws["A6"] = "產品處 廣告總營收"
                target_ws["E6"] = "目標"
                target_ws["F6"] = 1
                target_ws["G6"] = 2

                p4_ws = wb.create_sheet("績效追蹤 p4 5 (j)")
                p4_ws["F3"] = "Jan"
                p4_ws["G3"] = "Feb"
                p4_ws["B4"] = "產品處 廣告總營收"
                p4_ws["D4"] = "目標"
                p4_ws["F4"] = 100
                p4_ws["G4"] = 200
                p4_ws["D5"] = "實績"
                p4_ws["F5"] = 11
                p4_ws["G5"] = 22
                p4_ws["D6"] = "月 達成率"
                p4_ws["F6"] = "11%"
                p4_ws["G6"] = "22%"
                p4_ws["D26"] = "mltiFORCE 總目標"
                p4_ws["F26"] = 8628000
                p4_ws["G26"] = 6708000
                p4_ws["D37"] = "其他 營收總目標"
                p4_ws["F37"] = 219250
                p4_ws["G37"] = 219250

                buffer = BytesIO()
                wb.save(buffer)
            finally:
                wb.close()

            snapshot = svc._parse_monthly_p4_workbook_snapshot(
                buffer.getvalue(),
                filename="月報.xlsx",
            )

            self.assertEqual(snapshot["sheet"], "績效追蹤 p4 5 (j)")
            self.assertGreater(snapshot["entryCount"], 0)
            self.assertEqual(snapshot["entries"]["product_total.actual.2026-02"]["value"], 22.0)
            self.assertEqual(snapshot["entries"]["mf_total.target.2026-02"]["value"], 6708000.0)
            self.assertEqual(snapshot["entries"]["other_total.target.2026-02"]["value"], 219250.0)

    def test_monthly_p4_closed_workbook_feeds_media_cost_investment(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            repo = self._setup_project(root)
            svc = CanonicalService(repo)
            workbook_path = root / "closed.xlsx"
            wb = Workbook()
            try:
                ws = wb.active
                ws.title = "績效追蹤 p4 5 (j)"
                ws["F3"] = "Jan"
                ws["G3"] = "Feb"
                ws["H3"] = "Mar"
                ws["I3"] = "Apr"
                ws["F27"] = 5869138
                ws["G27"] = 4376941
                ws["H27"] = 4102932.4999999986
                ws["I27"] = 3696097
                wb.save(workbook_path)
            finally:
                wb.close()

            with repo.connect_monthly_report() as conn:
                repo.save_monthly_report_rows(
                    conn,
                    run_id="run-2026-04",
                    report_kind="ssp_regular_monthly_zone_campaign_size",
                    start_day="2026-04-01",
                    end_day="2026-04-30",
                    report_id=1,
                    records_total=1,
                    source="test",
                    pb=0,
                    request_payload={},
                    response_payload={},
                    sum_row={},
                    rows=[
                        {
                            "month": "2026-04",
                            "date": "2026-04",
                            "zone_id": 1,
                            "zone_name": "測試版位",
                            "profit": 1943968,
                            "advertiser_mu": 999,
                        }
                    ],
                )

            out = svc.import_monthly_p4_closed_workbook(
                workbook_path=workbook_path,
                through_month="2026-04",
                template_version="v1",
                rule_version="v1",
            )
            self.assertEqual(out["metric_count"], 4)
            self.assertEqual(out["metrics"][2]["value"], 4102933.0)

            snapshot = svc.build_monthly_media_cost_analysis(month="2026-04")
            self.assertEqual(snapshot["metrics"]["totalInvestment"], 3696097.0)
            self.assertEqual(snapshot["metrics"]["p4MfActual"], 3696097.0)
            self.assertAlmostEqual(float(snapshot["metrics"]["mediaCostRate"]), 52.595166, places=4)

    def test_monthly_charts_use_runtime_p4_before_close_and_closed_after_close(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            repo = self._setup_project(root)
            svc = CanonicalService(repo)
            svc.save(
                workflow="dsp",
                rows=[self._row(日期時間="2026-04-01", 執行金額=100.0)],
                template_version="v1",
                rule_version="v1",
            )
            svc.save_monthly_p4_manual_inputs(
                month="2026-04",
                inputs={
                    "hb_revenue": 0,
                    "external_beiliu_io": 0,
                    "remaining_traffic_revenue": 0,
                    "data_monetization_adjustment": 0,
                },
                template_version="v1",
                rule_version="v1",
            )
            with repo.connect_monthly_report() as conn:
                repo.save_monthly_report_rows(
                    conn,
                    run_id="run-2026-04",
                    report_kind="ssp_regular_monthly_zone_campaign_size",
                    start_day="2026-04-01",
                    end_day="2026-04-30",
                    report_id=1,
                    records_total=1,
                    source="test",
                    pb=0,
                    request_payload={},
                    response_payload={},
                    sum_row={},
                    rows=[
                        {
                            "month": "2026-04",
                            "date": "2026-04",
                            "zone_id": 1,
                            "zone_name": "測試版位",
                            "profit": 10,
                            "advertiser_mu": 999,
                        }
                    ],
                )

            runtime = svc.build_monthly_p4_snapshot(week_start="2026-04-01", week_end="2026-04-30")
            april_runtime = next(item for item in runtime["monthPayloads"] if item["month"] == "2026-04")
            self.assertEqual(float(april_runtime["actuals"]["mf_total"]), 100.0)

            charts = svc.build_monthly_charts_snapshot(months=["2026-04"])
            april = charts["monthly"][0]
            self.assertEqual(april["p4MfActual"], 100.0)
            self.assertEqual(april["mediaCostInvestment"], 100.0)
            self.assertFalse(april["p4Closed"])
            self.assertEqual(april["p4InvestmentSource"], "monthly_p4_runtime")

            media_cost = svc.build_monthly_media_cost_analysis(month="2026-04")
            self.assertEqual(media_cost["metrics"]["totalInvestment"], 100.0)
            self.assertEqual(media_cost["metrics"]["p4MfActual"], 100.0)
            self.assertFalse(media_cost["p4Closed"])
            self.assertEqual(media_cost["p4InvestmentSource"], "monthly_p4_runtime")

            svc.close_monthly_p4_month(month="2026-04", template_version="v1", rule_version="v1")
            closed_charts = svc.build_monthly_charts_snapshot(months=["2026-04"])
            closed_april = closed_charts["monthly"][0]
            self.assertEqual(closed_april["p4MfActual"], 100.0)
            self.assertEqual(closed_april["mediaCostInvestment"], 100.0)
            self.assertTrue(closed_april["p4Closed"])
            self.assertEqual(closed_april["p4InvestmentSource"], "monthly_p4_closed_metrics")

    def test_monthly_p4_close_rejects_missing_manual_inputs(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            repo = self._setup_project(Path(td))
            svc = CanonicalService(repo)
            svc.save(
                workflow="dsp",
                rows=[self._row(日期時間="2026-04-01", 執行金額=100.0)],
                template_version="v1",
                rule_version="v1",
            )
            svc.save_monthly_p4_manual_inputs(
                month="2026-04",
                inputs={
                    "hb_revenue": 0,
                    "remaining_traffic_revenue": 0,
                    "data_monetization_adjustment": 0,
                },
                template_version="v1",
                rule_version="v1",
            )

            with self.assertRaisesRegex(ValueError, "external_beiliu_io"):
                svc.close_monthly_p4_month(month="2026-04", template_version="v1", rule_version="v1")

            svc.save_monthly_p4_manual_inputs(
                month="2026-04",
                inputs={
                    "hb_revenue": 0,
                    "external_beiliu_io": 0,
                    "remaining_traffic_revenue": 0,
                    "data_monetization_adjustment": 0,
                },
                template_version="v1",
                rule_version="v1",
            )
            closed = svc.close_monthly_p4_month(month="2026-04", template_version="v1", rule_version="v1")
            self.assertEqual(closed["status"], "ok")
            self.assertEqual(closed["mf_total_actual"], 100.0)

    def test_monthly_p4_diff_uses_union_and_statuses_are_semantic(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            repo = self._setup_project(Path(td))
            svc = CanonicalService(repo)
            candidate = {
                "entries": {
                    "product_total.target.2026-03": {
                        "itemKey": "product_total",
                        "metric": "target",
                        "month": "2026-03",
                        "value": 100.0,
                    },
                    "mf_total.actual.2026-03": {
                        "itemKey": "mf_total",
                        "metric": "actual",
                        "month": "2026-03",
                        "value": 50.0,
                    },
                }
            }
            answer = {
                "entries": {
                    "product_total.target.2026-03": {
                        "itemKey": "product_total",
                        "metric": "target",
                        "month": "2026-03",
                        "value": 100.0,
                    },
                    "data_fee.actual.2026-03": {
                        "itemKey": "data_fee",
                        "metric": "actual",
                        "month": "2026-03",
                        "value": 7.0,
                        "cell": "P4J!F20",
                    },
                }
            }

            diff = svc._monthly_p4_diff(candidate, answer)
            matched = svc._monthly_p4_diff(candidate, candidate)

            self.assertEqual(diff["status"], "mismatch")
            self.assertEqual(diff["diffCount"], 2)
            self.assertEqual(
                {item["reason"] for item in diff["diffs"]},
                {"missing_in_check_template", "missing_in_candidate"},
            )
            self.assertEqual(matched["status"], "matched")
            self.assertEqual(matched["diffCount"], 0)


if __name__ == "__main__":
    unittest.main()
