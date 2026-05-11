from __future__ import annotations

import argparse
import json
import mimetypes
import sqlite3
import sys
from dataclasses import dataclass
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, quote, urlparse

from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from domain.services import (
    CanonicalService,
    DETAIL_INPUT_ROWS,
    DETAIL_YEAR_ROWS,
    MONTH_AMOUNT_COL_START,
    MONTH_COUNT,
    _pick_category,
    _resolve_year_month,
    _to_number,
)
from infra.sqlite.bootstrap import (
    AcceptanceGateError,
    bootstrap_health,
    bootstrap_init,
    build_config,
    ensure_acceptance_gate,
    get_feature_flags,
)
from infra.sqlite.repository import SQLiteRepository

FRONTEND_DIST_DIR = ROOT_DIR / "frontend" / "dist"


@dataclass(frozen=True)
class UiContext:
    root: Path
    manifest_rel: str
    workflow: str
    template_version: str
    rule_version: str
    artifact_root: Path


def _read_recent_rows(conn: sqlite3.Connection, sql: str, params: tuple[Any, ...] = ()) -> list[dict[str, Any]]:
    cur = conn.execute(sql, params)
    cols = [str(c[0]) for c in cur.description]
    out: list[dict[str, Any]] = []
    for row in cur.fetchall():
        out.append({cols[idx]: row[idx] for idx in range(len(cols))})
    return out


def _build_service(root: Path, manifest_rel: str) -> CanonicalService:
    ensure_acceptance_gate(root, manifest_rel)
    cfg = build_config(root, manifest_rel)
    feature_flags = get_feature_flags(root, manifest_rel)
    repo = SQLiteRepository(cfg.db_path, project_root=root)
    return CanonicalService(repo, feature_flags=feature_flags)


def _float_or_zero(value: object) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value or "").replace(",", "").replace("$", "").strip())
    except Exception:
        return 0.0


def _rate_or_none(value: object) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    return None


def _build_dsp_tab4_preview_summary(rows: list[dict[str, Any]]) -> dict[str, Any]:
    month_totals = [0.0 for _ in range(MONTH_COUNT)]
    row_amounts: dict[str, list[float]] = {
        "r3": [0.0 for _ in range(MONTH_COUNT)],
        "r4": [0.0 for _ in range(MONTH_COUNT)],
        "r5": [0.0 for _ in range(MONTH_COUNT)],
        "r6": [0.0 for _ in range(MONTH_COUNT)],
        "r7": [0.0 for _ in range(MONTH_COUNT)],
        "r8": [0.0 for _ in range(MONTH_COUNT)],
        "r9": [0.0 for _ in range(MONTH_COUNT)],
        "r10": [0.0 for _ in range(MONTH_COUNT)],
        "r11": [0.0 for _ in range(MONTH_COUNT)],
        "r12": [0.0 for _ in range(MONTH_COUNT)],
        "r13": [0.0 for _ in range(MONTH_COUNT)],
        "r14": [0.0 for _ in range(MONTH_COUNT)],
        "r15": [0.0 for _ in range(MONTH_COUNT)],
    }

    def contains_any(text: str, tokens: tuple[str, ...]) -> bool:
        return any(token in text for token in tokens)

    for row in rows:
        resolved = _resolve_year_month(row)
        if resolved is None:
            continue
        year, month_idx = resolved
        amount = _to_number(row.get("執行金額"))
        month_totals[month_idx] += amount

        b = _pick_category(row, ["分類層級B", "最終經銷商", "經銷商"])
        c = _pick_category(row, ["分類層級C", "最終廣告形式", "廣告形式"])
        d = _pick_category(row, ["分類層級D", "素材樣板", "素材", "訂單"])
        distributor = _pick_category(row, ["最終經銷商", "經銷商", "原始經銷商"])
        ad_format = _pick_category(row, ["最終廣告形式", "廣告形式", "素材樣板"])
        haystack = f"{b} {c} {d} {distributor} {ad_format}"
        haystack_lower = haystack.lower()

        if contains_any(distributor, ("策略",)) or contains_any(ad_format, ("策略",)):
            row_amounts["r4"][month_idx] += amount
        elif contains_any(b, ("外經銷商", "外部經銷商", "外部")) and (
            contains_any(c, ("外部經銷推廣", "經銷推廣")) or contains_any(distributor, ("玩藝", "春樹", "ADGeek", "ADgeek"))
        ):
            row_amounts["r5"][month_idx] += amount
        elif contains_any(c, ("IO委刊",)) or contains_any(distributor, ("IO委刊", "momo")):
            row_amounts["r6"][month_idx] += amount
        elif contains_any(b, ("HB串接",)) or contains_any(haystack, ("HB", "串接")):
            row_amounts["r7"][month_idx] += amount

        if contains_any(ad_format, ("創意",)) or contains_any(d, ("蓋板", "置底", "文中")):
            row_amounts["r10"][month_idx] += amount
        elif contains_any(ad_format, ("影音摩天", "outstream")):
            row_amounts["r11"][month_idx] += amount
        elif contains_any(ad_format.lower(), ("preroll", "pre roll", "instream")):
            row_amounts["r12"][month_idx] += amount
        elif contains_any(haystack, ("DOOH外部", "presco", "前線媒體")):
            row_amounts["r13"][month_idx] += amount
        elif contains_any(haystack, ("DOOH北流", "北流")):
            row_amounts["r14"][month_idx] += amount
        elif contains_any(haystack_upper := haystack.upper(), ("CTV",)):
            row_amounts["r15"][month_idx] += amount

    annual_total = sum(month_totals)

    def build_row(excel_row: int, row_id: str, note_only: bool = False) -> dict[str, Any]:
        monthly_amounts = month_totals if row_id in {"r3", "r9"} else row_amounts[row_id]
        annual_amount = sum(monthly_amounts)
        monthly_rates = [value / total if total > 0 else 0.0 for value, total in zip(monthly_amounts, month_totals)]
        annual_rate = annual_amount / annual_total if annual_total > 0 else 0.0
        return {
            "excelRow": excel_row,
            "monthlyAmounts": monthly_amounts,
            "monthlyRates": [None for _ in range(MONTH_COUNT)] if note_only else monthly_rates,
            "annualAmount": annual_amount,
            "annualRate": None if note_only else annual_rate,
        }

    rows_out = [
        build_row(3, "r3"),
        build_row(4, "r4"),
        build_row(5, "r5"),
        build_row(6, "r6"),
        build_row(7, "r7"),
        build_row(8, "r8", note_only=True),
        build_row(9, "r9"),
        build_row(10, "r10"),
        build_row(11, "r11"),
        build_row(12, "r12"),
        build_row(13, "r13"),
        build_row(14, "r14"),
        build_row(15, "r15"),
    ]

    return {
        "source": "canonical rows",
        "year": max((resolved[0] for row in rows if (resolved := _resolve_year_month(row)) is not None), default=None),
        "monthTotals": month_totals,
        "monthTotalRates": [1.0 if total > 0 else 0.0 for total in month_totals],
        "annualTotal": annual_total,
        "annualRate": 1.0 if annual_total > 0 else 0.0,
        "rows": rows_out,
    }


def _build_dsp_tab4_preview_detail(service: CanonicalService, rows: list[dict[str, Any]]) -> dict[str, Any] | None:
    try:
        template_path = service._resolve_dsp_export_template_path()
        wb = load_workbook(template_path, read_only=True, data_only=True)
    except Exception:
        return None

    try:
        if "各經銷商明細" not in wb.sheetnames:
            return None
        ws = wb["各經銷商明細"]
        month_amount_cols = [MONTH_AMOUNT_COL_START + (idx * 2) for idx in range(MONTH_COUNT)]
        month_labels = [str(ws.cell(row=82, column=col).value or f"{idx + 1}月") for idx, col in enumerate(month_amount_cols)]

        preview_year, detail_monthly_amounts = service._build_detail_matrix_values(
            rows=rows,
            fallback_year=max((resolved[0] for row in rows if (resolved := _resolve_year_month(row)) is not None), default=0) or 0,
        )

        def sum_rows(row_indices: list[int]) -> list[float]:
            return [
                sum(detail_monthly_amounts.get(row_idx, [0.0 for _ in range(MONTH_COUNT)])[month_idx] for row_idx in row_indices)
                for month_idx in range(MONTH_COUNT)
            ]

        def rate_rows(monthly_amounts: list[float], denominator: list[float]) -> list[float]:
            out: list[float] = []
            for amount, total in zip(monthly_amounts, denominator):
                out.append(amount / total if total > 0 else 0.0)
            return out

        section_specs = [
            {"id": "marketing", "year_row": 5, "total_row": 6},
            {"id": "strategy", "year_row": 24, "total_row": 25},
            {"id": "external_self", "year_row": 44, "total_row": 45},
            {"id": "external_io", "year_row": 63, "total_row": 64},
            {"id": "hb_bridge", "year_row": 82, "total_row": 83},
        ]

        sections: list[dict[str, Any]] = []
        for spec in section_specs:
            year_row = int(spec["year_row"])
            total_row = int(spec["total_row"])
            detail_row_indices = list(range(total_row + 1, total_row + 8))
            total_monthly_amounts = sum_rows(detail_row_indices)
            total_annual_amount = sum(total_monthly_amounts)
            total_monthly_rates = [0.0 for _ in range(MONTH_COUNT)]
            if total_annual_amount > 0:
                for month_idx in range(MONTH_COUNT):
                    total_monthly_rates[month_idx] = sum(
                        detail_monthly_amounts.get(row_idx, [0.0 for _ in range(MONTH_COUNT)])[month_idx]
                        / total_monthly_amounts[month_idx]
                        if total_monthly_amounts[month_idx] > 0
                        else 0.0
                        for row_idx in detail_row_indices
                    )
            total_annual_rate = 1.0 if total_annual_amount > 0 else 0.0
            detail_rows: list[dict[str, Any]] = []
            for row_idx in detail_row_indices:
                row_monthly_amounts = detail_monthly_amounts.get(row_idx, [0.0 for _ in range(MONTH_COUNT)])
                row_annual_amount = sum(row_monthly_amounts)
                row_monthly_rates = rate_rows(row_monthly_amounts, total_monthly_amounts)
                row_annual_rate = row_annual_amount / total_annual_amount if total_annual_amount > 0 else 0.0
                detail_rows.append(
                    {
                        "excelRow": row_idx,
                        "labelA": str(ws.cell(row=row_idx, column=1).value or ""),
                        "labelB": str(ws.cell(row=row_idx, column=2).value or ""),
                        "labelC": str(ws.cell(row=row_idx, column=3).value or ""),
                        "labelD": str(ws.cell(row=row_idx, column=4).value or ""),
                        "monthlyAmounts": row_monthly_amounts,
                        "monthlyRates": row_monthly_rates,
                        "annualAmount": row_annual_amount,
                        "annualRate": row_annual_rate,
                    }
                )

            sections.append(
                {
                    "id": str(spec["id"]),
                    "year": int(_float_or_zero(ws.cell(row=year_row, column=1).value) or 0) or preview_year or None,
                    "monthLabels": [str(ws.cell(row=year_row, column=col).value or label) for col, label in zip(month_amount_cols, month_labels)],
                    "total": {
                        "excelRow": total_row,
                        "labelA": str(ws.cell(row=total_row, column=1).value or ""),
                        "labelB": str(ws.cell(row=total_row, column=2).value or ""),
                        "labelC": str(ws.cell(row=total_row, column=3).value or ""),
                        "labelD": str(ws.cell(row=total_row, column=4).value or ""),
                        "monthlyAmounts": total_monthly_amounts,
                        "monthlyRates": total_monthly_rates,
                        "annualAmount": total_annual_amount,
                        "annualRate": total_annual_rate,
                    },
                    "rows": detail_rows,
                }
            )

        row6_amounts = sections[0]["total"]["monthlyAmounts"] if sections else [0.0 for _ in range(MONTH_COUNT)]
        row12_amounts = sections[0]["rows"][5]["monthlyAmounts"] if sections and len(sections[0]["rows"]) > 5 else [0.0 for _ in range(MONTH_COUNT)]
        row3_targets = [_float_or_zero(ws.cell(row=3, column=col).value) for col in month_amount_cols]
        row4_targets = [_float_or_zero(ws.cell(row=4, column=col).value) for col in month_amount_cols]
        row3_rates = [actual / target if target > 0 else 0.0 for actual, target in zip(row6_amounts, row3_targets)]
        row4_rates = [actual / target if target > 0 else 0.0 for actual, target in zip(row12_amounts, row4_targets)]
        row2_amounts = [sum(section["total"]["monthlyAmounts"][idx] for section in sections) for idx in range(MONTH_COUNT)]
        row2_rates = [1.0 if amount > 0 else 0.0 for amount in row2_amounts]
        row2_annual_amount = sum(row2_amounts)

        kpi_rows = [
            {
                "excelRow": 2,
                "label": str(ws.cell(row=2, column=1).value or ""),
                "monthlyAmounts": row2_amounts,
                "monthlyRates": row2_rates,
                "annualAmount": row2_annual_amount,
                "annualRate": 1.0 if row2_annual_amount > 0 else 0.0,
            },
            {
                "excelRow": 3,
                "label": str(ws.cell(row=3, column=1).value or ""),
                "monthlyAmounts": row3_targets,
                "monthlyRates": row3_rates,
                "annualAmount": sum(row3_targets),
                "annualRate": sum(row3_rates) / MONTH_COUNT if MONTH_COUNT > 0 else 0.0,
            },
            {
                "excelRow": 4,
                "label": str(ws.cell(row=4, column=1).value or ""),
                "monthlyAmounts": row4_targets,
                "monthlyRates": row4_rates,
                "annualAmount": sum(row4_targets),
                "annualRate": sum(row4_rates) / MONTH_COUNT if MONTH_COUNT > 0 else 0.0,
            },
        ]

        return {
            "source": str(template_path),
            "monthLabels": month_labels,
            "kpiRows": kpi_rows,
            "sections": sections,
        }
    finally:
        wb.close()


def _collect_dsp_tab4_template_summary(service: CanonicalService) -> dict[str, Any] | None:
    try:
        template_path = service._resolve_dsp_export_template_path()
        wb = load_workbook(template_path, read_only=True, data_only=True)
    except Exception:
        return None

    try:
        if "mF投資量_總表" not in wb.sheetnames:
            return None
        ws = wb["mF投資量_總表"]
        month_amount_cols = [5 + (idx * 2) for idx in range(12)]
        year_value = int(_float_or_zero(ws["A1"].value) or 0) or None
        rows: list[dict[str, Any]] = []
        for row_idx in range(3, 16):
            monthly_amounts = [_float_or_zero(ws.cell(row=row_idx, column=col).value) for col in month_amount_cols]
            monthly_rates = [_rate_or_none(ws.cell(row=row_idx, column=col + 1).value) for col in month_amount_cols]
            rows.append(
                {
                    "excelRow": row_idx,
                    "monthlyAmounts": monthly_amounts,
                    "monthlyRates": monthly_rates,
                    "annualAmount": _float_or_zero(ws.cell(row=row_idx, column=29).value),
                    "annualRate": _rate_or_none(ws.cell(row=row_idx, column=30).value),
                }
            )
        return {
            "source": str(template_path),
            "year": year_value,
            "monthTotals": [_float_or_zero(ws.cell(row=2, column=col).value) for col in month_amount_cols],
            "monthTotalRates": [_rate_or_none(ws.cell(row=2, column=col + 1).value) for col in month_amount_cols],
            "annualTotal": _float_or_zero(ws["AC2"].value),
            "annualRate": _rate_or_none(ws["AD2"].value),
            "rows": rows,
        }
    finally:
        wb.close()


def _collect_dsp_tab4_template_detail(service: CanonicalService) -> dict[str, Any] | None:
    try:
        template_path = service._resolve_dsp_export_template_path()
        wb = load_workbook(template_path, read_only=True, data_only=True)
    except Exception:
        return None

    try:
        if "各經銷商明細" not in wb.sheetnames:
            return None
        ws = wb["各經銷商明細"]
        month_amount_cols = [5 + (idx * 2) for idx in range(12)]

        def _collect_monthly(row_idx: int) -> tuple[list[float], list[float | None], float, float | None]:
            return (
                [_float_or_zero(ws.cell(row=row_idx, column=col).value) for col in month_amount_cols],
                [_rate_or_none(ws.cell(row=row_idx, column=col + 1).value) for col in month_amount_cols],
                _float_or_zero(ws.cell(row=row_idx, column=29).value),
                _rate_or_none(ws.cell(row=row_idx, column=30).value),
            )

        kpi_rows: list[dict[str, Any]] = []
        for row_idx in (2, 3, 4):
            monthly_amounts, monthly_rates, annual_amount, annual_rate = _collect_monthly(row_idx)
            kpi_rows.append(
                {
                    "excelRow": row_idx,
                    "label": str(ws.cell(row=row_idx, column=1).value or ""),
                    "monthlyAmounts": monthly_amounts,
                    "monthlyRates": monthly_rates,
                    "annualAmount": annual_amount,
                    "annualRate": annual_rate,
                }
            )

        section_specs = [
            {"id": "marketing", "year_row": 5, "total_row": 6},
            {"id": "strategy", "year_row": 24, "total_row": 25},
            {"id": "external_self", "year_row": 44, "total_row": 45},
            {"id": "external_io", "year_row": 63, "total_row": 64},
            {"id": "hb_bridge", "year_row": 82, "total_row": 83},
        ]
        sections: list[dict[str, Any]] = []
        for spec in section_specs:
            year_row = int(spec["year_row"])
            total_row = int(spec["total_row"])
            monthly_amounts, monthly_rates, annual_amount, annual_rate = _collect_monthly(total_row)
            detail_rows: list[dict[str, Any]] = []
            for row_idx in range(total_row + 1, total_row + 8):
                row_monthly_amounts, row_monthly_rates, row_annual_amount, row_annual_rate = _collect_monthly(row_idx)
                detail_rows.append(
                    {
                        "excelRow": row_idx,
                        "labelA": str(ws.cell(row=row_idx, column=1).value or ""),
                        "labelB": str(ws.cell(row=row_idx, column=2).value or ""),
                        "labelC": str(ws.cell(row=row_idx, column=3).value or ""),
                        "labelD": str(ws.cell(row=row_idx, column=4).value or ""),
                        "monthlyAmounts": row_monthly_amounts,
                        "monthlyRates": row_monthly_rates,
                        "annualAmount": row_annual_amount,
                        "annualRate": row_annual_rate,
                    }
                )
            sections.append(
                {
                    "id": str(spec["id"]),
                    "year": int(_float_or_zero(ws.cell(row=year_row, column=1).value) or 0) or None,
                    "monthLabels": [str(ws.cell(row=year_row, column=col).value or f"{idx + 1}月") for idx, col in enumerate(month_amount_cols)],
                    "total": {
                        "excelRow": total_row,
                        "labelA": str(ws.cell(row=total_row, column=1).value or ""),
                        "labelB": str(ws.cell(row=total_row, column=2).value or ""),
                        "labelC": str(ws.cell(row=total_row, column=3).value or ""),
                        "labelD": str(ws.cell(row=total_row, column=4).value or ""),
                        "monthlyAmounts": monthly_amounts,
                        "monthlyRates": monthly_rates,
                        "annualAmount": annual_amount,
                        "annualRate": annual_rate,
                    },
                    "rows": detail_rows,
                }
            )

        return {
            "source": str(template_path),
            "monthLabels": [str(ws.cell(row=82, column=col).value or f"{idx + 1}月") for idx, col in enumerate(month_amount_cols)],
            "kpiRows": kpi_rows,
            "sections": sections,
        }
    finally:
        wb.close()


def collect_runtime_status(ctx: UiContext) -> dict[str, Any]:
    health = bootstrap_health(ctx.root, ctx.manifest_rel)
    summary: dict[str, Any] = {
        "root": str(ctx.root),
        "manifest": ctx.manifest_rel,
        "canonical_source": "sqlite",
        "workflow": ctx.workflow,
        "template_version": ctx.template_version,
        "rule_version": ctx.rule_version,
        "artifact_root": str(ctx.artifact_root),
        "health": health,
        "recent": {
            "run_log": [],
            "audit_log": [],
            "publish_runs": [],
            "evidence_index": [],
        },
        "tab4_delivery": {
            "ready": False,
            "reason": "pending_delivery",
            "updated_at": "",
            "last_delivery_run_id": "",
            "last_change_run_id": "",
            "delivery_snapshot_token": "",
            "delivery_row_count": 0,
            "delivery_source_db_hash": "",
            "delivery_template_version": "",
            "delivery_rule_version": "",
        },
    }

    checks = health.get("checks") if isinstance(health, dict) else None
    db_path_text = ""
    if isinstance(checks, dict):
        db_path_text = str(checks.get("db_path") or "")
    if not db_path_text:
        return summary

    db_path = Path(db_path_text)
    if not db_path.exists():
        return summary

    repo = SQLiteRepository(db_path, project_root=ctx.root)
    conn = sqlite3.connect(str(db_path))
    try:
        summary["recent"]["run_log"] = _read_recent_rows(
            conn,
            """
            SELECT run_id, run_type, workflow, status, created_at
            FROM run_log
            ORDER BY created_at DESC
            LIMIT 8
            """,
        )
        summary["recent"]["audit_log"] = _read_recent_rows(
            conn,
            """
            SELECT event_type, scope, status, created_at
            FROM audit_log
            ORDER BY id DESC
            LIMIT 8
            """,
        )
        summary["recent"]["publish_runs"] = _read_recent_rows(
            conn,
            """
            SELECT run_id, template_id, template_version, output_path, status, created_at
            FROM publish_runs
            ORDER BY id DESC
            LIMIT 8
            """,
        )
        summary["recent"]["evidence_index"] = _read_recent_rows(
            conn,
            """
            SELECT run_id, scope, path, status, created_at
            FROM evidence_index
            ORDER BY id DESC
            LIMIT 8
            """,
        )
        if ctx.workflow == "dsp":
            # 讓前端可從 status 直接判斷 Tab4 是否已完成交付。
            summary["tab4_delivery"] = repo.get_tab4_delivery_state(conn, ctx.workflow)
    finally:
        conn.close()
    return summary


def collect_workflow_frame(ctx: UiContext) -> dict[str, Any]:
    health = bootstrap_health(ctx.root, ctx.manifest_rel)
    summary: dict[str, Any] = {
        "root": str(ctx.root),
        "manifest": ctx.manifest_rel,
        "canonical_source": "sqlite",
        "workflow": ctx.workflow,
        "template_version": ctx.template_version,
        "rule_version": ctx.rule_version,
        "artifact_root": str(ctx.artifact_root),
        "health": health,
        "columns": [],
        "rows": [],
        "row_count": 0,
        "pivot_preview": [],
        "tab4_preview_contract": {
            "kind": "template_preview",
            "note": "template layout rendered from canonical rows",
        },
    }

    checks = health.get("checks") if isinstance(health, dict) else None
    db_path_text = ""
    if isinstance(checks, dict):
        db_path_text = str(checks.get("db_path") or "")
    if not db_path_text:
        return summary

    db_path = Path(db_path_text)
    if not db_path.exists():
        return summary

    try:
        repo = SQLiteRepository(db_path, project_root=ctx.root)
        if ctx.workflow == "ssp":
            rows = repo.read_ssp_raw_rows()
            columns = ["row_order", *repo.workflow_columns(ctx.workflow), "updated_at"]
        else:
            rows = repo.read_canonical_rows(ctx.workflow)
            columns = ["row_order", *repo.canonical_columns, "updated_at"]
        summary["columns"] = columns
        summary["field_names"] = list(repo.workflow_columns(ctx.workflow)) if ctx.workflow == "ssp" else list(repo.canonical_columns)
        summary["manual_fields"] = [] if ctx.workflow == "ssp" else list(repo.modify_allowed_columns)
        summary["rows"] = rows
        summary["row_count"] = len(rows)
        summary["pivot_preview"] = [
            {"label": "row_count", "value": len(rows)},
            {"label": "workflow", "value": ctx.workflow},
            {"label": "template_version", "value": ctx.template_version},
            {"label": "rule_version", "value": ctx.rule_version},
        ]
        if ctx.workflow == "dsp":
            with repo.connect() as conn:
                # Tab4 解鎖狀態要跟最新 run_log 同步，供前端鎖定/解鎖顯示。
                tab4_delivery_state = repo.get_tab4_delivery_state(conn, ctx.workflow)
                summary["tab4_delivery"] = tab4_delivery_state
                summary["tab4_delivery_snapshot"] = {
                    "delivery_snapshot_token": str(tab4_delivery_state.get("delivery_snapshot_token") or ""),
                    "delivery_run_id": str(tab4_delivery_state.get("last_delivery_run_id") or ""),
                    "delivery_row_count": int(tab4_delivery_state.get("delivery_row_count") or 0),
                    "delivery_ready": bool(tab4_delivery_state.get("ready")),
                    "delivery_reason": str(tab4_delivery_state.get("reason") or ""),
                }
            service = CanonicalService(repo, feature_flags=get_feature_flags(ctx.root, ctx.manifest_rel))
            canonical_rows = repo.read_canonical_rows(ctx.workflow)
            template_summary = _build_dsp_tab4_preview_summary(canonical_rows)
            template_detail = _build_dsp_tab4_preview_detail(service, canonical_rows)
            if template_detail is None:
                template_detail = _collect_dsp_tab4_template_detail(service)
            if template_summary is None:
                template_summary = _collect_dsp_tab4_template_summary(service)
            summary["tab4_preview_template_summary"] = template_summary
            summary["tab4_preview_template_detail"] = template_detail
    except Exception as exc:
        summary["frame_error"] = str(exc)
    return summary


def dispatch_action(ctx: UiContext, payload: dict[str, Any]) -> dict[str, Any]:
    action = str(payload.get("action") or "").strip()
    if not action:
        raise ValueError("action required")

    workflow = str(payload.get("workflow") or ctx.workflow)
    template_version = str(payload.get("template_version") or ctx.template_version)
    rule_version = str(payload.get("rule_version") or ctx.rule_version)
    main_tab = str(payload.get("main_tab") or "")
    sub_tab = str(payload.get("sub_tab") or "")

    if action == "bootstrap":
        return bootstrap_init(ctx.root, ctx.manifest_rel)
    if action == "health":
        return bootstrap_health(ctx.root, ctx.manifest_rel)

    service = _build_service(ctx.root, ctx.manifest_rel)
    if action == "save":
        rows = payload.get("rows")
        if not isinstance(rows, list):
            raise ValueError("rows must be list")
        return service.save(
            workflow=workflow,
            rows=rows,
            template_version=template_version,
            rule_version=rule_version,
        )

    if action == "modify":
        updates = payload.get("updates")
        if not isinstance(updates, list):
            raise ValueError("updates must be list")
        return service.modify(
            workflow=workflow,
            updates=updates,
            template_version=template_version,
            rule_version=rule_version,
        )

    if action == "export":
        artifact_root = payload.get("artifact_root")
        resolved_artifact_root = ctx.artifact_root
        if isinstance(artifact_root, str) and artifact_root.strip():
            resolved_artifact_root = (ctx.root / artifact_root).resolve()
        request_week_start = payload.get("period_week_start")
        request_week_end = payload.get("period_week_end")
        if not isinstance(request_week_start, str):
            request_week_start = payload.get("week_start")
        if not isinstance(request_week_end, str):
            request_week_end = payload.get("week_end")
        if workflow == "dsp":
            if not main_tab:
                main_tab = "dsp_tab4"
            if not sub_tab:
                sub_tab = "overview"
            if main_tab != "dsp_tab4":
                raise PermissionError("dsp export must be triggered from dsp_tab4")
            if sub_tab not in {"overview"}:
                raise PermissionError("dsp export sub_tab out of scope")
        return service.export(
            workflow=workflow,
            artifact_root=resolved_artifact_root,
            template_version=template_version,
            rule_version=rule_version,
            main_tab=main_tab or None,
            sub_tab=sub_tab or None,
            week_start=request_week_start.strip() if isinstance(request_week_start, str) and request_week_start.strip() else None,
            week_end=request_week_end.strip() if isinstance(request_week_end, str) and request_week_end.strip() else None,
        )

    if action == "tab4_delivery":
        return service.mark_tab4_delivery(
            workflow=workflow,
            main_tab=main_tab,
            sub_tab=sub_tab,
            template_version=template_version,
            rule_version=rule_version,
        )

    raise ValueError(f"unsupported action: {action}")


def _resolve_frontend_asset(path: str) -> Path | None:
    # 只允許 frontend/dist 內檔案，避免 path traversal。
    relative = "index.html" if path == "/" else path.lstrip("/")
    candidate = (FRONTEND_DIST_DIR / relative).resolve()
    if not str(candidate).startswith(str(FRONTEND_DIST_DIR.resolve())):
        return None
    if not candidate.exists() or not candidate.is_file():
        return None
    return candidate


def _frontend_unavailable_page() -> str:
    return """<!doctype html>
<html lang="zh-Hant">
  <head>
    <meta charset="utf-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1" />
    <title>MDREP Frontend Unavailable</title>
    <style>
      body { margin: 24px; font-family: "Avenir Next", "Segoe UI", sans-serif; color: #1f2937; }
      code { background: #eef2ff; border-radius: 6px; padding: 2px 6px; }
    </style>
  </head>
  <body>
    <h1>React frontend artifact not found</h1>
    <p>請先在 <code>/Users/matt/MDREPROT2/frontend</code> 執行 <code>pnpm build</code>，再由本 UI shell 提供靜態前端入口。</p>
    <p>backend runtime API 仍可用：<code>/api/status</code>、<code>/api/frame</code>、<code>/api/action</code></p>
  </body>
</html>
"""


class UiRequestHandler(BaseHTTPRequestHandler):
    server_version = "MDREPUIShell/0.1"

    def _json(self, status: int, payload: dict[str, Any]) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _html_response(self, html: str, *, status: int = HTTPStatus.OK) -> None:
        body = html.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _file_response(self, path: Path, *, as_attachment: bool = False) -> None:
        body = path.read_bytes()
        content_type, _ = mimetypes.guess_type(str(path))
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type or "application/octet-stream")
        if as_attachment:
            safe_name = path.name.replace("\"", "")
            ascii_name = safe_name.encode("ascii", "ignore").decode("ascii") or "export.xlsx"
            self.send_header(
                "Content-Disposition",
                f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{quote(safe_name)}",
            )
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _resolve_download_artifact(self, *, ctx: UiContext, artifact_path_raw: str) -> Path:
        candidate = Path(artifact_path_raw)
        resolved = (ctx.root / candidate).resolve() if not candidate.is_absolute() else candidate.resolve()
        artifact_root_resolved = ctx.artifact_root.resolve()
        try:
            resolved.relative_to(artifact_root_resolved)
        except ValueError as exc:
            raise PermissionError("artifact_path out of artifact_root scope") from exc
        if not resolved.exists() or not resolved.is_file():
            raise FileNotFoundError(f"artifact not found: {resolved}")
        if resolved.suffix.lower() != ".xlsx":
            raise ValueError("download only supports .xlsx artifact")
        return resolved

    def do_GET(self) -> None:  # noqa: N802
        parsed = urlparse(self.path)
        path = parsed.path
        if path.startswith("/api/export/download"):
            params = parse_qs(parsed.query)
            try:
                root = Path(str(params.get("root", ["."])[0])).resolve()
                artifact_root_rel = str(params.get("artifact_root", ["artifacts"])[0])
                ctx = UiContext(
                    root=root,
                    manifest_rel=str(params.get("manifest", ["bootstrap.manifest.json"])[0]),
                    workflow=str(params.get("workflow", ["dsp"])[0]),
                    template_version=str(params.get("template_version", ["v1"])[0]),
                    rule_version=str(params.get("rule_version", ["v1"])[0]),
                    artifact_root=(root / artifact_root_rel).resolve(),
                )
                artifact_path_raw = str(params.get("artifact_path", [""])[0]).strip()
                if not artifact_path_raw:
                    raise ValueError("artifact_path required")
                artifact_path = self._resolve_download_artifact(ctx=ctx, artifact_path_raw=artifact_path_raw)
                self._file_response(artifact_path, as_attachment=True)
            except Exception as exc:
                self._json(
                    HTTPStatus.BAD_REQUEST,
                    {
                        "status": "error",
                        "error_code": "DOWNLOAD_FAILED",
                        "message": str(exc),
                    },
                )
            return
        if path.startswith("/api/status"):
            params = parse_qs(parsed.query)
            ctx = UiContext(
                root=Path(str(params.get("root", ["."])[0])).resolve(),
                manifest_rel=str(params.get("manifest", ["bootstrap.manifest.json"])[0]),
                workflow=str(params.get("workflow", ["dsp"])[0]),
                template_version=str(params.get("template_version", ["v1"])[0]),
                rule_version=str(params.get("rule_version", ["v1"])[0]),
                artifact_root=(Path(str(params.get("root", ["."])[0])).resolve() / str(params.get("artifact_root", ["artifacts"])[0])).resolve(),
            )
            payload = collect_runtime_status(ctx)
            self._json(HTTPStatus.OK, {"status": "ok", "result": payload})
            return
        if path.startswith("/api/frame"):
            params = parse_qs(parsed.query)
            ctx = UiContext(
                root=Path(str(params.get("root", ["."])[0])).resolve(),
                manifest_rel=str(params.get("manifest", ["bootstrap.manifest.json"])[0]),
                workflow=str(params.get("workflow", ["dsp"])[0]),
                template_version=str(params.get("template_version", ["v1"])[0]),
                rule_version=str(params.get("rule_version", ["v1"])[0]),
                artifact_root=(Path(str(params.get("root", ["."])[0])).resolve() / str(params.get("artifact_root", ["artifacts"])[0])).resolve(),
            )
            payload = collect_workflow_frame(ctx)
            self._json(HTTPStatus.OK, {"status": "ok", "result": payload})
            return
        if path.startswith("/api/"):
            self._json(HTTPStatus.NOT_FOUND, {"status": "error", "error": "NOT_FOUND"})
            return

        asset = _resolve_frontend_asset(path)
        if asset is not None:
            self._file_response(asset)
            return
        if path == "/":
            self._html_response(_frontend_unavailable_page(), status=HTTPStatus.SERVICE_UNAVAILABLE)
            return
        self._json(HTTPStatus.NOT_FOUND, {"status": "error", "error": "NOT_FOUND"})

    def do_POST(self) -> None:  # noqa: N802
        if self.path != "/api/action":
            self._json(HTTPStatus.NOT_FOUND, {"status": "error", "error": "NOT_FOUND"})
            return

        length = int(self.headers.get("Content-Length", "0"))
        raw = self.rfile.read(length)
        try:
            payload = json.loads(raw.decode("utf-8"))
            if not isinstance(payload, dict):
                raise ValueError("payload must be object")
            root = Path(str(payload.get("root") or ".")).resolve()
            manifest_rel = str(payload.get("manifest") or "bootstrap.manifest.json")
            artifact_root = (root / str(payload.get("artifact_root") or "artifacts")).resolve()
            ctx = UiContext(
                root=root,
                manifest_rel=manifest_rel,
                workflow=str(payload.get("workflow") or "dsp"),
                template_version=str(payload.get("template_version") or "v1"),
                rule_version=str(payload.get("rule_version") or "v1"),
                artifact_root=artifact_root,
            )
            result = dispatch_action(ctx, payload)
            self._json(HTTPStatus.OK, {"status": "ok", "result": result})
        except AcceptanceGateError as exc:
            self._json(
                HTTPStatus.BAD_REQUEST,
                {
                    "status": "error",
                    "error_code": "STRICT_ACCEPTANCE_GATE_FAILED",
                    "message": str(exc),
                    "details": {"reason_code": exc.reason_code, "checks": exc.checks},
                },
            )
        except Exception as exc:
            self._json(
                HTTPStatus.BAD_REQUEST,
                {
                    "status": "error",
                    "error_code": "UI_ACTION_FAILED",
                    "message": str(exc),
                },
            )


def _parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="MDREP runtime UI shell")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", default=8510, type=int)
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = _parse_args(argv)
    server = ThreadingHTTPServer((args.host, args.port), UiRequestHandler)
    print(f"MDREP UI shell running at http://{args.host}:{args.port}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
