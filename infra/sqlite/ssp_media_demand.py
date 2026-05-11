from __future__ import annotations

import json
import os
from pathlib import Path

import openpyxl

SSP_MEDIA_DEMAND_CATEGORIES = [
    "蓋板",
    "置底",
    "置底展開",
    "文中300x250",
    "文中320x480",
]
MASTER_PLACEMENT_SHEET = "版位編號"

DEFAULT_TEMPLATE_CANDIDATES = [
    "templates/ssp_template.xlsx",
    "templates/template.xlsx",
]


def _safe_int(value: object) -> int:
    if isinstance(value, bool):
        return 0
    if isinstance(value, (int, float)):
        try:
            return int(float(value))
        except Exception:
            return 0
    raw = _safe_text(value)
    if not raw:
        return 0
    try:
        return int(float(raw))
    except Exception:
        return 0


def _safe_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _safe_float(value: object) -> float:
    if isinstance(value, (int, float)):
        out = float(value)
        if out == out and out not in (float("inf"), float("-inf")):
            return out
        return 0.0
    raw = _safe_text(value).replace(",", "")
    if not raw:
        return 0.0
    try:
        out = float(raw)
    except Exception:
        return 0.0
    if out == out and out not in (float("inf"), float("-inf")):
        return out
    return 0.0


def _safe_int_string(value: object) -> str:
    raw = _safe_text(value)
    if not raw:
        return ""
    try:
        return str(int(float(raw)))
    except Exception:
        return ""


def _safe_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return float(value) != 0.0
    raw = _safe_text(value).lower()
    return raw in {"1", "true", "yes", "y", "on"}


def _safe_bool_text(value: object) -> str:
    raw = _safe_text(value)
    if not raw:
        return ""
    lowered = raw.lower()
    if lowered in {"true", "1", "yes", "y", "on"}:
        return "true"
    if lowered in {"false", "0", "no", "n", "off"}:
        return "false"
    return raw


def resolve_ssp_template_path(project_root: Path) -> Path | None:
    explicit = _safe_text(os.getenv("MDREP_SSP_TEMPLATE_PATH"))
    candidates: list[Path] = []
    if explicit:
        candidates.append(Path(explicit).expanduser())
    for rel_path in DEFAULT_TEMPLATE_CANDIDATES:
        candidates.append((project_root / rel_path).resolve())
    for path in candidates:
        if path.exists() and path.is_file():
            return path.resolve()
    return None


def _compute_compliance_rate(actual_request: float, media_target: float) -> float:
    if actual_request <= 0 or media_target <= 0:
        return 0.0
    return (actual_request / media_target) * 100


def _extract_int_candidates(ws, *, max_rows: int = 5000, max_cols: int = 20) -> set[int]:
    out: set[int] = set()
    for row_idx in range(1, min(ws.max_row, max_rows) + 1):
        for col_idx in range(1, min(ws.max_column, max_cols) + 1):
            raw = ws.cell(row_idx, col_idx).value
            value = _safe_int(raw)
            if value > 0:
                out.add(value)
    return out


def _load_template_groups(template_path: Path) -> list[dict[str, object]]:
    wb = openpyxl.load_workbook(template_path, data_only=False)
    try:
        master_ids: set[int] = set()
        if MASTER_PLACEMENT_SHEET in wb.sheetnames:
            master_ids = _extract_int_candidates(wb[MASTER_PLACEMENT_SHEET], max_rows=5000, max_cols=10)

        groups: list[dict[str, object]] = []
        for category in SSP_MEDIA_DEMAND_CATEGORIES:
            if category not in wb.sheetnames:
                continue
            ws = wb[category]
            header_row = None
            for row_idx in range(1, min(ws.max_row, 60) + 1):
                col_b = _safe_text(ws.cell(row_idx, 2).value)
                col_c = _safe_text(ws.cell(row_idx, 3).value)
                if col_b == "版位" and col_c == "版位名稱":
                    header_row = row_idx
                    break
            placement_ids: set[int] = set()
            default_rows: list[dict[str, object]] = []
            if header_row is not None:
                for row_idx in range(header_row + 1, min(ws.max_row, 5000) + 1):
                    placement_id = _safe_int(ws.cell(row_idx, 2).value)
                    if placement_id <= 0:
                        continue
                    if master_ids and placement_id not in master_ids:
                        continue
                    placement_ids.add(placement_id)
                    default_rows.append(
                        {
                            "placement_id": str(placement_id),
                            "placement_name": _safe_text(ws.cell(row_idx, 3).value),
                            "media_quality": _safe_text(ws.cell(row_idx, 4).value),
                            "need_call": _safe_bool(ws.cell(row_idx, 5).value),
                            "target_fr": _safe_text(ws.cell(row_idx, 6).value),
                            "estimated_request_0722": _safe_float(ws.cell(row_idx, 7).value),
                            "remark": _safe_text(ws.cell(row_idx, 1).value),
                        }
                    )
            if placement_ids:
                groups.append(
                    {
                        "name": category,
                        "placement_ids": placement_ids,
                        "default_rows": default_rows,
                    }
                )
        return groups
    finally:
        wb.close()


def _load_template_slots(template_path: Path) -> dict[str, list[dict[str, object]]]:
    slots_by_category: dict[str, list[dict[str, object]]] = {
        category: [] for category in SSP_MEDIA_DEMAND_CATEGORIES
    }
    for group in _load_template_groups(template_path):
        category = _safe_text(group.get("name"))
        default_rows = group.get("default_rows")
        if category not in slots_by_category or not isinstance(default_rows, list):
            continue
        for slot_order, row in enumerate(default_rows):
            if not isinstance(row, dict):
                continue
            placement_id = _safe_int_string(row.get("placement_id"))
            if not placement_id:
                continue
            slots_by_category[category].append(
                {
                    "category": category,
                    "slot_order": slot_order,
                    "placement_id": placement_id,
                    "placement_name": _safe_text(row.get("placement_name")),
                    "media_quality": _safe_text(row.get("media_quality")),
                    "need_call": _safe_bool(row.get("need_call")),
                    "target_fr": _safe_text(row.get("target_fr")),
                    "remark": _safe_text(row.get("remark")),
                    "media_target": _safe_float(row.get("estimated_request_0722", row.get("media_target"))),
                    "is_active": True,
                }
            )
    return slots_by_category


def _load_json_slots(group_overrides_path: Path) -> dict[str, list[dict[str, object]]]:
    if not group_overrides_path.exists():
        return {category: [] for category in SSP_MEDIA_DEMAND_CATEGORIES}
    payload = json.loads(group_overrides_path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        return {category: [] for category in SSP_MEDIA_DEMAND_CATEGORIES}
    slots_by_category: dict[str, list[dict[str, object]]] = {
        category: [] for category in SSP_MEDIA_DEMAND_CATEGORIES
    }
    for category in SSP_MEDIA_DEMAND_CATEGORIES:
        rows = payload.get(category)
        if not isinstance(rows, list):
            continue
        slot_order = 0
        for row in rows:
            if not isinstance(row, dict):
                continue
            placement_id = _safe_int_string(row.get("placement_id"))
            if not placement_id:
                continue
            slots_by_category[category].append(
                {
                    "category": category,
                    "slot_order": slot_order,
                    "placement_id": placement_id,
                    "placement_name": _safe_text(row.get("placement_name")),
                    "media_quality": _safe_text(row.get("media_quality")),
                    "need_call": _safe_bool(row.get("need_call")),
                    "target_fr": _safe_text(row.get("target_fr")),
                    "remark": _safe_text(row.get("remark")),
                    "media_target": _safe_float(row.get("estimated_request_0722", row.get("media_target"))),
                    "is_active": True,
                }
            )
            slot_order += 1
    return slots_by_category


def normalize_ssp_media_slot(slot: dict[str, object], *, fallback_category: str, fallback_order: int) -> dict[str, object]:
    category = _safe_text(slot.get("category"))
    if category not in SSP_MEDIA_DEMAND_CATEGORIES:
        category = fallback_category
    return {
        "category": category,
        "slot_order": int(slot.get("slot_order") or fallback_order),
        "placement_id": _safe_text(slot.get("placement_id")),
        "placement_name": _safe_text(slot.get("placement_name")),
        "media_quality": _safe_text(slot.get("media_quality")),
        "need_call": _safe_bool(slot.get("need_call")),
        "need_call_text": _safe_bool_text(slot.get("need_call")),
        "target_fr": _safe_text(slot.get("target_fr")),
        "remark": _safe_text(slot.get("remark")),
        "media_target": _safe_float(slot.get("estimated_request_0722", slot.get("media_target"))),
        "is_active": _safe_bool(slot.get("is_active", True)),
    }


def resolve_default_ssp_media_slots(project_root: Path, data_seed_root: Path) -> dict[str, object]:
    template_path = resolve_ssp_template_path(project_root)
    template_groups = _load_template_groups(template_path) if template_path is not None else []
    template_slots = (
        _load_template_slots(template_path)
        if template_path is not None
        else {category: [] for category in SSP_MEDIA_DEMAND_CATEGORIES}
    )
    group_overrides_path = (data_seed_root / "templates_rules_mapping" / "group_overrides.json").resolve()
    json_slots = _load_json_slots(group_overrides_path)

    categories = [str(group.get("name") or "").strip() for group in template_groups if str(group.get("name") or "").strip()]
    if not categories:
        categories = list(SSP_MEDIA_DEMAND_CATEGORIES)

    merged_slots: list[dict[str, object]] = []
    for category in categories:
        template_rows = template_slots.get(category) or []
        json_rows = json_slots.get(category) or []
        json_by_pid = {
            str(row.get("placement_id") or ""): row
            for row in json_rows
            if _safe_text(row.get("placement_id"))
        }
        used_pids: set[str] = set()

        if template_rows:
            for idx, template_row in enumerate(template_rows):
                placement_id = _safe_text(template_row.get("placement_id"))
                backup = json_by_pid.get(placement_id, {})
                merged_slots.append(
                    normalize_ssp_media_slot(
                        {
                            "category": category,
                            "slot_order": idx,
                            "placement_id": placement_id,
                            "placement_name": _safe_text(template_row.get("placement_name")) or _safe_text(backup.get("placement_name")),
                            "media_quality": _safe_text(template_row.get("media_quality")) or _safe_text(backup.get("media_quality")),
                            "need_call": backup.get("need_call", template_row.get("need_call", False)),
                            "target_fr": _safe_text(template_row.get("target_fr")) or _safe_text(backup.get("target_fr")),
                            "remark": _safe_text(template_row.get("remark")) or _safe_text(backup.get("remark")),
                            "estimated_request_0722": backup.get("estimated_request_0722", backup.get("media_target", template_row.get("estimated_request_0722", template_row.get("media_target", 0.0)))),
                            "is_active": True,
                        },
                        fallback_category=category,
                        fallback_order=idx,
                    )
                )
                if placement_id:
                    used_pids.add(placement_id)
            next_order = len(template_rows)
            for row in json_rows:
                placement_id = _safe_text(row.get("placement_id"))
                if not placement_id or placement_id in used_pids:
                    continue
                merged_slots.append(
                    normalize_ssp_media_slot(
                        {
                            **row,
                            "category": category,
                            "slot_order": next_order,
                        },
                        fallback_category=category,
                        fallback_order=next_order,
                    )
                )
                next_order += 1
        else:
            for idx, row in enumerate(json_rows):
                merged_slots.append(
                    normalize_ssp_media_slot(
                        {
                            **row,
                            "category": category,
                            "slot_order": idx,
                        },
                        fallback_category=category,
                        fallback_order=idx,
                    )
                )

    source = "template+json" if template_path is not None else "json"
    return {
        "categories": categories,
        "slots": merged_slots,
        "defaults_source": source,
        "template_path": str(template_path) if template_path is not None else "",
        "group_overrides_path": str(group_overrides_path) if group_overrides_path.exists() else "",
    }


def build_ssp_media_demand_view(
    *,
    categories: list[str],
    slots: list[dict[str, object]],
    matrix_rows: list[dict[str, object]],
    source_options: list[str],
    active_source: str,
    active_category: str,
    scope_mode: str,
    day_limit: int,
    threshold: float,
    only_unmet: bool,
) -> dict[str, object]:
    normalized_categories = [item for item in categories if _safe_text(item)]
    selected_category = active_category if active_category in normalized_categories else (normalized_categories[0] if normalized_categories else "")
    normalized_scope = "07-22" if scope_mode == "07-22" else "all"
    normalized_day_limit = max(1, int(day_limit or 7))
    normalized_threshold = _safe_float(threshold)

    date_keys = sorted(
        {_safe_text(row.get("date")) for row in matrix_rows if _safe_text(row.get("date"))},
        reverse=True,
    )[:normalized_day_limit]
    latest_date = date_keys[0] if date_keys else ""
    aggregate_map: dict[tuple[str, str], dict[str, float]] = {}
    for row in matrix_rows:
        date_key = _safe_text(row.get("date"))
        placement_id = _safe_int_string(row.get("placement_id"))
        if not date_key or not placement_id or date_key not in date_keys:
            continue
        aggregate_map[(placement_id, date_key)] = {
            "request_all": _safe_float(row.get("request_all")),
            "impression_all": _safe_float(row.get("impression_all")),
            "clicks_all": _safe_float(row.get("clicks_all")),
            "revenue_all": _safe_float(row.get("revenue_all")),
            "dsp_amount_all": _safe_float(row.get("dsp_amount_all")),
            "request_0722": _safe_float(row.get("request_0722")),
            "impression_0722": _safe_float(row.get("impression_0722")),
            "clicks_0722": _safe_float(row.get("clicks_0722")),
            "revenue_0722": _safe_float(row.get("revenue_0722")),
            "dsp_amount_0722": _safe_float(row.get("dsp_amount_0722")),
        }

    category_slots = sorted(
        [slot for slot in slots if _safe_text(slot.get("category")) == selected_category],
        key=lambda item: int(item.get("slot_order") or 0),
    )
    demand_rows: list[dict[str, object]] = []
    for slot in category_slots:
        placement_id = _safe_int_string(slot.get("placement_id"))
        media_target = _safe_float(slot.get("media_target"))
        metrics_by_date: dict[str, dict[str, dict[str, float]]] = {}
        latest_request = 0.0
        latest_compliance_rate = 0.0
        has_latest_date_data = False

        for date_key in date_keys:
            aggregate = aggregate_map.get((placement_id, date_key), {})
            request_all = _safe_float(aggregate.get("request_all"))
            impression_all = _safe_float(aggregate.get("impression_all"))
            clicks_all = _safe_float(aggregate.get("clicks_all"))
            revenue_all = _safe_float(aggregate.get("revenue_all"))
            dsp_amount_all = _safe_float(aggregate.get("dsp_amount_all"))
            request_0722 = _safe_float(aggregate.get("request_0722"))
            impression_0722 = _safe_float(aggregate.get("impression_0722"))
            clicks_0722 = _safe_float(aggregate.get("clicks_0722"))
            revenue_0722 = _safe_float(aggregate.get("revenue_0722"))
            dsp_amount_0722 = _safe_float(aggregate.get("dsp_amount_0722"))
            compliance_all = _compute_compliance_rate(request_all, media_target)
            compliance_0722 = _compute_compliance_rate(request_0722, media_target)
            fr_all = (impression_all / request_all) * 100 if request_all > 0 else 0.0
            fr_0722 = (impression_0722 / request_0722) * 100 if request_0722 > 0 else 0.0
            ctr_all = (clicks_all / request_all) * 100 if request_all > 0 else 0.0
            ctr_0722 = (clicks_0722 / request_0722) * 100 if request_0722 > 0 else 0.0
            ecpm_all = (revenue_all / impression_all) * 1000 if impression_all > 0 else 0.0
            ecpm_0722 = (revenue_0722 / impression_0722) * 1000 if impression_0722 > 0 else 0.0
            metrics_by_date[date_key] = {
                "all": {
                    "complianceRate": compliance_all,
                    "request": request_all,
                    "impression": impression_all,
                    "clicks": clicks_all,
                    "revenue": revenue_all,
                    "dspAmount": dsp_amount_all,
                    "fr": fr_all,
                    "ctr": ctr_all,
                    "ecpm": ecpm_all,
                },
                "07-22": {
                    "complianceRate": compliance_0722,
                    "request": request_0722,
                    "impression": impression_0722,
                    "clicks": clicks_0722,
                    "revenue": revenue_0722,
                    "dspAmount": dsp_amount_0722,
                    "fr": fr_0722,
                    "ctr": ctr_0722,
                    "ecpm": ecpm_0722,
                },
            }
            if date_key == latest_date:
                latest_scope_metrics = metrics_by_date[date_key][normalized_scope]
                latest_request = _safe_float(latest_scope_metrics.get("request"))
                latest_compliance_rate = _safe_float(latest_scope_metrics.get("complianceRate"))
                has_latest_date_data = (
                    _safe_float(latest_scope_metrics.get("request")) > 0
                    or _safe_float(latest_scope_metrics.get("impression")) > 0
                    or _safe_float(latest_scope_metrics.get("clicks")) > 0
                    or _safe_float(latest_scope_metrics.get("revenue")) > 0
                )

        demand_rows.append(
            {
                "slot": slot,
                "latest_request": latest_request,
                "latest_compliance_rate": latest_compliance_rate,
                "has_latest_date_data": has_latest_date_data,
                "metrics_by_date": metrics_by_date,
            }
        )

    demand_rows.sort(
        key=lambda item: (
            -int(bool(item["has_latest_date_data"])),
            -float(item["latest_request"]),
            int(((item.get("slot") or {}) or {}).get("slot_order") or 0),
        )
    )
    visible_rows = [
        row for row in demand_rows
        if not only_unmet
        or (float(row["latest_compliance_rate"]) > 0 and float(row["latest_compliance_rate"]) < normalized_threshold)
    ]
    unmet_count = sum(
        1 for row in demand_rows
        if float(row["latest_compliance_rate"]) > 0 and float(row["latest_compliance_rate"]) < normalized_threshold
    )
    latest_total_request = 0.0
    for row in demand_rows:
        metrics = ((row.get("metrics_by_date") or {}).get(latest_date) or {}).get(normalized_scope) or {}
        latest_total_request += _safe_float(metrics.get("request"))

    return {
        "category": selected_category,
        "source": active_source,
        "scope_mode": normalized_scope,
        "day_limit": normalized_day_limit,
        "threshold": normalized_threshold,
        "only_unmet": bool(only_unmet),
        "date_keys": date_keys,
        "latest_date": latest_date,
        "latest_total_request": latest_total_request,
        "unmet_count": unmet_count,
        "source_options": [item for item in source_options if _safe_text(item)],
        "rows": visible_rows,
    }
